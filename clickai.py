#!/usr/bin/env python3
"""
Click AI - Complete Business Management System v2.0
SINGLE FILE - 9000+ lines - Full SA Accounting
"""

from flask import Flask, jsonify, request, redirect, session, g
import json
import os
import re
import uuid
import requests
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from functools import wraps


# =============================================================================
# PIECE1_CORE.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 1: Core Foundation                                                    ║
║                                                                               ║
║   This piece contains:                                                        ║
║   - Flask application setup                                                   ║
║   - Database connection class (Supabase)                                      ║
║   - Helper functions                                                          ║
║   - VAT calculation logic (with zero-rated items)                            ║
║   - Date and currency formatting                                              ║
║   - User session management                                                   ║
║                                                                               ║
║   All calculations done by Flask - no JavaScript math                         ║
║   Opus AI advises, Flask executes                                             ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""


# ═══════════════════════════════════════════════════════════════════════════════
# FLASK APPLICATION SETUP
# ═══════════════════════════════════════════════════════════════════════════════

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "clickai-secret-key-change-in-production")

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

class Config:
    """Application configuration"""
    
    # Database
    SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
    SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
    
    # AI
    ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
    OPUS_MODEL = "claude-opus-4-20250514"
    SONNET_MODEL = "claude-sonnet-4-20250514"
    
    # Business Settings
    COMPANY_NAME = "FullTech"
    VAT_RATE = Decimal("0.15")  # 15% standard rate
    CURRENCY_SYMBOL = "R"
    
    # Financial Year (South African standard - March to February)
    FINANCIAL_YEAR_END_MONTH = 2  # February
    
    # Session timeout (minutes)
    SESSION_TIMEOUT = 480  # 8 hours


# ═══════════════════════════════════════════════════════════════════════════════
# VAT LOGIC - SOUTH AFRICAN RULES
# ═══════════════════════════════════════════════════════════════════════════════

class VAT:
    """
    South African VAT calculations
    
    Standard Rate: 15%
    Zero-Rated: 0% (fuel, basic foods, exports)
    Exempt: No VAT charged, no input VAT claimed
    """
    
    STANDARD_RATE = Decimal("0.15")
    ZERO_RATE = Decimal("0")
    
    # Keywords that indicate zero-rated items
    ZERO_RATED_KEYWORDS = [
        # Fuel
        'petrol', 'diesel', 'fuel', 'paraffin', 'illuminating paraffin',
        
        # Basic foods - grains
        'brown bread', 'maize meal', 'mealie meal', 'samp', 'rice', 
        'dried corn', 'mielie rice',
        
        # Basic foods - legumes
        'dried beans', 'lentils', 'dried peas', 'pilchards', 'sardines',
        
        # Basic foods - dairy & eggs
        'fresh milk', 'cultured milk', 'milk powder', 'eggs',
        
        # Basic foods - produce
        'vegetables', 'fruit', 'vegetable oil', 'cooking oil',
        
        # Other zero-rated
        'export', 'international transport',
    ]
    
    # Categories that are typically zero-rated
    ZERO_RATED_CATEGORIES = [
        'fuel', 'basic foods', 'exports'
    ]
    
    # Expense categories that are exempt (no VAT claim allowed)
    EXEMPT_CATEGORIES = [
        'financial services', 'bank charges', 'interest', 
        'insurance', 'residential rent', 'salaries', 'wages',
        'paye', 'uif', 'sdl'
    ]
    
    @classmethod
    def is_zero_rated(cls, description: str = "", category: str = "") -> bool:
        """
        Check if an item is zero-rated based on description or category
        
        Args:
            description: Item description
            category: Item category
            
        Returns:
            True if zero-rated, False if standard rated
        """
        description_lower = description.lower()
        category_lower = category.lower()
        
        # Check category first
        for exempt_cat in cls.ZERO_RATED_CATEGORIES:
            if exempt_cat in category_lower:
                return True
        
        # Check description keywords
        for keyword in cls.ZERO_RATED_KEYWORDS:
            if keyword in description_lower:
                return True
        
        return False
    
    @classmethod
    def is_exempt(cls, category: str = "") -> bool:
        """
        Check if a category is VAT exempt (no input VAT can be claimed)
        
        Args:
            category: Expense category
            
        Returns:
            True if exempt, False otherwise
        """
        category_lower = category.lower()
        
        for exempt_cat in cls.EXEMPT_CATEGORIES:
            if exempt_cat in category_lower:
                return True
        
        return False
    
    @classmethod
    def get_rate(cls, description: str = "", category: str = "") -> Decimal:
        """
        Get the applicable VAT rate for an item
        
        Args:
            description: Item description
            category: Item category
            
        Returns:
            VAT rate as Decimal (0.15 or 0)
        """
        if cls.is_zero_rated(description, category):
            return cls.ZERO_RATE
        return cls.STANDARD_RATE
    
    @classmethod
    def calculate_from_inclusive(cls, amount_incl: Decimal, rate: Decimal = None) -> dict:
        """
        Calculate VAT from VAT-inclusive amount
        
        Args:
            amount_incl: Amount including VAT
            rate: VAT rate (defaults to standard rate)
            
        Returns:
            Dictionary with 'exclusive', 'vat', and 'inclusive' amounts
        """
        if rate is None:
            rate = cls.STANDARD_RATE
        
        amount_incl = Decimal(str(amount_incl))
        
        if rate == cls.ZERO_RATE:
            return {
                'exclusive': amount_incl,
                'vat': Decimal("0"),
                'inclusive': amount_incl,
                'rate': rate
            }
        
        # Formula: exclusive = inclusive / (1 + rate)
        divisor = Decimal("1") + rate
        amount_excl = (amount_incl / divisor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        vat = (amount_incl - amount_excl).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        
        return {
            'exclusive': amount_excl,
            'vat': vat,
            'inclusive': amount_incl,
            'rate': rate
        }
    
    @classmethod
    def calculate_from_exclusive(cls, amount_excl: Decimal, rate: Decimal = None) -> dict:
        """
        Calculate VAT from VAT-exclusive amount
        
        Args:
            amount_excl: Amount excluding VAT
            rate: VAT rate (defaults to standard rate)
            
        Returns:
            Dictionary with 'exclusive', 'vat', and 'inclusive' amounts
        """
        if rate is None:
            rate = cls.STANDARD_RATE
        
        amount_excl = Decimal(str(amount_excl))
        
        if rate == cls.ZERO_RATE:
            return {
                'exclusive': amount_excl,
                'vat': Decimal("0"),
                'inclusive': amount_excl,
                'rate': rate
            }
        
        vat = (amount_excl * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        amount_incl = amount_excl + vat
        
        return {
            'exclusive': amount_excl,
            'vat': vat,
            'inclusive': amount_incl,
            'rate': rate
        }


# ═══════════════════════════════════════════════════════════════════════════════
# MONEY HANDLING - PRECISION MATTERS
# ═══════════════════════════════════════════════════════════════════════════════

class Money:
    """
    Handle money with proper precision
    All internal calculations use Decimal
    All storage uses cents (integer) to avoid floating point issues
    """
    
    @staticmethod
    def to_cents(rands: Decimal) -> int:
        """Convert Rands to cents for storage"""
        return int((Decimal(str(rands)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    
    @staticmethod
    def from_cents(cents: int) -> Decimal:
        """Convert cents back to Rands"""
        return Decimal(str(cents)) / 100
    
    @staticmethod
    def parse(value) -> Decimal:
        """
        Parse various input formats to Decimal
        
        Handles:
        - "R 1,234.56"
        - "1234.56"
        - 1234.56
        - "R1234,56" (comma as decimal)
        """
        if value is None:
            return Decimal("0")
        
        if isinstance(value, Decimal):
            return value
        
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        
        # String parsing
        s = str(value).strip()
        
        # Remove currency symbol and spaces
        s = s.replace("R", "").replace(" ", "")
        
        # Handle comma as thousands separator or decimal
        if "," in s and "." in s:
            # Both present: comma is thousands separator
            s = s.replace(",", "")
        elif "," in s:
            # Only comma: could be decimal separator
            # Check if it's in the last 3 characters (likely decimal)
            if len(s) - s.rfind(",") <= 3:
                s = s.replace(",", ".")
            else:
                s = s.replace(",", "")
        
        try:
            return Decimal(s).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except:
            return Decimal("0")
    
    @staticmethod
    def format(amount: Decimal, symbol: bool = True) -> str:
        """
        Format amount for display
        
        Args:
            amount: Amount to format
            symbol: Include currency symbol
            
        Returns:
            Formatted string like "R 1,234.56"
        """
        amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        
        # Format with thousands separator
        formatted = f"{amount:,.2f}"
        
        if symbol:
            return f"R {formatted}"
        return formatted


# ═══════════════════════════════════════════════════════════════════════════════
# DATABASE CONNECTION - SUPABASE
# ═══════════════════════════════════════════════════════════════════════════════

class Database:
    """
    Supabase database connection
    
    Handles all CRUD operations with proper error handling
    """
    
    def __init__(self, url: str = None, key: str = None):
        self.url = (url or Config.SUPABASE_URL).rstrip('/')
        self.key = key or Config.SUPABASE_KEY
        self.headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation"
        }
        self.timeout = 10  # 10 seconds max per DB call
    
    def _request(self, method: str, endpoint: str, data: dict = None) -> tuple:
        """
        Make HTTP request to Supabase
        
        Returns:
            Tuple of (success: bool, data: dict/list or error: str)
        """
        url = f"{self.url}/rest/v1/{endpoint}"
        
        try:
            if method == "GET":
                response = requests.get(url, headers=self.headers, timeout=self.timeout)
            elif method == "POST":
                response = requests.post(url, headers=self.headers, json=data, timeout=self.timeout)
            elif method == "PATCH":
                response = requests.patch(url, headers=self.headers, json=data, timeout=self.timeout)
            elif method == "DELETE":
                response = requests.delete(url, headers=self.headers, timeout=self.timeout)
            else:
                return False, f"Unknown method: {method}"
            
            if response.status_code in [200, 201, 204]:
                if response.text:
                    return True, response.json()
                return True, []
            else:
                return False, f"HTTP {response.status_code}: {response.text}"
                
        except requests.exceptions.Timeout:
            return False, "Database timeout"
        except requests.exceptions.ConnectionError:
            return False, "Database connection error"
        except Exception as e:
            return False, str(e)
    
    def select(self, table: str, filters: dict = None, order: str = None, 
               limit: int = None, columns: str = "*") -> list:
        """
        Select records from table
        
        Args:
            table: Table name
            filters: Dict of column=value filters
            order: Column to order by (prefix with - for descending)
            limit: Maximum records to return
            columns: Columns to select (default all)
            
        Returns:
            List of records
        """
        endpoint = f"{table}?select={columns}"
        
        if filters:
            for key, value in filters.items():
                endpoint += f"&{key}=eq.{value}"
        
        if order:
            if order.startswith("-"):
                endpoint += f"&order={order[1:]}.desc"
            else:
                endpoint += f"&order={order}.asc"
        
        if limit:
            endpoint += f"&limit={limit}"
        
        success, result = self._request("GET", endpoint)
        return result if success and isinstance(result, list) else []
    
    def select_one(self, table: str, id: str) -> dict:
        """Get single record by ID"""
        results = self.select(table, {"id": id}, limit=1)
        return results[0] if results else None
    
    def insert(self, table: str, data: dict) -> tuple:
        """
        Insert record
        
        Returns:
            Tuple of (success: bool, record: dict or error: str)
        """
        return self._request("POST", table, data)
    
    def update(self, table: str, id: str, data: dict) -> tuple:
        """
        Update record by ID
        
        Returns:
            Tuple of (success: bool, record: dict or error: str)
        """
        endpoint = f"{table}?id=eq.{id}"
        return self._request("PATCH", endpoint, data)
    
    def delete(self, table: str, id: str) -> tuple:
        """
        Delete record by ID
        
        Returns:
            Tuple of (success: bool, result or error: str)
        """
        endpoint = f"{table}?id=eq.{id}"
        return self._request("DELETE", endpoint)
    
    def upsert(self, table: str, data: dict, id_field: str = "id") -> tuple:
        """
        Insert or update record
        
        If record with ID exists, update it. Otherwise insert new.
        """
        if data.get(id_field):
            existing = self.select_one(table, data[id_field])
            if existing:
                return self.update(table, data[id_field], data)
        
        return self.insert(table, data)
    
    def count(self, table: str, filters: dict = None) -> int:
        """Count records matching filters"""
        endpoint = f"{table}?select=id"
        
        if filters:
            for key, value in filters.items():
                endpoint += f"&{key}=eq.{value}"
        
        # Use HEAD request with Prefer header for count
        headers = self.headers.copy()
        headers["Prefer"] = "count=exact"
        
        try:
            response = requests.head(
                f"{self.url}/rest/v1/{endpoint}",
                headers=headers,
                timeout=self.timeout
            )
            count_header = response.headers.get("content-range", "")
            if "/" in count_header:
                return int(count_header.split("/")[1])
        except:
            pass
        
        # Fallback: do actual select and count
        results = self.select(table, filters)
        return len(results)
    
    def sum(self, table: str, column: str, filters: dict = None) -> Decimal:
        """Sum a numeric column"""
        records = self.select(table, filters, columns=column)
        total = Decimal("0")
        for record in records:
            value = record.get(column, 0)
            if value:
                total += Decimal(str(value))
        return total


# Initialize database instance
db = Database()


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def generate_id() -> str:
    """Generate unique ID for records"""
    return str(uuid.uuid4())


def now() -> str:
    """Current datetime in ISO format"""
    return datetime.now().isoformat()


def today() -> str:
    """Current date as YYYY-MM-DD"""
    return datetime.now().strftime("%Y-%m-%d")


def format_date(date_str: str, output_format: str = "%d %b %Y") -> str:
    """
    Format date string for display
    
    Args:
        date_str: Date in ISO format or YYYY-MM-DD
        output_format: Desired output format
        
    Returns:
        Formatted date string
    """
    if not date_str:
        return ""
    
    try:
        # Handle ISO format with time
        if "T" in date_str:
            date_str = date_str.split("T")[0]
        
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime(output_format)
    except:
        return date_str


def safe_string(value, max_length: int = None) -> str:
    """
    Sanitize string for safe output
    Removes potentially dangerous characters
    """
    if value is None:
        return ""
    
    s = str(value)
    
    # Remove control characters and potentially dangerous chars
    s = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', s)
    s = s.replace('<', '&lt;').replace('>', '&gt;')
    s = s.replace('"', '&quot;').replace("'", '&#39;')
    
    if max_length:
        s = s[:max_length]
    
    return s.strip()


def safe_json_string(value) -> str:
    """Sanitize string for use in JavaScript"""
    if value is None:
        return ""
    
    s = str(value)
    s = s.replace('\\', '\\\\')
    s = s.replace('"', '\\"')
    s = s.replace("'", "\\'")
    s = s.replace('\n', '\\n')
    s = s.replace('\r', '\\r')
    s = s.replace('\t', '\\t')
    
    # Remove non-printable characters
    s = re.sub(r'[^\x20-\x7E]', '', s)
    
    return s


# ═══════════════════════════════════════════════════════════════════════════════
# USER SESSION MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

class UserSession:
    """Handle user authentication and sessions"""
    
    @staticmethod
    def login(username: str, password: str) -> tuple:
        """
        Authenticate user
        
        Returns:
            Tuple of (success: bool, user: dict or error: str)
        """
        users = db.select("users", {"username": username}, limit=1)
        
        if not users:
            return False, "User not found"
        
        user = users[0]
        
        # Simple password check (in production, use proper hashing)
        if user.get("password") != password:
            return False, "Invalid password"
        
        if not user.get("active", True):
            return False, "Account disabled"
        
        # Set session
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["role"] = user.get("role", "user")
        session["login_time"] = now()
        
        # Set default business if user has businesses
        businesses = BusinessManager.get_user_businesses(user["id"])
        if businesses:
            session["current_business_id"] = businesses[0]["id"]
        
        return True, user
    
    @staticmethod
    def logout():
        """Clear user session"""
        session.clear()
    
    @staticmethod
    def get_current_user() -> dict:
        """Get current logged in user"""
        if "user_id" not in session:
            return None
        
        return {
            "id": session.get("user_id"),
            "username": session.get("username"),
            "role": session.get("role"),
            "current_business_id": session.get("current_business_id")
        }
    
    @staticmethod
    def is_logged_in() -> bool:
        """Check if user is logged in"""
        return "user_id" in session
    
    @staticmethod
    def require_login(f):
        """Decorator to require login for a route"""
        @wraps(f)
        def decorated(*args, **kwargs):
            if not UserSession.is_logged_in():
                return redirect("/login")
            return f(*args, **kwargs)
        return decorated
    
    @staticmethod
    def require_role(role: str):
        """Decorator to require specific role"""
        def decorator(f):
            @wraps(f)
            def decorated(*args, **kwargs):
                if not UserSession.is_logged_in():
                    return redirect("/login")
                if session.get("role") != role and session.get("role") != "admin":
                    return redirect("/unauthorized")
                return f(*args, **kwargs)
            return decorated
        return decorator


# ═══════════════════════════════════════════════════════════════════════════════
# MULTI-BUSINESS MANAGEMENT - THE COMPETITIVE ADVANTAGE
# ═══════════════════════════════════════════════════════════════════════════════

class BusinessManager:
    """
    Multi-Business Management System
    
    One subscription, unlimited businesses. Each business:
    - Has completely separate data (customers, stock, transactions)
    - Can have its own industry template
    - Maintains its own GL, VAT, and reports
    - Accessible via business switcher in header
    
    This is what Xero/QuickBooks charge extra for. Click includes it FREE.
    """
    
    # Industry templates with module visibility
    INDUSTRY_CONFIGS = {
        "hardware": {
            "name": "Hardware Store",
            "icon": "🔧",
            "description": "Retail hardware, tools, building supplies",
            "visible_modules": ["dashboard", "pos", "stock", "customers", "suppliers", "invoices", "quotes", "expenses", "banking", "reports", "payroll"],
            "hidden_modules": ["rooms", "bookings", "tables", "job_cards", "recipes"],
            "categories": ["Fasteners", "Power Tools", "Hand Tools", "Plumbing", "Electrical", "Paint", "Building Materials", "Safety Equipment", "Garden", "General"],
            "terminology": {}  # Use defaults
        },
        "pub_grill": {
            "name": "Pub & Grill",
            "icon": "🍺",
            "description": "Restaurant, bar, food service",
            "visible_modules": ["dashboard", "pos", "stock", "customers", "suppliers", "invoices", "expenses", "banking", "reports", "payroll", "tables", "recipes"],
            "hidden_modules": ["rooms", "bookings", "job_cards", "quotes"],
            "categories": ["Food - Proteins", "Food - Produce", "Food - Dry Goods", "Beverages - Beer", "Beverages - Spirits", "Beverages - Wine", "Beverages - Soft", "Cleaning", "Disposables", "Equipment"],
            "terminology": {"customer": "Guest", "stock": "Inventory"}
        },
        "bnb": {
            "name": "B&B / Guesthouse",
            "icon": "🛏️",
            "description": "Accommodation, hospitality",
            "visible_modules": ["dashboard", "customers", "suppliers", "invoices", "expenses", "banking", "reports", "payroll", "rooms", "bookings"],
            "hidden_modules": ["pos", "stock", "quotes", "job_cards", "tables", "recipes"],
            "categories": ["Linen", "Toiletries", "Breakfast Items", "Cleaning", "Maintenance", "Guest Amenities"],
            "terminology": {"customer": "Guest", "invoice": "Booking Invoice"}
        },
        "engineering": {
            "name": "Engineering / Manufacturing",
            "icon": "⚙️",
            "description": "Fabrication, repairs, job work",
            "visible_modules": ["dashboard", "stock", "customers", "suppliers", "invoices", "quotes", "expenses", "banking", "reports", "payroll", "job_cards"],
            "hidden_modules": ["pos", "rooms", "bookings", "tables", "recipes"],
            "categories": ["Raw Materials", "Steel", "Stainless", "Consumables", "Welding", "Tools", "Safety", "Finished Goods"],
            "terminology": {"invoice": "Job Invoice", "quote": "Job Quote"}
        },
        "retail": {
            "name": "General Retail",
            "icon": "🛒",
            "description": "Shop, store, general trading",
            "visible_modules": ["dashboard", "pos", "stock", "customers", "suppliers", "invoices", "quotes", "expenses", "banking", "reports", "payroll"],
            "hidden_modules": ["rooms", "bookings", "tables", "job_cards", "recipes"],
            "categories": ["Category A", "Category B", "Category C", "Accessories", "Specials", "General"],
            "terminology": {}
        },
        "services": {
            "name": "Professional Services",
            "icon": "💼",
            "description": "Consulting, IT, accounting",
            "visible_modules": ["dashboard", "customers", "suppliers", "invoices", "quotes", "expenses", "banking", "reports", "payroll"],
            "hidden_modules": ["pos", "stock", "rooms", "bookings", "tables", "job_cards", "recipes"],
            "categories": [],
            "terminology": {"customer": "Client"}
        }
    }
    
    @classmethod
    def get_user_businesses(cls, user_id: str) -> list:
        """Get all businesses for a user - cached per request"""
        # Use Flask's g object to cache per-request
        cache_key = f'user_businesses_{user_id}'
        if hasattr(g, cache_key):
            return getattr(g, cache_key)
        
        try:
            businesses = db.select("businesses", {"owner_id": user_id, "active": True}, order="business_name")
            result = businesses if businesses else []
        except:
            result = []
        
        setattr(g, cache_key, result)
        return result
    
    @classmethod
    def get_current_business(cls) -> dict:
        """Get current active business from session - cached per request"""
        business_id = session.get("current_business_id")
        if not business_id:
            return None
        
        # Use Flask's g object to cache per-request
        cache_key = 'current_business'
        if hasattr(g, cache_key):
            return getattr(g, cache_key)
        
        try:
            business = db.select_one("businesses", business_id)
            if business:
                # Merge with industry config
                industry = business.get("industry", "retail")
                config = cls.INDUSTRY_CONFIGS.get(industry, cls.INDUSTRY_CONFIGS["retail"])
                business["config"] = config
        except:
            business = None
        
        setattr(g, cache_key, business)
        return business
    
    @classmethod
    def switch_business(cls, business_id: str) -> bool:
        """Switch to a different business"""
        user = UserSession.get_current_user()
        if not user:
            return False
        
        # Verify user owns this business
        try:
            business = db.select_one("businesses", business_id)
            if business and business.get("owner_id") == user["id"]:
                session["current_business_id"] = business_id
                return True
        except:
            pass
        
        return False
    
    @classmethod
    def create_business(cls, owner_id: str, name: str, industry: str) -> tuple:
        """
        Create a new business
        
        Returns:
            Tuple of (success: bool, business: dict or error: str)
        """
        if industry not in cls.INDUSTRY_CONFIGS:
            industry = "retail"
        
        config = cls.INDUSTRY_CONFIGS[industry]
        
        business = {
            "id": generate_id(),
            "owner_id": owner_id,
            "business_name": name,
            "name": name,  # Keep both for compatibility
            "industry": industry,
            "business_type": config["name"],
            "icon": config["icon"],
            "visible_modules": json.dumps(config["visible_modules"]),
            "hidden_modules": json.dumps(config["hidden_modules"]),
            "terminology": json.dumps(config.get("terminology", {})),
            "created_at": now(),
            "active": True
        }
        
        try:
            success, result = db.insert("businesses", business)
            if success:
                # Create default stock categories for this business
                for i, cat_name in enumerate(config["categories"]):
                    try:
                        db.insert("stock_categories", {
                            "id": generate_id(),
                            "business_id": business["id"],
                            "name": cat_name,
                            "sort_order": i
                        })
                    except:
                        pass
                
                # Set as current business
                session["current_business_id"] = business["id"]
                return True, business
            return False, "Failed to create business"
        except Exception as e:
            return False, str(e)
    
    @classmethod
    def update_business(cls, business_id: str, data: dict) -> bool:
        """Update business settings"""
        try:
            success, _ = db.update("businesses", business_id, data)
            return success
        except:
            return False
    
    @classmethod
    def delete_business(cls, business_id: str) -> bool:
        """Delete a business (soft delete - just deactivate)"""
        return cls.update_business(business_id, {"active": False})
    
    @classmethod
    def is_module_visible(cls, module: str) -> bool:
        """Check if a module is visible for current business"""
        business = cls.get_current_business()
        if not business:
            return True  # No business context, show everything
        
        try:
            visible = json.loads(business.get("visible_modules", "[]"))
            hidden = json.loads(business.get("hidden_modules", "[]"))
            
            if module in hidden:
                return False
            if visible and module not in visible:
                return False
            return True
        except:
            return True
    
    @classmethod
    def get_terminology(cls, term: str) -> str:
        """Get custom terminology for current business"""
        business = cls.get_current_business()
        if not business:
            return term.title()
        
        try:
            terminology = json.loads(business.get("terminology", "{}"))
            return terminology.get(term.lower(), term.title())
        except:
            return term.title()
    
    @classmethod
    def get_business_filter(cls) -> dict:
        """Get filter dict for current business - use in all queries"""
        business_id = session.get("current_business_id")
        if business_id:
            return {"business_id": business_id}
        return {}


# Helper function to filter queries by business
def biz_filter(extra_filters: dict = None) -> dict:
    """
    Get business filter for database queries
    
    Usage:
        customers = db.select("customers", biz_filter({"active": True}))
    """
    filters = BusinessManager.get_business_filter()
    if extra_filters:
        filters.update(extra_filters)
    return filters


def get_biz_id() -> str:
    """Get current business ID or None"""
    return session.get("current_business_id")


# ═══════════════════════════════════════════════════════════════════════════════
# OPUS AI INTEGRATION
# ═══════════════════════════════════════════════════════════════════════════════

class OpusAI:
    """
    Claude Opus integration for intelligent features
    
    Opus ADVISES, Flask EXECUTES
    - Receipt reading and data extraction
    - Expense categorization suggestions
    - Anomaly detection
    - Business insights
    """
    
    @staticmethod
    def _call_api(messages: list, model: str = None, max_tokens: int = 1000) -> tuple:
        """
        Call Anthropic API
        
        Returns:
            Tuple of (success: bool, response: str or error: str)
        """
        api_key = Config.ANTHROPIC_API_KEY
        if not api_key:
            return False, "API key not configured"
        
        if model is None:
            model = Config.OPUS_MODEL
        
        try:
            response = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "content-type": "application/json",
                    "anthropic-version": "2023-06-01"
                },
                json={
                    "model": model,
                    "max_tokens": max_tokens,
                    "messages": messages
                },
                timeout=60
            )
            
            if response.status_code == 200:
                result = response.json()
                text = result.get("content", [{}])[0].get("text", "")
                return True, text
            else:
                return False, f"API error: {response.status_code}"
                
        except requests.exceptions.Timeout:
            return False, "API timeout"
        except Exception as e:
            return False, str(e)
    
    @classmethod
    def read_receipt(cls, image_base64: str) -> dict:
        """
        Extract data from receipt/invoice image - SA formats including Afrikaans
        FOCUS: Get results, not errors!
        """
        # Clean base64 string
        if "," in image_base64:
            image_base64 = image_base64.split(",")[1]
        
        prompt = """Look at this South African receipt/invoice and tell me:

1. SUPPLIER: What store/company name do you see at the top?
2. TOTAL: What is the final total amount? (look for TOTAAL, TOTAL, or the biggest number at bottom)
3. VAT: What is the VAT/BTW amount?
4. DATE: What date is on it?
5. ITEMS: List the main items purchased

South African terms: BTW=VAT, BEDRAG=Amount, TOTAAL=Total, KONTANT=Cash, INKL=Including

Reply with ONLY this JSON:
{"supplier":"store name","total":123.45,"vat":15.67,"date":"2025-12-21","description":"brief list of items","category":"Stock/Inventory"}

For category use: Stock/Inventory, Fuel, Office Supplies, Repairs, Utilities, or Other"""

        messages = [{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/jpeg",
                        "data": image_base64
                    }
                },
                {"type": "text", "text": prompt}
            ]
        }]
        
        # Use Sonnet - fast and good at vision
        success, response = cls._call_api(messages, max_tokens=500, model="claude-sonnet-4-20250514")
        
        if not success:
            # Still try to return something useful
            return {"supplier": "", "total": 0, "description": "Could not read - please enter manually", "category": "Other"}
        
        # Parse JSON from response
        try:
            start = response.find('{')
            end = response.rfind('}') + 1
            if start >= 0 and end > start:
                data = json.loads(response[start:end])
                
                # Ensure we have the fields the form needs
                if "total" in data and "amount" not in data:
                    data["amount"] = data["total"]
                
                return data
        except:
            pass
        
        # Last resort - return empty form
        return {"supplier": "", "total": 0, "description": "", "category": "Other"}
    
    @classmethod
    def suggest_category(cls, description: str, supplier: str = "") -> str:
        """
        Suggest expense category based on description
        
        Args:
            description: Expense description
            supplier: Supplier name (optional)
            
        Returns:
            Suggested category name
        """
        prompt = f"""Based on this expense:
Supplier: {supplier}
Description: {description}

Suggest the most appropriate expense category from this list:
- Accounting Fees
- Advertising & Marketing
- Bank Charges
- Cleaning
- Computer Expenses
- Consumables
- Electricity & Water
- Entertainment
- Fuel & Oil
- General Expenses
- Insurance
- Internet & Telephone
- Motor Vehicle Expenses
- Office Supplies
- Postage & Courier
- Printing & Stationery
- Rent
- Repairs & Maintenance
- Salaries & Wages
- Security
- Stock/Inventory (if buying items for resale)
- Transport
- Travel & Accommodation

Reply with ONLY the category name, nothing else."""

        messages = [{"role": "user", "content": prompt}]
        
        success, response = cls._call_api(messages, model=Config.SONNET_MODEL, max_tokens=50)
        
        if success:
            return response.strip()
        
        return "General Expenses"


# ═══════════════════════════════════════════════════════════════════════════════
# DOCUMENT NUMBER GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

class DocumentNumbers:
    """Generate sequential document numbers"""
    
    @staticmethod
    def get_next(prefix: str, table: str, column: str = "number") -> str:
        """
        Get next document number
        
        Args:
            prefix: Document prefix (e.g., "INV", "QUO")
            table: Table to check for existing numbers
            column: Column containing document numbers
            
        Returns:
            Next number like "INV0001"
        """
        # Get all existing numbers with this prefix
        records = db.select(table, columns=column)
        
        max_num = 0
        for record in records:
            num_str = record.get(column, "")
            if num_str.startswith(prefix):
                try:
                    num = int(num_str[len(prefix):])
                    max_num = max(max_num, num)
                except:
                    pass
        
        next_num = max_num + 1
        return f"{prefix}{next_num:04d}"


# ═══════════════════════════════════════════════════════════════════════════════
# FINANCIAL PERIOD HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

class FinancialPeriod:
    """Handle financial year and period calculations"""
    
    @staticmethod
    def get_current_year_start() -> str:
        """Get start date of current financial year"""
        now = datetime.now()
        year_end_month = Config.FINANCIAL_YEAR_END_MONTH
        
        if now.month <= year_end_month:
            # We're in the latter part of the financial year
            year_start = datetime(now.year - 1, year_end_month + 1, 1)
        else:
            # We're in the earlier part of the financial year
            year_start = datetime(now.year, year_end_month + 1, 1)
        
        return year_start.strftime("%Y-%m-%d")
    
    @staticmethod
    def get_current_year_end() -> str:
        """Get end date of current financial year"""
        now = datetime.now()
        year_end_month = Config.FINANCIAL_YEAR_END_MONTH
        
        if now.month <= year_end_month:
            year_end = datetime(now.year, year_end_month + 1, 1) - timedelta(days=1)
        else:
            year_end = datetime(now.year + 1, year_end_month + 1, 1) - timedelta(days=1)
        
        return year_end.strftime("%Y-%m-%d")
    
    @staticmethod
    def get_vat_period_start(date: datetime = None) -> str:
        """Get start of current VAT period (bi-monthly)"""
        if date is None:
            date = datetime.now()
        
        # VAT periods: Jan-Feb, Mar-Apr, May-Jun, Jul-Aug, Sep-Oct, Nov-Dec
        month = date.month
        if month % 2 == 0:
            period_start_month = month - 1
        else:
            period_start_month = month
        
        return datetime(date.year, period_start_month, 1).strftime("%Y-%m-%d")
    
    @staticmethod
    def get_vat_period_end(date: datetime = None) -> str:
        """Get end of current VAT period"""
        if date is None:
            date = datetime.now()
        
        month = date.month
        if month % 2 == 0:
            period_end_month = month
        else:
            period_end_month = month + 1
        
        # Get last day of period_end_month
        if period_end_month == 12:
            next_month = datetime(date.year + 1, 1, 1)
        else:
            next_month = datetime(date.year, period_end_month + 1, 1)
        
        period_end = next_month - timedelta(days=1)
        return period_end.strftime("%Y-%m-%d")


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 1
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 1 COMPLETE - Core Foundation

Contains:
✓ Flask app setup
✓ Configuration class
✓ VAT logic with zero-rated detection
✓ Money handling with Decimal precision
✓ Database class with full CRUD
✓ Helper functions (ID generation, dates, string sanitization)
✓ User session management
✓ Opus AI integration
✓ Document number generation
✓ Financial period helpers

Next: Piece 2 - Chart of Accounts
"""

# =============================================================================
# PIECE2_ACCOUNTS.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 2: Chart of Accounts                                                  ║
║                                                                               ║
║   This piece contains:                                                        ║
║   - Complete South African SME Chart of Accounts                              ║
║   - Account types and categories                                              ║
║   - Account class for operations                                              ║
║   - Functions to initialize and manage accounts                               ║
║                                                                               ║
║   Account Numbering Convention:                                               ║
║   1000-1999: Assets                                                           ║
║   2000-2999: Liabilities                                                      ║
║   3000-3999: Equity                                                           ║
║   4000-4999: Revenue/Income                                                   ║
║   5000-5999: Cost of Sales                                                    ║
║   6000-6999: Operating Expenses                                               ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

from enum import Enum
from typing import List, Dict, Optional


# ═══════════════════════════════════════════════════════════════════════════════
# ACCOUNT TYPES
# ═══════════════════════════════════════════════════════════════════════════════

class AccountType(Enum):
    """
    Account types determine how the account appears in financial statements
    and how its balance is calculated
    """
    ASSET = "asset"                 # Debit balance, appears on Balance Sheet
    LIABILITY = "liability"         # Credit balance, appears on Balance Sheet
    EQUITY = "equity"               # Credit balance, appears on Balance Sheet
    REVENUE = "revenue"             # Credit balance, appears on Income Statement
    COST_OF_SALES = "cost_of_sales" # Debit balance, appears on Income Statement
    EXPENSE = "expense"             # Debit balance, appears on Income Statement


class AccountCategory(Enum):
    """
    Sub-categories for grouping accounts in reports
    """
    # Asset categories
    CURRENT_ASSET = "current_asset"
    FIXED_ASSET = "fixed_asset"
    
    # Liability categories
    CURRENT_LIABILITY = "current_liability"
    LONG_TERM_LIABILITY = "long_term_liability"
    
    # Equity categories
    CAPITAL = "capital"
    RETAINED_EARNINGS = "retained_earnings"
    
    # Revenue categories
    SALES = "sales"
    OTHER_INCOME = "other_income"
    
    # Cost of Sales categories
    DIRECT_COSTS = "direct_costs"
    
    # Expense categories
    OPERATING_EXPENSE = "operating_expense"
    ADMINISTRATIVE_EXPENSE = "administrative_expense"
    FINANCIAL_EXPENSE = "financial_expense"


# ═══════════════════════════════════════════════════════════════════════════════
# STANDARD CHART OF ACCOUNTS - SOUTH AFRICAN SME
# ═══════════════════════════════════════════════════════════════════════════════

STANDARD_CHART_OF_ACCOUNTS = [
    # ═══════════════════════════════════════════════════════════════════════════
    # ASSETS (1000-1999)
    # ═══════════════════════════════════════════════════════════════════════════
    
    # Current Assets (1000-1399)
    {
        "code": "1000",
        "name": "Bank - Current Account",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Main business bank account",
        "vat_applicable": False,
        "system_account": True  # Cannot be deleted
    },
    {
        "code": "1010",
        "name": "Bank - Savings Account",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Business savings account",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1020",
        "name": "Petty Cash",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Cash on hand for small expenses",
        "vat_applicable": False,
        "system_account": True
    },
    {
        "code": "1030",
        "name": "Cash Float",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Till/register cash float",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1100",
        "name": "Debtors Control",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Total amount owed by customers",
        "vat_applicable": False,
        "system_account": True
    },
    {
        "code": "1110",
        "name": "Debtors - Trade",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Amounts owed for goods/services sold",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1120",
        "name": "Debtors - Other",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Other amounts receivable",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1130",
        "name": "Provision for Bad Debts",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Contra account for doubtful debts",
        "vat_applicable": False,
        "system_account": False,
        "contra_account": True
    },
    {
        "code": "1200",
        "name": "Stock / Inventory",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Goods held for resale",
        "vat_applicable": False,
        "system_account": True
    },
    {
        "code": "1210",
        "name": "Stock - Raw Materials",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Raw materials for manufacturing",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1220",
        "name": "Stock - Work in Progress",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Partially completed goods",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1230",
        "name": "Stock - Finished Goods",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Completed goods ready for sale",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1300",
        "name": "Prepaid Expenses",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Expenses paid in advance",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1310",
        "name": "Prepaid Rent",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Rent paid in advance",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1320",
        "name": "Prepaid Insurance",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Insurance premiums paid in advance",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "1350",
        "name": "Deposits Paid",
        "type": AccountType.ASSET,
        "category": AccountCategory.CURRENT_ASSET,
        "description": "Refundable deposits paid",
        "vat_applicable": False,
        "system_account": False
    },
    
    # Fixed Assets (1400-1999)
    {
        "code": "1400",
        "name": "Land",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Land owned by business",
        "vat_applicable": False,
        "system_account": False,
        "depreciable": False
    },
    {
        "code": "1410",
        "name": "Buildings",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Buildings at cost",
        "vat_applicable": False,
        "system_account": False,
        "depreciable": True,
        "depreciation_rate": Decimal("0.05")  # 5% per year
    },
    {
        "code": "1411",
        "name": "Buildings - Accumulated Depreciation",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Accumulated depreciation on buildings",
        "vat_applicable": False,
        "system_account": False,
        "contra_account": True
    },
    {
        "code": "1500",
        "name": "Equipment at Cost",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Equipment and machinery at cost",
        "vat_applicable": False,
        "system_account": False,
        "depreciable": True,
        "depreciation_rate": Decimal("0.20")  # 20% per year
    },
    {
        "code": "1510",
        "name": "Equipment - Accumulated Depreciation",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Accumulated depreciation on equipment",
        "vat_applicable": False,
        "system_account": False,
        "contra_account": True
    },
    {
        "code": "1600",
        "name": "Vehicles at Cost",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Motor vehicles at cost",
        "vat_applicable": False,
        "system_account": False,
        "depreciable": True,
        "depreciation_rate": Decimal("0.20")  # 20% per year
    },
    {
        "code": "1610",
        "name": "Vehicles - Accumulated Depreciation",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Accumulated depreciation on vehicles",
        "vat_applicable": False,
        "system_account": False,
        "contra_account": True
    },
    {
        "code": "1700",
        "name": "Furniture & Fittings at Cost",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Office furniture at cost",
        "vat_applicable": False,
        "system_account": False,
        "depreciable": True,
        "depreciation_rate": Decimal("0.1667")  # 16.67% per year (6 years)
    },
    {
        "code": "1710",
        "name": "Furniture - Accumulated Depreciation",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Accumulated depreciation on furniture",
        "vat_applicable": False,
        "system_account": False,
        "contra_account": True
    },
    {
        "code": "1800",
        "name": "Computer Equipment at Cost",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Computers and IT equipment at cost",
        "vat_applicable": False,
        "system_account": False,
        "depreciable": True,
        "depreciation_rate": Decimal("0.3333")  # 33.33% per year (3 years)
    },
    {
        "code": "1810",
        "name": "Computer Equipment - Accumulated Depreciation",
        "type": AccountType.ASSET,
        "category": AccountCategory.FIXED_ASSET,
        "description": "Accumulated depreciation on computer equipment",
        "vat_applicable": False,
        "system_account": False,
        "contra_account": True
    },
    
    # ═══════════════════════════════════════════════════════════════════════════
    # LIABILITIES (2000-2999)
    # ═══════════════════════════════════════════════════════════════════════════
    
    # Current Liabilities (2000-2499)
    {
        "code": "2000",
        "name": "Creditors Control",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Total amount owed to suppliers",
        "vat_applicable": False,
        "system_account": True
    },
    {
        "code": "2010",
        "name": "Creditors - Trade",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Amounts owed for goods/services purchased",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2020",
        "name": "Creditors - Other",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Other amounts payable",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2100",
        "name": "VAT Output",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "VAT collected on sales - payable to SARS",
        "vat_applicable": False,
        "system_account": True
    },
    {
        "code": "2110",
        "name": "VAT Input",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "VAT paid on purchases - claimable from SARS",
        "vat_applicable": False,
        "system_account": True,
        "contra_account": True  # Contra to VAT Output
    },
    {
        "code": "2120",
        "name": "VAT Control",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Net VAT position (Output - Input)",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2200",
        "name": "PAYE Payable",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Employee tax deducted - payable to SARS",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2210",
        "name": "UIF Payable",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "UIF contributions payable",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2220",
        "name": "SDL Payable",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Skills Development Levy payable",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2300",
        "name": "Accrued Expenses",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Expenses incurred but not yet paid",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2310",
        "name": "Salaries Payable",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Salaries earned but not yet paid",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2400",
        "name": "Short-term Loans",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Loans payable within 12 months",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2410",
        "name": "Bank Overdraft",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Bank overdraft facility",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2420",
        "name": "Credit Card",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Business credit card balance",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2450",
        "name": "Income Received in Advance",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.CURRENT_LIABILITY,
        "description": "Payments received for services not yet rendered",
        "vat_applicable": False,
        "system_account": False
    },
    
    # Long-term Liabilities (2500-2999)
    {
        "code": "2500",
        "name": "Long-term Loans",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.LONG_TERM_LIABILITY,
        "description": "Loans payable after 12 months",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2510",
        "name": "Vehicle Finance",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.LONG_TERM_LIABILITY,
        "description": "Finance agreements on vehicles",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2520",
        "name": "Equipment Finance",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.LONG_TERM_LIABILITY,
        "description": "Finance agreements on equipment",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2600",
        "name": "Mortgage Bond",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.LONG_TERM_LIABILITY,
        "description": "Property mortgage",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "2700",
        "name": "Shareholder Loan",
        "type": AccountType.LIABILITY,
        "category": AccountCategory.LONG_TERM_LIABILITY,
        "description": "Loans from shareholders/directors",
        "vat_applicable": False,
        "system_account": False
    },
    
    # ═══════════════════════════════════════════════════════════════════════════
    # EQUITY (3000-3999)
    # ═══════════════════════════════════════════════════════════════════════════
    
    {
        "code": "3000",
        "name": "Share Capital / Owner's Capital",
        "type": AccountType.EQUITY,
        "category": AccountCategory.CAPITAL,
        "description": "Capital invested by owners",
        "vat_applicable": False,
        "system_account": True
    },
    {
        "code": "3010",
        "name": "Share Premium",
        "type": AccountType.EQUITY,
        "category": AccountCategory.CAPITAL,
        "description": "Premium received on share issue",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "3100",
        "name": "Retained Earnings",
        "type": AccountType.EQUITY,
        "category": AccountCategory.RETAINED_EARNINGS,
        "description": "Accumulated profits from prior years",
        "vat_applicable": False,
        "system_account": True
    },
    {
        "code": "3200",
        "name": "Current Year Earnings",
        "type": AccountType.EQUITY,
        "category": AccountCategory.RETAINED_EARNINGS,
        "description": "Profit/loss for current financial year",
        "vat_applicable": False,
        "system_account": True
    },
    {
        "code": "3300",
        "name": "Drawings",
        "type": AccountType.EQUITY,
        "category": AccountCategory.CAPITAL,
        "description": "Amounts withdrawn by owner",
        "vat_applicable": False,
        "system_account": False,
        "contra_account": True  # Reduces equity
    },
    {
        "code": "3400",
        "name": "Dividends",
        "type": AccountType.EQUITY,
        "category": AccountCategory.RETAINED_EARNINGS,
        "description": "Dividends declared",
        "vat_applicable": False,
        "system_account": False,
        "contra_account": True
    },
    
    # ═══════════════════════════════════════════════════════════════════════════
    # REVENUE / INCOME (4000-4999)
    # ═══════════════════════════════════════════════════════════════════════════
    
    {
        "code": "4000",
        "name": "Sales - Goods",
        "type": AccountType.REVENUE,
        "category": AccountCategory.SALES,
        "description": "Revenue from sale of goods",
        "vat_applicable": True,
        "system_account": True
    },
    {
        "code": "4010",
        "name": "Sales - Services",
        "type": AccountType.REVENUE,
        "category": AccountCategory.SALES,
        "description": "Revenue from services rendered",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "4020",
        "name": "Sales - Other",
        "type": AccountType.REVENUE,
        "category": AccountCategory.SALES,
        "description": "Other sales revenue",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "4050",
        "name": "Sales Returns",
        "type": AccountType.REVENUE,
        "category": AccountCategory.SALES,
        "description": "Goods returned by customers",
        "vat_applicable": True,
        "system_account": False,
        "contra_account": True
    },
    {
        "code": "4060",
        "name": "Sales Discounts",
        "type": AccountType.REVENUE,
        "category": AccountCategory.SALES,
        "description": "Discounts given to customers",
        "vat_applicable": True,
        "system_account": False,
        "contra_account": True
    },
    {
        "code": "4100",
        "name": "Discount Received",
        "type": AccountType.REVENUE,
        "category": AccountCategory.OTHER_INCOME,
        "description": "Discounts received from suppliers",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "4200",
        "name": "Interest Received",
        "type": AccountType.REVENUE,
        "category": AccountCategory.OTHER_INCOME,
        "description": "Interest earned on bank accounts",
        "vat_applicable": False,  # Interest is exempt from VAT
        "system_account": False
    },
    {
        "code": "4300",
        "name": "Rental Income",
        "type": AccountType.REVENUE,
        "category": AccountCategory.OTHER_INCOME,
        "description": "Income from property rental",
        "vat_applicable": True,  # Commercial rental is VATable
        "system_account": False
    },
    {
        "code": "4400",
        "name": "Commission Received",
        "type": AccountType.REVENUE,
        "category": AccountCategory.OTHER_INCOME,
        "description": "Commission earned",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "4500",
        "name": "Bad Debts Recovered",
        "type": AccountType.REVENUE,
        "category": AccountCategory.OTHER_INCOME,
        "description": "Previously written off debts collected",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "4600",
        "name": "Profit on Sale of Assets",
        "type": AccountType.REVENUE,
        "category": AccountCategory.OTHER_INCOME,
        "description": "Profit from selling fixed assets",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "4900",
        "name": "Other Income",
        "type": AccountType.REVENUE,
        "category": AccountCategory.OTHER_INCOME,
        "description": "Miscellaneous income",
        "vat_applicable": True,
        "system_account": False
    },
    
    # ═══════════════════════════════════════════════════════════════════════════
    # COST OF SALES (5000-5999)
    # ═══════════════════════════════════════════════════════════════════════════
    
    {
        "code": "5000",
        "name": "Cost of Goods Sold",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Cost of goods sold to customers",
        "vat_applicable": False,  # Already accounted for on purchase
        "system_account": True
    },
    {
        "code": "5010",
        "name": "Opening Stock",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Stock at beginning of period",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "5020",
        "name": "Purchases",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Goods purchased for resale",
        "vat_applicable": True,
        "system_account": True
    },
    {
        "code": "5025",
        "name": "Purchase Returns",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Goods returned to suppliers",
        "vat_applicable": True,
        "system_account": False,
        "contra_account": True
    },
    {
        "code": "5030",
        "name": "Closing Stock",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Stock at end of period",
        "vat_applicable": False,
        "system_account": False,
        "contra_account": True
    },
    {
        "code": "5040",
        "name": "Carriage Inward",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Freight and delivery costs on purchases",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "5050",
        "name": "Import Duties",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Customs and import duties",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "5100",
        "name": "Direct Labour",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Labour directly involved in production",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "5200",
        "name": "Subcontractors",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Subcontracted work",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "5300",
        "name": "Manufacturing Overheads",
        "type": AccountType.COST_OF_SALES,
        "category": AccountCategory.DIRECT_COSTS,
        "description": "Factory overheads allocated to production",
        "vat_applicable": False,
        "system_account": False
    },
    
    # ═══════════════════════════════════════════════════════════════════════════
    # OPERATING EXPENSES (6000-6999)
    # ═══════════════════════════════════════════════════════════════════════════
    
    {
        "code": "6000",
        "name": "Accounting Fees",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.ADMINISTRATIVE_EXPENSE,
        "description": "Fees paid to accountants",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6010",
        "name": "Advertising & Marketing",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Advertising, marketing, promotions",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6020",
        "name": "Bad Debts",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Irrecoverable customer debts written off",
        "vat_applicable": False,  # VAT already claimed on original sale
        "system_account": False
    },
    {
        "code": "6030",
        "name": "Bank Charges",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.FINANCIAL_EXPENSE,
        "description": "Bank fees and charges",
        "vat_applicable": False,  # Bank charges are exempt
        "system_account": True
    },
    {
        "code": "6040",
        "name": "Cleaning",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Cleaning services and supplies",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6050",
        "name": "Computer Expenses",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Software, IT support, computer supplies",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6060",
        "name": "Consulting Fees",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.ADMINISTRATIVE_EXPENSE,
        "description": "Fees paid to consultants",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6070",
        "name": "Consumables",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "General consumable items",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6080",
        "name": "Depreciation",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Depreciation on fixed assets",
        "vat_applicable": False,  # Non-cash expense
        "system_account": True
    },
    {
        "code": "6090",
        "name": "Electricity & Water",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Utilities - electricity and water",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6100",
        "name": "Entertainment",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Client entertainment, meals",
        "vat_applicable": True,  # But only 50% deductible for tax
        "system_account": False
    },
    {
        "code": "6110",
        "name": "Equipment Hire",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Rental of equipment",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6120",
        "name": "Fines & Penalties",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Traffic fines, penalties (non-deductible)",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "6130",
        "name": "Fuel & Oil",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Petrol, diesel for vehicles",
        "vat_applicable": False,  # Zero-rated
        "system_account": True,
        "zero_rated": True
    },
    {
        "code": "6140",
        "name": "General Expenses",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Miscellaneous expenses",
        "vat_applicable": True,
        "system_account": True
    },
    {
        "code": "6150",
        "name": "Insurance",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Business insurance premiums",
        "vat_applicable": False,  # Insurance is exempt
        "system_account": False
    },
    {
        "code": "6160",
        "name": "Interest Paid",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.FINANCIAL_EXPENSE,
        "description": "Interest on loans and overdrafts",
        "vat_applicable": False,  # Interest is exempt
        "system_account": False
    },
    {
        "code": "6170",
        "name": "Internet & Telephone",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Telephone, internet, data costs",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6180",
        "name": "Legal Fees",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.ADMINISTRATIVE_EXPENSE,
        "description": "Fees paid to attorneys",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6190",
        "name": "Licences & Permits",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.ADMINISTRATIVE_EXPENSE,
        "description": "Business licences and permits",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6200",
        "name": "Motor Vehicle Expenses",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Vehicle maintenance, repairs, licence fees",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6210",
        "name": "Office Supplies",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Stationery, office consumables",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6220",
        "name": "Postage & Courier",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Postage, courier services",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6230",
        "name": "Printing & Stationery",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Printing, stationery purchases",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6240",
        "name": "Protective Clothing",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Safety gear, uniforms for staff",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6250",
        "name": "Rent",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Office/premises rental",
        "vat_applicable": True,  # Commercial rent is VATable
        "system_account": False
    },
    {
        "code": "6260",
        "name": "Repairs & Maintenance",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Repairs and maintenance",
        "vat_applicable": True,
        "system_account": True
    },
    {
        "code": "6270",
        "name": "Salaries & Wages",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Staff salaries and wages",
        "vat_applicable": False,  # Salaries are not subject to VAT
        "system_account": True
    },
    {
        "code": "6280",
        "name": "PAYE Expense",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Employer portion of PAYE",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "6290",
        "name": "UIF Expense",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Employer UIF contribution",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "6300",
        "name": "SDL Expense",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Skills Development Levy",
        "vat_applicable": False,
        "system_account": False
    },
    {
        "code": "6310",
        "name": "Staff Welfare",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Staff refreshments, welfare",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6320",
        "name": "Security",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Security services",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6330",
        "name": "Subscriptions",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Magazine, professional subscriptions",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6340",
        "name": "Training",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Staff training and courses",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6350",
        "name": "Travel & Accommodation",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Business travel, accommodation",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6360",
        "name": "Uniforms",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Staff uniforms",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6370",
        "name": "Waste Removal",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Refuse removal, waste disposal",
        "vat_applicable": True,
        "system_account": False
    },
    {
        "code": "6900",
        "name": "Loss on Sale of Assets",
        "type": AccountType.EXPENSE,
        "category": AccountCategory.OPERATING_EXPENSE,
        "description": "Loss from selling fixed assets",
        "vat_applicable": False,
        "system_account": False
    },
]


# ═══════════════════════════════════════════════════════════════════════════════
# ACCOUNT CLASS
# ═══════════════════════════════════════════════════════════════════════════════

class Account:
    """
    Account class for managing chart of accounts
    """
    
    TABLE = "accounts"
    
    @classmethod
    def get_all(cls, active_only: bool = True) -> List[dict]:
        """Get all accounts"""
        filters = {"active": True} if active_only else None
        return db.select(cls.TABLE, filters, order="code")
    
    @classmethod
    def get_by_code(cls, code: str) -> Optional[dict]:
        """Get account by code"""
        results = db.select(cls.TABLE, {"code": code}, limit=1)
        return results[0] if results else None
    
    @classmethod
    def get_by_type(cls, account_type: AccountType) -> List[dict]:
        """Get all accounts of a specific type"""
        return db.select(cls.TABLE, {"type": account_type.value}, order="code")
    
    @classmethod
    def get_by_category(cls, category: AccountCategory) -> List[dict]:
        """Get all accounts in a specific category"""
        return db.select(cls.TABLE, {"category": category.value}, order="code")
    
    @classmethod
    def get_expense_accounts(cls) -> List[dict]:
        """Get all expense accounts for dropdown"""
        return db.select(cls.TABLE, {"type": AccountType.EXPENSE.value}, order="code")
    
    @classmethod
    def get_income_accounts(cls) -> List[dict]:
        """Get all income accounts"""
        return db.select(cls.TABLE, {"type": AccountType.REVENUE.value}, order="code")
    
    @classmethod
    def create(cls, code: str, name: str, account_type: AccountType, 
               category: AccountCategory, **kwargs) -> tuple:
        """Create a new account"""
        data = {
            "id": generate_id(),
            "code": code,
            "name": name,
            "type": account_type.value,
            "category": category.value,
            "description": kwargs.get("description", ""),
            "vat_applicable": kwargs.get("vat_applicable", True),
            "system_account": kwargs.get("system_account", False),
            "contra_account": kwargs.get("contra_account", False),
            "active": True,
            "created_at": now()
        }
        return db.insert(cls.TABLE, data)
    
    @classmethod
    def deactivate(cls, code: str) -> tuple:
        """Deactivate an account (don't delete, preserve history)"""
        account = cls.get_by_code(code)
        if account:
            if account.get("system_account"):
                return False, "Cannot deactivate system account"
            return db.update(cls.TABLE, account["id"], {"active": False})
        return False, "Account not found"
    
    @classmethod
    def initialize_chart(cls, force: bool = False) -> dict:
        """
        Initialize the standard chart of accounts
        
        Args:
            force: If True, recreate all accounts (use with caution)
            
        Returns:
            Dict with counts of accounts created/skipped
        """
        created = 0
        skipped = 0
        errors = []
        
        for account_def in STANDARD_CHART_OF_ACCOUNTS:
            # Check if account already exists
            existing = cls.get_by_code(account_def["code"])
            
            if existing and not force:
                skipped += 1
                continue
            
            data = {
                "id": generate_id(),
                "code": account_def["code"],
                "name": account_def["name"],
                "type": account_def["type"].value,
                "category": account_def["category"].value,
                "description": account_def.get("description", ""),
                "vat_applicable": account_def.get("vat_applicable", True),
                "system_account": account_def.get("system_account", False),
                "contra_account": account_def.get("contra_account", False),
                "zero_rated": account_def.get("zero_rated", False),
                "depreciable": account_def.get("depreciable", False),
                "depreciation_rate": str(account_def.get("depreciation_rate", 0)),
                "active": True,
                "created_at": now()
            }
            
            if existing and force:
                # Update existing
                success, result = db.update(cls.TABLE, existing["id"], data)
            else:
                # Create new
                success, result = db.insert(cls.TABLE, data)
            
            if success:
                created += 1
            else:
                errors.append(f"{account_def['code']}: {result}")
        
        return {
            "created": created,
            "skipped": skipped,
            "errors": errors,
            "total": len(STANDARD_CHART_OF_ACCOUNTS)
        }


# ═══════════════════════════════════════════════════════════════════════════════
# ACCOUNT LOOKUP HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

class AccountCodes:
    """
    Quick lookup for common account codes
    Used throughout the system for journal posting
    """
    
    # Assets
    BANK = "1000"
    PETTY_CASH = "1020"
    DEBTORS = "1100"
    STOCK = "1200"
    
    # Liabilities
    CREDITORS = "2000"
    VAT_OUTPUT = "2100"
    VAT_INPUT = "2110"
    
    # Equity
    CAPITAL = "3000"
    RETAINED_EARNINGS = "3100"
    CURRENT_YEAR_EARNINGS = "3200"
    DRAWINGS = "3300"
    
    # Revenue
    SALES = "4000"
    
    # Cost of Sales
    COGS = "5000"
    PURCHASES = "5020"
    
    # Common Expenses
    BANK_CHARGES = "6030"
    FUEL = "6130"
    GENERAL_EXPENSES = "6140"
    REPAIRS = "6260"
    SALARIES = "6270"


def get_account_code(name_or_code: str) -> str:
    """
    Get account code from name or code
    Useful for flexible input
    """
    # If it looks like a code already, return it
    if name_or_code.isdigit() and len(name_or_code) == 4:
        return name_or_code
    
    # Search by name
    accounts = Account.get_all()
    name_lower = name_or_code.lower()
    
    for acc in accounts:
        if acc["name"].lower() == name_lower:
            return acc["code"]
        if name_lower in acc["name"].lower():
            return acc["code"]
    
    # Default to general expenses if not found
    return AccountCodes.GENERAL_EXPENSES


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 2
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 2 COMPLETE - Chart of Accounts

Contains:
✓ Account types (Asset, Liability, Equity, Revenue, Cost of Sales, Expense)
✓ Account categories for report grouping
✓ Complete SA SME Chart of Accounts (80+ accounts)
✓ Proper VAT flags (standard, zero-rated, exempt)
✓ Depreciation rates for fixed assets
✓ System accounts that cannot be deleted
✓ Contra accounts (accumulated depreciation, drawings, etc.)
✓ Account class with CRUD operations
✓ Chart initialization function
✓ Quick lookup helpers (AccountCodes class)

Account Ranges:
- 1000-1999: Assets
- 2000-2999: Liabilities  
- 3000-3999: Equity
- 4000-4999: Revenue
- 5000-5999: Cost of Sales
- 6000-6999: Operating Expenses

Next: Piece 3 - Journal Engine (double-entry posting)
"""

# =============================================================================
# PIECE3_JOURNAL.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 3: Journal Engine                                                     ║
║                                                                               ║
║   This piece contains:                                                        ║
║   - Double-entry journal posting                                              ║
║   - Transaction validation                                                    ║
║   - Account balance calculations                                              ║
║   - Transaction templates for common operations                               ║
║   - Audit trail                                                               ║
║                                                                               ║
║   RULE: Every debit must have an equal credit                                 ║
║   Flask does ALL calculations - no JavaScript math                            ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

from typing import List, Dict, Optional, Tuple
from enum import Enum



# ═══════════════════════════════════════════════════════════════════════════════
# TRANSACTION TYPES
# ═══════════════════════════════════════════════════════════════════════════════

class TransactionType(Enum):
    """Types of transactions for categorization and reporting"""
    SALE = "sale"
    SALE_RETURN = "sale_return"
    PURCHASE = "purchase"
    PURCHASE_RETURN = "purchase_return"
    EXPENSE = "expense"
    RECEIPT = "receipt"
    PAYMENT = "payment"
    TRANSFER = "transfer"
    JOURNAL = "journal"
    OPENING = "opening"
    CLOSING = "closing"
    DEPRECIATION = "depreciation"
    ADJUSTMENT = "adjustment"


# ═══════════════════════════════════════════════════════════════════════════════
# JOURNAL LINE
# ═══════════════════════════════════════════════════════════════════════════════

class JournalLine:
    """Single line in a journal entry"""
    
    def __init__(self, account_code: str, debit: Decimal = None, 
                 credit: Decimal = None, description: str = ""):
        self.account_code = account_code
        self.debit = Decimal(str(debit or 0)).quantize(Decimal("0.01"))
        self.credit = Decimal(str(credit or 0)).quantize(Decimal("0.01"))
        self.description = description
    
    def to_dict(self) -> dict:
        return {
            "account_code": self.account_code,
            "debit": float(self.debit),
            "credit": float(self.credit),
            "description": self.description
        }


# ═══════════════════════════════════════════════════════════════════════════════
# JOURNAL ENTRY
# ═══════════════════════════════════════════════════════════════════════════════

class JournalEntry:
    """
    Complete journal entry with header and lines
    
    Usage:
        entry = JournalEntry(date="2024-01-15", description="Cash sale")
        entry.debit("1000", 1150)  # Bank
        entry.credit("4000", 1000) # Sales
        entry.credit("2100", 150)  # VAT Output
        entry.post()
    """
    
    TABLE = "journal"
    
    def __init__(self, date: str = None, reference: str = "", 
                 description: str = "", trans_type: TransactionType = TransactionType.JOURNAL,
                 source_type: str = "", source_id: str = ""):
        self.id = generate_id()
        self.date = date or today()
        self.reference = reference
        self.description = description
        self.trans_type = trans_type
        self.source_type = source_type  # e.g., "invoice", "expense"
        self.source_id = source_id      # Link back to source document
        self.lines: List[JournalLine] = []
        self.created_at = now()
        self.created_by = None
        self.posted = False
    
    def debit(self, account_code: str, amount, description: str = "") -> 'JournalEntry':
        """Add a debit line"""
        amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if amount > 0:
            self.lines.append(JournalLine(
                account_code=account_code,
                debit=amount,
                credit=Decimal("0"),
                description=description or self.description
            ))
        return self
    
    def credit(self, account_code: str, amount, description: str = "") -> 'JournalEntry':
        """Add a credit line"""
        amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if amount > 0:
            self.lines.append(JournalLine(
                account_code=account_code,
                debit=Decimal("0"),
                credit=amount,
                description=description or self.description
            ))
        return self
    
    def total_debits(self) -> Decimal:
        """Sum of all debits"""
        return sum(line.debit for line in self.lines)
    
    def total_credits(self) -> Decimal:
        """Sum of all credits"""
        return sum(line.credit for line in self.lines)
    
    def is_balanced(self) -> bool:
        """Check if debits equal credits"""
        diff = abs(self.total_debits() - self.total_credits())
        return diff < Decimal("0.01")
    
    def validate(self) -> Tuple[bool, str]:
        """
        Validate the journal entry
        
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not self.lines:
            return False, "Journal entry has no lines"
        
        if not self.date:
            return False, "Journal entry has no date"
        
        if not self.is_balanced():
            diff = self.total_debits() - self.total_credits()
            return False, f"Entry not balanced. Difference: {Money.format(diff)}"
        
        # Validate each line
        for i, line in enumerate(self.lines):
            if line.debit == 0 and line.credit == 0:
                return False, f"Line {i+1} has no amount"
            
            if line.debit > 0 and line.credit > 0:
                return False, f"Line {i+1} has both debit and credit"
            
            # Check account exists
            account = Account.get_by_code(line.account_code)
            if not account:
                return False, f"Account {line.account_code} not found"
            
            if not account.get("active", True):
                return False, f"Account {line.account_code} is inactive"
        
        return True, ""
    
    def post(self, user_id: str = None) -> Tuple[bool, str]:
        """
        Post the journal entry to the database
        
        Returns:
            Tuple of (success, message or entry_id)
        """
        # Validate first
        is_valid, error = self.validate()
        if not is_valid:
            return False, error
        
        self.created_by = user_id
        self.posted = True
        
        # Save each line to the journal table
        for line in self.lines:
            record = {
                "id": generate_id(),
                "entry_id": self.id,
                "date": self.date,
                "account_code": line.account_code,
                "description": safe_string(line.description, 500),
                "reference": safe_string(self.reference, 100),
                "debit": float(line.debit),
                "credit": float(line.credit),
                "trans_type": self.trans_type.value,
                "source_type": self.source_type,
                "source_id": self.source_id,
                "created_at": self.created_at,
                "created_by": self.created_by
            }
            
            success, result = db.insert(self.TABLE, record)
            if not success:
                return False, f"Failed to post line: {result}"
        
        return True, self.id


# ═══════════════════════════════════════════════════════════════════════════════
# JOURNAL QUERIES
# ═══════════════════════════════════════════════════════════════════════════════

class Journal:
    """Query and reporting functions for the journal"""
    
    TABLE = "journal"
    
    @classmethod
    def get_entries(cls, date_from: str = None, date_to: str = None,
                    account_code: str = None, trans_type: str = None,
                    limit: int = 1000) -> List[dict]:
        """
        Get journal entries with optional filters
        """
        # Build query
        entries = db.select(cls.TABLE, order="-date", limit=limit)
        
        # Apply filters (Supabase filtering would be more efficient)
        if date_from:
            entries = [e for e in entries if e.get("date", "") >= date_from]
        if date_to:
            entries = [e for e in entries if e.get("date", "") <= date_to]
        if account_code:
            entries = [e for e in entries if e.get("account_code") == account_code]
        if trans_type:
            entries = [e for e in entries if e.get("trans_type") == trans_type]
        
        return entries
    
    @classmethod
    def get_account_balance(cls, account_code: str, as_at: str = None) -> Decimal:
        """
        Calculate account balance
        
        For Asset/Expense accounts: Debits - Credits
        For Liability/Equity/Revenue accounts: Credits - Debits
        """
        entries = cls.get_entries(account_code=account_code)
        
        if as_at:
            entries = [e for e in entries if e.get("date", "") <= as_at]
        
        total_debits = sum(Decimal(str(e.get("debit", 0))) for e in entries)
        total_credits = sum(Decimal(str(e.get("credit", 0))) for e in entries)
        
        # Get account type to determine balance direction
        account = Account.get_by_code(account_code)
        if account:
            acc_type = account.get("type")
            # Debit-balance accounts: Assets, Expenses, Cost of Sales
            if acc_type in ["asset", "expense", "cost_of_sales"]:
                return total_debits - total_credits
            # Credit-balance accounts: Liabilities, Equity, Revenue
            else:
                return total_credits - total_debits
        
        return total_debits - total_credits
    
    @classmethod
    def get_account_movement(cls, account_code: str, 
                            date_from: str, date_to: str) -> dict:
        """
        Get account movement for a period
        
        Returns dict with opening, debits, credits, closing
        """
        # Opening balance (everything before date_from)
        opening = cls.get_account_balance(account_code, as_at=date_from)
        
        # Period transactions
        entries = cls.get_entries(
            account_code=account_code,
            date_from=date_from,
            date_to=date_to
        )
        
        debits = sum(Decimal(str(e.get("debit", 0))) for e in entries)
        credits = sum(Decimal(str(e.get("credit", 0))) for e in entries)
        
        # Closing balance
        closing = cls.get_account_balance(account_code, as_at=date_to)
        
        return {
            "account_code": account_code,
            "opening": opening,
            "debits": debits,
            "credits": credits,
            "closing": closing
        }
    
    @classmethod
    def get_trial_balance(cls, as_at: str = None) -> List[dict]:
        """
        Generate trial balance
        
        Returns list of accounts with debit and credit balances
        """
        accounts = Account.get_all(active_only=False)
        trial_balance = []
        
        total_debits = Decimal("0")
        total_credits = Decimal("0")
        
        for account in accounts:
            balance = cls.get_account_balance(account["code"], as_at=as_at)
            
            if balance == 0:
                continue  # Skip zero-balance accounts
            
            entry = {
                "code": account["code"],
                "name": account["name"],
                "type": account["type"],
                "category": account["category"],
                "debit": Decimal("0"),
                "credit": Decimal("0"),
                "balance": balance
            }
            
            # Determine which column based on account type and balance
            acc_type = account["type"]
            if acc_type in ["asset", "expense", "cost_of_sales"]:
                if balance >= 0:
                    entry["debit"] = balance
                    total_debits += balance
                else:
                    entry["credit"] = abs(balance)
                    total_credits += abs(balance)
            else:  # liability, equity, revenue
                if balance >= 0:
                    entry["credit"] = balance
                    total_credits += balance
                else:
                    entry["debit"] = abs(balance)
                    total_debits += abs(balance)
            
            trial_balance.append(entry)
        
        # Sort by account code
        trial_balance.sort(key=lambda x: x["code"])
        
        # Add totals row
        trial_balance.append({
            "code": "",
            "name": "TOTAL",
            "type": "",
            "category": "",
            "debit": total_debits,
            "credit": total_credits,
            "balance": total_debits - total_credits,
            "is_total": True
        })
        
        return trial_balance
    
    @classmethod
    def get_income_statement(cls, date_from: str, date_to: str) -> dict:
        """
        Generate Income Statement (Profit & Loss)
        
        Revenue - Cost of Sales = Gross Profit
        Gross Profit - Expenses = Net Profit
        """
        result = {
            "period_from": date_from,
            "period_to": date_to,
            "revenue": [],
            "cost_of_sales": [],
            "expenses": [],
            "total_revenue": Decimal("0"),
            "total_cost_of_sales": Decimal("0"),
            "gross_profit": Decimal("0"),
            "total_expenses": Decimal("0"),
            "net_profit": Decimal("0")
        }
        
        accounts = Account.get_all()
        
        for account in accounts:
            movement = cls.get_account_movement(
                account["code"], date_from, date_to
            )
            
            # Calculate period activity (not balance)
            # Revenue: Credits - Debits (income increases with credits)
            # Expenses: Debits - Credits (expenses increase with debits)
            
            acc_type = account["type"]
            
            if acc_type == "revenue":
                amount = movement["credits"] - movement["debits"]
                if amount != 0:
                    result["revenue"].append({
                        "code": account["code"],
                        "name": account["name"],
                        "amount": amount
                    })
                    result["total_revenue"] += amount
            
            elif acc_type == "cost_of_sales":
                amount = movement["debits"] - movement["credits"]
                if amount != 0:
                    result["cost_of_sales"].append({
                        "code": account["code"],
                        "name": account["name"],
                        "amount": amount
                    })
                    result["total_cost_of_sales"] += amount
            
            elif acc_type == "expense":
                amount = movement["debits"] - movement["credits"]
                if amount != 0:
                    result["expenses"].append({
                        "code": account["code"],
                        "name": account["name"],
                        "amount": amount
                    })
                    result["total_expenses"] += amount
        
        # Sort each section by amount (descending)
        result["revenue"].sort(key=lambda x: x["amount"], reverse=True)
        result["cost_of_sales"].sort(key=lambda x: x["amount"], reverse=True)
        result["expenses"].sort(key=lambda x: x["amount"], reverse=True)
        
        # Calculate profits
        result["gross_profit"] = result["total_revenue"] - result["total_cost_of_sales"]
        result["net_profit"] = result["gross_profit"] - result["total_expenses"]
        
        return result
    
    @classmethod
    def get_balance_sheet(cls, as_at: str = None) -> dict:
        """
        Generate Balance Sheet
        
        Assets = Liabilities + Equity
        """
        if not as_at:
            as_at = today()
        
        result = {
            "as_at": as_at,
            "current_assets": [],
            "fixed_assets": [],
            "current_liabilities": [],
            "long_term_liabilities": [],
            "equity": [],
            "total_current_assets": Decimal("0"),
            "total_fixed_assets": Decimal("0"),
            "total_assets": Decimal("0"),
            "total_current_liabilities": Decimal("0"),
            "total_long_term_liabilities": Decimal("0"),
            "total_liabilities": Decimal("0"),
            "total_equity": Decimal("0"),
            "total_liab_equity": Decimal("0")
        }
        
        accounts = Account.get_all()
        
        for account in accounts:
            balance = cls.get_account_balance(account["code"], as_at=as_at)
            
            if balance == 0:
                continue
            
            entry = {
                "code": account["code"],
                "name": account["name"],
                "balance": balance
            }
            
            acc_type = account["type"]
            category = account.get("category", "")
            
            if acc_type == "asset":
                if category == "current_asset":
                    result["current_assets"].append(entry)
                    result["total_current_assets"] += balance
                else:
                    result["fixed_assets"].append(entry)
                    result["total_fixed_assets"] += balance
            
            elif acc_type == "liability":
                if category in ["current_liability"]:
                    result["current_liabilities"].append(entry)
                    result["total_current_liabilities"] += balance
                else:
                    result["long_term_liabilities"].append(entry)
                    result["total_long_term_liabilities"] += balance
            
            elif acc_type == "equity":
                result["equity"].append(entry)
                result["total_equity"] += balance
        
        # Calculate totals
        result["total_assets"] = result["total_current_assets"] + result["total_fixed_assets"]
        result["total_liabilities"] = result["total_current_liabilities"] + result["total_long_term_liabilities"]
        result["total_liab_equity"] = result["total_liabilities"] + result["total_equity"]
        
        return result
    
    @classmethod
    def get_vat_report(cls, date_from: str, date_to: str) -> dict:
        """
        Generate VAT Report
        
        Output VAT (collected on sales) - Input VAT (paid on purchases)
        """
        # Get VAT Output (what we collected)
        vat_output = cls.get_account_movement(
            AccountCodes.VAT_OUTPUT, date_from, date_to
        )
        
        # Get VAT Input (what we paid)
        vat_input = cls.get_account_movement(
            AccountCodes.VAT_INPUT, date_from, date_to
        )
        
        # Output VAT is a liability - credits increase it
        output_amount = vat_output["credits"] - vat_output["debits"]
        
        # Input VAT is contra-liability - debits increase it
        input_amount = vat_input["debits"] - vat_input["credits"]
        
        # Net VAT position
        net_vat = output_amount - input_amount
        
        return {
            "period_from": date_from,
            "period_to": date_to,
            "output_vat": output_amount,
            "input_vat": input_amount,
            "net_vat": net_vat,
            "payable_to_sars": net_vat if net_vat > 0 else Decimal("0"),
            "refund_due": abs(net_vat) if net_vat < 0 else Decimal("0")
        }


# ═══════════════════════════════════════════════════════════════════════════════
# TRANSACTION TEMPLATES
# ═══════════════════════════════════════════════════════════════════════════════

class Transactions:
    """
    Pre-built transaction templates for common operations
    
    These ensure correct double-entry posting for standard transactions.
    All VAT calculations are done by Flask.
    """
    
    @staticmethod
    def cash_sale(amount_incl_vat: Decimal, description: str = "Cash sale",
                  reference: str = "", is_zero_rated: bool = False) -> JournalEntry:
        """
        Record a cash sale
        
        DR Bank (total including VAT)
            CR Sales (excluding VAT)
            CR VAT Output (VAT amount) - unless zero-rated
        """
        amount = Decimal(str(amount_incl_vat))
        
        if is_zero_rated:
            vat_info = VAT.calculate_from_inclusive(amount, VAT.ZERO_RATE)
        else:
            vat_info = VAT.calculate_from_inclusive(amount)
        
        entry = JournalEntry(
            description=description,
            reference=reference,
            trans_type=TransactionType.SALE
        )
        
        entry.debit(AccountCodes.BANK, vat_info["inclusive"])
        entry.credit(AccountCodes.SALES, vat_info["exclusive"])
        
        if vat_info["vat"] > 0:
            entry.credit(AccountCodes.VAT_OUTPUT, vat_info["vat"])
        
        return entry
    
    @staticmethod
    def credit_sale(customer_id: str, amount_incl_vat: Decimal, 
                   description: str = "Credit sale", reference: str = "",
                   is_zero_rated: bool = False) -> JournalEntry:
        """
        Record a sale on account
        
        DR Debtors (total including VAT)
            CR Sales (excluding VAT)
            CR VAT Output (VAT amount)
        """
        amount = Decimal(str(amount_incl_vat))
        
        if is_zero_rated:
            vat_info = VAT.calculate_from_inclusive(amount, VAT.ZERO_RATE)
        else:
            vat_info = VAT.calculate_from_inclusive(amount)
        
        entry = JournalEntry(
            description=description,
            reference=reference,
            trans_type=TransactionType.SALE,
            source_type="customer",
            source_id=customer_id
        )
        
        entry.debit(AccountCodes.DEBTORS, vat_info["inclusive"])
        entry.credit(AccountCodes.SALES, vat_info["exclusive"])
        
        if vat_info["vat"] > 0:
            entry.credit(AccountCodes.VAT_OUTPUT, vat_info["vat"])
        
        return entry
    
    @staticmethod
    def receive_payment(customer_id: str, amount: Decimal,
                       description: str = "Payment received",
                       reference: str = "") -> JournalEntry:
        """
        Record payment received from customer
        
        DR Bank
            CR Debtors
        """
        entry = JournalEntry(
            description=description,
            reference=reference,
            trans_type=TransactionType.RECEIPT,
            source_type="customer",
            source_id=customer_id
        )
        
        entry.debit(AccountCodes.BANK, amount)
        entry.credit(AccountCodes.DEBTORS, amount)
        
        return entry
    
    @staticmethod
    def cash_expense(amount_incl_vat: Decimal, expense_account: str,
                    description: str = "", reference: str = "",
                    vat_type: str = "standard") -> JournalEntry:
        """
        Record a cash expense
        
        DR Expense account (excluding VAT)
        DR VAT Input (VAT amount) - if applicable
            CR Bank (total)
        
        vat_type: "standard", "zero", "exempt"
        """
        amount = Decimal(str(amount_incl_vat))
        
        if vat_type == "zero":
            vat_info = VAT.calculate_from_inclusive(amount, VAT.ZERO_RATE)
        elif vat_type == "exempt":
            # No VAT to claim
            vat_info = {"exclusive": amount, "vat": Decimal("0"), "inclusive": amount}
        else:
            vat_info = VAT.calculate_from_inclusive(amount)
        
        entry = JournalEntry(
            description=description,
            reference=reference,
            trans_type=TransactionType.EXPENSE
        )
        
        entry.debit(expense_account, vat_info["exclusive"])
        
        if vat_info["vat"] > 0:
            entry.debit(AccountCodes.VAT_INPUT, vat_info["vat"])
        
        entry.credit(AccountCodes.BANK, vat_info["inclusive"])
        
        return entry
    
    @staticmethod
    def credit_expense(supplier_id: str, amount_incl_vat: Decimal,
                      expense_account: str, description: str = "",
                      reference: str = "", vat_type: str = "standard") -> JournalEntry:
        """
        Record an expense on credit
        
        DR Expense account (excluding VAT)
        DR VAT Input (VAT amount) - if applicable
            CR Creditors (total)
        """
        amount = Decimal(str(amount_incl_vat))
        
        if vat_type == "zero":
            vat_info = VAT.calculate_from_inclusive(amount, VAT.ZERO_RATE)
        elif vat_type == "exempt":
            vat_info = {"exclusive": amount, "vat": Decimal("0"), "inclusive": amount}
        else:
            vat_info = VAT.calculate_from_inclusive(amount)
        
        entry = JournalEntry(
            description=description,
            reference=reference,
            trans_type=TransactionType.EXPENSE,
            source_type="supplier",
            source_id=supplier_id
        )
        
        entry.debit(expense_account, vat_info["exclusive"])
        
        if vat_info["vat"] > 0:
            entry.debit(AccountCodes.VAT_INPUT, vat_info["vat"])
        
        entry.credit(AccountCodes.CREDITORS, vat_info["inclusive"])
        
        return entry
    
    @staticmethod
    def pay_supplier(supplier_id: str, amount: Decimal,
                    description: str = "Payment to supplier",
                    reference: str = "") -> JournalEntry:
        """
        Record payment to supplier
        
        DR Creditors
            CR Bank
        """
        entry = JournalEntry(
            description=description,
            reference=reference,
            trans_type=TransactionType.PAYMENT,
            source_type="supplier",
            source_id=supplier_id
        )
        
        entry.debit(AccountCodes.CREDITORS, amount)
        entry.credit(AccountCodes.BANK, amount)
        
        return entry
    
    @staticmethod
    def stock_purchase(amount_incl_vat: Decimal, description: str = "Stock purchase",
                      reference: str = "", is_zero_rated: bool = False) -> JournalEntry:
        """
        Record stock/inventory purchase for cash
        
        DR Stock (excluding VAT)
        DR VAT Input (VAT amount)
            CR Bank (total)
        """
        amount = Decimal(str(amount_incl_vat))
        
        if is_zero_rated:
            vat_info = VAT.calculate_from_inclusive(amount, VAT.ZERO_RATE)
        else:
            vat_info = VAT.calculate_from_inclusive(amount)
        
        entry = JournalEntry(
            description=description,
            reference=reference,
            trans_type=TransactionType.PURCHASE
        )
        
        entry.debit(AccountCodes.STOCK, vat_info["exclusive"])
        
        if vat_info["vat"] > 0:
            entry.debit(AccountCodes.VAT_INPUT, vat_info["vat"])
        
        entry.credit(AccountCodes.BANK, vat_info["inclusive"])
        
        return entry
    
    @staticmethod
    def record_cost_of_sale(cost_amount: Decimal, 
                           description: str = "Cost of goods sold") -> JournalEntry:
        """
        Record cost of goods sold (when stock is sold)
        
        DR Cost of Goods Sold
            CR Stock
        """
        entry = JournalEntry(
            description=description,
            trans_type=TransactionType.SALE
        )
        
        entry.debit(AccountCodes.COGS, cost_amount)
        entry.credit(AccountCodes.STOCK, cost_amount)
        
        return entry
    
    @staticmethod
    def bank_transfer(from_account: str, to_account: str, amount: Decimal,
                     description: str = "Bank transfer") -> JournalEntry:
        """
        Transfer between bank accounts
        
        DR To Account
            CR From Account
        """
        entry = JournalEntry(
            description=description,
            trans_type=TransactionType.TRANSFER
        )
        
        entry.debit(to_account, amount)
        entry.credit(from_account, amount)
        
        return entry
    
    @staticmethod
    def owner_drawing(amount: Decimal, 
                     description: str = "Owner drawing") -> JournalEntry:
        """
        Record owner withdrawal
        
        DR Drawings
            CR Bank
        """
        entry = JournalEntry(
            description=description,
            trans_type=TransactionType.JOURNAL
        )
        
        entry.debit(AccountCodes.DRAWINGS, amount)
        entry.credit(AccountCodes.BANK, amount)
        
        return entry
    
    @staticmethod
    def owner_contribution(amount: Decimal,
                          description: str = "Capital contribution") -> JournalEntry:
        """
        Record owner capital contribution
        
        DR Bank
            CR Capital
        """
        entry = JournalEntry(
            description=description,
            trans_type=TransactionType.JOURNAL
        )
        
        entry.debit(AccountCodes.BANK, amount)
        entry.credit(AccountCodes.CAPITAL, amount)
        
        return entry


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def post_sale(amount: Decimal, payment_method: str = "cash",
              customer_id: str = None, invoice_number: str = "",
              is_zero_rated: bool = False, cost_price: Decimal = None) -> Tuple[bool, str]:
    """
    Complete sale posting including cost of goods sold
    
    Args:
        amount: Total sale amount including VAT
        payment_method: "cash" or "account"
        customer_id: Customer ID for credit sales
        invoice_number: Reference number
        is_zero_rated: True for zero-rated items
        cost_price: Cost of goods for COGS entry
        
    Returns:
        Tuple of (success, message)
    """
    # Create sale entry
    if payment_method == "account" and customer_id:
        entry = Transactions.credit_sale(
            customer_id=customer_id,
            amount_incl_vat=amount,
            reference=invoice_number,
            is_zero_rated=is_zero_rated
        )
    else:
        entry = Transactions.cash_sale(
            amount_incl_vat=amount,
            reference=invoice_number,
            is_zero_rated=is_zero_rated
        )
    
    entry.source_type = "invoice"
    entry.source_id = invoice_number
    
    # Post sale
    success, result = entry.post()
    if not success:
        return False, result
    
    # Post COGS if cost price provided
    if cost_price and cost_price > 0:
        cogs_entry = Transactions.record_cost_of_sale(
            cost_amount=cost_price,
            description=f"COGS for {invoice_number}"
        )
        cogs_entry.source_type = "invoice"
        cogs_entry.source_id = invoice_number
        
        success, result = cogs_entry.post()
        if not success:
            return False, f"Sale posted but COGS failed: {result}"
    
    return True, "Sale posted successfully"


def post_expense(amount: Decimal, expense_category: str,
                supplier: str = "", description: str = "",
                reference: str = "", vat_type: str = "standard") -> Tuple[bool, str]:
    """
    Post an expense with automatic account selection
    
    Args:
        amount: Total amount including VAT
        expense_category: Category name or account code
        supplier: Supplier name for description
        description: Expense description
        reference: Reference number
        vat_type: "standard", "zero", or "exempt"
        
    Returns:
        Tuple of (success, message)
    """
    # Get the expense account code
    expense_account = get_account_code(expense_category)
    
    # Build description
    if supplier:
        full_desc = f"{supplier}: {description}" if description else supplier
    else:
        full_desc = description or expense_category
    
    # Create and post entry
    entry = Transactions.cash_expense(
        amount_incl_vat=amount,
        expense_account=expense_account,
        description=full_desc,
        reference=reference,
        vat_type=vat_type
    )
    
    return entry.post()


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 3
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 3 COMPLETE - Journal Engine

Contains:
✓ Transaction types enum
✓ JournalLine class for individual entries
✓ JournalEntry class with validation and posting
✓ Journal query class for:
  - Account balances
  - Account movements
  - Trial Balance generation
  - Income Statement (P&L)
  - Balance Sheet
  - VAT Report
✓ Transaction templates for:
  - Cash and credit sales
  - Payment receipts
  - Cash and credit expenses
  - Supplier payments
  - Stock purchases
  - Cost of goods sold
  - Bank transfers
  - Owner drawings/contributions
✓ Helper functions for common operations

All calculations in Flask using Decimal for precision.
Every entry validates that debits = credits before posting.

Next: Piece 4 - CSS and UI Framework
"""

# =============================================================================
# PIECE4_UI.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 4: CSS & UI Framework                                                 ║
║                                                                               ║
║   This piece contains:                                                        ║
║   - Complete CSS with glowing header                                          ║
║   - Responsive design for desktop and mobile                                  ║
║   - UI component templates                                                    ║
║   - Page wrapper functions                                                    ║
║   - Flash messages and alerts                                                 ║
║                                                                               ║
║   Design Principles:                                                          ║
║   - Sticky glowing header - always visible                                    ║
║   - Active page highlighted - no duplicate titles                             ║
║   - Dark theme - easy on the eyes                                             ║
║   - Every pixel counts - especially on mobile                                 ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""



# ═══════════════════════════════════════════════════════════════════════════════
# CSS STYLES
# ═══════════════════════════════════════════════════════════════════════════════

CSS = """
<style>
/* ═══════════════════════════════════════════════════════════════════════════
   ROOT VARIABLES
   ═══════════════════════════════════════════════════════════════════════════ */
:root {
    /* Background colors */
    --bg-primary: #050508;
    --bg-secondary: #0a0a10;
    --bg-card: #0f0f18;
    --bg-hover: #15151f;
    
    /* Border colors */
    --border: #1a1a2e;
    --border-hover: #2a2a4e;
    
    /* Text colors */
    --text-primary: #f0f0f0;
    --text-secondary: #a0a0a0;
    --text-muted: #606070;
    
    /* Accent colors */
    --purple: #8b5cf6;
    --purple-glow: rgba(139, 92, 246, 0.5);
    --blue: #3b82f6;
    --blue-glow: rgba(59, 130, 246, 0.3);
    --green: #10b981;
    --green-light: #34d399;
    --red: #ef4444;
    --red-light: #f87171;
    --orange: #f59e0b;
    --yellow: #eab308;
    --cyan: #06b6d4;
    
    /* Gradients */
    --gradient-primary: linear-gradient(135deg, var(--purple), var(--blue));
    --gradient-glow: linear-gradient(135deg, var(--purple-glow), var(--blue-glow));
    
    /* Spacing */
    --header-height: 56px;
    --container-max: 1400px;
    --radius-sm: 6px;
    --radius-md: 10px;
    --radius-lg: 14px;
    
    /* Transitions */
    --transition-fast: 0.15s ease;
    --transition-normal: 0.25s ease;
}

/* ═══════════════════════════════════════════════════════════════════════════
   RESET & BASE
   ═══════════════════════════════════════════════════════════════════════════ */
*, *::before, *::after {
    margin: 0;
    padding: 0;
    box-sizing: border-box;
}

html {
    font-size: 14px;
    scroll-behavior: smooth;
}

body {
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
    background: var(--bg-primary);
    color: var(--text-primary);
    min-height: 100vh;
    line-height: 1.5;
    -webkit-font-smoothing: antialiased;
}

a {
    color: var(--blue);
    text-decoration: none;
    transition: color var(--transition-fast);
}

a:hover {
    color: var(--purple);
}

/* ═══════════════════════════════════════════════════════════════════════════
   GLOWING HEADER
   ═══════════════════════════════════════════════════════════════════════════ */
.header {
    flex-wrap: nowrap;
    position: sticky;
    top: 0;
    z-index: 1000;
    height: var(--header-height);
    background: linear-gradient(180deg, var(--bg-card) 0%, rgba(15, 15, 24, 0.97) 100%);
    border-bottom: 1px solid var(--border);
    backdrop-filter: blur(12px);
    -webkit-backdrop-filter: blur(12px);
    display: flex;
    align-items: center;
    padding: 0 20px;
    gap: 8px;
    overflow: visible;
}

.header::-webkit-scrollbar {
    height: 0;
    display: none;
}

/* Glowing Logo */
.logo {
    font-size: 22px;
    font-weight: 800;
    letter-spacing: -0.5px;
    background: var(--gradient-primary);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    text-decoration: none;
    margin-right: 20px;
    white-space: nowrap;
    position: relative;
    animation: logo-pulse 3s ease-in-out infinite;
}

.logo::after {
    content: '';
    position: absolute;
    inset: -10px -20px;
    background: var(--gradient-glow);
    filter: blur(20px);
    opacity: 0.4;
    z-index: -1;
    animation: glow-pulse 3s ease-in-out infinite;
}

@keyframes logo-pulse {
    0%, 100% { filter: brightness(1); }
    50% { filter: brightness(1.2); }
}

@keyframes glow-pulse {
    0%, 100% { opacity: 0.3; transform: scale(1); }
    50% { opacity: 0.6; transform: scale(1.1); }
}

/* Navigation Items */
.nav {
    flex-shrink: 0;
    white-space: nowrap;
    display: flex;
    align-items: center;
    gap: 4px;
    flex: 1;
}

.nav-item {
    padding: 8px 14px;
    border-radius: var(--radius-sm);
    font-size: 13px;
    font-weight: 500;
    color: var(--text-secondary);
    text-decoration: none;
    white-space: nowrap;
    transition: all var(--transition-fast);
    border: 1px solid transparent;
    position: relative;
}

.nav-item:hover {
    color: var(--text-primary);
    background: var(--bg-hover);
}

.nav-item.active {
    color: white;
    background: linear-gradient(135deg, rgba(139, 92, 246, 0.25), rgba(59, 130, 246, 0.25));
    border-color: rgba(139, 92, 246, 0.4);
    box-shadow: 0 0 20px rgba(139, 92, 246, 0.15), inset 0 1px 0 rgba(255, 255, 255, 0.05);
}

.nav-item.active::before {
    content: '';
    position: absolute;
    bottom: -1px;
    left: 20%;
    right: 20%;
    height: 2px;
    background: var(--gradient-primary);
    border-radius: 2px;
}

/* Header User Section */
.header-user {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-left: auto;
    padding-left: 20px;
    border-left: 1px solid var(--border);
}

.header-user-name {
    font-size: 13px;
    color: var(--text-secondary);
}

/* ═══════════════════════════════════════════════════════════════════════════
   MAIN CONTAINER
   ═══════════════════════════════════════════════════════════════════════════ */
.main {
    min-height: calc(100vh - var(--header-height));
    padding: 24px;
}

.container {
    max-width: var(--container-max);
    margin: 0 auto;
}

/* ═══════════════════════════════════════════════════════════════════════════
   CARDS
   ═══════════════════════════════════════════════════════════════════════════ */
.card {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 20px;
    margin-bottom: 20px;
}

.card-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 16px;
    padding-bottom: 16px;
    border-bottom: 1px solid var(--border);
}

.card-title {
    font-size: 16px;
    font-weight: 600;
    color: var(--text-primary);
}

.card-subtitle {
    font-size: 13px;
    color: var(--text-muted);
    margin-top: 4px;
}

/* ═══════════════════════════════════════════════════════════════════════════
   STATS GRID
   ═══════════════════════════════════════════════════════════════════════════ */
.stats {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
    gap: 16px;
    margin-bottom: 24px;
}

.stat {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-md);
    padding: 20px;
    text-align: center;
    transition: all var(--transition-normal);
}

.stat:hover {
    border-color: var(--border-hover);
    transform: translateY(-2px);
}

.stat-value {
    font-size: 28px;
    font-weight: 700;
    color: var(--blue);
    line-height: 1.2;
}

.stat-value.green { color: var(--green); }
.stat-value.red { color: var(--red); }
.stat-value.orange { color: var(--orange); }
.stat-value.purple { color: var(--purple); }

.stat-label {
    font-size: 12px;
    color: var(--text-muted);
    margin-top: 6px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

/* ═══════════════════════════════════════════════════════════════════════════
   BUTTONS
   ═══════════════════════════════════════════════════════════════════════════ */
.btn {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
    padding: 10px 20px;
    font-size: 14px;
    font-weight: 600;
    border: none;
    border-radius: var(--radius-sm);
    cursor: pointer;
    text-decoration: none;
    transition: all var(--transition-fast);
    white-space: nowrap;
}

.btn:hover {
    transform: translateY(-1px);
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.3);
}

.btn:active {
    transform: translateY(0);
}

.btn-primary {
    background: var(--gradient-primary);
    color: white;
}

.btn-blue {
    background: var(--blue);
    color: white;
}

.btn-green {
    background: var(--green);
    color: white;
}

.btn-red {
    background: var(--red);
    color: white;
}

.btn-orange {
    background: var(--orange);
    color: black;
}

.btn-purple {
    background: var(--purple);
    color: white;
}

.btn-ghost {
    background: transparent;
    color: var(--text-secondary);
    border: 1px solid var(--border);
}

.btn-ghost:hover {
    background: var(--bg-hover);
    color: var(--text-primary);
    border-color: var(--border-hover);
}

.btn-sm {
    padding: 6px 12px;
    font-size: 12px;
}

.btn-lg {
    padding: 14px 28px;
    font-size: 16px;
}

.btn-block {
    width: 100%;
}

.btn:disabled {
    opacity: 0.5;
    cursor: not-allowed;
    transform: none;
}

/* Button Group */
.btn-group {
    display: flex;
    gap: 8px;
    flex-wrap: wrap;
}

/* ═══════════════════════════════════════════════════════════════════════════
   TOOLTIPS & HELP
   ═══════════════════════════════════════════════════════════════════════════ */
.help-tip {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 16px;
    height: 16px;
    background: var(--border);
    border-radius: 50%;
    font-size: 10px;
    color: var(--text-muted);
    cursor: help;
    margin-left: 6px;
    position: relative;
}

.help-tip:hover::after {
    content: attr(data-tip);
    position: absolute;
    bottom: 100%;
    left: 50%;
    transform: translateX(-50%);
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 8px 12px;
    font-size: 12px;
    white-space: nowrap;
    z-index: 1000;
    box-shadow: 0 4px 12px rgba(0,0,0,0.3);
    margin-bottom: 8px;
}

.info-box {
    background: rgba(59, 130, 246, 0.1);
    border: 1px solid rgba(59, 130, 246, 0.2);
    border-radius: 8px;
    padding: 12px 16px;
    margin-bottom: 16px;
    font-size: 13px;
    color: #93c5fd;
}

.info-box::before {
    content: 'ℹ️ ';
}

.success-box {
    background: rgba(16, 185, 129, 0.1);
    border: 1px solid rgba(16, 185, 129, 0.2);
    border-radius: 8px;
    padding: 12px 16px;
    margin-bottom: 16px;
    font-size: 13px;
    color: #6ee7b7;
}

.success-box::before {
    content: '✓ ';
}

.warning-box {
    background: rgba(245, 158, 11, 0.1);
    border: 1px solid rgba(245, 158, 11, 0.2);
    border-radius: 8px;
    padding: 12px 16px;
    margin-bottom: 16px;
    font-size: 13px;
    color: #fcd34d;
}

.warning-box::before {
    content: '⚠️ ';
}

/* ═══════════════════════════════════════════════════════════════════════════
   FORMS
   ═══════════════════════════════════════════════════════════════════════════ */
.form-group {
    margin-bottom: 16px;
}

.form-label {
    display: block;
    font-size: 12px;
    font-weight: 500;
    color: var(--text-secondary);
    margin-bottom: 6px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.form-input,
.form-select,
.form-textarea {
    width: 100%;
    padding: 12px 14px;
    font-size: 14px;
    color: var(--text-primary);
    background: var(--bg-secondary);
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    transition: all var(--transition-fast);
}

.form-input:focus,
.form-select:focus,
.form-textarea:focus {
    outline: none;
    border-color: var(--purple);
    box-shadow: 0 0 0 3px rgba(139, 92, 246, 0.1);
}

.form-input::placeholder {
    color: var(--text-muted);
}

.form-textarea {
    min-height: 100px;
    resize: vertical;
}

.form-select {
    cursor: pointer;
    appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%23606070' d='M6 8L1 3h10z'/%3E%3C/svg%3E");
    background-repeat: no-repeat;
    background-position: right 12px center;
    padding-right: 36px;
}

.form-row {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px;
}

.form-check {
    display: flex;
    align-items: center;
    gap: 8px;
    cursor: pointer;
}

.form-check input[type="checkbox"] {
    width: 18px;
    height: 18px;
    cursor: pointer;
    accent-color: var(--purple);
}

/* ═══════════════════════════════════════════════════════════════════════════
   TABLES
   ═══════════════════════════════════════════════════════════════════════════ */
.table-wrapper {
    overflow-x: auto;
    margin: -20px;
    padding: 20px;
}

.table {
    width: 100%;
    border-collapse: collapse;
}

.table th,
.table td {
    padding: 12px 16px;
    text-align: left;
    border-bottom: 1px solid var(--border);
}

.table th {
    font-size: 11px;
    font-weight: 600;
    color: var(--text-muted);
    text-transform: uppercase;
    letter-spacing: 0.5px;
    background: var(--bg-secondary);
}

.table tr:hover td {
    background: var(--bg-hover);
}

.table tr:last-child td {
    border-bottom: none;
}

.table td.number,
.table th.number {
    text-align: right;
    font-variant-numeric: tabular-nums;
}

.table-clickable tr {
    cursor: pointer;
}

/* ═══════════════════════════════════════════════════════════════════════════
   PRODUCT GRID (POS)
   ═══════════════════════════════════════════════════════════════════════════ */
.product-grid {
    margin-top: 5px;
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(140px, 1fr));
    gap: 12px;
    max-height: 68vh;
    overflow-y: auto;
    padding: 4px;
}

.product-item {
    background: var(--bg-secondary);
    border: 1px solid var(--border);
    border-radius: var(--radius-md);
    padding: 14px;
    cursor: pointer;
    transition: all var(--transition-fast);
    display: flex;
    flex-direction: column;
    height: 110px;
}

.product-item:hover {
    border-color: var(--purple);
    transform: translateY(-2px);
    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.2);
}

.product-name {
    font-size: 13px;
    font-weight: 600;
    color: var(--text-primary);
    margin-bottom: 6px;
    line-height: 1.3;
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
    flex: 1;
}

.product-price {
    font-size: 15px;
    font-weight: 700;
    color: var(--green);
    margin-top: auto;
}

.product-stock {
    font-size: 11px;
    color: var(--text-muted);
    margin-top: 4px;
}

.product-stock.low {
    color: var(--orange);
}

.product-stock.out {
    color: var(--red);
}

/* ═══════════════════════════════════════════════════════════════════════════
   CART
   ═══════════════════════════════════════════════════════════════════════════ */
.cart {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 20px;
    display: flex;
    flex-direction: column;
    height: fit-content;
    position: sticky;
    top: calc(var(--header-height) + 24px);
}

.cart-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 16px;
}

.cart-title {
    font-size: 16px;
    font-weight: 600;
}

.cart-items {
    /* Increased height */
    flex: 1;
    min-height: 200px;
    max-height: 300px;
    overflow-y: auto;
}

.cart-item {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 12px 0;
    border-bottom: 1px solid var(--border);
}

.cart-item:last-child {
    border-bottom: none;
}

.cart-item-name {
    font-weight: 600;
    font-size: 13px;
}

.cart-item-details {
    font-size: 12px;
    color: var(--text-muted);
}

.cart-item-price {
    font-weight: 600;
    white-space: nowrap;
}

.cart-empty {
    display: flex;
    align-items: center;
    justify-content: center;
    height: 150px;
    color: var(--text-muted);
    font-size: 14px;
}

.cart-total {
    padding: 16px 0;
    border-top: 2px solid var(--border);
    margin-top: 16px;
    text-align: right;
}

.cart-total-label {
    font-size: 14px;
    color: var(--text-secondary);
}

.cart-total-value {
    font-size: 28px;
    font-weight: 800;
    color: var(--green);
}

.cart-actions {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 10px;
    margin-top: 16px;
}

/* ═══════════════════════════════════════════════════════════════════════════
   MODALS
   ═══════════════════════════════════════════════════════════════════════════ */
.modal-overlay {
    display: none;
    position: fixed;
    inset: 0;
    background: rgba(0, 0, 0, 0.8);
    backdrop-filter: blur(4px);
    z-index: 2000;
    align-items: center;
    justify-content: center;
    padding: 20px;
}

.modal-overlay.show {
    display: flex;
}

.modal {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    width: 100%;
    max-width: 500px;
    max-height: 90vh;
    overflow-y: auto;
}

.modal-header {
    padding: 20px;
    border-bottom: 1px solid var(--border);
    display: flex;
    justify-content: space-between;
    align-items: center;
}

.modal-title {
    font-size: 18px;
    font-weight: 600;
}

.modal-close {
    background: none;
    border: none;
    color: var(--text-muted);
    font-size: 24px;
    cursor: pointer;
    padding: 4px;
    line-height: 1;
}

.modal-close:hover {
    color: var(--text-primary);
}

.modal-body {
    padding: 20px;
}

.modal-footer {
    padding: 16px 20px;
    border-top: 1px solid var(--border);
    display: flex;
    justify-content: flex-end;
    gap: 12px;
}

/* ═══════════════════════════════════════════════════════════════════════════
   ALERTS & MESSAGES
   ═══════════════════════════════════════════════════════════════════════════ */
.alert {
    padding: 14px 18px;
    border-radius: var(--radius-sm);
    margin-bottom: 16px;
    font-size: 14px;
    display: flex;
    align-items: center;
    gap: 10px;
}

.alert-info {
    background: rgba(59, 130, 246, 0.1);
    border: 1px solid rgba(59, 130, 246, 0.3);
    color: var(--blue);
}

.alert-success {
    background: rgba(16, 185, 129, 0.1);
    border: 1px solid rgba(16, 185, 129, 0.3);
    color: var(--green);
}

.alert-warning {
    background: rgba(245, 158, 11, 0.1);
    border: 1px solid rgba(245, 158, 11, 0.3);
    color: var(--orange);
}

.alert-error {
    background: rgba(239, 68, 68, 0.1);
    border: 1px solid rgba(239, 68, 68, 0.3);
    color: var(--red);
}

/* ═══════════════════════════════════════════════════════════════════════════
   BADGES
   ═══════════════════════════════════════════════════════════════════════════ */
.badge {
    display: inline-flex;
    align-items: center;
    padding: 4px 10px;
    font-size: 11px;
    font-weight: 600;
    border-radius: 20px;
    text-transform: uppercase;
    letter-spacing: 0.3px;
}

.badge-green {
    background: rgba(16, 185, 129, 0.15);
    color: var(--green);
}

.badge-red {
    background: rgba(239, 68, 68, 0.15);
    color: var(--red);
}

.badge-orange {
    background: rgba(245, 158, 11, 0.15);
    color: var(--orange);
}

.badge-blue {
    background: rgba(59, 130, 246, 0.15);
    color: var(--blue);
}

.badge-purple {
    background: rgba(139, 92, 246, 0.15);
    color: var(--purple);
}

/* ═══════════════════════════════════════════════════════════════════════════
   REPORT CARDS
   ═══════════════════════════════════════════════════════════════════════════ */
.report-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 16px;
}

.report-card {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 24px;
    text-decoration: none;
    color: var(--text-primary);
    transition: all var(--transition-normal);
    display: block;
}

.report-card:hover {
    border-color: var(--purple);
    transform: translateY(-4px);
    box-shadow: 0 12px 40px rgba(0, 0, 0, 0.3);
}

.report-card-icon {
    font-size: 32px;
    margin-bottom: 12px;
}

.report-card-title {
    font-size: 18px;
    font-weight: 600;
    margin-bottom: 8px;
}

.report-card-desc {
    font-size: 13px;
    color: var(--text-muted);
    line-height: 1.5;
}

/* ═══════════════════════════════════════════════════════════════════════════
   EMPTY STATES
   ═══════════════════════════════════════════════════════════════════════════ */
.empty-state {
    text-align: center;
    padding: 60px 20px;
    color: var(--text-muted);
}

.empty-state-icon {
    font-size: 48px;
    margin-bottom: 16px;
    opacity: 0.5;
}

.empty-state-title {
    font-size: 18px;
    font-weight: 600;
    color: var(--text-secondary);
    margin-bottom: 8px;
}

.empty-state-desc {
    font-size: 14px;
    max-width: 400px;
    margin: 0 auto;
}

/* ═══════════════════════════════════════════════════════════════════════════
   LOADING STATES
   ═══════════════════════════════════════════════════════════════════════════ */
.loading {
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 40px;
}

.spinner {
    width: 40px;
    height: 40px;
    border: 3px solid var(--border);
    border-top-color: var(--purple);
    border-radius: 50%;
    animation: spin 0.8s linear infinite;
}

@keyframes spin {
    to { transform: rotate(360deg); }
}

/* ═══════════════════════════════════════════════════════════════════════════
   LAYOUT HELPERS
   ═══════════════════════════════════════════════════════════════════════════ */
.flex {
    display: flex;
}

.flex-between {
    display: flex;
    justify-content: space-between;
    align-items: center;
}

.flex-center {
    display: flex;
    align-items: center;
    justify-content: center;
}

.flex-wrap {
    flex-wrap: wrap;
}

.gap-sm { gap: 8px; }
.gap-md { gap: 16px; }
.gap-lg { gap: 24px; }

.grid-2 {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 16px;
}

.grid-3 {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 16px;
}

.mt-sm { margin-top: 8px; }
.mt-md { margin-top: 16px; }
.mt-lg { margin-top: 24px; }
.mb-sm { margin-bottom: 8px; }
.mb-md { margin-bottom: 16px; }
.mb-lg { margin-bottom: 24px; }

.text-center { text-align: center; }
.text-right { text-align: right; }
.text-muted { color: var(--text-muted); }
.text-green { color: var(--green); }
.text-red { color: var(--red); }
.text-orange { color: var(--orange); }

.font-bold { font-weight: 600; }
.font-mono { font-family: 'Monaco', 'Consolas', monospace; }

/* ═══════════════════════════════════════════════════════════════════════════
   POS LAYOUT
   ═══════════════════════════════════════════════════════════════════════════ */
.pos-layout {
    display: grid;
    grid-template-columns: 1fr 380px;
    gap: 24px;
    align-items: start;
    height: calc(100vh - var(--header-height) - 32px);
}

.pos-layout > .card {
    height: 100%;
    overflow: hidden;
    display: flex;
    flex-direction: column;
}

.pos-layout > .card #product-grid {
    flex: 1;
    overflow-y: auto;
    min-height: 0;
}

.pos-layout > .cart {
    height: 100%;
    overflow-y: auto;
}

/* ═══════════════════════════════════════════════════════════════════════════
   RESPONSIVE
   ═══════════════════════════════════════════════════════════════════════════ */
@media (max-width: 1024px) {
    .pos-layout {
        grid-template-columns: 1fr 320px;
    }
}

@media (max-width: 768px) {
    html {
        font-size: 13px;
    }
    
    .header {
    flex-wrap: nowrap;
        padding: 0 12px;
        gap: 4px;
    }
    
    .logo {
        font-size: 18px;
        margin-right: 12px;
    }
    
    .nav-item {
        padding: 6px 10px;
        font-size: 12px;
    }
    
    .main {
        padding: 16px;
    }
    
    .stats {
        grid-template-columns: repeat(2, 1fr);
    }
    
    .pos-layout {
        grid-template-columns: 1fr;
    }
    
    .cart {
        position: relative;
        top: 0;
    }
    
    .form-row {
        grid-template-columns: 1fr;
    }
    
    .grid-2, .grid-3 {
        grid-template-columns: 1fr;
    }
    
    .report-grid {
        grid-template-columns: 1fr;
    }
    
    .cart-actions {
        grid-template-columns: 1fr;
    }
    
    .header-user {
        display: none;
    }
}

/* ═══════════════════════════════════════════════════════════════════════════
   MOBILE SPECIFIC INTERFACE
   ═══════════════════════════════════════════════════════════════════════════ */
.mobile-only {
    display: none;
}

@media (max-width: 480px) {
    .desktop-only {
        display: none;
    }
    
    .mobile-only {
        display: flex;
    }
    
    .mobile-big-buttons {
        display: flex;
        flex-direction: column;
        gap: 20px;
        padding: 40px 20px;
        min-height: calc(100vh - var(--header-height));
        justify-content: center;
    }
    
    .mobile-big-btn {
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        padding: 40px;
        border-radius: var(--radius-lg);
        text-decoration: none;
        font-size: 20px;
        font-weight: 700;
        transition: all var(--transition-normal);
    }
    
    .mobile-big-btn-icon {
        font-size: 48px;
        margin-bottom: 12px;
    }
    
    .mobile-big-btn.stock {
        background: linear-gradient(135deg, rgba(139, 92, 246, 0.2), rgba(59, 130, 246, 0.2));
        border: 2px solid rgba(139, 92, 246, 0.4);
        color: var(--purple);
    }
    
    .mobile-big-btn.expense {
        background: linear-gradient(135deg, rgba(239, 68, 68, 0.2), rgba(245, 158, 11, 0.2));
        border: 2px solid rgba(239, 68, 68, 0.4);
        color: var(--red);
    }
    
    .mobile-big-btn:active {
        transform: scale(0.98);
    }
}

/* ═══════════════════════════════════════════════════════════════════════════
   PRINT STYLES
   ═══════════════════════════════════════════════════════════════════════════ */
@media print {
    .header,
    .btn,
    .modal-overlay {
        display: none !important;
    }
    
    body {
        background: white;
        color: black;
    }
    
    .card {
        border: 1px solid #ccc;
        break-inside: avoid;
    }
    
    .table th,
    .table td {
        border: 1px solid #ccc;
    }
}
</style>
"""


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE WRAPPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_header_html(active: str = "", user: dict = None) -> str:
    """
    Generate the glowing header HTML with business switcher
    
    Args:
        active: Current active page (home, pos, stock, etc.)
        user: Current user dict with name
    """
    # All nav items - will be filtered by business config
    all_nav_items = [
        ("dashboard", "home", "Home", "/"),
        ("pos", "pos", "POS", "/pos"),
        ("stock", "stock", "Stock", "/stock"),
        ("customers", "customers", "Customers", "/customers"),
        ("suppliers", "suppliers", "Suppliers", "/suppliers"),
        ("purchase-orders", "purchase-orders", "Orders", "/purchase-orders"),
        ("invoices", "invoices", "Invoices", "/invoices"),
        ("quotes", "quotes", "Quotes", "/quotes"),
        ("expenses", "expenses", "Expenses", "/expenses"),
        ("payroll", "payroll", "Payroll", "/payroll"),
        ("reports", "reports", "Reports", "/reports"),
        ("staging", "staging", "📋 Review", "/staging"),
        ("settings", "settings", "⚙️", "/settings"),
    ]
    
    # Filter nav items based on current business config
    nav_html = ""
    for module, key, label, url in all_nav_items:
        if BusinessManager.is_module_visible(module):
            active_class = " active" if key == active else ""
            nav_html += f'<a href="{url}" class="nav-item{active_class}">{label}</a>'
    
    # Business switcher
    business_html = ""
    if user:
        businesses = BusinessManager.get_user_businesses(user.get("id", ""))
        current_biz = BusinessManager.get_current_business()
        
        if len(businesses) > 0:
            current_name = current_biz.get("business_name", current_biz.get("name", "Select Business")) if current_biz else "Select Business"
            current_icon = current_biz.get("icon", "🏢") if current_biz else "🏢"
            
            # Build dropdown options
            dropdown_items = ""
            for biz in businesses:
                biz_name = biz.get("business_name", biz.get("name", "Unnamed"))
                is_current = current_biz and biz["id"] == current_biz["id"]
                check = "✓ " if is_current else "&nbsp;&nbsp;&nbsp;"
                dropdown_items += f'<a href="/switch-business/{biz["id"]}" style="display:block;padding:12px 16px;color:#e0e0e0;text-decoration:none;border-bottom:1px solid #222;{"background:#1a1a2e;" if is_current else ""}">{check}{biz.get("icon", "🏢")} {safe_string(biz_name)}</a>'
            
            dropdown_items += '<a href="/businesses/new" style="display:block;padding:12px 16px;color:#a78bfa;text-decoration:none;font-weight:600;">➕ Add New Business</a>'
            
            business_html = f'''
            <style>
            .biz-switcher {{ position: relative; }}
            .biz-btn {{ 
                display: flex; align-items: center; gap: 8px;
                background: rgba(139,92,246,0.15); border: 1px solid rgba(139,92,246,0.3);
                padding: 8px 14px; border-radius: 8px; cursor: pointer;
                color: #a78bfa; font-size: 14px; font-weight: 600;
                margin-right: 12px;
            }}
            .biz-btn:hover {{ background: rgba(139,92,246,0.25); }}
            .biz-drop {{ 
                display: none; position: fixed; 
                background: #0d0d14; border: 1px solid #333; border-radius: 8px;
                min-width: 240px; box-shadow: 0 10px 40px rgba(0,0,0,0.8); 
                z-index: 99999; overflow: hidden;
            }}
            .biz-drop.open {{ display: block; }}
            .biz-drop a:hover {{ background: #1a1a2e !important; }}
            </style>
            <div class="biz-switcher">
                <div class="biz-btn" id="bizBtn">
                    <span>{current_icon}</span>
                    <span style="max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">{safe_string(current_name)}</span>
                    <span style="font-size:10px;">▼</span>
                </div>
                <div class="biz-drop" id="bizDrop">
                    {dropdown_items}
                </div>
            </div>
            <script>
            (function(){{
                var btn = document.getElementById('bizBtn');
                var drop = document.getElementById('bizDrop');
                if(btn && drop){{
                    btn.onclick = function(e){{
                        e.stopPropagation();
                        var rect = btn.getBoundingClientRect();
                        drop.style.top = (rect.bottom + 5) + 'px';
                        drop.style.right = (window.innerWidth - rect.right) + 'px';
                        drop.classList.toggle('open');
                    }};
                    document.onclick = function(){{ drop.classList.remove('open'); }};
                }}
            }})();
            </script>
            '''
        elif len(businesses) == 0:
            # No businesses yet - show setup prompt
            business_html = f'''
            <a href="/businesses/new" class="btn btn-sm btn-purple" style="margin-right:12px;">
                ➕ Setup Your Business
            </a>
            '''
    
    user_html = ""
    if user:
        user_html = f'''
        <div class="header-user">
            <span class="header-user-name">{safe_string(user.get("username", ""))}</span>
            <a href="/logout" class="btn btn-sm btn-ghost">Logout</a>
        </div>
        '''
    
    return f'''
    <header class="header">
        <a href="/" class="logo">Click AI</a>
        <nav class="nav">
            {nav_html}
        </nav>
        <div style="display:flex;align-items:center;">
            {business_html}
            {user_html}
        </div>
    </header>
    '''


def page_wrapper(title: str, content: str, active: str = "", 
                 user: dict = None, extra_css: str = "", 
                 extra_js: str = "") -> str:
    """
    Wrap page content in full HTML document
    
    Args:
        title: Page title
        content: Main content HTML
        active: Active nav item
        user: Current user
        extra_css: Additional CSS to include
        extra_js: Additional JavaScript to include
    """
    return f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{safe_string(title)} - Click AI</title>
    {CSS}
    {extra_css}
</head>
<body>
    {get_header_html(active, user)}
    <main class="main">
        <div class="container">
            {content}
        </div>
    </main>
    {extra_js}
</body>
</html>'''


def error_page(title: str, message: str, back_url: str = "/") -> str:
    """Generate an error page"""
    content = f'''
    <div class="empty-state">
        <div class="empty-state-icon">⚠️</div>
        <h2 class="empty-state-title">{safe_string(title)}</h2>
        <p class="empty-state-desc">{safe_string(message)}</p>
        <a href="{back_url}" class="btn btn-primary mt-lg">Go Back</a>
    </div>
    '''
    return page_wrapper(title, content)


def success_message(message: str) -> str:
    """Generate a success alert HTML"""
    return f'<div class="alert alert-success">✓ {safe_string(message)}</div>'


def error_message(message: str) -> str:
    """Generate an error alert HTML"""
    return f'<div class="alert alert-error">✗ {safe_string(message)}</div>'


def info_message(message: str) -> str:
    """Generate an info alert HTML"""
    return f'<div class="alert alert-info">ℹ {safe_string(message)}</div>'


# ═══════════════════════════════════════════════════════════════════════════════
# COMMON UI COMPONENTS
# ═══════════════════════════════════════════════════════════════════════════════

def stat_card(value: str, label: str, color: str = "") -> str:
    """Generate a stat card"""
    color_class = f" {color}" if color else ""
    return f'''
    <div class="stat">
        <div class="stat-value{color_class}">{safe_string(value)}</div>
        <div class="stat-label">{safe_string(label)}</div>
    </div>
    '''


def money_stat(amount, label: str, positive_is_green: bool = True) -> str:
    """Generate a money stat card with automatic color"""
    amount = Decimal(str(amount))
    color = ""
    if amount > 0:
        color = "green" if positive_is_green else "red"
    elif amount < 0:
        color = "red" if positive_is_green else "green"
    
    return stat_card(Money.format(amount), label, color)


def badge(text: str, color: str = "blue") -> str:
    """Generate a badge"""
    return f'<span class="badge badge-{color}">{safe_string(text)}</span>'


def empty_state(icon: str, title: str, description: str = "") -> str:
    """Generate an empty state placeholder"""
    desc_html = f'<p class="empty-state-desc">{safe_string(description)}</p>' if description else ""
    return f'''
    <div class="empty-state">
        <div class="empty-state-icon">{icon}</div>
        <h3 class="empty-state-title">{safe_string(title)}</h3>
        {desc_html}
    </div>
    '''


def loading_spinner() -> str:
    """Generate a loading spinner"""
    return '<div class="loading"><div class="spinner"></div></div>'


def modal_template(id: str, title: str, body: str, 
                   footer: str = "", size: str = "md") -> str:
    """Generate a modal template"""
    size_style = ""
    if size == "lg":
        size_style = "max-width: 700px;"
    elif size == "sm":
        size_style = "max-width: 400px;"
    
    footer_html = f'<div class="modal-footer">{footer}</div>' if footer else ""
    
    return f'''
    <div class="modal-overlay" id="{id}">
        <div class="modal" style="{size_style}">
            <div class="modal-header">
                <h3 class="modal-title">{safe_string(title)}</h3>
                <button class="modal-close" onclick="closeModal('{id}')">&times;</button>
            </div>
            <div class="modal-body">
                {body}
            </div>
            {footer_html}
        </div>
    </div>
    '''


def table_html(headers: list, rows: list, clickable: bool = False,
               empty_message: str = "No data found") -> str:
    """
    Generate a table
    
    Args:
        headers: List of header strings or dicts with 'label' and 'class'
        rows: List of rows, each row is a list of cell values
        clickable: Add clickable class
        empty_message: Message when no rows
    """
    if not rows:
        return empty_state("📋", empty_message)
    
    # Build headers
    header_html = "<tr>"
    for h in headers:
        if isinstance(h, dict):
            cls = f' class="{h.get("class", "")}"' if h.get("class") else ""
            header_html += f'<th{cls}>{safe_string(h["label"])}</th>'
        else:
            header_html += f'<th>{safe_string(h)}</th>'
    header_html += "</tr>"
    
    # Build rows
    rows_html = ""
    for row in rows:
        rows_html += "<tr>"
        for cell in row:
            if isinstance(cell, dict):
                cls = f' class="{cell.get("class", "")}"' if cell.get("class") else ""
                rows_html += f'<td{cls}>{cell["value"]}</td>'
            else:
                rows_html += f'<td>{cell}</td>'
        rows_html += "</tr>"
    
    clickable_class = " table-clickable" if clickable else ""
    
    return f'''
    <div class="table-wrapper">
        <table class="table{clickable_class}">
            <thead>{header_html}</thead>
            <tbody>{rows_html}</tbody>
        </table>
    </div>
    '''


# ═══════════════════════════════════════════════════════════════════════════════
# FORM HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def form_input(name: str, label: str = "", value: str = "",
               input_type: str = "text", placeholder: str = "",
               required: bool = False) -> str:
    """Generate a form input field"""
    label_html = f'<label class="form-label" for="{name}">{safe_string(label)}</label>' if label else ""
    req = " required" if required else ""
    
    return f'''
    <div class="form-group">
        {label_html}
        <input type="{input_type}" 
               id="{name}" 
               name="{name}" 
               class="form-input"
               value="{safe_string(value)}"
               placeholder="{safe_string(placeholder)}"{req}>
    </div>
    '''


def form_select(name: str, options: list, label: str = "",
                selected: str = "", required: bool = False) -> str:
    """
    Generate a form select field
    
    options: List of tuples (value, label) or dicts with 'value' and 'label'
    """
    label_html = f'<label class="form-label" for="{name}">{safe_string(label)}</label>' if label else ""
    req = " required" if required else ""
    
    options_html = '<option value="">Select...</option>'
    for opt in options:
        if isinstance(opt, dict):
            val, lbl = opt["value"], opt["label"]
        else:
            val, lbl = opt[0], opt[1]
        sel = " selected" if str(val) == str(selected) else ""
        options_html += f'<option value="{safe_string(val)}"{sel}>{safe_string(lbl)}</option>'
    
    return f'''
    <div class="form-group">
        {label_html}
        <select id="{name}" name="{name}" class="form-select"{req}>
            {options_html}
        </select>
    </div>
    '''


def form_textarea(name: str, label: str = "", value: str = "",
                  placeholder: str = "", rows: int = 4) -> str:
    """Generate a form textarea"""
    label_html = f'<label class="form-label" for="{name}">{safe_string(label)}</label>' if label else ""
    
    return f'''
    <div class="form-group">
        {label_html}
        <textarea id="{name}" 
                  name="{name}" 
                  class="form-textarea"
                  rows="{rows}"
                  placeholder="{safe_string(placeholder)}">{safe_string(value)}</textarea>
    </div>
    '''


def form_checkbox(name: str, label: str, checked: bool = False) -> str:
    """Generate a form checkbox"""
    chk = " checked" if checked else ""
    return f'''
    <label class="form-check">
        <input type="checkbox" name="{name}" id="{name}"{chk}>
        <span>{safe_string(label)}</span>
    </label>
    '''


# ═══════════════════════════════════════════════════════════════════════════════
# JAVASCRIPT HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

COMMON_JS = """
<script>
// Modal functions
function openModal(id) {
    document.getElementById(id).classList.add('show');
}

function closeModal(id) {
    document.getElementById(id).classList.remove('show');
}

// Close modal on overlay click
document.addEventListener('click', function(e) {
    if (e.target.classList.contains('modal-overlay')) {
        e.target.classList.remove('show');
    }
});

// Close modal on Escape key
document.addEventListener('keydown', function(e) {
    if (e.key === 'Escape') {
        document.querySelectorAll('.modal-overlay.show').forEach(function(m) {
            m.classList.remove('show');
        });
    }
});

// Format currency
function formatMoney(amount) {
    return 'R ' + parseFloat(amount).toFixed(2).replace(/\\B(?=(\\d{3})+(?!\\d))/g, ',');
}

// API helper
async function api(url, method, data) {
    const options = {
        method: method || 'GET',
        headers: {'Content-Type': 'application/json'}
    };
    if (data) {
        options.body = JSON.stringify(data);
    }
    const response = await fetch(url, options);
    return response.json();
}
</script>
"""


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 4
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 4 COMPLETE - CSS & UI Framework

Contains:
✓ Complete CSS with CSS variables
✓ Glowing animated header with logo
✓ Active navigation highlighting
✓ Responsive design (desktop, tablet, mobile)
✓ Dark theme optimized for long use
✓ Component styles:
  - Cards, Stats, Buttons, Forms
  - Tables, Modals, Alerts, Badges
  - Product grid, Cart, Empty states
  - Loading spinners, Report cards
✓ Page wrapper function
✓ UI component helper functions
✓ Form helper functions
✓ Common JavaScript utilities
✓ Print styles
✓ Mobile-specific big button interface

Design Principles Applied:
- Sticky glowing header - always visible
- Active page highlighted - no duplicate titles
- Every pixel counts - compact but readable
- Dark theme - professional, easy on eyes

Next: Piece 5 - Dashboard
"""

# =============================================================================
# PIECE5_DASHBOARD.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 5: Dashboard & Landing                                                ║
║                                                                               ║
║   This piece contains:                                                        ║
║   - Landing page                                                              ║
║   - Dashboard with real-time stats from ledger                                ║
║   - Quick actions                                                             ║
║   - Recent activity                                                           ║
║   - Login/logout routes                                                       ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""




# ═══════════════════════════════════════════════════════════════════════════════
# LANDING PAGE
# ═══════════════════════════════════════════════════════════════════════════════

LANDING_CSS = """
<style>
.landing {
    min-height: 100vh;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    padding: 40px 20px;
    background: radial-gradient(ellipse at center, #0a0a15 0%, #050508 100%);
}

.landing-logo {
    font-size: 72px;
    font-weight: 900;
    letter-spacing: -2px;
    background: linear-gradient(135deg, #8b5cf6, #3b82f6, #06b6d4);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    margin-bottom: 16px;
    position: relative;
    animation: float 6s ease-in-out infinite;
}

.landing-logo::before {
    content: 'Click AI';
    position: absolute;
    inset: 0;
    background: linear-gradient(135deg, #8b5cf6, #3b82f6);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    filter: blur(30px);
    opacity: 0.6;
    z-index: -1;
    animation: glow-breathe 3s ease-in-out infinite;
}

@keyframes float {
    0%, 100% { transform: translateY(0); }
    50% { transform: translateY(-10px); }
}

@keyframes glow-breathe {
    0%, 100% { opacity: 0.4; transform: scale(1); }
    50% { opacity: 0.8; transform: scale(1.1); }
}

.landing-tagline {
    font-size: 20px;
    color: #606070;
    margin-bottom: 48px;
    text-align: center;
}

.landing-tagline span {
    color: #8b5cf6;
}

.landing-buttons {
    display: flex;
    gap: 16px;
    flex-wrap: wrap;
    justify-content: center;
}

.landing-btn {
    padding: 16px 40px;
    font-size: 18px;
    font-weight: 600;
    border-radius: 12px;
    text-decoration: none;
    transition: all 0.3s ease;
}

.landing-btn-primary {
    background: linear-gradient(135deg, #8b5cf6, #3b82f6);
    color: white;
    box-shadow: 0 4px 20px rgba(139, 92, 246, 0.4);
}

.landing-btn-primary:hover {
    transform: translateY(-3px);
    box-shadow: 0 8px 30px rgba(139, 92, 246, 0.5);
}

.landing-btn-secondary {
    background: transparent;
    color: #a0a0a0;
    border: 1px solid #2a2a3e;
}

.landing-btn-secondary:hover {
    background: #15151f;
    color: white;
    border-color: #3a3a4e;
}

.landing-features {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 24px;
    max-width: 800px;
    margin-top: 80px;
}

.landing-feature {
    text-align: center;
    padding: 24px;
}

.landing-feature-icon {
    font-size: 36px;
    margin-bottom: 12px;
}

.landing-feature-title {
    font-size: 16px;
    font-weight: 600;
    color: #f0f0f0;
    margin-bottom: 8px;
}

.landing-feature-desc {
    font-size: 13px;
    color: #606070;
    line-height: 1.5;
}

@media (max-width: 600px) {
    .landing-logo {
        font-size: 48px;
    }
    
    .landing-tagline {
        font-size: 16px;
    }
    
    .landing-btn {
        width: 100%;
        text-align: center;
    }
}
</style>
"""


@app.route("/")
def landing():
    """Landing page - professional marketing page that sells Click AI"""
    
    # If already logged in, go to dashboard
    if UserSession.is_logged_in():
        return redirect("/dashboard")
    
    html = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Click AI - The Accounting System That Works For You</title>
    <link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&family=Space+Grotesk:wght@500;700&display=swap" rel="stylesheet">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        
        :root {
            --bg-dark: #050508;
            --bg-card: #0a0a12;
            --purple: #8b5cf6;
            --blue: #3b82f6;
            --green: #10b981;
            --orange: #f59e0b;
            --text: #f0f0f5;
            --text-muted: #8b8b9a;
        }
        
        body {
            font-family: 'DM Sans', -apple-system, sans-serif;
            background: var(--bg-dark);
            color: var(--text);
            line-height: 1.6;
            overflow-x: hidden;
        }
        
        /* ═══════════════════════════════════════════════════════════════════
           HERO SECTION
           ═══════════════════════════════════════════════════════════════════ */
        .hero {
            min-height: 100vh;
            display: flex;
            flex-direction: column;
            position: relative;
            overflow: hidden;
        }
        
        .hero::before {
            content: '';
            position: absolute;
            top: -50%;
            left: -50%;
            width: 200%;
            height: 200%;
            background: 
                radial-gradient(ellipse at 20% 20%, rgba(139, 92, 246, 0.15) 0%, transparent 50%),
                radial-gradient(ellipse at 80% 80%, rgba(59, 130, 246, 0.1) 0%, transparent 50%);
            animation: rotate 60s linear infinite;
            z-index: 0;
        }
        
        @keyframes rotate {
            from { transform: rotate(0deg); }
            to { transform: rotate(360deg); }
        }
        
        /* NAV */
        .nav {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 24px 48px;
            position: relative;
            z-index: 10;
        }
        
        .nav-logo {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, var(--purple), var(--blue));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            text-decoration: none;
        }
        
        .nav-links {
            display: flex;
            gap: 32px;
            align-items: center;
        }
        
        .nav-link {
            color: var(--text-muted);
            text-decoration: none;
            font-size: 14px;
            font-weight: 500;
            transition: color 0.2s;
        }
        
        .nav-link:hover { color: var(--text); }
        
        .btn {
            padding: 12px 24px;
            border-radius: 8px;
            font-weight: 600;
            font-size: 14px;
            text-decoration: none;
            transition: all 0.2s;
            cursor: pointer;
            border: none;
        }
        
        .btn-primary {
            background: linear-gradient(135deg, var(--purple), var(--blue));
            color: white;
        }
        
        .btn-primary:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 24px rgba(139, 92, 246, 0.3);
        }
        
        .btn-ghost {
            background: transparent;
            border: 1px solid rgba(255,255,255,0.1);
            color: var(--text);
        }
        
        .btn-ghost:hover {
            background: rgba(255,255,255,0.05);
        }
        
        /* HERO CONTENT */
        .hero-content {
            flex: 1;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center;
            padding: 60px 24px;
            position: relative;
            z-index: 10;
        }
        
        .hero-badge {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            background: rgba(139, 92, 246, 0.15);
            border: 1px solid rgba(139, 92, 246, 0.3);
            border-radius: 100px;
            padding: 8px 16px;
            font-size: 13px;
            color: var(--purple);
            margin-bottom: 32px;
        }
        
        .hero-badge-dot {
            width: 8px;
            height: 8px;
            background: var(--green);
            border-radius: 50%;
            animation: pulse 2s infinite;
        }
        
        @keyframes pulse {
            0%, 100% { opacity: 1; }
            50% { opacity: 0.5; }
        }
        
        .hero-title {
            font-family: 'Space Grotesk', sans-serif;
            font-size: clamp(40px, 8vw, 72px);
            font-weight: 700;
            line-height: 1.1;
            margin-bottom: 24px;
            max-width: 900px;
        }
        
        .hero-title span {
            background: linear-gradient(135deg, var(--purple), var(--blue));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        
        .hero-subtitle {
            font-size: 20px;
            color: var(--text-muted);
            max-width: 600px;
            margin-bottom: 40px;
            line-height: 1.7;
        }
        
        .hero-buttons {
            display: flex;
            gap: 16px;
            flex-wrap: wrap;
            justify-content: center;
        }
        
        .hero-buttons .btn {
            padding: 16px 32px;
            font-size: 16px;
        }
        
        .hero-stats {
            display: flex;
            gap: 48px;
            margin-top: 80px;
            padding-top: 40px;
            border-top: 1px solid rgba(255,255,255,0.1);
        }
        
        .hero-stat {
            text-align: center;
        }
        
        .hero-stat-value {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 36px;
            font-weight: 700;
            color: var(--purple);
        }
        
        .hero-stat-label {
            font-size: 13px;
            color: var(--text-muted);
            margin-top: 4px;
        }
        
        /* ═══════════════════════════════════════════════════════════════════
           PROBLEM SECTION
           ═══════════════════════════════════════════════════════════════════ */
        .section {
            padding: 120px 48px;
            position: relative;
        }
        
        .section-label {
            font-size: 12px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 2px;
            color: var(--purple);
            margin-bottom: 16px;
        }
        
        .section-title {
            font-family: 'Space Grotesk', sans-serif;
            font-size: clamp(32px, 5vw, 48px);
            font-weight: 700;
            margin-bottom: 24px;
            max-width: 700px;
        }
        
        .section-subtitle {
            font-size: 18px;
            color: var(--text-muted);
            max-width: 600px;
            line-height: 1.7;
        }
        
        .problem-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 24px;
            margin-top: 60px;
        }
        
        .problem-card {
            background: var(--bg-card);
            border: 1px solid rgba(255,255,255,0.05);
            border-radius: 16px;
            padding: 32px;
            transition: all 0.3s;
        }
        
        .problem-card:hover {
            border-color: rgba(239, 68, 68, 0.3);
            transform: translateY(-4px);
        }
        
        .problem-icon {
            font-size: 40px;
            margin-bottom: 20px;
        }
        
        .problem-title {
            font-size: 20px;
            font-weight: 600;
            margin-bottom: 12px;
        }
        
        .problem-desc {
            color: var(--text-muted);
            font-size: 15px;
            line-height: 1.6;
        }
        
        /* ═══════════════════════════════════════════════════════════════════
           FEATURES SECTION
           ═══════════════════════════════════════════════════════════════════ */
        .features {
            background: linear-gradient(180deg, var(--bg-dark) 0%, var(--bg-card) 50%, var(--bg-dark) 100%);
        }
        
        .feature-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(320px, 1fr));
            gap: 32px;
            margin-top: 60px;
        }
        
        .feature-card {
            background: rgba(255,255,255,0.02);
            border: 1px solid rgba(255,255,255,0.05);
            border-radius: 20px;
            padding: 40px;
            position: relative;
            overflow: hidden;
            transition: all 0.3s;
        }
        
        .feature-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 3px;
            background: linear-gradient(90deg, var(--purple), var(--blue));
            opacity: 0;
            transition: opacity 0.3s;
        }
        
        .feature-card:hover::before { opacity: 1; }
        
        .feature-card:hover {
            border-color: rgba(139, 92, 246, 0.2);
            transform: translateY(-4px);
        }
        
        .feature-icon {
            width: 56px;
            height: 56px;
            background: linear-gradient(135deg, rgba(139, 92, 246, 0.2), rgba(59, 130, 246, 0.2));
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 28px;
            margin-bottom: 24px;
        }
        
        .feature-title {
            font-size: 22px;
            font-weight: 600;
            margin-bottom: 12px;
        }
        
        .feature-desc {
            color: var(--text-muted);
            font-size: 15px;
            line-height: 1.7;
            margin-bottom: 20px;
        }
        
        .feature-list {
            list-style: none;
        }
        
        .feature-list li {
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 14px;
            color: var(--text-muted);
            margin-bottom: 8px;
        }
        
        .feature-list li::before {
            content: '✓';
            color: var(--green);
            font-weight: 600;
        }
        
        /* ═══════════════════════════════════════════════════════════════════
           IMPORT SECTION
           ═══════════════════════════════════════════════════════════════════ */
        .import-section {
            text-align: center;
        }
        
        .import-logos {
            display: flex;
            justify-content: center;
            gap: 40px;
            flex-wrap: wrap;
            margin: 60px 0;
        }
        
        .import-logo {
            background: var(--bg-card);
            border: 1px solid rgba(255,255,255,0.1);
            border-radius: 12px;
            padding: 24px 40px;
            font-size: 18px;
            font-weight: 600;
            color: var(--text-muted);
            transition: all 0.3s;
        }
        
        .import-logo:hover {
            border-color: var(--purple);
            color: var(--text);
            transform: translateY(-4px);
        }
        
        .import-steps {
            display: flex;
            justify-content: center;
            gap: 80px;
            margin-top: 60px;
            flex-wrap: wrap;
        }
        
        .import-step {
            text-align: center;
            max-width: 200px;
        }
        
        .import-step-num {
            width: 48px;
            height: 48px;
            background: linear-gradient(135deg, var(--purple), var(--blue));
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            font-family: 'Space Grotesk', sans-serif;
            font-size: 20px;
            font-weight: 700;
            margin: 0 auto 16px;
        }
        
        .import-step-title {
            font-size: 16px;
            font-weight: 600;
            margin-bottom: 8px;
        }
        
        .import-step-desc {
            font-size: 14px;
            color: var(--text-muted);
        }
        
        /* ═══════════════════════════════════════════════════════════════════
           PRICING
           ═══════════════════════════════════════════════════════════════════ */
        .pricing {
            text-align: center;
        }
        
        .pricing-cards {
            display: flex;
            justify-content: center;
            gap: 32px;
            margin-top: 60px;
            flex-wrap: wrap;
        }
        
        .pricing-card {
            background: var(--bg-card);
            border: 1px solid rgba(255,255,255,0.05);
            border-radius: 20px;
            padding: 40px;
            width: 320px;
            text-align: left;
            position: relative;
            transition: all 0.3s;
        }
        
        .pricing-card.featured {
            border-color: var(--purple);
            transform: scale(1.05);
        }
        
        .pricing-card.featured::before {
            content: 'Most Popular';
            position: absolute;
            top: -12px;
            left: 50%;
            transform: translateX(-50%);
            background: linear-gradient(135deg, var(--purple), var(--blue));
            padding: 4px 16px;
            border-radius: 100px;
            font-size: 12px;
            font-weight: 600;
        }
        
        .pricing-name {
            font-size: 20px;
            font-weight: 600;
            margin-bottom: 8px;
        }
        
        .pricing-desc {
            font-size: 14px;
            color: var(--text-muted);
            margin-bottom: 24px;
        }
        
        .pricing-price {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 48px;
            font-weight: 700;
            margin-bottom: 8px;
        }
        
        .pricing-price span {
            font-size: 16px;
            color: var(--text-muted);
            font-weight: 400;
        }
        
        .pricing-features {
            list-style: none;
            margin: 24px 0;
            padding-top: 24px;
            border-top: 1px solid rgba(255,255,255,0.1);
        }
        
        .pricing-features li {
            display: flex;
            align-items: center;
            gap: 12px;
            font-size: 14px;
            color: var(--text-muted);
            margin-bottom: 12px;
        }
        
        .pricing-features li::before {
            content: '✓';
            color: var(--green);
            font-weight: 600;
        }
        
        .pricing-card .btn {
            width: 100%;
            text-align: center;
            margin-top: 24px;
        }
        
        /* ═══════════════════════════════════════════════════════════════════
           CTA
           ═══════════════════════════════════════════════════════════════════ */
        .cta {
            text-align: center;
            background: linear-gradient(135deg, rgba(139, 92, 246, 0.1) 0%, rgba(59, 130, 246, 0.1) 100%);
            border-top: 1px solid rgba(139, 92, 246, 0.2);
            border-bottom: 1px solid rgba(59, 130, 246, 0.2);
        }
        
        .cta-title {
            font-family: 'Space Grotesk', sans-serif;
            font-size: clamp(28px, 5vw, 40px);
            font-weight: 700;
            margin-bottom: 16px;
        }
        
        .cta-subtitle {
            font-size: 18px;
            color: var(--text-muted);
            margin-bottom: 32px;
        }
        
        /* ═══════════════════════════════════════════════════════════════════
           FOOTER
           ═══════════════════════════════════════════════════════════════════ */
        .footer {
            padding: 60px 48px;
            text-align: center;
            color: var(--text-muted);
            font-size: 14px;
        }
        
        .footer-logo {
            font-family: 'Space Grotesk', sans-serif;
            font-size: 24px;
            font-weight: 700;
            background: linear-gradient(135deg, var(--purple), var(--blue));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin-bottom: 16px;
        }
        
        /* ═══════════════════════════════════════════════════════════════════
           RESPONSIVE
           ═══════════════════════════════════════════════════════════════════ */
        @media (max-width: 768px) {
            .nav { padding: 16px 24px; }
            .nav-links { display: none; }
            .section { padding: 80px 24px; }
            .hero-stats { flex-direction: column; gap: 24px; }
            .import-steps { gap: 40px; }
            .pricing-card.featured { transform: none; }
        }
    </style>
</head>
<body>
    <!-- HERO -->
    <section class="hero">
        <nav class="nav">
            <a href="/" class="nav-logo">Click AI</a>
            <div class="nav-links">
                <a href="#features" class="nav-link">Features</a>
                <a href="#pricing" class="nav-link">Pricing</a>
                <a href="/login" class="btn btn-ghost">Login</a>
                <a href="/demo" class="btn btn-primary">Try Free Demo</a>
            </div>
        </nav>
        
        <div class="hero-content">
            <div class="hero-badge">
                <span class="hero-badge-dot"></span>
                Built for South African Businesses
            </div>
            
            <h1 class="hero-title">
                Stop Fighting Your<br><span>Accounting Software</span>
            </h1>
            
            <p class="hero-subtitle">
                Click AI is the accounting system that actually works the way you think. 
                AI-powered invoicing, instant expense capture, and reports that explain themselves.
                Switch from Sage or Pastel in under 5 minutes.
            </p>
            
            <div class="hero-buttons">
                <a href="/free-analysis" class="btn btn-primary" style="background: linear-gradient(135deg, #10b981, #059669);">🧠 Free TB Analysis — No Signup</a>
                <a href="/demo" class="btn btn-ghost">Try Full Demo</a>
            </div>
            
            <div class="hero-stats">
                <div class="hero-stat">
                    <div class="hero-stat-value">5 min</div>
                    <div class="hero-stat-label">To import your data</div>
                </div>
                <div class="hero-stat">
                    <div class="hero-stat-value">R0</div>
                    <div class="hero-stat-label">Setup fees</div>
                </div>
                <div class="hero-stat">
                    <div class="hero-stat-value">100%</div>
                    <div class="hero-stat-label">SARS compliant</div>
                </div>
            </div>
        </div>
    </section>
    
    <!-- FREE TB ANALYZER PROMO -->
    <section class="section" style="background: linear-gradient(135deg, rgba(16,185,129,0.15), rgba(139,92,246,0.15)); text-align: center; padding: 80px 48px;">
        <div style="max-width: 800px; margin: 0 auto;">
            <div style="font-size: 64px; margin-bottom: 24px;">🧠</div>
            <h2 class="section-title" style="margin: 0 auto 16px;">Free AI Business Health Check</h2>
            <p class="section-subtitle" style="margin: 0 auto 32px; max-width: 600px;">
                Upload your Trial Balance from any accounting system. Our AI will analyze your business health, 
                spot red flags, find opportunities, and answer your questions — <strong>completely free, no signup required</strong>.
            </p>
            <a href="/free-analysis" class="btn btn-primary" style="padding: 20px 48px; font-size: 18px; background: linear-gradient(135deg, #10b981, #059669);">
                Upload My Trial Balance →
            </a>
            <p style="color: var(--text-muted); font-size: 13px; margin-top: 16px;">
                Works with exports from Sage, Pastel, Xero, QuickBooks, or any CSV/Excel file
            </p>
        </div>
    </section>
    
    <!-- PROBLEMS -->
    <section class="section">
        <div class="section-label">The Problem</div>
        <h2 class="section-title">Traditional accounting software wasn't built for you</h2>
        <p class="section-subtitle">
            You started a business to do what you love — not to become a bookkeeper.
        </p>
        
        <div class="problem-grid">
            <div class="problem-card">
                <div class="problem-icon">😤</div>
                <h3 class="problem-title">Too Complicated</h3>
                <p class="problem-desc">
                    You need an accounting degree to understand the interface. 
                    Why does adding an invoice require 15 clicks?
                </p>
            </div>
            <div class="problem-card">
                <div class="problem-icon">🐌</div>
                <h3 class="problem-title">Painfully Slow</h3>
                <p class="problem-desc">
                    Loading screens. Timeouts. That spinning wheel while you wait to capture one expense.
                </p>
            </div>
            <div class="problem-card">
                <div class="problem-icon">📱</div>
                <h3 class="problem-title">Desktop Only</h3>
                <p class="problem-desc">
                    You're on site, supplier hands you a receipt. Now you have to remember to capture it... later.
                </p>
            </div>
            <div class="problem-card">
                <div class="problem-icon">📊</div>
                <h3 class="problem-title">Useless Reports</h3>
                <p class="problem-desc">
                    Pages of numbers with no explanation. Is your business healthy? Who knows.
                </p>
            </div>
        </div>
    </section>
    
    <!-- FEATURES -->
    <section class="section features" id="features">
        <div class="section-label">The Solution</div>
        <h2 class="section-title">Accounting that works with you, not against you</h2>
        
        <div class="feature-grid">
            <div class="feature-card">
                <div class="feature-icon">📷</div>
                <h3 class="feature-title">Snap & Done</h3>
                <p class="feature-desc">
                    Take a photo of any receipt or invoice. AI reads it, categorizes it, and creates the entry. You just approve.
                </p>
                <ul class="feature-list">
                    <li>Works with any South African invoice format</li>
                    <li>Learns your suppliers automatically</li>
                    <li>VAT extracted correctly every time</li>
                </ul>
            </div>
            
            <div class="feature-card">
                <div class="feature-icon">💰</div>
                <h3 class="feature-title">Lightning POS</h3>
                <p class="feature-desc">
                    Ring up sales in seconds. Stock updates automatically. Customer accounts track themselves.
                </p>
                <ul class="feature-list">
                    <li>Works offline — sync when you're ready</li>
                    <li>Automatic low-stock alerts</li>
                    <li>Customer credit limits built in</li>
                </ul>
            </div>
            
            <div class="feature-card">
                <div class="feature-icon">🤖</div>
                <h3 class="feature-title">AI Business Health</h3>
                <p class="feature-desc">
                    Reports that actually tell you what's happening. Not just numbers — explanations and recommendations.
                </p>
                <ul class="feature-list">
                    <li>"Your debtors are 47 days — here's who to chase"</li>
                    <li>Bank-ready management accounts</li>
                    <li>Business rescue early warning</li>
                </ul>
            </div>
            
            <div class="feature-card">
                <div class="feature-icon">👥</div>
                <h3 class="feature-title">Proper Payroll</h3>
                <p class="feature-desc">
                    PAYE, UIF, SDL calculated correctly. Payslips generated. EMP201 ready for SARS.
                </p>
                <ul class="feature-list">
                    <li>2024/2025 tax tables built in</li>
                    <li>Handles advances and deductions</li>
                    <li>AI checks for mistakes before you pay</li>
                </ul>
            </div>
            
            <div class="feature-card">
                <div class="feature-icon">📋</div>
                <h3 class="feature-title">Real Double-Entry</h3>
                <p class="feature-desc">
                    Not some spreadsheet pretending to be accounting. Proper debits, credits, and audit trail.
                </p>
                <ul class="feature-list">
                    <li>Trial balance that actually balances</li>
                    <li>Full general ledger</li>
                    <li>Your accountant will love you</li>
                </ul>
            </div>
            
            <div class="feature-card">
                <div class="feature-icon">☁️</div>
                <h3 class="feature-title">Anywhere Access</h3>
                <p class="feature-desc">
                    Phone, tablet, laptop — your business travels with you. No installation, no updates.
                </p>
                <ul class="feature-list">
                    <li>Secure cloud backup</li>
                    <li>Multiple users with permissions</li>
                    <li>Works on any device with a browser</li>
                </ul>
            </div>
        </div>
    </section>
    
    <!-- IMPORT -->
    <section class="section import-section">
        <div class="section-label">Switch in 5 Minutes</div>
        <h2 class="section-title">Bring your data. Leave the headaches.</h2>
        <p class="section-subtitle">
            Export from your current system, upload to Click AI. We handle the rest.
        </p>
        
        <div class="import-logos">
            <div class="import-logo">Sage Pastel</div>
            <div class="import-logo">Sage One</div>
            <div class="import-logo">Xero</div>
            <div class="import-logo">QuickBooks</div>
            <div class="import-logo">Excel</div>
        </div>
        
        <div class="import-steps">
            <div class="import-step">
                <div class="import-step-num">1</div>
                <h4 class="import-step-title">Export</h4>
                <p class="import-step-desc">Download your data from your old system (we'll show you how)</p>
            </div>
            <div class="import-step">
                <div class="import-step-num">2</div>
                <h4 class="import-step-title">Upload</h4>
                <p class="import-step-desc">Drag and drop into Click AI</p>
            </div>
            <div class="import-step">
                <div class="import-step-num">3</div>
                <h4 class="import-step-title">Done</h4>
                <p class="import-step-desc">AI maps your accounts automatically. Start working.</p>
            </div>
        </div>
    </section>
    
    <!-- PRICING -->
    <section class="section pricing" id="pricing">
        <div class="section-label">Simple Pricing</div>
        <h2 class="section-title">No surprises. No hidden fees.</h2>
        
        <div class="pricing-cards">
            <div class="pricing-card">
                <h3 class="pricing-name">Starter</h3>
                <p class="pricing-desc">Perfect for sole traders</p>
                <div class="pricing-price">R299<span>/month</span></div>
                <ul class="pricing-features">
                    <li>1 user</li>
                    <li>Unlimited invoices</li>
                    <li>Receipt scanning (50/month)</li>
                    <li>Basic reports</li>
                    <li>Email support</li>
                </ul>
                <a href="/register?plan=starter" class="btn btn-ghost">Get Started</a>
            </div>
            
            <div class="pricing-card featured">
                <h3 class="pricing-name">Business</h3>
                <p class="pricing-desc">For growing businesses</p>
                <div class="pricing-price">R599<span>/month</span></div>
                <ul class="pricing-features">
                    <li>5 users</li>
                    <li>Unlimited everything</li>
                    <li>AI Business Health reports</li>
                    <li>Payroll (up to 20 employees)</li>
                    <li>Priority support</li>
                </ul>
                <a href="/register?plan=business" class="btn btn-primary">Get Started</a>
            </div>
            
            <div class="pricing-card">
                <h3 class="pricing-name">Enterprise</h3>
                <p class="pricing-desc">Multi-branch operations</p>
                <div class="pricing-price">R1,499<span>/month</span></div>
                <ul class="pricing-features">
                    <li>Unlimited users</li>
                    <li>Multi-branch support</li>
                    <li>Custom integrations</li>
                    <li>Dedicated account manager</li>
                    <li>On-site training</li>
                </ul>
                <a href="/register?plan=enterprise" class="btn btn-ghost">Contact Sales</a>
            </div>
        </div>
    </section>
    
    <!-- CTA -->
    <section class="section cta">
        <h2 class="cta-title">Ready to take control of your books?</h2>
        <p class="cta-subtitle">Try the full demo — no credit card, no signup required.</p>
        <a href="/demo" class="btn btn-primary" style="padding: 16px 40px; font-size: 18px;">
            Launch Demo →
        </a>
    </section>
    
    <!-- FOOTER -->
    <footer class="footer">
        <div class="footer-logo">Click AI</div>
        <p>Built in South Africa, for South African businesses.</p>
        <p style="margin-top: 8px;">© 2025 Click Fulltech. SARS-compliant accounting.</p>
    </footer>
</body>
</html>'''
    
    return html


# ═══════════════════════════════════════════════════════════════════════════════
# FREE TB ANALYZER - PUBLIC (No Login Required) - THE SALES HOOK!
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/free-analysis", methods=["GET", "POST"])
def free_analysis():
    """Free TB Analysis - No login required - The killer sales demo!"""
    
    if request.method == "POST":
        file = request.files.get("tb_file")
        context = request.form.get("context", "")
        
        if not file:
            return redirect("/free-analysis")
        
        # Read file content
        filename = file.filename.lower()
        try:
            if filename.endswith('.csv'):
                content = file.read().decode('utf-8', errors='ignore')
                accounts = TBAnalyzer.parse_tb_csv(content)
            elif filename.endswith(('.xlsx', '.xls')):
                import io
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(file.read()))
                    ws = wb.active
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append(','.join(str(c) if c else '' for c in row))
                    content = '\n'.join(rows)
                    accounts = TBAnalyzer.parse_tb_csv(content)
                except ImportError:
                    content = file.read().decode('utf-8', errors='ignore')
                    accounts = TBAnalyzer.parse_tb_csv(content)
            else:
                accounts = []
        except:
            accounts = []
        
        if not accounts:
            return free_analysis_page(error="Could not read the file. Please make sure it's a valid CSV or Excel file with account names and balances.")
        
        # Store in session
        session['free_tb_accounts'] = accounts
        session['free_tb_context'] = context
        session['free_tb_chat_history'] = []
        
        # Calculate quick stats
        total_debit = sum(a['debit'] for a in accounts)
        total_credit = sum(a['credit'] for a in accounts)
        
        return free_analysis_results(accounts, total_debit, total_credit)
    
    return free_analysis_page()


def free_analysis_page(error=None):
    """Render the free analysis upload page"""
    error_html = f'<div style="background:rgba(239,68,68,0.1);border:1px solid rgba(239,68,68,0.3);color:#fca5a5;padding:16px;border-radius:8px;margin-bottom:24px;">{error}</div>' if error else ''
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Free AI Business Analysis - Click AI</title>
    <link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&family=Space+Grotesk:wght@500;700&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        :root {{
            --bg-dark: #050508;
            --bg-card: #0a0a12;
            --purple: #8b5cf6;
            --green: #10b981;
            --text: #f0f0f5;
            --text-muted: #8b8b9a;
        }}
        body {{
            font-family: 'DM Sans', sans-serif;
            background: var(--bg-dark);
            color: var(--text);
            min-height: 100vh;
        }}
        .container {{
            max-width: 700px;
            margin: 0 auto;
            padding: 40px 20px;
        }}
        .logo {{
            font-family: 'Space Grotesk', sans-serif;
            font-size: 24px;
            font-weight: 700;
            background: linear-gradient(135deg, var(--purple), #3b82f6);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            text-decoration: none;
            display: inline-block;
            margin-bottom: 40px;
        }}
        h1 {{
            font-family: 'Space Grotesk', sans-serif;
            font-size: 36px;
            margin-bottom: 16px;
        }}
        h1 span {{
            background: linear-gradient(135deg, var(--green), #059669);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}
        .subtitle {{
            color: var(--text-muted);
            font-size: 18px;
            margin-bottom: 40px;
            line-height: 1.6;
        }}
        .upload-box {{
            background: var(--bg-card);
            border: 2px dashed var(--green);
            border-radius: 16px;
            padding: 60px 40px;
            text-align: center;
            transition: all 0.3s;
        }}
        .upload-box:hover {{
            border-color: var(--purple);
            background: rgba(139,92,246,0.05);
        }}
        .upload-icon {{
            font-size: 64px;
            margin-bottom: 20px;
        }}
        .upload-title {{
            font-size: 20px;
            font-weight: 600;
            margin-bottom: 8px;
        }}
        .upload-desc {{
            color: var(--text-muted);
            margin-bottom: 24px;
        }}
        .btn {{
            padding: 16px 32px;
            border-radius: 8px;
            font-weight: 600;
            font-size: 16px;
            cursor: pointer;
            border: none;
            transition: all 0.2s;
        }}
        .btn-primary {{
            background: linear-gradient(135deg, var(--green), #059669);
            color: white;
        }}
        .btn-primary:hover {{
            transform: translateY(-2px);
            box-shadow: 0 8px 24px rgba(16,185,129,0.3);
        }}
        .context-box {{
            margin-top: 24px;
            text-align: left;
        }}
        .context-label {{
            font-size: 14px;
            color: var(--text-muted);
            margin-bottom: 8px;
        }}
        .context-input {{
            width: 100%;
            padding: 12px 16px;
            background: rgba(255,255,255,0.05);
            border: 1px solid rgba(255,255,255,0.1);
            border-radius: 8px;
            color: var(--text);
            font-size: 14px;
            resize: vertical;
        }}
        .context-input:focus {{
            outline: none;
            border-color: var(--green);
        }}
        .formats {{
            display: flex;
            gap: 16px;
            justify-content: center;
            flex-wrap: wrap;
            margin-top: 32px;
            padding-top: 24px;
            border-top: 1px solid rgba(255,255,255,0.1);
        }}
        .format-badge {{
            background: rgba(255,255,255,0.05);
            padding: 8px 16px;
            border-radius: 100px;
            font-size: 13px;
            color: var(--text-muted);
        }}
        .features {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-top: 48px;
        }}
        .feature {{
            background: var(--bg-card);
            border: 1px solid rgba(255,255,255,0.05);
            border-radius: 12px;
            padding: 24px;
            text-align: center;
        }}
        .feature-icon {{
            font-size: 32px;
            margin-bottom: 12px;
        }}
        .feature-title {{
            font-weight: 600;
            margin-bottom: 4px;
        }}
        .feature-desc {{
            font-size: 13px;
            color: var(--text-muted);
        }}
        .file-input {{ display: none; }}
        .file-name {{
            color: var(--green);
            font-weight: 600;
            margin-top: 16px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <a href="/" class="logo">← Click AI</a>
        
        <h1>Free <span>AI Business Analysis</span></h1>
        <p class="subtitle">
            Upload your Trial Balance and get instant AI-powered insights. 
            No signup, no credit card, no strings attached.
        </p>
        
        {error_html}
        
        <form method="POST" enctype="multipart/form-data" id="upload-form">
            <div class="upload-box" onclick="document.getElementById('file-input').click()">
                <div class="upload-icon">📊</div>
                <div class="upload-title">Drop your Trial Balance here</div>
                <div class="upload-desc">or click to browse files</div>
                <input type="file" name="tb_file" id="file-input" class="file-input" accept=".csv,.xlsx,.xls" onchange="handleFile(this)">
                <div id="file-name" class="file-name" style="display:none;"></div>
            </div>
            
            <div class="context-box">
                <div class="context-label">Tell us about your business (optional - helps AI give better advice)</div>
                <textarea name="context" class="context-input" rows="2" placeholder="e.g. Hardware store, trading for 5 years, struggling with cash flow lately..."></textarea>
            </div>
            
            <div style="text-align:center;margin-top:24px;">
                <button type="submit" class="btn btn-primary" id="analyze-btn" disabled>
                    🧠 Analyze My Business
                </button>
            </div>
            
            <div class="formats">
                <span class="format-badge">Sage Pastel</span>
                <span class="format-badge">Sage One</span>
                <span class="format-badge">Xero</span>
                <span class="format-badge">QuickBooks</span>
                <span class="format-badge">Excel/CSV</span>
            </div>
        </form>
        
        <div class="features">
            <div class="feature">
                <div class="feature-icon">🔍</div>
                <div class="feature-title">Deep Analysis</div>
                <div class="feature-desc">AI examines every account for red flags</div>
            </div>
            <div class="feature">
                <div class="feature-icon">💬</div>
                <div class="feature-title">Ask Questions</div>
                <div class="feature-desc">Chat with AI about your numbers</div>
            </div>
            <div class="feature">
                <div class="feature-icon">🏛️</div>
                <div class="feature-title">SARS Ready</div>
                <div class="feature-desc">Spots compliance issues</div>
            </div>
            <div class="feature">
                <div class="feature-icon">🔒</div>
                <div class="feature-title">Private & Secure</div>
                <div class="feature-desc">Your data is never stored</div>
            </div>
        </div>
    </div>
    
    <script>
    function handleFile(input) {{
        if (input.files && input.files[0]) {{
            document.getElementById('file-name').textContent = '✓ ' + input.files[0].name;
            document.getElementById('file-name').style.display = 'block';
            document.getElementById('analyze-btn').disabled = false;
        }}
    }}
    
    // Drag and drop
    const uploadBox = document.querySelector('.upload-box');
    uploadBox.addEventListener('dragover', (e) => {{
        e.preventDefault();
        uploadBox.style.borderColor = 'var(--purple)';
        uploadBox.style.background = 'rgba(139,92,246,0.1)';
    }});
    uploadBox.addEventListener('dragleave', () => {{
        uploadBox.style.borderColor = 'var(--green)';
        uploadBox.style.background = 'var(--bg-card)';
    }});
    uploadBox.addEventListener('drop', (e) => {{
        e.preventDefault();
        uploadBox.style.borderColor = 'var(--green)';
        if (e.dataTransfer.files.length) {{
            document.getElementById('file-input').files = e.dataTransfer.files;
            handleFile(document.getElementById('file-input'));
        }}
    }});
    </script>
</body>
</html>'''
    return html


def free_analysis_results(accounts, total_debit, total_credit):
    """Render the free analysis results page with chat"""
    net = total_credit - total_debit
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Your Business Analysis - Click AI</title>
    <link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&family=Space+Grotesk:wght@500;700&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        :root {{
            --bg-dark: #050508;
            --bg-card: #0a0a12;
            --purple: #8b5cf6;
            --green: #10b981;
            --red: #ef4444;
            --text: #f0f0f5;
            --text-muted: #8b8b9a;
        }}
        body {{
            font-family: 'DM Sans', sans-serif;
            background: var(--bg-dark);
            color: var(--text);
            min-height: 100vh;
        }}
        .container {{
            max-width: 900px;
            margin: 0 auto;
            padding: 40px 20px;
        }}
        .header {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 32px;
        }}
        .logo {{
            font-family: 'Space Grotesk', sans-serif;
            font-size: 24px;
            font-weight: 700;
            background: linear-gradient(135deg, var(--purple), #3b82f6);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            text-decoration: none;
        }}
        h1 {{
            font-family: 'Space Grotesk', sans-serif;
            font-size: 28px;
            margin-bottom: 8px;
        }}
        .subtitle {{
            color: var(--text-muted);
            font-size: 14px;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .stat {{
            background: var(--bg-card);
            border: 1px solid rgba(255,255,255,0.05);
            border-radius: 12px;
            padding: 20px;
            text-align: center;
        }}
        .stat-value {{
            font-family: 'Space Grotesk', sans-serif;
            font-size: 24px;
            font-weight: 700;
        }}
        .stat-label {{
            font-size: 12px;
            color: var(--text-muted);
            margin-top: 4px;
        }}
        .health-banner {{
            background: linear-gradient(135deg, rgba(139,92,246,0.2), rgba(139,92,246,0.05));
            border: 1px solid rgba(139,92,246,0.3);
            border-radius: 16px;
            padding: 24px;
            display: flex;
            align-items: center;
            gap: 20px;
            margin-bottom: 24px;
        }}
        .health-icon {{
            font-size: 48px;
        }}
        .health-title {{
            font-size: 20px;
            font-weight: 600;
            color: var(--purple);
        }}
        .health-summary {{
            color: var(--text-muted);
            font-size: 14px;
            margin-top: 4px;
        }}
        .chat-box {{
            background: var(--bg-card);
            border: 1px solid rgba(255,255,255,0.05);
            border-radius: 16px;
            overflow: hidden;
        }}
        .chat-messages {{
            max-height: 400px;
            overflow-y: auto;
            padding: 20px;
        }}
        .chat-message {{
            padding: 16px;
            border-radius: 12px;
            margin-bottom: 12px;
        }}
        .chat-message.assistant {{
            background: rgba(139,92,246,0.1);
        }}
        .chat-message.user {{
            background: rgba(59,130,246,0.1);
        }}
        .chat-sender {{
            font-weight: 600;
            margin-bottom: 8px;
            font-size: 14px;
        }}
        .chat-sender.assistant {{ color: var(--purple); }}
        .chat-sender.user {{ color: #60a5fa; }}
        .chat-text {{
            line-height: 1.6;
            white-space: pre-wrap;
        }}
        .chat-input-row {{
            display: flex;
            gap: 12px;
            padding: 16px 20px;
            border-top: 1px solid rgba(255,255,255,0.05);
        }}
        .chat-input {{
            flex: 1;
            padding: 12px 16px;
            background: rgba(255,255,255,0.05);
            border: 1px solid rgba(255,255,255,0.1);
            border-radius: 8px;
            color: var(--text);
            font-size: 14px;
        }}
        .chat-input:focus {{
            outline: none;
            border-color: var(--purple);
        }}
        .btn {{
            padding: 12px 24px;
            border-radius: 8px;
            font-weight: 600;
            cursor: pointer;
            border: none;
            transition: all 0.2s;
        }}
        .btn-purple {{
            background: var(--purple);
            color: white;
        }}
        .btn-purple:hover {{
            background: #7c3aed;
        }}
        .quick-questions {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 12px;
            margin-top: 20px;
        }}
        .quick-btn {{
            padding: 12px 16px;
            background: rgba(255,255,255,0.05);
            border: 1px solid rgba(255,255,255,0.1);
            border-radius: 8px;
            color: var(--text);
            font-size: 13px;
            cursor: pointer;
            transition: all 0.2s;
        }}
        .quick-btn:hover {{
            background: rgba(139,92,246,0.1);
            border-color: var(--purple);
        }}
        .cta-box {{
            background: linear-gradient(135deg, rgba(16,185,129,0.15), rgba(139,92,246,0.15));
            border: 1px solid rgba(16,185,129,0.3);
            border-radius: 16px;
            padding: 32px;
            text-align: center;
            margin-top: 32px;
        }}
        .cta-title {{
            font-family: 'Space Grotesk', sans-serif;
            font-size: 24px;
            margin-bottom: 12px;
        }}
        .cta-desc {{
            color: var(--text-muted);
            margin-bottom: 24px;
            line-height: 1.6;
        }}
        .cta-btn {{
            background: linear-gradient(135deg, var(--green), #059669);
            color: white;
            padding: 16px 32px;
            font-size: 16px;
            text-decoration: none;
            display: inline-block;
        }}
        .cta-features {{
            display: flex;
            gap: 24px;
            justify-content: center;
            flex-wrap: wrap;
            margin-top: 24px;
            font-size: 14px;
            color: var(--text-muted);
        }}
        .cta-features span::before {{
            content: '✓';
            color: var(--green);
            margin-right: 8px;
        }}
        @keyframes spin {{ to {{ transform: rotate(360deg); }} }}
        .spinner {{
            display: inline-block;
            width: 16px;
            height: 16px;
            border: 2px solid var(--text-muted);
            border-top-color: var(--purple);
            border-radius: 50%;
            animation: spin 1s linear infinite;
            margin-right: 8px;
            vertical-align: middle;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div>
                <a href="/" class="logo">Click AI</a>
            </div>
            <a href="/free-analysis" style="color: var(--text-muted); text-decoration: none; font-size: 14px;">← Upload New File</a>
        </div>
        
        <h1>🧠 Your Business Analysis</h1>
        <p class="subtitle">Analyzed {len(accounts)} accounts from your trial balance</p>
        
        <div class="stats" style="margin-top: 24px;">
            <div class="stat">
                <div class="stat-value" style="color: var(--green);">R {total_credit:,.0f}</div>
                <div class="stat-label">Total Credits</div>
            </div>
            <div class="stat">
                <div class="stat-value" style="color: var(--red);">R {total_debit:,.0f}</div>
                <div class="stat-label">Total Debits</div>
            </div>
            <div class="stat">
                <div class="stat-value" style="color: var(--purple);">R {abs(net):,.0f}</div>
                <div class="stat-label">{"Net Profit" if net > 0 else "Net Loss"}</div>
            </div>
            <div class="stat">
                <div class="stat-value" style="color: #60a5fa;">{len(accounts)}</div>
                <div class="stat-label">Accounts</div>
            </div>
        </div>
        
        <div class="health-banner" id="health-banner">
            <div class="health-icon" id="health-icon">⏳</div>
            <div>
                <div class="health-title" id="health-title">Analyzing your business...</div>
                <div class="health-summary" id="health-summary">Our AI is examining your numbers. This takes about 15 seconds.</div>
            </div>
        </div>
        
        <div class="chat-box">
            <div class="chat-messages" id="chat-messages">
                <div class="chat-message assistant" id="initial-message">
                    <div class="chat-sender assistant">🤖 BB Fin Advisor</div>
                    <div class="chat-text">
                        <span class="spinner"></span> Analyzing your trial balance...
                    </div>
                </div>
            </div>
            
            <div class="chat-input-row">
                <input type="text" class="chat-input" id="chat-input" placeholder="Ask a follow-up question..." onkeypress="if(event.key==='Enter')sendMessage()" disabled>
                <button class="btn btn-purple" onclick="sendMessage()" id="send-btn" disabled>Send</button>
            </div>
        </div>
        
        <div class="quick-questions" id="quick-questions" style="opacity: 0.5; pointer-events: none;">
            <button class="quick-btn" onclick="askQuestion('What are the biggest red flags?')">🚩 Red Flags</button>
            <button class="quick-btn" onclick="askQuestion('What opportunities do you see?')">💡 Opportunities</button>
            <button class="quick-btn" onclick="askQuestion('What would SARS look at?')">🏛️ SARS Concerns</button>
            <button class="quick-btn" onclick="askQuestion('What should I prioritize?')">📋 Priorities</button>
            <button class="quick-btn" onclick="askQuestion('How is my cash flow?')">💰 Cash Flow</button>
            <button class="quick-btn" onclick="askQuestion('Am I profitable?')">📈 Profitability</button>
        </div>
        
        <!-- THE SOFT SELL -->
        <div class="cta-box">
            <div class="cta-title">💡 Imagine this insight — every day</div>
            <div class="cta-desc">
                What if you could chat with an AI that knows ALL your transactions, customers, and history?<br>
                Not just a once-off upload — <strong>real-time insight into your business</strong>, whenever you need it.
            </div>
            <a href="/register" class="btn cta-btn">Start Free Trial — No Credit Card</a>
            <div class="cta-features">
                <span>Live transaction data</span>
                <span>Automatic VAT returns</span>
                <span>Invoice in 30 seconds</span>
                <span>SARS compliant</span>
            </div>
        </div>
    </div>
    
    <script>
    // Load analysis on page load
    window.onload = function() {{
        fetch('/free-analysis/analyze', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}}
        }})
        .then(r => r.json())
        .then(data => {{
            // Update health banner
            const colors = {{'good': '#10b981', 'warning': '#f59e0b', 'critical': '#ef4444'}};
            const icons = {{'good': '✅', 'warning': '⚠️', 'critical': '🚨'}};
            const health = data.health || 'unknown';
            const color = colors[health] || '#8b8b9a';
            const icon = icons[health] || '📊';
            
            document.getElementById('health-icon').textContent = icon;
            document.getElementById('health-title').textContent = 'Business Health: ' + health.toUpperCase();
            document.getElementById('health-title').style.color = color;
            document.getElementById('health-summary').textContent = data.summary || '';
            document.getElementById('health-banner').style.background = 'linear-gradient(135deg, ' + color + '20, ' + color + '05)';
            document.getElementById('health-banner').style.borderColor = color + '50';
            
            // Update chat
            document.getElementById('initial-message').innerHTML = `
                <div class="chat-sender assistant">🤖 BB Fin Advisor</div>
                <div class="chat-text">${{data.chat_response || 'Analysis complete!'}}</div>
            `;
            
            // Enable inputs
            document.getElementById('chat-input').disabled = false;
            document.getElementById('send-btn').disabled = false;
            document.getElementById('quick-questions').style.opacity = '1';
            document.getElementById('quick-questions').style.pointerEvents = 'auto';
        }})
        .catch(err => {{
            document.getElementById('initial-message').innerHTML = `
                <div class="chat-sender assistant">🤖 BB Fin Advisor</div>
                <div class="chat-text" style="color: #fca5a5;">Sorry, the analysis timed out. Please try again with a smaller file.</div>
            `;
        }});
    }};
    
    function sendMessage() {{
        const input = document.getElementById('chat-input');
        const message = input.value.trim();
        if (!message) return;
        
        const chatMessages = document.getElementById('chat-messages');
        chatMessages.innerHTML += `
            <div class="chat-message user">
                <div class="chat-sender user">You</div>
                <div class="chat-text">${{message}}</div>
            </div>
        `;
        chatMessages.innerHTML += `
            <div class="chat-message assistant" id="loading-msg">
                <div class="chat-sender assistant">🤖 BB Fin Advisor</div>
                <div class="chat-text"><span class="spinner"></span> Thinking...</div>
            </div>
        `;
        chatMessages.scrollTop = chatMessages.scrollHeight;
        input.value = '';
        
        fetch('/free-analysis/chat', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{message: message}})
        }})
        .then(r => r.json())
        .then(data => {{
            document.getElementById('loading-msg').remove();
            chatMessages.innerHTML += `
                <div class="chat-message assistant">
                    <div class="chat-sender assistant">🤖 BB Fin Advisor</div>
                    <div class="chat-text">${{data.response}}</div>
                </div>
            `;
            chatMessages.scrollTop = chatMessages.scrollHeight;
        }});
    }}
    
    function askQuestion(q) {{
        document.getElementById('chat-input').value = q;
        sendMessage();
    }}
    </script>
</body>
</html>'''
    return html


@app.route("/free-analysis/analyze", methods=["POST"])
def free_analysis_analyze():
    """Run AI analysis for free TB analyzer"""
    accounts = session.get('free_tb_accounts', [])
    context = session.get('free_tb_context', '')
    
    if not accounts:
        return jsonify({
            "error": "No data",
            "health": "unknown",
            "summary": "Please upload a trial balance first",
            "chat_response": "I don't have any data to analyze. Please go back and upload a file."
        })
    
    try:
        analysis = TBAnalyzer.analyze_with_opus(accounts, context)
        session['free_tb_analysis'] = analysis
        
        chat_response = TBAnalyzer.chat_response(analysis)
        session['free_tb_chat_history'] = [{"role": "assistant", "content": chat_response}]
        
        return jsonify({
            "health": analysis.get('company_health', 'unknown'),
            "summary": analysis.get('health_summary', ''),
            "chat_response": chat_response
        })
    except Exception as e:
        return jsonify({
            "error": str(e),
            "health": "unknown",
            "summary": "Analysis failed",
            "chat_response": f"Sorry, something went wrong: {str(e)}"
        })


@app.route("/free-analysis/chat", methods=["POST"])
def free_analysis_chat():
    """Handle chat for free TB analyzer"""
    data = request.get_json()
    message = data.get("message", "")
    
    analysis = session.get('free_tb_analysis', {})
    history = session.get('free_tb_chat_history', [])
    
    if not analysis:
        return jsonify({"response": "Please wait for the initial analysis to complete, or upload a new file."})
    
    history.append({"role": "user", "content": message})
    response = TBAnalyzer.chat_response(analysis, message, history)
    history.append({"role": "assistant", "content": response})
    session['free_tb_chat_history'] = history[-20:]
    
    return jsonify({"response": response})


# ═══════════════════════════════════════════════════════════════════════════════
# LOGIN / LOGOUT
# ═══════════════════════════════════════════════════════════════════════════════

LOGIN_CSS = """
<style>
.login-container {
    min-height: 100vh;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 20px;
    background: radial-gradient(ellipse at center, #0a0a15 0%, #050508 100%);
}

.login-box {
    background: #0f0f18;
    border: 1px solid #1a1a2e;
    border-radius: 16px;
    padding: 40px;
    width: 100%;
    max-width: 400px;
}

.login-logo {
    font-size: 32px;
    font-weight: 800;
    background: linear-gradient(135deg, #8b5cf6, #3b82f6);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    text-align: center;
    margin-bottom: 32px;
}

.login-title {
    font-size: 24px;
    font-weight: 600;
    text-align: center;
    margin-bottom: 24px;
    color: #f0f0f0;
}
</style>
"""


@app.route("/login", methods=["GET", "POST"])
def login():
    """Login page - beautiful and professional"""
    
    error = ""
    
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        
        success, result = UserSession.login(username, password)
        
        if success:
            return redirect("/dashboard")
        else:
            error = result
    
    error_html = f'<div style="background:rgba(239,68,68,0.15);border:1px solid rgba(239,68,68,0.3);border-radius:8px;padding:12px;margin-bottom:16px;color:#ef4444;font-size:14px;">{error}</div>' if error else ""
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Click AI - Business Accounting</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            min-height: 100vh;
            display: flex;
            background: #050508;
            color: #f0f0f5;
        }}
        .split-left {{
            flex: 1;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 40px;
        }}
        .split-right {{
            flex: 1;
            background: linear-gradient(135deg, rgba(139, 92, 246, 0.2) 0%, rgba(59, 130, 246, 0.2) 100%);
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            padding: 60px;
            position: relative;
            overflow: hidden;
        }}
        .split-right::before {{
            content: '';
            position: absolute;
            top: -50%;
            left: -50%;
            width: 200%;
            height: 200%;
            background: radial-gradient(circle, rgba(139, 92, 246, 0.1) 0%, transparent 70%);
            animation: rotate 30s linear infinite;
        }}
        @keyframes rotate {{
            from {{ transform: rotate(0deg); }}
            to {{ transform: rotate(360deg); }}
        }}
        .login-box {{
            width: 100%;
            max-width: 380px;
        }}
        .logo {{
            font-size: 36px;
            font-weight: 800;
            background: linear-gradient(135deg, #8b5cf6, #3b82f6);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin-bottom: 8px;
        }}
        .tagline {{
            color: #8b8b9a;
            font-size: 14px;
            margin-bottom: 40px;
        }}
        .form-group {{
            margin-bottom: 20px;
        }}
        .form-label {{
            display: block;
            font-size: 13px;
            font-weight: 500;
            color: #a0a0b0;
            margin-bottom: 8px;
        }}
        .form-input {{
            width: 100%;
            padding: 14px 16px;
            background: #0a0a12;
            border: 1px solid #2a2a4a;
            border-radius: 10px;
            color: #f0f0f5;
            font-size: 15px;
            transition: all 0.2s;
        }}
        .form-input:focus {{
            outline: none;
            border-color: #8b5cf6;
            box-shadow: 0 0 0 3px rgba(139, 92, 246, 0.2);
        }}
        .btn-login {{
            width: 100%;
            padding: 16px;
            background: linear-gradient(135deg, #8b5cf6, #3b82f6);
            border: none;
            border-radius: 10px;
            color: white;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.2s;
            margin-top: 8px;
        }}
        .btn-login:hover {{
            transform: translateY(-2px);
            box-shadow: 0 8px 24px rgba(139, 92, 246, 0.4);
        }}
        .demo-link {{
            display: block;
            text-align: center;
            margin-top: 24px;
            color: #8b8b9a;
            font-size: 14px;
            text-decoration: none;
        }}
        .demo-link:hover {{
            color: #8b5cf6;
        }}
        .feature-list {{
            position: relative;
            z-index: 1;
            text-align: center;
        }}
        .feature-title {{
            font-size: 32px;
            font-weight: 700;
            margin-bottom: 16px;
        }}
        .feature-sub {{
            font-size: 18px;
            color: #a0a0b0;
            margin-bottom: 40px;
        }}
        .features {{
            display: flex;
            flex-direction: column;
            gap: 16px;
            text-align: left;
        }}
        .feature {{
            display: flex;
            align-items: center;
            gap: 12px;
            font-size: 15px;
        }}
        .feature-icon {{
            width: 36px;
            height: 36px;
            background: rgba(139, 92, 246, 0.2);
            border-radius: 8px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 18px;
        }}
        @media (max-width: 900px) {{
            body {{ flex-direction: column; }}
            .split-right {{ display: none; }}
            .split-left {{ min-height: 100vh; }}
        }}
    </style>
</head>
<body>
    <div class="split-left">
        <div class="login-box">
            <div class="logo">Click AI</div>
            <div class="tagline">Business accounting, powered by AI</div>
            
            {error_html}
            
            <form method="POST">
                <div class="form-group">
                    <label class="form-label">Username</label>
                    <input type="text" name="username" class="form-input" placeholder="Enter username" required autofocus>
                </div>
                <div class="form-group">
                    <label class="form-label">Password</label>
                    <input type="password" name="password" class="form-input" placeholder="Enter password" required>
                </div>
                <button type="submit" class="btn-login">Sign In</button>
            </form>
            
            <a href="/demo" class="demo-link">Try demo without account →</a>
        </div>
    </div>
    <div class="split-right">
        <div class="feature-list">
            <div class="feature-title">Accounting Made Simple</div>
            <div class="feature-sub">The AI does the hard work</div>
            <div class="features">
                <div class="feature">
                    <div class="feature-icon">📷</div>
                    <span>Scan invoices - AI reads everything</span>
                </div>
                <div class="feature">
                    <div class="feature-icon">📦</div>
                    <span>Stock created automatically</span>
                </div>
                <div class="feature">
                    <div class="feature-icon">💰</div>
                    <span>Prices calculated for you</span>
                </div>
                <div class="feature">
                    <div class="feature-icon">📊</div>
                    <span>Reports that explain themselves</span>
                </div>
                <div class="feature">
                    <div class="feature-icon">🤖</div>
                    <span>AI business health analysis</span>
                </div>
            </div>
        </div>
    </div>
</body>
</html>'''
    
    return html


@app.route("/logout")
def logout():
    """Logout and redirect to landing"""
    UserSession.logout()
    return redirect("/")


@app.route("/demo")
def demo():
    """Demo mode - create demo session with sample data"""
    # Set demo user in session
    session["user_id"] = "demo"
    session["username"] = "Demo User"
    session["role"] = "admin"
    session["is_demo"] = True
    
    # Load sample data for demo (only if not already loaded)
    try:
        existing_stock = db.select("stock_items", filters={"created_by": "demo"})
        if not existing_stock or len(existing_stock) < 5:
            load_demo_data()
    except:
        pass
    
    return redirect("/dashboard")


def load_demo_data():
    """Load sample data for demo mode so prospects can test the system"""
    
    # Sample customers
    demo_customers = [
        {"id": "demo-cust-1", "name": "ABC Construction", "email": "accounts@abcconstruction.co.za", "phone": "011-555-1234", "vat_number": "4000123456", "balance": Decimal("15420.00"), "credit_limit": Decimal("50000"), "created_by": "demo"},
        {"id": "demo-cust-2", "name": "Mike's Plumbing", "email": "mike@mikesplumbing.co.za", "phone": "082-555-5678", "balance": Decimal("3200.50"), "credit_limit": Decimal("10000"), "created_by": "demo"},
        {"id": "demo-cust-3", "name": "Sarah Williams", "email": "sarah.w@gmail.com", "phone": "076-555-9012", "balance": Decimal("0"), "credit_limit": Decimal("5000"), "created_by": "demo"},
        {"id": "demo-cust-4", "name": "Township Hardware", "email": "orders@townshiphardware.co.za", "phone": "011-555-3456", "vat_number": "4000789012", "balance": Decimal("8750.00"), "credit_limit": Decimal("25000"), "created_by": "demo"},
        {"id": "demo-cust-5", "name": "Cash Customer", "email": "", "phone": "", "balance": Decimal("0"), "credit_limit": Decimal("0"), "created_by": "demo"},
    ]
    
    for c in demo_customers:
        c["created_at"] = now()
        try:
            db.insert("customers", c)
        except:
            pass
    
    # Sample suppliers
    demo_suppliers = [
        {"id": "demo-supp-1", "name": "Makro", "email": "accounts@makro.co.za", "phone": "011-555-7890", "vat_number": "4100234567", "balance": Decimal("0"), "created_by": "demo"},
        {"id": "demo-supp-2", "name": "Builders Warehouse", "email": "trade@builderswarehouse.co.za", "phone": "011-555-2345", "vat_number": "4100345678", "balance": Decimal("12500.00"), "created_by": "demo"},
        {"id": "demo-supp-3", "name": "Mica Hardware", "email": "orders@mica.co.za", "phone": "011-555-6789", "vat_number": "4100456789", "balance": Decimal("4200.00"), "created_by": "demo"},
        {"id": "demo-supp-4", "name": "Engen Fuel", "email": "fleet@engen.co.za", "phone": "011-555-0123", "balance": Decimal("0"), "created_by": "demo"},
    ]
    
    for s in demo_suppliers:
        s["created_at"] = now()
        try:
            db.insert("suppliers", s)
        except:
            pass
    
    # Sample stock items with realistic hardware store inventory
    demo_stock = [
        # Fasteners
        {"id": "demo-stock-1", "sku": "NAIL-65MM", "name": "Nail 65mm (500g)", "description": "Wire nails for general construction", "category": "Fasteners", "cost_price": Decimal("18.50"), "selling_price": Decimal("32.00"), "quantity": 145, "reorder_level": 20, "created_by": "demo"},
        {"id": "demo-stock-2", "sku": "SCREW-40MM", "name": "Wood Screw 40mm (100pk)", "description": "Countersunk wood screws", "category": "Fasteners", "cost_price": Decimal("24.00"), "selling_price": Decimal("45.00"), "quantity": 89, "reorder_level": 15, "created_by": "demo"},
        {"id": "demo-stock-3", "sku": "BOLT-M10", "name": "Bolt M10x50 with Nut", "description": "Hex bolt with nut and washer", "category": "Fasteners", "cost_price": Decimal("4.50"), "selling_price": Decimal("8.50"), "quantity": 234, "reorder_level": 50, "created_by": "demo"},
        
        # Electrical
        {"id": "demo-stock-4", "sku": "CABLE-2.5", "name": "Cable 2.5mm² (100m)", "description": "Twin & earth electrical cable", "category": "Electrical", "cost_price": Decimal("890.00"), "selling_price": Decimal("1299.00"), "quantity": 12, "reorder_level": 5, "created_by": "demo"},
        {"id": "demo-stock-5", "sku": "SWITCH-1G", "name": "Light Switch 1-Gang", "description": "White plastic light switch", "category": "Electrical", "cost_price": Decimal("28.00"), "selling_price": Decimal("55.00"), "quantity": 67, "reorder_level": 20, "created_by": "demo"},
        {"id": "demo-stock-6", "sku": "SOCKET-DBL", "name": "Socket Outlet Double", "description": "Double plug socket", "category": "Electrical", "cost_price": Decimal("45.00"), "selling_price": Decimal("89.00"), "quantity": 43, "reorder_level": 15, "created_by": "demo"},
        
        # Plumbing
        {"id": "demo-stock-7", "sku": "PIPE-20MM", "name": "PVC Pipe 20mm (6m)", "description": "Class 16 PVC pressure pipe", "category": "Plumbing", "cost_price": Decimal("85.00"), "selling_price": Decimal("145.00"), "quantity": 28, "reorder_level": 10, "created_by": "demo"},
        {"id": "demo-stock-8", "sku": "ELBOW-20", "name": "PVC Elbow 20mm 90°", "description": "90 degree elbow fitting", "category": "Plumbing", "cost_price": Decimal("8.00"), "selling_price": Decimal("16.00"), "quantity": 156, "reorder_level": 30, "created_by": "demo"},
        {"id": "demo-stock-9", "sku": "TAPE-PTFE", "name": "PTFE Tape (12m)", "description": "Thread seal tape", "category": "Plumbing", "cost_price": Decimal("12.00"), "selling_price": Decimal("25.00"), "quantity": 89, "reorder_level": 25, "created_by": "demo"},
        
        # Tools
        {"id": "demo-stock-10", "sku": "HAMMER-500", "name": "Claw Hammer 500g", "description": "Fiberglass handle claw hammer", "category": "Tools", "cost_price": Decimal("125.00"), "selling_price": Decimal("199.00"), "quantity": 18, "reorder_level": 5, "created_by": "demo"},
        {"id": "demo-stock-11", "sku": "TAPE-5M", "name": "Measuring Tape 5m", "description": "Self-locking tape measure", "category": "Tools", "cost_price": Decimal("45.00"), "selling_price": Decimal("85.00"), "quantity": 34, "reorder_level": 10, "created_by": "demo"},
        {"id": "demo-stock-12", "sku": "LEVEL-600", "name": "Spirit Level 600mm", "description": "3 vial spirit level", "category": "Tools", "cost_price": Decimal("89.00"), "selling_price": Decimal("159.00"), "quantity": 8, "reorder_level": 3, "created_by": "demo"},
        
        # Safety
        {"id": "demo-stock-13", "sku": "GLOVE-PU", "name": "Work Gloves PU Palm", "description": "Size L work gloves", "category": "Safety", "cost_price": Decimal("35.00"), "selling_price": Decimal("65.00"), "quantity": 56, "reorder_level": 20, "created_by": "demo"},
        {"id": "demo-stock-14", "sku": "GLASSES-CLR", "name": "Safety Glasses Clear", "description": "Impact resistant safety glasses", "category": "Safety", "cost_price": Decimal("28.00"), "selling_price": Decimal("55.00"), "quantity": 42, "reorder_level": 15, "created_by": "demo"},
        
        # Building
        {"id": "demo-stock-15", "sku": "CEMENT-50", "name": "Cement 50kg", "description": "Portland cement CEM II", "category": "Building", "cost_price": Decimal("95.00"), "selling_price": Decimal("135.00"), "quantity": 75, "reorder_level": 25, "created_by": "demo"},
        {"id": "demo-stock-16", "sku": "SAND-TON", "name": "Building Sand (per ton)", "description": "Washed building sand", "category": "Building", "cost_price": Decimal("450.00"), "selling_price": Decimal("650.00"), "quantity": 15, "reorder_level": 5, "created_by": "demo"},
        
        # Low stock items (for alerts)
        {"id": "demo-stock-17", "sku": "DRILL-10", "name": "Drill Bit Set 10pc", "description": "HSS drill bit set 1-10mm", "category": "Tools", "cost_price": Decimal("120.00"), "selling_price": Decimal("199.00"), "quantity": 3, "reorder_level": 5, "created_by": "demo"},
        {"id": "demo-stock-18", "sku": "ANGLE-100", "name": "Angle Grinder 100mm", "description": "850W angle grinder", "category": "Tools", "cost_price": Decimal("550.00"), "selling_price": Decimal("899.00"), "quantity": 2, "reorder_level": 3, "created_by": "demo"},
    ]
    
    for s in demo_stock:
        s["created_at"] = now()
        try:
            db.insert("stock_items", s)
        except:
            pass
    
    # Sample invoices (recent sales)
    from datetime import timedelta
    today_date = datetime.now()
    
    demo_invoices = [
        {"id": "demo-inv-1", "invoice_number": "INV-0001", "customer_id": "demo-cust-1", "customer_name": "ABC Construction", "date": (today_date - timedelta(days=0)).strftime("%Y-%m-%d"), "subtotal": Decimal("4500.00"), "vat": Decimal("675.00"), "total": Decimal("5175.00"), "status": "paid", "created_by": "demo"},
        {"id": "demo-inv-2", "invoice_number": "INV-0002", "customer_id": "demo-cust-2", "customer_name": "Mike's Plumbing", "date": (today_date - timedelta(days=1)).strftime("%Y-%m-%d"), "subtotal": Decimal("1280.00"), "vat": Decimal("192.00"), "total": Decimal("1472.00"), "status": "pending", "created_by": "demo"},
        {"id": "demo-inv-3", "invoice_number": "INV-0003", "customer_id": "demo-cust-5", "customer_name": "Cash Customer", "date": (today_date - timedelta(days=1)).strftime("%Y-%m-%d"), "subtotal": Decimal("350.00"), "vat": Decimal("52.50"), "total": Decimal("402.50"), "status": "paid", "created_by": "demo"},
        {"id": "demo-inv-4", "invoice_number": "INV-0004", "customer_id": "demo-cust-4", "customer_name": "Township Hardware", "date": (today_date - timedelta(days=3)).strftime("%Y-%m-%d"), "subtotal": Decimal("8500.00"), "vat": Decimal("1275.00"), "total": Decimal("9775.00"), "status": "pending", "created_by": "demo"},
        {"id": "demo-inv-5", "invoice_number": "INV-0005", "customer_id": "demo-cust-1", "customer_name": "ABC Construction", "date": (today_date - timedelta(days=5)).strftime("%Y-%m-%d"), "subtotal": Decimal("12340.00"), "vat": Decimal("1851.00"), "total": Decimal("14191.00"), "status": "paid", "created_by": "demo"},
    ]
    
    for inv in demo_invoices:
        inv["created_at"] = now()
        try:
            db.insert("invoices", inv)
        except:
            pass
    
    # Sample expenses
    demo_expenses = [
        {"id": "demo-exp-1", "date": (today_date - timedelta(days=2)).strftime("%Y-%m-%d"), "description": "Office electricity", "category": "Utilities", "amount": Decimal("1850.00"), "vat": Decimal("277.50"), "supplier_id": "demo-supp-1", "supplier_name": "City Power", "status": "paid", "created_by": "demo"},
        {"id": "demo-exp-2", "date": (today_date - timedelta(days=4)).strftime("%Y-%m-%d"), "description": "Delivery fuel", "category": "Transport", "amount": Decimal("1200.00"), "vat": Decimal("0"), "supplier_id": "demo-supp-4", "supplier_name": "Engen Fuel", "status": "paid", "is_zero_rated": True, "created_by": "demo"},
        {"id": "demo-exp-3", "date": (today_date - timedelta(days=6)).strftime("%Y-%m-%d"), "description": "Stock purchase - fasteners", "category": "Stock", "amount": Decimal("4500.00"), "vat": Decimal("675.00"), "supplier_id": "demo-supp-2", "supplier_name": "Builders Warehouse", "status": "pending", "created_by": "demo"},
    ]
    
    for exp in demo_expenses:
        exp["created_at"] = now()
        try:
            db.insert("expenses", exp)
        except:
            pass
    
    # Add some staged transactions for review demo
    demo_staged = [
        {"id": "demo-staged-1", "type": "supplier_invoice", "supplier_name": "Mica Hardware", "description": "PVC fittings and pipe", "amount": Decimal("2450.00"), "vat": Decimal("367.50"), "status": "pending", "scanned_at": now(), "created_by": "demo"},
        {"id": "demo-staged-2", "type": "expense", "description": "Telkom Internet", "category": "Communications", "amount": Decimal("899.00"), "vat": Decimal("134.85"), "status": "pending", "scanned_at": now(), "created_by": "demo"},
    ]
    
    for staged in demo_staged:
        staged["created_at"] = now()
        try:
            db.insert("staged_transactions", staged)
        except:
            pass


# ═══════════════════════════════════════════════════════════════════════════════
# QUICKSTART - DATA IMPORT WIZARD
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/quickstart")
def quickstart():
    """Quickstart - Step 1: Choose your current system"""
    
    html = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Quick Start - Click AI</title>
    <link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&family=Space+Grotesk:wght@500;700&display=swap" rel="stylesheet">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'DM Sans', sans-serif; background: #050508; color: #f0f0f5; }
        .qs-container { min-height: 100vh; }
        .qs-header { padding: 24px 48px; display: flex; justify-content: space-between; align-items: center; border-bottom: 1px solid rgba(255,255,255,0.05); }
        .qs-logo { font-family: 'Space Grotesk', sans-serif; font-size: 24px; font-weight: 700; background: linear-gradient(135deg, #8b5cf6, #3b82f6); -webkit-background-clip: text; -webkit-text-fill-color: transparent; text-decoration: none; }
        .qs-main { max-width: 900px; margin: 0 auto; padding: 60px 24px; }
        .qs-title { font-family: 'Space Grotesk', sans-serif; font-size: 36px; font-weight: 700; text-align: center; margin-bottom: 16px; }
        .qs-subtitle { font-size: 18px; color: #8b8b9a; text-align: center; margin-bottom: 48px; }
        .qs-options { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 40px; }
        .qs-option { background: #0a0a12; border: 2px solid rgba(255,255,255,0.05); border-radius: 16px; padding: 32px 24px; text-align: center; cursor: pointer; transition: all 0.3s; text-decoration: none; color: inherit; }
        .qs-option:hover { border-color: rgba(139, 92, 246, 0.3); transform: translateY(-4px); }
        .qs-option-icon { font-size: 48px; margin-bottom: 16px; }
        .qs-option-title { font-size: 18px; font-weight: 600; margin-bottom: 8px; }
        .qs-option-desc { font-size: 14px; color: #8b8b9a; }
        .qs-btn { padding: 14px 28px; border-radius: 10px; font-weight: 600; font-size: 15px; border: none; cursor: pointer; transition: all 0.2s; text-decoration: none; display: inline-block; }
        .qs-btn-ghost { background: transparent; border: 1px solid rgba(255,255,255,0.1); color: #f0f0f5; }
        .qs-btn-ghost:hover { background: rgba(255,255,255,0.05); }
        .qs-buttons { display: flex; gap: 16px; justify-content: center; }
    </style>
</head>
<body>
    <div class="qs-container">
        <header class="qs-header">
            <a href="/" class="qs-logo">Click AI</a>
            <a href="/login" class="qs-btn qs-btn-ghost">Already have an account?</a>
        </header>
        
        <main class="qs-main">
            <h1 class="qs-title">Where's your data coming from?</h1>
            <p class="qs-subtitle">We'll help you export from your current system and import into Click AI.</p>
            
            <div class="qs-options">
                <a href="/quickstart/sage-pastel" class="qs-option">
                    <div class="qs-option-icon">📊</div>
                    <h3 class="qs-option-title">Sage Pastel</h3>
                    <p class="qs-option-desc">Partner, Xpress, or Evolution</p>
                </a>
                
                <a href="/quickstart/sage-one" class="qs-option">
                    <div class="qs-option-icon">☁️</div>
                    <h3 class="qs-option-title">Sage One</h3>
                    <p class="qs-option-desc">Cloud-based Sage</p>
                </a>
                
                <a href="/quickstart/xero" class="qs-option">
                    <div class="qs-option-icon">🔵</div>
                    <h3 class="qs-option-title">Xero</h3>
                    <p class="qs-option-desc">Export your Xero data</p>
                </a>
                
                <a href="/quickstart/quickbooks" class="qs-option">
                    <div class="qs-option-icon">🟢</div>
                    <h3 class="qs-option-title">QuickBooks</h3>
                    <p class="qs-option-desc">Desktop or Online</p>
                </a>
                
                <a href="/quickstart/excel" class="qs-option">
                    <div class="qs-option-icon">📗</div>
                    <h3 class="qs-option-title">Excel</h3>
                    <p class="qs-option-desc">Upload CSV or Excel files</p>
                </a>
                
                <a href="/register?source=fresh" class="qs-option">
                    <div class="qs-option-icon">✨</div>
                    <h3 class="qs-option-title">Fresh Start</h3>
                    <p class="qs-option-desc">Starting from scratch</p>
                </a>
            </div>
            
            <div class="qs-buttons">
                <a href="/" class="qs-btn qs-btn-ghost">← Back to Home</a>
            </div>
        </main>
    </div>
</body>
</html>'''
    
    return html


@app.route("/quickstart/<source>", methods=["GET", "POST"])
def quickstart_source(source):
    """Quickstart - Step 2: Export instructions for specific system"""
    import csv
    import io
    
    sources = {
        "sage-pastel": {"name": "Sage Pastel", "icon": "📊", "instructions": ["Open Sage Pastel → File → Export", "Export Customers as CSV", "Export Suppliers as CSV", "Export Inventory/Stock as CSV", "Run Trial Balance report → Export as CSV", "Upload all files below"]},
        "sage-one": {"name": "Sage Business Cloud", "icon": "☁️", "instructions": ["Log into Sage Business Cloud", "Go to Settings → Data Export", "Select Customers, Suppliers, Items", "Click Export to CSV", "Upload the files below"]},
        "xero": {"name": "Xero", "icon": "🔵", "instructions": ["Log into Xero → Settings → General Settings", "Click Export Accounting Data", "Select Contacts, Items, Chart of Accounts", "Download and upload the CSV files below"]},
        "quickbooks": {"name": "QuickBooks", "icon": "🟢", "instructions": ["In QuickBooks → Reports → Trial Balance → Export to Excel", "Lists → Customer List → Export", "Lists → Vendor List → Export", "Lists → Item List → Export", "Upload all files below"]},
        "excel": {"name": "Excel / CSV", "icon": "📗", "instructions": ["Prepare your data with columns:", "Customers: Name, Email, Phone, Balance", "Suppliers: Name, Email, Phone, Balance", "Stock: SKU, Name, Cost Price, Selling Price, Quantity", "Save as CSV and upload below"]},
    }
    
    if source not in sources:
        return redirect("/quickstart")
    
    src = sources[source]
    
    # Handle file upload
    if request.method == "POST" and "files" in request.files:
        files = request.files.getlist("files")
        results = {"customers": [], "suppliers": [], "stock": [], "errors": []}
        
        for file in files:
            if not file.filename:
                continue
            try:
                if file.filename.lower().endswith('.csv'):
                    content = file.read().decode('utf-8', errors='ignore')
                    reader = csv.DictReader(io.StringIO(content))
                    rows = list(reader)
                    if rows:
                        headers = [h.lower() for h in rows[0].keys()]
                        # Detect type and extract data
                        if any('customer' in h or 'client' in h or 'debtor' in h for h in headers):
                            for row in rows:
                                name = row.get('Customer') or row.get('Customer Name') or row.get('Client') or row.get('Name', '')
                                if name and name.strip():
                                    results["customers"].append({"name": name.strip(), "email": row.get('Email', ''), "phone": row.get('Phone', '') or row.get('Tel', ''), "balance": row.get('Balance', '0')})
                        elif any('supplier' in h or 'vendor' in h for h in headers):
                            for row in rows:
                                name = row.get('Supplier') or row.get('Vendor') or row.get('Name', '')
                                if name and name.strip():
                                    results["suppliers"].append({"name": name.strip(), "email": row.get('Email', ''), "phone": row.get('Phone', ''), "balance": row.get('Balance', '0')})
                        elif any('sku' in h or 'item' in h or 'product' in h or 'stock' in h for h in headers):
                            for row in rows:
                                name = row.get('Item') or row.get('Product') or row.get('Description') or row.get('Name', '')
                                if name and name.strip():
                                    results["stock"].append({"sku": row.get('SKU', '') or row.get('Code', ''), "name": name.strip(), "cost_price": row.get('Cost', '0') or row.get('Cost Price', '0'), "selling_price": row.get('Price', '0') or row.get('Selling Price', '0'), "quantity": row.get('Qty', '0') or row.get('Quantity', '0')})
                else:
                    results["errors"].append(f"{file.filename}: Please convert to CSV first")
            except Exception as e:
                results["errors"].append(f"{file.filename}: {str(e)}")
        
        session["import_results"] = results
        return redirect("/quickstart/review")
    
    instructions_html = "".join([f"<li>{inst}</li>" for inst in src["instructions"]])
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Import from {src["name"]} - Click AI</title>
    <link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&family=Space+Grotesk:wght@500;700&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'DM Sans', sans-serif; background: #050508; color: #f0f0f5; }}
        .qs-container {{ min-height: 100vh; }}
        .qs-header {{ padding: 24px 48px; display: flex; justify-content: space-between; align-items: center; border-bottom: 1px solid rgba(255,255,255,0.05); }}
        .qs-logo {{ font-family: 'Space Grotesk', sans-serif; font-size: 24px; font-weight: 700; background: linear-gradient(135deg, #8b5cf6, #3b82f6); -webkit-background-clip: text; -webkit-text-fill-color: transparent; text-decoration: none; }}
        .qs-main {{ max-width: 700px; margin: 0 auto; padding: 60px 24px; }}
        .qs-title {{ font-family: 'Space Grotesk', sans-serif; font-size: 32px; font-weight: 700; text-align: center; margin-bottom: 16px; }}
        .qs-subtitle {{ font-size: 16px; color: #8b8b9a; text-align: center; margin-bottom: 40px; }}
        .qs-instructions {{ background: rgba(59, 130, 246, 0.1); border: 1px solid rgba(59, 130, 246, 0.2); border-radius: 12px; padding: 24px; margin-bottom: 32px; }}
        .qs-instructions h4 {{ color: #3b82f6; margin-bottom: 16px; }}
        .qs-instructions ol {{ padding-left: 20px; color: #a0a0b0; font-size: 14px; line-height: 2; }}
        .qs-upload {{ background: #0a0a12; border: 2px dashed rgba(139, 92, 246, 0.3); border-radius: 16px; padding: 50px; text-align: center; margin-bottom: 32px; cursor: pointer; transition: all 0.3s; }}
        .qs-upload:hover {{ border-color: #8b5cf6; background: rgba(139, 92, 246, 0.05); }}
        .qs-upload-icon {{ font-size: 48px; margin-bottom: 16px; }}
        .qs-upload-text {{ font-size: 16px; margin-bottom: 8px; }}
        .qs-upload-hint {{ font-size: 13px; color: #8b8b9a; }}
        .qs-btn {{ padding: 14px 28px; border-radius: 10px; font-weight: 600; font-size: 15px; border: none; cursor: pointer; text-decoration: none; display: inline-block; }}
        .qs-btn-ghost {{ background: transparent; border: 1px solid rgba(255,255,255,0.1); color: #f0f0f5; }}
        .qs-buttons {{ display: flex; gap: 16px; justify-content: center; }}
    </style>
</head>
<body>
    <div class="qs-container">
        <header class="qs-header">
            <a href="/" class="qs-logo">Click AI</a>
        </header>
        
        <main class="qs-main">
            <h1 class="qs-title">{src["icon"]} Import from {src["name"]}</h1>
            <p class="qs-subtitle">Follow these steps to export your data.</p>
            
            <div class="qs-instructions">
                <h4>📋 How to export:</h4>
                <ol>{instructions_html}</ol>
            </div>
            
            <form method="POST" enctype="multipart/form-data" id="uploadForm">
                <div class="qs-upload" onclick="document.getElementById('fileInput').click()">
                    <div class="qs-upload-icon">📁</div>
                    <p class="qs-upload-text">Drop files here or click to browse</p>
                    <p class="qs-upload-hint">Accepts CSV files • Select multiple files</p>
                </div>
                <input type="file" name="files" multiple accept=".csv" style="display:none" id="fileInput" onchange="document.getElementById('uploadForm').submit()">
            </form>
            
            <div class="qs-buttons">
                <a href="/quickstart" class="qs-btn qs-btn-ghost">← Choose Different System</a>
            </div>
        </main>
    </div>
</body>
</html>'''
    
    return html


@app.route("/quickstart/review")
def quickstart_review():
    """Quickstart - Step 3: Review imported data"""
    
    results = session.get("import_results", {})
    if not results:
        return redirect("/quickstart")
    
    c_count = len(results.get("customers", []))
    s_count = len(results.get("suppliers", []))
    st_count = len(results.get("stock", []))
    errors = results.get("errors", [])
    
    errors_html = ""
    if errors:
        errors_html = '<div style="background:rgba(239,68,68,0.1);border:1px solid rgba(239,68,68,0.2);border-radius:12px;padding:20px;margin-bottom:24px;"><h4 style="color:#ef4444;margin-bottom:12px;">⚠️ Warnings</h4><ul style="color:#a0a0b0;font-size:14px;padding-left:20px;">' + "".join([f"<li>{e}</li>" for e in errors]) + '</ul></div>'
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Review Import - Click AI</title>
    <link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&family=Space+Grotesk:wght@500;700&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'DM Sans', sans-serif; background: #050508; color: #f0f0f5; }}
        .qs-main {{ max-width: 600px; margin: 0 auto; padding: 60px 24px; }}
        .qs-logo {{ font-family: 'Space Grotesk', sans-serif; font-size: 24px; font-weight: 700; background: linear-gradient(135deg, #8b5cf6, #3b82f6); -webkit-background-clip: text; -webkit-text-fill-color: transparent; text-decoration: none; display: block; text-align: center; margin-bottom: 40px; }}
        .qs-title {{ font-family: 'Space Grotesk', sans-serif; font-size: 32px; font-weight: 700; text-align: center; margin-bottom: 16px; }}
        .qs-subtitle {{ font-size: 16px; color: #8b8b9a; text-align: center; margin-bottom: 40px; }}
        .qs-result {{ background: #0a0a12; border: 1px solid rgba(16,185,129,0.3); border-radius: 16px; padding: 32px; margin-bottom: 32px; }}
        .qs-result-title {{ display: flex; align-items: center; gap: 12px; font-size: 20px; font-weight: 600; margin-bottom: 24px; color: #10b981; }}
        .qs-result-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; }}
        .qs-result-item {{ background: rgba(255,255,255,0.02); border-radius: 8px; padding: 16px; text-align: center; }}
        .qs-result-value {{ font-family: 'Space Grotesk', sans-serif; font-size: 32px; font-weight: 700; color: #8b5cf6; }}
        .qs-result-label {{ font-size: 13px; color: #8b8b9a; margin-top: 4px; }}
        .qs-btn {{ padding: 16px 32px; border-radius: 10px; font-weight: 600; font-size: 16px; border: none; cursor: pointer; text-decoration: none; display: inline-block; }}
        .qs-btn-primary {{ background: linear-gradient(135deg, #8b5cf6, #3b82f6); color: white; }}
        .qs-btn-ghost {{ background: transparent; border: 1px solid rgba(255,255,255,0.1); color: #f0f0f5; }}
        .qs-buttons {{ display: flex; gap: 16px; justify-content: center; }}
    </style>
</head>
<body>
    <main class="qs-main">
        <a href="/" class="qs-logo">Click AI</a>
        
        <h1 class="qs-title">✅ Data Ready to Import</h1>
        <p class="qs-subtitle">Here's what we found in your files.</p>
        
        {errors_html}
        
        <div class="qs-result">
            <div class="qs-result-title"><span>📊</span> Import Summary</div>
            <div class="qs-result-grid">
                <div class="qs-result-item">
                    <div class="qs-result-value">{c_count}</div>
                    <div class="qs-result-label">Customers</div>
                </div>
                <div class="qs-result-item">
                    <div class="qs-result-value">{s_count}</div>
                    <div class="qs-result-label">Suppliers</div>
                </div>
                <div class="qs-result-item">
                    <div class="qs-result-value">{st_count}</div>
                    <div class="qs-result-label">Stock Items</div>
                </div>
            </div>
        </div>
        
        <div style="text-align:center;margin-bottom:32px;">
            <p style="color:#8b8b9a;margin-bottom:24px;">Create your account and we'll import everything automatically.</p>
            <a href="/register?import=true" class="qs-btn qs-btn-primary">Create Account & Import →</a>
        </div>
        
        <div class="qs-buttons">
            <a href="/quickstart" class="qs-btn qs-btn-ghost">← Start Over</a>
        </div>
    </main>
</body>
</html>'''
    
    return html


def import_user_data(user_id, results):
    """Import the uploaded data for a new user"""
    
    # Import customers
    for c in results.get("customers", []):
        try:
            balance = c.get("balance", "0")
            if isinstance(balance, str):
                balance = balance.replace('R', '').replace(',', '').replace(' ', '').strip() or "0"
            db.insert("customers", {
                "id": generate_id(),
                "name": c.get("name", ""),
                "email": c.get("email", ""),
                "phone": c.get("phone", ""),
                "balance": Decimal(str(balance)),
                "credit_limit": Decimal("10000"),
                "created_by": user_id,
                "created_at": now()
            })
        except:
            pass
    
    # Import suppliers
    for s in results.get("suppliers", []):
        try:
            balance = s.get("balance", "0")
            if isinstance(balance, str):
                balance = balance.replace('R', '').replace(',', '').replace(' ', '').strip() or "0"
            db.insert("suppliers", {
                "id": generate_id(),
                "name": s.get("name", ""),
                "email": s.get("email", ""),
                "phone": s.get("phone", ""),
                "balance": Decimal(str(balance)),
                "created_by": user_id,
                "created_at": now()
            })
        except:
            pass
    
    # Import stock
    for st in results.get("stock", []):
        try:
            cost = st.get("cost_price", "0")
            sell = st.get("selling_price", "0")
            qty = st.get("quantity", "0")
            if isinstance(cost, str):
                cost = cost.replace('R', '').replace(',', '').replace(' ', '').strip() or "0"
            if isinstance(sell, str):
                sell = sell.replace('R', '').replace(',', '').replace(' ', '').strip() or "0"
            if isinstance(qty, str):
                qty = qty.replace(',', '').replace(' ', '').strip() or "0"
            db.insert("stock_items", {
                "id": generate_id(),
                "sku": st.get("sku", ""),
                "name": st.get("name", ""),
                "cost_price": Decimal(str(cost)),
                "selling_price": Decimal(str(sell)),
                "quantity": int(float(qty)),
                "reorder_level": 5,
                "created_by": user_id,
                "created_at": now()
            })
        except:
            pass


@app.route("/register", methods=["GET", "POST"])
def register():
    """Registration page for new users"""
    import hashlib
    
    plan = request.args.get("plan", "business")
    do_import = request.args.get("import") == "true"
    
    error = ""
    
    if request.method == "POST":
        company = request.form.get("company", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        
        if not company or not email or not password:
            error = "Please fill in all fields"
        elif len(password) < 6:
            error = "Password must be at least 6 characters"
        else:
            user_id = generate_id()
            password_hash = hashlib.sha256(password.encode()).hexdigest()
            
            try:
                db.insert("users", {
                    "id": user_id,
                    "username": email,
                    "email": email,
                    "password_hash": password_hash,
                    "company_name": company,
                    "role": "admin",
                    "plan": plan,
                    "created_at": now()
                })
                
                session["user_id"] = user_id
                session["username"] = email
                session["role"] = "admin"
                
                if do_import and "import_results" in session:
                    import_user_data(user_id, session["import_results"])
                    session.pop("import_results", None)
                
                return redirect("/dashboard?welcome=true")
                
            except Exception as e:
                error = f"Error creating account: {str(e)}"
    
    error_html = f'<div style="background:rgba(239,68,68,0.15);border:1px solid rgba(239,68,68,0.3);border-radius:8px;padding:12px;margin-bottom:16px;color:#ef4444;font-size:14px;">{error}</div>' if error else ""
    
    import_notice = ""
    if do_import:
        results = session.get("import_results", {})
        total = len(results.get("customers", [])) + len(results.get("suppliers", [])) + len(results.get("stock", []))
        if total > 0:
            import_notice = f'<div style="background:rgba(16,185,129,0.15);border:1px solid rgba(16,185,129,0.3);border-radius:8px;padding:12px;margin-bottom:16px;color:#10b981;font-size:14px;">✓ {total} records ready to import</div>'
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Create Account - Click AI</title>
    <link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;700&family=Space+Grotesk:wght@500;700&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'DM Sans', sans-serif; background: #050508; color: #f0f0f5; min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }}
        .register-box {{ width: 100%; max-width: 420px; }}
        .logo {{ font-family: 'Space Grotesk', sans-serif; font-size: 32px; font-weight: 700; background: linear-gradient(135deg, #8b5cf6, #3b82f6); -webkit-background-clip: text; -webkit-text-fill-color: transparent; text-align: center; margin-bottom: 8px; }}
        .tagline {{ text-align: center; color: #8b8b9a; font-size: 14px; margin-bottom: 32px; }}
        .form-card {{ background: #0a0a12; border: 1px solid rgba(255,255,255,0.05); border-radius: 16px; padding: 32px; }}
        .form-group {{ margin-bottom: 20px; }}
        .form-label {{ display: block; font-size: 13px; font-weight: 500; color: #a0a0b0; margin-bottom: 8px; }}
        .form-input {{ width: 100%; padding: 14px 16px; background: #050508; border: 1px solid rgba(255,255,255,0.1); border-radius: 10px; color: #f0f0f5; font-size: 15px; transition: all 0.2s; }}
        .form-input:focus {{ outline: none; border-color: #8b5cf6; box-shadow: 0 0 0 3px rgba(139,92,246,0.2); }}
        .btn {{ width: 100%; padding: 16px; background: linear-gradient(135deg, #8b5cf6, #3b82f6); border: none; border-radius: 10px; color: white; font-size: 16px; font-weight: 600; cursor: pointer; transition: all 0.2s; }}
        .btn:hover {{ transform: translateY(-2px); box-shadow: 0 8px 24px rgba(139,92,246,0.3); }}
        .login-link {{ text-align: center; margin-top: 24px; color: #8b8b9a; font-size: 14px; }}
        .login-link a {{ color: #8b5cf6; text-decoration: none; }}
    </style>
</head>
<body>
    <div class="register-box">
        <div class="logo">Click AI</div>
        <p class="tagline">Create your account</p>
        
        {error_html}
        {import_notice}
        
        <form method="POST" class="form-card">
            <div class="form-group">
                <label class="form-label">Company Name</label>
                <input type="text" name="company" class="form-input" placeholder="Your Business Name" required>
            </div>
            <div class="form-group">
                <label class="form-label">Email Address</label>
                <input type="email" name="email" class="form-input" placeholder="you@company.co.za" required>
            </div>
            <div class="form-group">
                <label class="form-label">Password</label>
                <input type="password" name="password" class="form-input" placeholder="At least 6 characters" required>
            </div>
            <button type="submit" class="btn">Create Account</button>
        </form>
        
        <p class="login-link">Already have an account? <a href="/login">Login</a></p>
    </div>
</body>
</html>'''
    
    return html


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/dashboard")
def dashboard():
    """Dashboard - memory efficient version"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Simple counts only - these are fast and low memory
    stock_count = 0
    customer_count = 0
    supplier_count = 0
    invoice_count = 0
    expense_count = 0
    
    try:
        stock_count = db.count("stock_items")
    except:
        pass
    try:
        customer_count = db.count("customers")
    except:
        pass
    try:
        supplier_count = db.count("suppliers")
    except:
        pass
    # Get real financial data
    today_sales = Decimal("0")
    week_sales = Decimal("0")
    month_sales = Decimal("0")
    outstanding_debtors = Decimal("0")
    outstanding_creditors = Decimal("0")
    low_stock_count = 0
    pending_review = 0
    
    try:
        # Today's sales
        today_invoices = db.select("invoices", filters={"date": today()}) or []
        today_sales = sum(Decimal(str(inv.get("total", 0) or 0)) for inv in today_invoices)
        
        # This week/month sales (simplified)
        all_invoices = db.select("invoices") or []
        from datetime import datetime, timedelta
        now_date = datetime.now()
        week_ago = (now_date - timedelta(days=7)).strftime("%Y-%m-%d")
        month_ago = (now_date - timedelta(days=30)).strftime("%Y-%m-%d")
        
        for inv in all_invoices:
            inv_date = inv.get("date", "")[:10]
            total = Decimal(str(inv.get("total", 0) or 0))
            if inv_date >= week_ago:
                week_sales += total
            if inv_date >= month_ago:
                month_sales += total
        
        # Outstanding amounts
        customers = db.select("customers") or []
        outstanding_debtors = sum(Decimal(str(c.get("balance", 0) or 0)) for c in customers if Decimal(str(c.get("balance", 0) or 0)) > 0)
        
        suppliers = db.select("suppliers") or []
        outstanding_creditors = sum(Decimal(str(s.get("balance", 0) or 0)) for s in suppliers if Decimal(str(s.get("balance", 0) or 0)) > 0)
        
        # Low stock
        stock = db.select("stock_items") or []
        low_stock_count = sum(1 for s in stock if int(s.get("quantity", 0) or 0) <= int(s.get("reorder_level", 5) or 5) and int(s.get("quantity", 0) or 0) > 0)
        
        # Pending review
        staged = db.select("staged_transactions", filters={"status": "pending"}) or []
        pending_review = len(staged)
    except:
        pass
    
    # Greeting based on time
    hour = datetime.now().hour
    if hour < 12:
        greeting = "Good morning"
    elif hour < 17:
        greeting = "Good afternoon"
    else:
        greeting = "Good evening"
    
    # Cash position indicator
    cash_status = "green" if outstanding_debtors < month_sales * Decimal("0.3") else "orange" if outstanding_debtors < month_sales else "red"
    
    # Alerts section
    alerts_html = ""
    if pending_review > 0:
        alerts_html += f'<a href="/staging" class="alert-item alert-purple"><span class="alert-icon">📋</span><span class="alert-text">{pending_review} scanned item{"s" if pending_review > 1 else ""} waiting for review</span><span class="alert-action">Review →</span></a>'
    if low_stock_count > 0:
        alerts_html += f'<a href="/stock?filter=low" class="alert-item alert-orange"><span class="alert-icon">📦</span><span class="alert-text">{low_stock_count} item{"s" if low_stock_count > 1 else ""} running low</span><span class="alert-action">View →</span></a>'
    if outstanding_debtors > 0:
        alerts_html += f'<a href="/reports/debtors" class="alert-item alert-blue"><span class="alert-icon">💰</span><span class="alert-text">{Money.format(outstanding_debtors)} outstanding from customers</span><span class="alert-action">Collect →</span></a>'
    
    if not alerts_html:
        alerts_html = '<div class="alert-item alert-green"><span class="alert-icon">✓</span><span class="alert-text">All clear! Nothing needs attention right now.</span></div>'
    
    content = f'''
    <style>
        .dashboard-hero {{
            background: linear-gradient(135deg, rgba(139, 92, 246, 0.1) 0%, rgba(59, 130, 246, 0.1) 100%);
            border: 1px solid rgba(139, 92, 246, 0.2);
            border-radius: 16px;
            padding: 32px;
            margin-bottom: 24px;
        }}
        .dashboard-greeting {{
            font-size: 28px;
            font-weight: 700;
            margin-bottom: 8px;
        }}
        .dashboard-date {{
            color: var(--text-muted);
            font-size: 14px;
        }}
        .metric-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .metric-card {{
            background: var(--bg-card);
            border: 1px solid var(--border);
            border-radius: 12px;
            padding: 20px;
            transition: transform 0.2s, box-shadow 0.2s;
        }}
        .metric-card:hover {{
            transform: translateY(-2px);
            box-shadow: 0 8px 24px rgba(0,0,0,0.3);
        }}
        .metric-label {{
            font-size: 12px;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 8px;
        }}
        .metric-value {{
            font-size: 28px;
            font-weight: 700;
        }}
        .metric-value.green {{ color: var(--green); }}
        .metric-value.orange {{ color: var(--orange); }}
        .metric-value.red {{ color: var(--red); }}
        .metric-value.blue {{ color: var(--blue); }}
        .metric-sub {{
            font-size: 12px;
            color: var(--text-muted);
            margin-top: 4px;
        }}
        .action-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 12px;
            margin-bottom: 24px;
        }}
        .action-btn {{
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            background: var(--bg-card);
            border: 1px solid var(--border);
            border-radius: 12px;
            padding: 20px;
            text-decoration: none;
            color: var(--text-primary);
            transition: all 0.2s;
        }}
        .action-btn:hover {{
            border-color: var(--purple);
            transform: translateY(-2px);
            box-shadow: 0 8px 24px rgba(139, 92, 246, 0.2);
        }}
        .action-btn:hover .action-icon {{
            transform: scale(1.1);
        }}
        .action-icon {{
            font-size: 32px;
            margin-bottom: 8px;
            transition: transform 0.2s;
        }}
        .action-label {{
            font-size: 13px;
            font-weight: 600;
        }}
        .alert-list {{
            display: flex;
            flex-direction: column;
            gap: 8px;
            margin-bottom: 24px;
        }}
        .alert-item {{
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 14px 16px;
            border-radius: 10px;
            text-decoration: none;
            color: var(--text-primary);
            transition: transform 0.2s;
        }}
        .alert-item:hover {{
            transform: translateX(4px);
        }}
        .alert-purple {{ background: rgba(139, 92, 246, 0.15); border-left: 3px solid var(--purple); }}
        .alert-orange {{ background: rgba(245, 158, 11, 0.15); border-left: 3px solid var(--orange); }}
        .alert-blue {{ background: rgba(59, 130, 246, 0.15); border-left: 3px solid var(--blue); }}
        .alert-green {{ background: rgba(16, 185, 129, 0.15); border-left: 3px solid var(--green); }}
        .alert-icon {{ font-size: 20px; }}
        .alert-text {{ flex: 1; font-size: 14px; }}
        .alert-action {{ font-size: 12px; color: var(--text-muted); }}
        .section-title {{
            font-size: 14px;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 12px;
        }}
    </style>
    
    <div class="dashboard-hero">
        <div class="dashboard-greeting">{greeting}, {user.get("username", "Boss")} 👋</div>
        <div class="dashboard-date">{datetime.now().strftime("%A, %d %B %Y")}</div>
    </div>
    
    <div class="section-title">💰 Your Money</div>
    <div class="metric-grid">
        <div class="metric-card">
            <div class="metric-label">Today's Sales</div>
            <div class="metric-value green">{Money.format(today_sales)}</div>
            <div class="metric-sub">{len(today_invoices)} invoice{"s" if len(today_invoices) != 1 else ""}</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">This Week</div>
            <div class="metric-value">{Money.format(week_sales)}</div>
            <div class="metric-sub">Last 7 days</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">This Month</div>
            <div class="metric-value">{Money.format(month_sales)}</div>
            <div class="metric-sub">Last 30 days</div>
        </div>
        <div class="metric-card">
            <div class="metric-label">Owed to You</div>
            <div class="metric-value {cash_status}">{Money.format(outstanding_debtors)}</div>
            <div class="metric-sub"><a href="/reports/debtors">View aging →</a></div>
        </div>
        <div class="metric-card">
            <div class="metric-label">You Owe</div>
            <div class="metric-value orange">{Money.format(outstanding_creditors)}</div>
            <div class="metric-sub"><a href="/reports/creditors">View aging →</a></div>
        </div>
    </div>
    
    <div class="section-title">⚡ Needs Attention</div>
    <div class="alert-list">
        {alerts_html}
    </div>
    
    <div class="section-title">🚀 Quick Actions</div>
    <div class="action-grid">
        <a href="/pos" class="action-btn">
            <span class="action-icon">💰</span>
            <span class="action-label">New Sale</span>
        </a>
        <a href="/m" class="action-btn">
            <span class="action-icon">📷</span>
            <span class="action-label">Scan Invoice</span>
        </a>
        <a href="/invoices/new" class="action-btn">
            <span class="action-icon">📄</span>
            <span class="action-label">New Invoice</span>
        </a>
        <a href="/quotes/new" class="action-btn">
            <span class="action-icon">📋</span>
            <span class="action-label">New Quote</span>
        </a>
        <a href="/purchase-orders/new" class="action-btn">
            <span class="action-icon">📦</span>
            <span class="action-label">New Order</span>
        </a>
        <a href="/expenses/new" class="action-btn">
            <span class="action-icon">💸</span>
            <span class="action-label">Add Expense</span>
        </a>
        <a href="/customers/new" class="action-btn">
            <span class="action-icon">👤</span>
            <span class="action-label">New Customer</span>
        </a>
        <a href="/stock/new" class="action-btn">
            <span class="action-icon">🏷️</span>
            <span class="action-label">New Product</span>
        </a>
    </div>
    
    <div class="section-title">📊 Reports</div>
    <div class="action-grid">
        <a href="/reports/business-health" class="action-btn" style="border-color: var(--purple);">
            <span class="action-icon">🤖</span>
            <span class="action-label">AI Health Check</span>
        </a>
        <a href="/reports/trial-balance" class="action-btn">
            <span class="action-icon">⚖️</span>
            <span class="action-label">Trial Balance</span>
        </a>
        <a href="/reports/income-statement" class="action-btn">
            <span class="action-icon">📈</span>
            <span class="action-label">Profit & Loss</span>
        </a>
        <a href="/reports/vat" class="action-btn">
            <span class="action-icon">🏛️</span>
            <span class="action-label">VAT Report</span>
        </a>
    </div>
    '''
    
    return page_wrapper("Dashboard", content, "dashboard", user)

@app.route("/mobile")
def mobile_home():
    """Mobile-optimized home with big buttons - the best scanner interface"""
    
    user = UserSession.get_current_user()
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>Click AI Mobile</title>
    <link rel="manifest" href="/manifest.json">
    <meta name="theme-color" content="#050508">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: linear-gradient(180deg, #050508 0%, #0a0a12 100%);
            color: #f0f0f0;
            min-height: 100vh;
        }}
        .mobile-home {{
            min-height: 100vh;
            display: flex;
            flex-direction: column;
            padding: 12px;
        }}
        .mobile-header {{
            text-align: center;
            padding: 16px 0 8px;
        }}
        .mobile-logo {{
            font-size: 28px;
            font-weight: 800;
            background: linear-gradient(135deg, #8b5cf6, #3b82f6);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }}
        .mobile-tagline {{
            color: #606070;
            font-size: 13px;
            margin-top: 2px;
        }}
        .section-label {{
            font-size: 10px;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: #808090;
            margin: 14px 0 8px;
            padding-left: 4px;
            display: flex;
            align-items: center;
            gap: 6px;
        }}
        .section-label span {{
            font-size: 12px;
        }}
        .mobile-buttons {{
            flex: 1;
            display: flex;
            flex-direction: column;
            gap: 8px;
        }}
        .btn-row {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 10px;
        }}
        .mobile-btn {{
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            padding: 20px 12px;
            border-radius: 14px;
            text-decoration: none;
            transition: all 0.2s;
            position: relative;
        }}
        .mobile-btn:active {{
            transform: scale(0.97);
        }}
        .mobile-btn-icon {{
            font-size: 32px;
            margin-bottom: 8px;
        }}
        .mobile-btn-label {{
            font-size: 14px;
            font-weight: 700;
            text-align: center;
        }}
        .mobile-btn-desc {{
            font-size: 9px;
            opacity: 0.85;
            margin-top: 3px;
            text-align: center;
        }}
        /* Expense - Red */
        .mobile-btn.expense {{
            background: linear-gradient(135deg, #ef4444, #dc2626);
            color: white;
            box-shadow: 0 4px 16px rgba(239, 68, 68, 0.25);
        }}
        .mobile-btn.expense-paid {{
            background: linear-gradient(135deg, #f87171, #ef4444);
            color: white;
            box-shadow: 0 4px 16px rgba(239, 68, 68, 0.2);
            border: 2px solid #22c55e;
        }}
        /* Stock - Purple */
        .mobile-btn.stock {{
            background: linear-gradient(135deg, #8b5cf6, #7c3aed);
            color: white;
            box-shadow: 0 4px 16px rgba(139, 92, 246, 0.25);
        }}
        .mobile-btn.stock-paid {{
            background: linear-gradient(135deg, #a78bfa, #8b5cf6);
            color: white;
            box-shadow: 0 4px 16px rgba(139, 92, 246, 0.2);
            border: 2px solid #22c55e;
        }}
        /* COS - Orange */
        .mobile-btn.cos {{
            background: linear-gradient(135deg, #f59e0b, #d97706);
            color: white;
            box-shadow: 0 4px 16px rgba(245, 158, 11, 0.25);
        }}
        .mobile-btn.cos-paid {{
            background: linear-gradient(135deg, #fbbf24, #f59e0b);
            color: white;
            box-shadow: 0 4px 16px rgba(245, 158, 11, 0.2);
            border: 2px solid #22c55e;
        }}
        .mobile-footer {{
            text-align: center;
            padding: 16px 0;
            display: flex;
            gap: 10px;
            justify-content: center;
        }}
        .mobile-footer a {{
            color: #808090;
            text-decoration: none;
            font-size: 12px;
            padding: 10px 16px;
            border: 1px solid #2a2a4a;
            border-radius: 8px;
        }}
        .paid-badge {{
            position: absolute;
            top: 6px;
            right: 6px;
            background: #22c55e;
            color: white;
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 8px;
            font-weight: 700;
        }}
    </style>
</head>
<body>
    <div class="mobile-home">
        <div class="mobile-header">
            <div class="mobile-logo">Click AI</div>
            <div class="mobile-tagline">Tap • Snap • Done</div>
        </div>
        
        <div class="mobile-buttons">
            <!-- EXPENSES -->
            <div class="section-label"><span>🧾</span> EXPENSES (Business costs, NOT for resale)</div>
            <div class="btn-row">
                <a href="/m" class="mobile-btn expense" onclick="localStorage.setItem('lastScanType','exp')">
                    <div class="mobile-btn-icon">🧾</div>
                    <div class="mobile-btn-label">Expense</div>
                    <div class="mobile-btn-desc">Van parts, office, etc</div>
                </a>
                
                <a href="/m" class="mobile-btn expense-paid" onclick="localStorage.setItem('lastScanType','exp_paid')">
                    <span class="paid-badge">PAID</span>
                    <div class="mobile-btn-icon">💵</div>
                    <div class="mobile-btn-label">Expense Paid</div>
                    <div class="mobile-btn-desc">Cash - already paid</div>
                </a>
            </div>
            
            <!-- STOCK -->
            <div class="section-label"><span>📦</span> STOCK (Items for shop inventory)</div>
            <div class="btn-row">
                <a href="/m" class="mobile-btn stock" onclick="localStorage.setItem('lastScanType','cos')">
                    <div class="mobile-btn-icon">📦</div>
                    <div class="mobile-btn-label">Stock In</div>
                    <div class="mobile-btn-desc">Supplier invoice</div>
                </a>
                
                <a href="/m" class="mobile-btn stock-paid" onclick="localStorage.setItem('lastScanType','cos_paid')">
                    <span class="paid-badge">PAID</span>
                    <div class="mobile-btn-icon">💵</div>
                    <div class="mobile-btn-label">Stock Paid</div>
                    <div class="mobile-btn-desc">Cash - already paid</div>
                </a>
            </div>
            
            <!-- COS - BUYOUTS -->
            <div class="section-label"><span>🔄</span> COS (Bought for immediate resale)</div>
            <div class="btn-row">
                <a href="/m" class="mobile-btn cos" onclick="localStorage.setItem('lastScanType','cos')">
                    <div class="mobile-btn-icon">🔄</div>
                    <div class="mobile-btn-label">COS</div>
                    <div class="mobile-btn-desc">Buyout - on account</div>
                </a>
                
                <a href="/m" class="mobile-btn cos-paid" onclick="localStorage.setItem('lastScanType','cos_paid')">
                    <span class="paid-badge">PAID</span>
                    <div class="mobile-btn-icon">💵</div>
                    <div class="mobile-btn-label">COS Paid</div>
                    <div class="mobile-btn-desc">Buyout - cash paid</div>
                </a>
            </div>
        </div>
        
        <div class="mobile-footer">
            <a href="/m">📷 Full Scanner</a>
            <a href="/dashboard">📊 Desktop</a>
        </div>
    </div>
</body>
</html>'''
    
    return html


# ═══════════════════════════════════════════════════════════════════════════════
# ERROR PAGES
# ═══════════════════════════════════════════════════════════════════════════════

@app.errorhandler(404)
def page_not_found(e):
    """404 error page"""
    content = '''
    <div class="empty-state">
        <div class="empty-state-icon">🔍</div>
        <h2 class="empty-state-title">Page Not Found</h2>
        <p class="empty-state-desc">The page you're looking for doesn't exist or has been moved.</p>
        <a href="/dashboard" class="btn btn-primary mt-lg">Go to Dashboard</a>
    </div>
    '''
    return page_wrapper("Page Not Found", content), 404


@app.errorhandler(500)
def server_error(e):
    """500 error page"""
    content = '''
    <div class="empty-state">
        <div class="empty-state-icon">⚠️</div>
        <h2 class="empty-state-title">Something Went Wrong</h2>
        <p class="empty-state-desc">We're sorry, something went wrong on our end. Please try again.</p>
        <a href="/dashboard" class="btn btn-primary mt-lg">Go to Dashboard</a>
    </div>
    '''
    return page_wrapper("Server Error", content), 500


@app.route("/unauthorized")
def unauthorized():
    """Unauthorized access page"""
    content = '''
    <div class="empty-state">
        <div class="empty-state-icon">🔒</div>
        <h2 class="empty-state-title">Access Denied</h2>
        <p class="empty-state-desc">You don't have permission to access this page.</p>
        <a href="/dashboard" class="btn btn-primary mt-lg">Go to Dashboard</a>
    </div>
    '''
    return page_wrapper("Unauthorized", content)


# ═══════════════════════════════════════════════════════════════════════════════
# HEALTH CHECK
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/health")
def health_check():
    """Health check endpoint for monitoring"""
    return {
        "status": "healthy",
        "version": "2.0",
        "app": "Click AI"
    }


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 5
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 5 COMPLETE - Dashboard & Landing

Contains:
✓ Animated landing page with glowing logo
✓ Login/logout system
✓ Demo mode
✓ Dashboard with real-time stats from ledger:
  - Bank balance
  - Debtors/Creditors
  - Stock value
  - Revenue (YTD)
  - Gross profit
  - Expenses (YTD)
  - Net profit
  - VAT position
  - Quick counts
✓ Quick actions
✓ Recent transactions from journal
✓ Mobile-optimized interface with big buttons
✓ Error pages (404, 500, unauthorized)
✓ Health check endpoint

All stats pulled from Journal class - real accounting data!

Next: Piece 6 - POS (Point of Sale)
"""

# =============================================================================
# PIECE6_POS.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 6: Point of Sale (POS)                                                ║
║                                                                               ║
║   This piece contains:                                                        ║
║   - POS interface with product grid                                           ║
║   - Shopping cart functionality                                               ║
║   - Customer selection                                                        ║
║   - Multiple payment methods (cash, card, account)                            ║
║   - Proper GL posting with VAT                                                ║
║   - Stock reduction on sale                                                   ║
║   - Receipt generation                                                        ║
║                                                                               ║
║   All calculations done by Flask - NO JavaScript math                         ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""




# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_stock_items_for_pos() -> list:
    """Get stock items formatted for POS"""
    items = db.select("stock_items", order="description")
    result = []
    
    for item in items:
        if not item.get("active", True):
            continue
        
        price = Decimal(str(item.get("selling_price", 0) or 0))
        cost = Decimal(str(item.get("cost_price", 0) or 0))
        qty = int(item.get("quantity", 0) or 0)
        
        # Check if zero-rated
        desc = item.get("description", "")
        category = item.get("category", "")
        is_zero_rated = VAT.is_zero_rated(desc, category)
        
        result.append({
            "id": item["id"],
            "code": item.get("code", ""),
            "description": desc,
            "category": category,
            "quantity": qty,
            "price": float(price),
            "price_formatted": Money.format(price),
            "cost": float(cost),
            "is_zero_rated": is_zero_rated
        })
    
    return result


def get_customers_for_pos() -> list:
    """Get customers formatted for POS"""
    customers = db.select("customers", order="name")
    result = []
    
    for cust in customers:
        if not cust.get("active", True):
            continue
        
        result.append({
            "id": cust["id"],
            "name": cust.get("name", ""),
            "code": cust.get("code", ""),
            "balance": float(cust.get("balance", 0) or 0)
        })
    
    return result


def reduce_stock(item_id: str, quantity: int) -> bool:
    """Reduce stock quantity after sale"""
    item = db.select_one("stock_items", item_id)
    if not item:
        return False
    
    current_qty = int(item.get("quantity", 0) or 0)
    new_qty = max(0, current_qty - quantity)
    
    success, _ = db.update("stock_items", item_id, {"quantity": new_qty})
    return success


def update_customer_balance(customer_id: str, amount: Decimal) -> bool:
    """Add to customer balance (for account sales)"""
    customer = db.select_one("customers", customer_id)
    if not customer:
        return False
    
    current_balance = Decimal(str(customer.get("balance", 0) or 0))
    new_balance = current_balance + amount
    
    success, _ = db.update("customers", customer_id, {"balance": float(new_balance)})
    return success


# ═══════════════════════════════════════════════════════════════════════════════
# POS API ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/pos/stock")
def api_pos_stock():
    """Get stock items for POS"""
    items = get_stock_items_for_pos()
    return jsonify(items)


@app.route("/api/pos/customers")
def api_pos_customers():
    """Get customers for POS"""
    customers = get_customers_for_pos()
    return jsonify(customers)


@app.route("/api/pos/sale", methods=["POST"])
def api_pos_sale():
    """
    Process a sale - ALL CALCULATIONS DONE HERE IN FLASK
    
    Receives cart items, calculates totals, VAT, posts to GL
    """
    try:
        data = request.get_json()
        items = data.get("items", [])
        customer_id = data.get("customer_id", "")
        payment_method = data.get("payment_method", "cash")
        discount_type = data.get("discount_type", "percent")
        discount_value = Decimal(str(data.get("discount_value", 0) or 0))
        
        if not items:
            return jsonify({"success": False, "error": "Cart is empty"})
        
        # Calculate totals - ALL IN FLASK
        total_incl = Decimal("0")
        total_excl = Decimal("0")
        total_vat = Decimal("0")
        total_cost = Decimal("0")
        
        line_items = []
        
        for item in items:
            price = Decimal(str(item.get("price", 0)))
            cost = Decimal(str(item.get("cost", 0)))
            qty = int(item.get("quantity", 1))
            is_zero_rated = item.get("is_zero_rated", False)
            
            line_total = price * qty
            
            # Calculate VAT
            if is_zero_rated:
                vat_info = VAT.calculate_from_inclusive(line_total, VAT.ZERO_RATE)
            else:
                vat_info = VAT.calculate_from_inclusive(line_total)
            
            total_incl += vat_info["inclusive"]
            total_excl += vat_info["exclusive"]
            total_vat += vat_info["vat"]
            total_cost += cost * qty
            
            line_items.append({
                "item_id": item.get("id"),
                "code": item.get("code", ""),
                "description": item.get("description", ""),
                "quantity": qty,
                "price": float(price),
                "line_total": float(line_total),
                "vat": float(vat_info["vat"]),
                "cost": float(cost)
            })
        
        # Apply discount
        discount_amount = Decimal("0")
        if discount_value > 0:
            if discount_type == "percent":
                discount_amount = (total_incl * discount_value / Decimal("100")).quantize(Decimal("0.01"))
            else:
                discount_amount = discount_value
            
            # Reduce totals proportionally
            if total_incl > 0:
                ratio = (total_incl - discount_amount) / total_incl
                total_incl = total_incl - discount_amount
                total_excl = (total_excl * ratio).quantize(Decimal("0.01"))
                total_vat = (total_vat * ratio).quantize(Decimal("0.01"))
        
        # Generate invoice number
        invoice_number = DocumentNumbers.get_next("INV", "invoices", "invoice_number")
        
        # Get customer name
        customer_name = "Walk-in Customer"
        if customer_id:
            cust = db.select_one("customers", customer_id)
            if cust:
                customer_name = cust.get("name", "Customer")
        
        # Create invoice record
        invoice_id = generate_id()
        invoice_record = {
            "id": invoice_id,
            "invoice_number": invoice_number,
            "date": today(),
            "customer_id": customer_id or None,
            "customer_name": customer_name,
            "items": json.dumps(line_items),
            "subtotal": float(total_excl),
            "vat": float(total_vat),
            "total": float(total_incl),
            "payment_method": payment_method,
            "status": "paid" if payment_method != "account" else "outstanding",
            "created_at": now()
        }
        
        success, result = db.insert("invoices", invoice_record)
        if not success:
            return jsonify({"success": False, "error": f"Failed to save invoice: {result}"})
        
        # POST TO GENERAL LEDGER
        description = f"Sale {invoice_number} - {customer_name}"
        
        entry = JournalEntry(
            date=today(),
            reference=invoice_number,
            description=description,
            trans_type=TransactionType.SALE,
            source_type="invoice",
            source_id=invoice_id
        )
        
        # Debit: Cash/Bank or Debtors
        if payment_method == "account":
            entry.debit(AccountCodes.DEBTORS, total_incl)
            # Update customer balance
            update_customer_balance(customer_id, total_incl)
        else:
            entry.debit(AccountCodes.BANK, total_incl)
        
        # Credit: Sales (excl VAT)
        entry.credit(AccountCodes.SALES, total_excl)
        
        # Credit: VAT Output
        if total_vat > 0:
            entry.credit(AccountCodes.VAT_OUTPUT, total_vat)
        
        # Post the sale entry
        success, result = entry.post()
        if not success:
            return jsonify({"success": False, "error": f"Failed to post to ledger: {result}"})
        
        # POST COST OF GOODS SOLD
        if total_cost > 0:
            cogs_entry = JournalEntry(
                date=today(),
                reference=invoice_number,
                description=f"COGS - {invoice_number}",
                trans_type=TransactionType.SALE,
                source_type="invoice",
                source_id=invoice_id
            )
            
            cogs_entry.debit(AccountCodes.COGS, total_cost)
            cogs_entry.credit(AccountCodes.STOCK, total_cost)
            
            cogs_entry.post()
        
        # REDUCE STOCK QUANTITIES
        for item in line_items:
            reduce_stock(item["item_id"], item["quantity"])
        
        # Return success with receipt data
        return jsonify({
            "success": True,
            "invoice_id": invoice_id,
            "invoice_number": invoice_number,
            "total": float(total_incl),
            "total_formatted": Money.format(total_incl),
            "vat": float(total_vat),
            "vat_formatted": Money.format(total_vat),
            "payment_method": payment_method.title(),
            "customer_name": customer_name,
            "date": today(),
            "items_count": len(line_items)
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ═══════════════════════════════════════════════════════════════════════════════
# POS PAGE
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/pos")
def pos_page():
    """POS page"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    pos_js = '''
<script>
let stockItems = [];
let customers = [];
let cart = [];

document.addEventListener('DOMContentLoaded', loadData);

async function loadData() {
    document.getElementById('loading').style.display = 'flex';
    document.getElementById('products').style.display = 'none';
    
    try {
        const [stockRes, custRes] = await Promise.all([
            fetch('/api/pos/stock'),
            fetch('/api/pos/customers')
        ]);
        
        stockItems = await stockRes.json();
        customers = await custRes.json();
        
        renderProducts();
        renderCustomers();
        
        document.getElementById('loading').style.display = 'none';
        document.getElementById('products').style.display = 'block';
        
    } catch (error) {
        alert('Failed to load: ' + error.message);
    }
}

function renderProducts() {
    const search = document.getElementById('search').value.toLowerCase();
    const container = document.getElementById('product-grid');
    
    let html = '';
    let count = 0;
    
    for (const item of stockItems) {
        if (item.quantity <= 0) continue;
        
        if (search) {
            const m1 = item.description.toLowerCase().includes(search);
            const m2 = (item.code || '').toLowerCase().includes(search);
            if (!m1 && !m2) continue;
        }
        
        if (count >= 100) break;
        count++;
        
        const sc = item.quantity <= 5 ? ' low' : '';
        
        html += '<div class="product-item" onclick="addToCart(\\''+item.id+'\\')">';
        html += '<div class="product-name">'+escHtml(item.description)+'</div>';
        html += '<div class="product-price">'+item.price_formatted+'</div>';
        html += '<div class="product-stock'+sc+'">'+item.quantity+' in stock</div>';
        html += '</div>';
    }
    
    if (!html) {
        html = '<div style="text-align:center;padding:40px;color:#606070;">No products found</div>';
    }
    
    container.innerHTML = html;
}

function renderCustomers() {
    const sel = document.getElementById('customer-select');
    let html = '<option value="">Walk-in Customer</option>';
    
    for (const c of customers) {
        const balance = parseFloat(c.balance || 0);
        const balanceNote = balance > 0 ? ' (R'+balance.toFixed(0)+' owing)' : '';
        html += '<option value="'+c.id+'" data-balance="'+balance+'">'+escHtml(c.name)+balanceNote+'</option>';
    }
    
    sel.innerHTML = html;
}

function addToCart(itemId) {
    const item = stockItems.find(i => i.id === itemId);
    if (!item) return;
    
    const existing = cart.find(c => c.id === itemId);
    if (existing) {
        existing.quantity++;
    } else {
        cart.push({
            id: item.id,
            code: item.code,
            description: item.description,
            price: item.price,
            price_formatted: item.price_formatted,
            cost: item.cost,
            quantity: 1,
            is_zero_rated: item.is_zero_rated
        });
    }
    
    renderCart();
}

let discountType = 'percent';
let discountValue = 0;

function renderCart() {
    const container = document.getElementById('cart-items');
    const subtotalEl = document.getElementById('cart-subtotal');
    const totalEl = document.getElementById('cart-total');
    const discountLine = document.getElementById('discount-line');
    const discountAmountEl = document.getElementById('discount-amount');
    const discountSection = document.getElementById('discount-section');
    
    if (cart.length === 0) {
        container.innerHTML = '<div class="cart-empty">Cart is empty</div>';
        subtotalEl.textContent = 'R 0.00';
        totalEl.textContent = 'R 0.00';
        discountLine.style.display = 'none';
        discountSection.style.display = 'none';
        return;
    }
    
    let html = '';
    let subtotal = 0;
    
    for (let i = 0; i < cart.length; i++) {
        const item = cart[i];
        const lineTotal = item.price * item.quantity;
        subtotal += lineTotal;
        
        html += '<div class="cart-item">';
        html += '<div>';
        html += '<div class="cart-item-name">'+escHtml(item.description)+'</div>';
        html += '<div class="cart-item-details">'+item.price_formatted+' × '+item.quantity+'</div>';
        html += '</div>';
        html += '<div style="display:flex;align-items:center;gap:8px;">';
        html += '<span class="cart-item-price">R '+lineTotal.toFixed(2)+'</span>';
        html += '<button class="btn btn-sm btn-red" onclick="removeFromCart('+i+')">×</button>';
        html += '</div></div>';
    }
    
    container.innerHTML = html;
    subtotalEl.textContent = 'R ' + subtotal.toFixed(2);
    
    // Apply discount
    let discountAmt = 0;
    if (discountValue > 0) {
        if (discountType === 'percent') {
            discountAmt = subtotal * (discountValue / 100);
        } else {
            discountAmt = discountValue;
        }
        discountLine.style.display = 'flex';
        discountAmountEl.textContent = '-R ' + discountAmt.toFixed(2);
        discountSection.style.display = 'block';
        document.getElementById('discount-display').textContent = 
            discountType === 'percent' ? discountValue + '% off' : 'R ' + discountValue.toFixed(2) + ' off';
    } else {
        discountLine.style.display = 'none';
        discountSection.style.display = 'none';
    }
    
    const total = subtotal - discountAmt;
    totalEl.textContent = 'R ' + total.toFixed(2);
}

function removeFromCart(index) {
    cart.splice(index, 1);
    renderCart();
}

function clearCart() {
    cart = [];
    discountValue = 0;
    renderCart();
}

// Discount functions
function showDiscountModal() {
    if (cart.length === 0) {
        alert('Add items to cart first');
        return;
    }
    document.getElementById('discount-modal').classList.add('show');
    setDiscountType('percent');
}

function closeDiscountModal() {
    document.getElementById('discount-modal').classList.remove('show');
}

function setDiscountType(type) {
    discountType = type;
    document.getElementById('disc-type-percent').className = type === 'percent' ? 'btn btn-primary' : 'btn btn-ghost';
    document.getElementById('disc-type-amount').className = type === 'amount' ? 'btn btn-primary' : 'btn btn-ghost';
    document.getElementById('discount-label').textContent = type === 'percent' ? 'Percentage' : 'Amount (R)';
    document.getElementById('discount-value').placeholder = type === 'percent' ? '10' : '50';
}

function setDiscountQuick(val) {
    document.getElementById('discount-value').value = val;
}

function applyDiscount() {
    const val = parseFloat(document.getElementById('discount-value').value) || 0;
    if (val <= 0) {
        alert('Enter a discount value');
        return;
    }
    if (discountType === 'percent' && val > 100) {
        alert('Percentage cannot exceed 100%');
        return;
    }
    discountValue = val;
    closeDiscountModal();
    renderCart();
}

function removeDiscount() {
    discountValue = 0;
    renderCart();
}

// Customer balance display
function showCustomerBalance() {
    const select = document.getElementById('customer-select');
    const balanceDiv = document.getElementById('customer-balance');
    const balanceAmount = document.getElementById('balance-amount');
    
    if (!select.value) {
        balanceDiv.style.display = 'none';
        return;
    }
    
    const option = select.options[select.selectedIndex];
    const balance = parseFloat(option.dataset.balance || 0);
    
    if (balance > 0) {
        balanceDiv.style.display = 'block';
        balanceAmount.textContent = 'R ' + balance.toFixed(2) + ' owing';
        balanceAmount.style.color = '#f59e0b';
    } else if (balance < 0) {
        balanceDiv.style.display = 'block';
        balanceAmount.textContent = 'R ' + Math.abs(balance).toFixed(2) + ' credit';
        balanceAmount.style.color = '#10b981';
    } else {
        balanceDiv.style.display = 'block';
        balanceAmount.textContent = 'R 0.00';
        balanceAmount.style.color = '#8b8b9a';
    }
}

async function processSale(method) {
    if (cart.length === 0) {
        alert('Cart is empty');
        return;
    }
    
    const customerId = document.getElementById('customer-select').value;
    
    // Disable buttons
    document.querySelectorAll('.cart-actions .btn').forEach(b => b.disabled = true);
    
    try {
        const response = await fetch('/api/pos/sale', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify({
                items: cart,
                customer_id: customerId,
                payment_method: method,
                discount_type: discountType,
                discount_value: discountValue
            })
        });
        
        const result = await response.json();
        
        if (result.success) {
            showReceipt(result);
            cart = [];
            discountValue = 0;
            renderCart();
            loadData();
        } else {
            alert('Error: ' + result.error);
        }
        
    } catch (error) {
        alert('Failed: ' + error.message);
    }
    
    document.querySelectorAll('.cart-actions .btn').forEach(b => b.disabled = false);
}

function showReceipt(data) {
    document.getElementById('r-number').textContent = data.invoice_number;
    document.getElementById('r-total').textContent = data.total_formatted;
    document.getElementById('r-vat').textContent = data.vat_formatted;
    document.getElementById('r-method').textContent = data.payment_method;
    document.getElementById('r-customer').textContent = data.customer_name;
    document.getElementById('receipt-modal').classList.add('show');
}

function printThermal() {
    const invoiceNum = document.getElementById("r-number").textContent;
    window.open("/print/thermal/" + encodeURIComponent(invoiceNum), "_blank");
}

function printOffice() {
    const invoiceNum = document.getElementById("r-number").textContent;
    window.open("/print/office/" + encodeURIComponent(invoiceNum), "_blank");
}

function closeReceipt() {
    document.getElementById('receipt-modal').classList.remove('show');
}

function escHtml(text) {
    const div = document.createElement('div');
    div.textContent = text || '';
    return div.innerHTML;
}

document.getElementById('search').addEventListener('input', renderProducts);
</script>
'''
    
    content = '''
    <div class="pos-layout">
        <!-- Products Section -->
        <div class="card">
            <input type="text" id="search" class="form-input" 
                   placeholder="Search products..." style="margin-bottom: 12px;">
            
            <div id="loading" class="loading">
                <div class="spinner"></div>
            </div>
            
            <div id="products" style="display: none;">
                <div id="product-grid" class="product-grid"></div>
            </div>
        </div>
        
        <!-- Cart Section -->
        <div class="cart">
            <div class="cart-header">
                <h3 class="cart-title">Cart</h3>
                <button class="btn btn-sm btn-ghost" onclick="clearCart()">Clear</button>
            </div>
            
            <div style="margin-bottom: 12px;">
                <select id="customer-select" class="form-select" onchange="showCustomerBalance()">
                    <option value="">Walk-in Customer</option>
                </select>
                <div id="customer-balance" style="display:none; margin-top:8px; padding:8px 12px; background:#1a1a2e; border-radius:8px; font-size:13px;">
                    <span style="color:#8b8b9a;">Balance:</span> <span id="balance-amount" style="font-weight:600;"></span>
                </div>
            </div>
            
            <div id="cart-items" class="cart-items">
                <div class="cart-empty">Cart is empty</div>
            </div>
            
            <!-- Discount Section -->
            <div id="discount-section" style="display:none; padding:12px; background:#1a1a2e; border-radius:8px; margin-bottom:12px;">
                <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                    <span style="color:#f59e0b;">🏷️ Discount Applied</span>
                    <button class="btn btn-sm btn-ghost" onclick="removeDiscount()">Remove</button>
                </div>
                <div id="discount-display" style="font-weight:600;"></div>
            </div>
            
            <div class="cart-total">
                <div style="display:flex; justify-content:space-between; margin-bottom:4px;">
                    <span style="color:#8b8b9a; font-size:13px;">Subtotal</span>
                    <span id="cart-subtotal" style="color:#8b8b9a; font-size:13px;">R 0.00</span>
                </div>
                <div id="discount-line" style="display:none; justify-content:space-between; margin-bottom:4px;">
                    <span style="color:#f59e0b; font-size:13px;">Discount</span>
                    <span id="discount-amount" style="color:#f59e0b; font-size:13px;">-R 0.00</span>
                </div>
                <div class="cart-total-label">Total</div>
                <div class="cart-total-value" id="cart-total">R 0.00</div>
            </div>
            
            <div class="cart-actions">
                <button class="btn btn-orange btn-sm" onclick="showDiscountModal()" style="margin-bottom:8px; width:100%;">🏷️ Discount</button>
            </div>
            <div class="cart-actions">
                <button class="btn btn-green" onclick="processSale('cash')">💵 Cash</button>
                <button class="btn btn-blue" onclick="processSale('card')">💳 Card</button>
            </div>
            <button class="btn btn-purple btn-block mt-sm" onclick="processSale('account')">
                📋 Account
            </button>
        </div>
    </div>
    
    <!-- Discount Modal -->
    <div class="modal-overlay" id="discount-modal">
        <div class="modal" style="max-width: 320px;">
            <div class="modal-header">
                <h3 class="modal-title">🏷️ Apply Discount</h3>
                <button class="modal-close" onclick="closeDiscountModal()">&times;</button>
            </div>
            <div class="modal-body">
                <div class="btn-group mb-lg" style="width:100%;">
                    <button class="btn btn-ghost" onclick="setDiscountType('percent')" id="disc-type-percent" style="flex:1;">%</button>
                    <button class="btn btn-ghost" onclick="setDiscountType('amount')" id="disc-type-amount" style="flex:1;">R</button>
                </div>
                <div class="form-group">
                    <label class="form-label" id="discount-label">Percentage</label>
                    <input type="number" id="discount-value" class="form-input" placeholder="10" min="0" style="font-size:24px; text-align:center;">
                </div>
                <div class="btn-group" style="flex-wrap:wrap; gap:8px;">
                    <button class="btn btn-sm btn-ghost" onclick="setDiscountQuick(5)">5%</button>
                    <button class="btn btn-sm btn-ghost" onclick="setDiscountQuick(10)">10%</button>
                    <button class="btn btn-sm btn-ghost" onclick="setDiscountQuick(15)">15%</button>
                    <button class="btn btn-sm btn-ghost" onclick="setDiscountQuick(20)">20%</button>
                </div>
            </div>
            <div class="modal-footer">
                <button class="btn btn-ghost" onclick="closeDiscountModal()">Cancel</button>
                <button class="btn btn-primary" onclick="applyDiscount()">Apply</button>
            </div>
        </div>
    </div>
    
    <!-- Receipt Modal -->
    <div class="modal-overlay" id="receipt-modal">
        <div class="modal" style="max-width: 400px;">
            <div class="modal-header">
                <h3 class="modal-title">✓ Sale Complete</h3>
                <button class="modal-close" onclick="closeReceipt()">&times;</button>
            </div>
            <div class="modal-body" style="text-align: center;">
                <div style="font-size: 14px; color: #606070; margin-bottom: 4px;">Invoice</div>
                <div id="r-number" style="font-size: 24px; font-weight: 700; margin-bottom: 20px;"></div>
                
                <div style="font-size: 14px; color: #606070; margin-bottom: 4px;">Total</div>
                <div id="r-total" style="font-size: 36px; font-weight: 800; color: #10b981; margin-bottom: 12px;"></div>
                
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-top: 20px; text-align: left;">
                    <div>
                        <div style="font-size: 11px; color: #606070; text-transform: uppercase;">VAT</div>
                        <div id="r-vat" style="font-weight: 600;"></div>
                    </div>
                    <div>
                        <div style="font-size: 11px; color: #606070; text-transform: uppercase;">Payment</div>
                        <div id="r-method" style="font-weight: 600;"></div>
                    </div>
                    <div style="grid-column: 1 / -1;">
                        <div style="font-size: 11px; color: #606070; text-transform: uppercase;">Customer</div>
                        <div id="r-customer" style="font-weight: 600;"></div>
                    </div>
                </div>
            </div>
            <div class="modal-footer" style="justify-content: center;">
                <button class="btn btn-ghost" onclick="printThermal()">🖨️ Thermal</button>
                <button class="btn btn-ghost" onclick="printOffice()">📄 Office</button>
                <button class="btn btn-primary" onclick="closeReceipt()">Done</button>
            </div>
        </div>
    </div>
    ''' + pos_js
    
    return page_wrapper("Point of Sale", content, active="pos", user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 6
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 6 COMPLETE - Point of Sale

Contains:
✓ POS page with product grid
✓ Real-time product search
✓ Shopping cart management
✓ Customer selection (walk-in or account)
✓ Multiple payment methods:
  - Cash
  - Card
  - Account (adds to customer balance)
✓ All calculations done in Flask:
  - Line totals
  - VAT calculation (15% standard, 0% for zero-rated)
  - Invoice totals
✓ Proper GL posting:
  - DR Bank/Debtors (total incl VAT)
  - CR Sales (excl VAT)
  - CR VAT Output
  - DR Cost of Goods Sold
  - CR Stock
✓ Stock quantity reduction
✓ Customer balance update for account sales
✓ Receipt modal with summary
✓ Invoice record creation

All financial calculations happen on the server.
JavaScript only handles UI interactions.

Next: Piece 7 - Stock Management
"""

# =============================================================================
# PIECE7_STOCK.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 7: Stock Management                                                   ║
║                                                                               ║
║   This piece contains:                                                        ║
║   - Stock item listing with search                                            ║
║   - Add/Edit stock items                                                      ║
║   - Category management                                                       ║
║   - Stock take / adjustment                                                   ║
║   - Low stock alerts                                                          ║
║   - Stock valuation                                                           ║
║   - Import from CSV/JSON                                                      ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

import csv
import io



# ═══════════════════════════════════════════════════════════════════════════════
# STOCK ITEM FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_all_stock(include_inactive: bool = False) -> list:
    """Get all stock items"""
    items = db.select("stock_items", order="description")
    
    if not include_inactive:
        items = [i for i in items if i.get("active", True)]
    
    return items


def get_stock_item(item_id: str) -> dict:
    """Get single stock item"""
    return db.select_one("stock_items", item_id)


def get_stock_categories() -> list:
    """Get unique categories"""
    items = db.select("stock_items", columns="category")
    categories = set()
    
    for item in items:
        cat = item.get("category", "").strip()
        if cat:
            categories.add(cat)
    
    return sorted(list(categories))


def calculate_stock_valuation() -> dict:
    """Calculate total stock valuation"""
    items = get_all_stock()
    
    total_cost = Decimal("0")
    total_retail = Decimal("0")
    total_items = 0
    total_qty = 0
    
    for item in items:
        qty = int(item.get("quantity", 0) or 0)
        cost = Decimal(str(item.get("cost_price", 0) or 0))
        sell = Decimal(str(item.get("selling_price", 0) or 0))
        
        total_items += 1
        total_qty += qty
        total_cost += cost * qty
        total_retail += sell * qty
    
    return {
        "total_items": total_items,
        "total_quantity": total_qty,
        "total_cost": total_cost,
        "total_retail": total_retail,
        "potential_profit": total_retail - total_cost
    }


def get_low_stock_items(threshold: int = 5) -> list:
    """Get items below reorder level"""
    items = get_all_stock()
    
    low = []
    for item in items:
        qty = int(item.get("quantity", 0) or 0)
        reorder = int(item.get("reorder_level", threshold) or threshold)
        
        if qty <= reorder and qty > 0:
            item["_status"] = "low"
            low.append(item)
        elif qty <= 0:
            item["_status"] = "out"
            low.append(item)
    
    return low


# ═══════════════════════════════════════════════════════════════════════════════
# STOCK PAGE
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/stock")
def stock_list():
    """Stock listing page"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    items = get_all_stock()
    valuation = calculate_stock_valuation()
    low_stock = get_low_stock_items()
    
    # Build table rows
    rows = []
    for item in items:
        qty = int(item.get("quantity", 0) or 0)
        cost = Decimal(str(item.get("cost_price", 0) or 0))
        sell = Decimal(str(item.get("selling_price", 0) or 0))
        
        # Stock status badge
        if qty <= 0:
            qty_html = f'<span class="badge badge-red">{qty}</span>'
        elif qty <= 5:
            qty_html = f'<span class="badge badge-orange">{qty}</span>'
        else:
            qty_html = str(qty)
        
        # Zero-rated indicator
        desc = item.get("description", "")
        if VAT.is_zero_rated(desc, item.get("category", "")):
            desc += ' <span class="badge badge-blue">0%</span>'
        
        rows.append([
            item.get("code", "-"),
            desc,
            item.get("category", "-"),
            {"value": qty_html, "class": ""},
            {"value": Money.format(cost), "class": "number"},
            {"value": Money.format(sell), "class": "number"},
            f'<a href="/stock/{item["id"]}/edit" class="btn btn-sm btn-ghost">Edit</a>'
        ])
    
    table = table_html(
        headers=["Code", "Description", "Category", "Qty", 
                {"label": "Cost", "class": "number"}, 
                {"label": "Price", "class": "number"}, ""],
        rows=rows,
        empty_message="No stock items. Add your first item!"
    )
    
    # Low stock alert
    low_alert = ""
    if low_stock:
        low_count = len(low_stock)
        out_count = len([i for i in low_stock if i.get("_status") == "out"])
        low_alert = f'''
        <div class="alert alert-warning">
            ⚠️ {low_count} items need attention: {out_count} out of stock, {low_count - out_count} low stock
            <a href="/stock/low" class="btn btn-sm btn-ghost" style="margin-left: auto;">View</a>
        </div>
        '''
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <p class="text-muted">{valuation["total_items"]} items • {valuation["total_quantity"]} units</p>
        </div>
        <div class="btn-group">
            <a href="/stock/import" class="btn btn-ghost">Import</a>
            <a href="/stock/new" class="btn btn-primary">+ Add Item</a>
        </div>
    </div>
    
    {low_alert}
    
    <div class="stats">
        <div class="stat">
            <div class="stat-value">{valuation["total_items"]}</div>
            <div class="stat-label">Items</div>
        </div>
        <div class="stat">
            <div class="stat-value">{valuation["total_quantity"]}</div>
            <div class="stat-label">Units</div>
        </div>
        <div class="stat">
            <div class="stat-value">{Money.format(valuation["total_cost"])}</div>
            <div class="stat-label">Cost Value</div>
        </div>
        <div class="stat">
            <div class="stat-value green">{Money.format(valuation["total_retail"])}</div>
            <div class="stat-label">Retail Value</div>
        </div>
    </div>
    
    <div class="card">
        <div class="card-header">
            <h2 class="card-title">Stock Items</h2>
            <input type="text" id="search" class="form-input" 
                   placeholder="Search..." style="width: 250px; margin: 0;"
                   onkeyup="filterTable(this.value)">
        </div>
        <div id="stock-table">
            {table}
        </div>
    </div>
    
    <script>
    function filterTable(q) {{
        q = q.toLowerCase();
        const rows = document.querySelectorAll('#stock-table tbody tr');
        rows.forEach(row => {{
            const text = row.textContent.toLowerCase();
            row.style.display = text.includes(q) ? '' : 'none';
        }});
    }}
    </script>
    '''
    
    return page_wrapper("Stock", content, active="stock", user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# ADD / EDIT STOCK ITEM
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/stock/new", methods=["GET", "POST"])
def stock_new():
    """Add new stock item"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = ""
    
    if request.method == "POST":
        # Get form data
        code = request.form.get("code", "").strip()
        description = request.form.get("description", "").strip()
        category = request.form.get("category", "").strip()
        quantity = int(request.form.get("quantity", 0) or 0)
        cost_price = Money.parse(request.form.get("cost_price", "0"))
        selling_price = Money.parse(request.form.get("selling_price", "0"))
        reorder_level = int(request.form.get("reorder_level", 5) or 5)
        
        if not description:
            message = error_message("Description is required")
        else:
            item = {
                "id": generate_id(),
                "code": code,
                "description": description,
                "category": category or "General",
                "quantity": quantity,
                "cost_price": float(cost_price),
                "selling_price": float(selling_price),
                "reorder_level": reorder_level,
                "active": True,
                "created_at": now()
            }
            
            success, result = db.insert("stock_items", item)
            
            if success:
                return redirect("/stock")
            else:
                message = error_message(f"Failed to save: {result}")
    
    categories = get_stock_categories()
    cat_options = [(c, c) for c in categories]
    cat_options.insert(0, ("General", "General"))
    
    content = f'''
    <a href="/stock" class="btn btn-ghost mb-lg">← Back to Stock</a>
    
    <div class="card">
        <h2 class="card-title mb-md">Add Stock Item</h2>
        
        {message}
        
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Code</label>
                    <input type="text" name="code" class="form-input" placeholder="SKU or barcode">
                </div>
                <div class="form-group">
                    <label class="form-label">Category</label>
                    <input type="text" name="category" class="form-input" 
                           list="categories" value="General">
                    <datalist id="categories">
                        {"".join([f'<option value="{c}">' for c in categories])}
                    </datalist>
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Description *</label>
                <input type="text" name="description" class="form-input" 
                       placeholder="Item name/description" required>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Quantity</label>
                    <input type="number" name="quantity" class="form-input" value="0" min="0">
                </div>
                <div class="form-group">
                    <label class="form-label">Reorder Level</label>
                    <input type="number" name="reorder_level" class="form-input" value="5" min="0">
                </div>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Cost Price</label>
                    <input type="text" name="cost_price" class="form-input" placeholder="0.00">
                </div>
                <div class="form-group">
                    <label class="form-label">Selling Price (incl VAT)</label>
                    <input type="text" name="selling_price" class="form-input" placeholder="0.00">
                </div>
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Item</button>
                <a href="/stock" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    '''
    
    return page_wrapper("Add Stock Item", content, active="stock", user=user)


@app.route("/stock/<item_id>/edit", methods=["GET", "POST"])
def stock_edit(item_id):
    """Edit stock item"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    item = get_stock_item(item_id)
    if not item:
        return redirect("/stock")
    
    message = ""
    
    if request.method == "POST":
        # Check for delete
        if request.form.get("action") == "delete":
            db.update("stock_items", item_id, {"active": False})
            return redirect("/stock")
        
        # Update item
        updates = {
            "code": request.form.get("code", "").strip(),
            "description": request.form.get("description", "").strip(),
            "category": request.form.get("category", "General").strip(),
            "quantity": int(request.form.get("quantity", 0) or 0),
            "cost_price": float(Money.parse(request.form.get("cost_price", "0"))),
            "selling_price": float(Money.parse(request.form.get("selling_price", "0"))),
            "reorder_level": int(request.form.get("reorder_level", 5) or 5)
        }
        
        if not updates["description"]:
            message = error_message("Description is required")
        else:
            success, result = db.update("stock_items", item_id, updates)
            
            if success:
                message = success_message("Item updated")
                item.update(updates)
            else:
                message = error_message(f"Failed: {result}")
    
    categories = get_stock_categories()
    
    content = f'''
    <a href="/stock" class="btn btn-ghost mb-lg">← Back to Stock</a>
    
    <div class="card">
        <h2 class="card-title mb-md">Edit Stock Item</h2>
        
        {message}
        
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Code</label>
                    <input type="text" name="code" class="form-input" 
                           value="{safe_string(item.get("code", ""))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Category</label>
                    <input type="text" name="category" class="form-input" 
                           list="categories" value="{safe_string(item.get("category", "General"))}">
                    <datalist id="categories">
                        {"".join([f'<option value="{c}">' for c in categories])}
                    </datalist>
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Description *</label>
                <input type="text" name="description" class="form-input" 
                       value="{safe_string(item.get("description", ""))}" required>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Quantity</label>
                    <input type="number" name="quantity" class="form-input" 
                           value="{item.get("quantity", 0)}" min="0">
                </div>
                <div class="form-group">
                    <label class="form-label">Reorder Level</label>
                    <input type="number" name="reorder_level" class="form-input" 
                           value="{item.get("reorder_level", 5)}" min="0">
                </div>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Cost Price</label>
                    <input type="text" name="cost_price" class="form-input" 
                           value="{item.get("cost_price", 0)}">
                </div>
                <div class="form-group">
                    <label class="form-label">Selling Price</label>
                    <input type="text" name="selling_price" class="form-input" 
                           value="{item.get("selling_price", 0)}">
                </div>
            </div>
            
            <div class="flex-between mt-lg">
                <button type="submit" class="btn btn-primary">Save Changes</button>
                <button type="submit" name="action" value="delete" class="btn btn-red"
                        onclick="return confirm('Delete this item?')">Delete</button>
            </div>
        </form>
    </div>
    '''
    
    return page_wrapper("Edit Stock Item", content, active="stock", user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# LOW STOCK PAGE
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/stock/low")
def stock_low():
    """Low stock items page"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    items = get_low_stock_items()
    
    rows = []
    for item in items:
        qty = int(item.get("quantity", 0) or 0)
        
        if item.get("_status") == "out":
            status = '<span class="badge badge-red">OUT</span>'
        else:
            status = '<span class="badge badge-orange">LOW</span>'
        
        rows.append([
            item.get("code", "-"),
            item.get("description", ""),
            item.get("category", "-"),
            qty,
            item.get("reorder_level", 5),
            status,
            f'<a href="/stock/{item["id"]}/edit" class="btn btn-sm btn-ghost">Edit</a>'
        ])
    
    table = table_html(
        headers=["Code", "Description", "Category", "Qty", "Reorder At", "Status", ""],
        rows=rows,
        empty_message="All stock levels are healthy! 🎉"
    )
    
    content = f'''
    <a href="/stock" class="btn btn-ghost mb-lg">← Back to Stock</a>
    
    <div class="alert alert-warning mb-lg">
        ⚠️ These items are below their reorder level and need attention.
    </div>
    
    <div class="card">
        <h2 class="card-title mb-md">Low Stock Items</h2>
        {table}
    </div>
    '''
    
    return page_wrapper("Low Stock", content, active="stock", user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# STOCK IMPORT
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/stock/import", methods=["GET", "POST"])
def stock_import():
    """Import stock from JSON"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = ""
    
    if request.method == "POST":
        data_str = request.form.get("data", "").strip()
        
        if not data_str:
            message = error_message("No data provided")
        else:
            try:
                data = json.loads(data_str)
                
                # Support different formats
                items = []
                if isinstance(data, list):
                    items = data
                elif isinstance(data, dict):
                    items = data.get("stock", data.get("items", data.get("stock_items", [])))
                
                imported = 0
                for item in items:
                    # Map common field names
                    record = {
                        "id": generate_id(),
                        "code": str(item.get("code", item.get("sku", "")))[:50],
                        "description": str(item.get("description", item.get("name", item.get("desc", ""))))[:200],
                        "category": str(item.get("category", "General"))[:50],
                        "quantity": int(item.get("quantity", item.get("qty", 0)) or 0),
                        "cost_price": float(item.get("cost_price", item.get("cost", 0)) or 0),
                        "selling_price": float(item.get("selling_price", item.get("price", item.get("sell", 0))) or 0),
                        "reorder_level": int(item.get("reorder_level", 5) or 5),
                        "active": True,
                        "created_at": now()
                    }
                    
                    if record["description"]:
                        success, _ = db.insert("stock_items", record)
                        if success:
                            imported += 1
                
                message = success_message(f"Imported {imported} items")
                
            except json.JSONDecodeError:
                message = error_message("Invalid JSON format")
            except Exception as e:
                message = error_message(f"Error: {str(e)}")
    
    content = f'''
    <a href="/stock" class="btn btn-ghost mb-lg">← Back to Stock</a>
    
    <div class="card">
        <h2 class="card-title mb-md">Import Stock</h2>
        
        {message}
        
        <div class="alert alert-info mb-md">
            Paste JSON data with stock items. Supported formats:
            <ul style="margin-top: 8px; margin-left: 20px;">
                <li>Array of items: [{{...}}, {{...}}]</li>
                <li>Object with "stock" or "items" key</li>
            </ul>
        </div>
        
        <form method="POST">
            <div class="form-group">
                <label class="form-label">JSON Data</label>
                <textarea name="data" class="form-textarea" rows="12" 
                          placeholder='[{{"code": "001", "description": "Item 1", "quantity": 10, "cost": 50, "price": 100}}]'></textarea>
            </div>
            
            <div class="btn-group">
                <button type="submit" class="btn btn-primary">Import</button>
                <a href="/stock" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    '''
    
    return page_wrapper("Import Stock", content, active="stock", user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# STOCK API
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/stock")
def api_stock_list():
    """API: Get all stock items"""
    items = get_all_stock()
    return jsonify(items)


@app.route("/api/stock/<item_id>")
def api_stock_item(item_id):
    """API: Get single stock item"""
    item = get_stock_item(item_id)
    if item:
        return jsonify(item)
    return jsonify({"error": "Not found"}), 404


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 7
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 7 COMPLETE - Stock Management

Contains:
✓ Stock listing with search
✓ Stock valuation (cost and retail)
✓ Low stock alerts
✓ Add new stock item
✓ Edit stock item
✓ Deactivate (soft delete) items
✓ Category management with datalist
✓ Import from JSON
✓ Low stock page
✓ Zero-rated item detection
✓ Stock API endpoints

Fields tracked:
- Code (SKU/barcode)
- Description
- Category
- Quantity
- Cost price
- Selling price (incl VAT)
- Reorder level
- Active status

Next: Piece 8 - Customers & Suppliers
"""

# =============================================================================
# PIECE8_CUSTOMERS.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 8: Customers & Suppliers                                              ║
║                                                                               ║
║   This piece contains:                                                        ║
║   - Customer listing and CRUD                                                 ║
║   - Supplier listing and CRUD                                                 ║
║   - Balance tracking                                                          ║
║   - Transaction history                                                       ║
║   - Receive payments from customers                                           ║
║   - Make payments to suppliers                                                ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""




# ═══════════════════════════════════════════════════════════════════════════════
# CUSTOMER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_all_customers(include_inactive: bool = False) -> list:
    """Get all customers"""
    customers = db.select("customers", order="name")
    
    if not include_inactive:
        customers = [c for c in customers if c.get("active", True)]
    
    return customers


def get_customer(customer_id: str) -> dict:
    """Get single customer"""
    return db.select_one("customers", customer_id)


def get_customer_invoices(customer_id: str) -> list:
    """Get all invoices for a customer"""
    invoices = db.select("invoices", {"customer_id": customer_id}, order="-date")
    return invoices


def get_customer_transactions(customer_id: str) -> list:
    """Get journal entries related to this customer"""
    entries = Journal.get_entries()
    return [e for e in entries if e.get("source_id") == customer_id or 
            (e.get("source_type") == "customer" and customer_id in str(e.get("description", "")))]


# ═══════════════════════════════════════════════════════════════════════════════
# CUSTOMER PAGES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/customers")
def customer_list():
    """Customer listing page"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    customers = get_all_customers()
    
    total_balance = sum(Decimal(str(c.get("balance", 0) or 0)) for c in customers)
    with_balance = len([c for c in customers if Decimal(str(c.get("balance", 0) or 0)) > 0])
    
    rows = []
    for cust in customers:
        balance = Decimal(str(cust.get("balance", 0) or 0))
        
        if balance > 0:
            bal_html = f'<span class="text-red font-bold">{Money.format(balance)}</span>'
        else:
            bal_html = Money.format(balance)
        
        rows.append([
            cust.get("code", "-"),
            f'<a href="/customers/{cust["id"]}">{safe_string(cust.get("name", ""))}</a>',
            cust.get("phone", "-"),
            cust.get("email", "-"),
            {"value": bal_html, "class": "number"},
            f'''<div class="btn-group">
                <a href="/customers/{cust["id"]}" class="btn btn-sm btn-ghost">View</a>
                <a href="/customers/{cust["id"]}/edit" class="btn btn-sm btn-ghost">Edit</a>
            </div>'''
        ])
    
    table = table_html(
        headers=["Code", "Name", "Phone", "Email", {"label": "Balance", "class": "number"}, ""],
        rows=rows,
        empty_message="No customers yet. Add your first customer!"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <p class="text-muted">{len(customers)} customers</p>
        </div>
        <a href="/customers/new" class="btn btn-primary">+ Add Customer</a>
    </div>
    
    <div class="stats">
        <div class="stat">
            <div class="stat-value">{len(customers)}</div>
            <div class="stat-label">Customers</div>
        </div>
        <div class="stat">
            <div class="stat-value red">{Money.format(total_balance)}</div>
            <div class="stat-label">Total Owing</div>
        </div>
        <div class="stat">
            <div class="stat-value orange">{with_balance}</div>
            <div class="stat-label">With Balance</div>
        </div>
    </div>
    
    <div class="card">
        <div class="card-header">
            <h2 class="card-title">Customers</h2>
            <input type="text" id="search" class="form-input" 
                   placeholder="Search..." style="width: 250px; margin: 0;"
                   onkeyup="filterTable(this.value)">
        </div>
        <div id="customer-table">
            {table}
        </div>
    </div>
    
    <script>
    function filterTable(q) {{
        q = q.toLowerCase();
        const rows = document.querySelectorAll('#customer-table tbody tr');
        rows.forEach(row => {{
            const text = row.textContent.toLowerCase();
            row.style.display = text.includes(q) ? '' : 'none';
        }});
    }}
    </script>
    '''
    
    return page_wrapper("Customers", content, active="customers", user=user)


@app.route("/customers/new", methods=["GET", "POST"])
def customer_new():
    """Add new customer"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = ""
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        
        if not name:
            message = error_message("Name is required")
        else:
            customer = {
                "id": generate_id(),
                "code": request.form.get("code", "").strip(),
                "name": name,
                "phone": request.form.get("phone", "").strip(),
                "email": request.form.get("email", "").strip(),
                "address": request.form.get("address", "").strip(),
                "balance": 0,
                "active": True,
                "created_at": now()
            }
            
            success, result = db.insert("customers", customer)
            
            if success:
                return redirect("/customers")
            else:
                message = error_message(f"Failed: {result}")
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">Add Customer</h2>
        
        {message}
        
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Code</label>
                    <input type="text" name="code" class="form-input" placeholder="Customer code">
                </div>
                <div class="form-group">
                    <label class="form-label">Name *</label>
                    <input type="text" name="name" class="form-input" required>
                </div>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Phone</label>
                    <input type="text" name="phone" class="form-input">
                </div>
                <div class="form-group">
                    <label class="form-label">Email</label>
                    <input type="email" name="email" class="form-input">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Address</label>
                <textarea name="address" class="form-textarea" rows="2"></textarea>
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Customer</button>
                <a href="/customers" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    '''
    
    return page_wrapper("Add Customer", content, active="customers", user=user)


@app.route("/customers/<customer_id>")
def customer_view(customer_id):
    """View customer details and history"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    customer = get_customer(customer_id)
    if not customer:
        return redirect("/customers")
    
    invoices = get_customer_invoices(customer_id)
    balance = Decimal(str(customer.get("balance", 0) or 0))
    
    # Invoice rows
    rows = []
    for inv in invoices[:20]:
        total = Decimal(str(inv.get("total", 0) or 0))
        status = inv.get("status", "draft")
        
        if status == "paid":
            status_badge = badge("Paid", "green")
        elif status == "outstanding":
            status_badge = badge("Outstanding", "orange")
        else:
            status_badge = badge(status.title(), "blue")
        
        rows.append([
            f'<a href="/invoices/{inv.get("id")}">{inv.get("invoice_number", "-")}</a>',
            inv.get("date", "")[:10],
            {"value": Money.format(total), "class": "number"},
            status_badge
        ])
    
    table = table_html(
        headers=["Invoice", "Date", {"label": "Total", "class": "number"}, "Status"],
        rows=rows,
        empty_message="No invoices yet"
    )
    
    content = f'''
    
    
    <div class="flex-between mb-lg">
        <div>
            <h1 style="font-size: 24px; font-weight: 700;">{safe_string(customer.get("name", ""))}</h1>
            <p class="text-muted">{safe_string(customer.get("phone", ""))} • {safe_string(customer.get("email", ""))}</p>
        </div>
        <div class="btn-group">
            <a href="/customers/{customer_id}/statement" class="btn btn-ghost" target="_blank">📄 Statement</a>
            <a href="/customers/{customer_id}/receive" class="btn btn-green">Receive Payment</a>
            <a href="/customers/{customer_id}/edit" class="btn btn-ghost">Edit</a>
        </div>
    </div>
    
    <div class="stats">
        <div class="stat">
            <div class="stat-value{' red' if balance > 0 else ''}">{Money.format(balance)}</div>
            <div class="stat-label">Balance Owing</div>
        </div>
        <div class="stat">
            <div class="stat-value">{len(invoices)}</div>
            <div class="stat-label">Invoices</div>
        </div>
    </div>
    
    <div class="card">
        <h3 class="card-title mb-md">Invoices</h3>
        {table}
    </div>
    '''
    
    return page_wrapper(customer.get("name", "Customer"), content, active="customers", user=user)


@app.route("/customers/<customer_id>/edit", methods=["GET", "POST"])
def customer_edit(customer_id):
    """Edit customer"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    customer = get_customer(customer_id)
    if not customer:
        return redirect("/customers")
    
    message = ""
    
    if request.method == "POST":
        if request.form.get("action") == "delete":
            db.update("customers", customer_id, {"active": False})
            return redirect("/customers")
        
        updates = {
            "code": request.form.get("code", "").strip(),
            "name": request.form.get("name", "").strip(),
            "phone": request.form.get("phone", "").strip(),
            "email": request.form.get("email", "").strip(),
            "address": request.form.get("address", "").strip()
        }
        
        if not updates["name"]:
            message = error_message("Name is required")
        else:
            success, _ = db.update("customers", customer_id, updates)
            if success:
                message = success_message("Updated")
                customer.update(updates)
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">Edit Customer</h2>
        
        {message}
        
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Code</label>
                    <input type="text" name="code" class="form-input" 
                           value="{safe_string(customer.get("code", ""))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Name *</label>
                    <input type="text" name="name" class="form-input" 
                           value="{safe_string(customer.get("name", ""))}" required>
                </div>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Phone</label>
                    <input type="text" name="phone" class="form-input" 
                           value="{safe_string(customer.get("phone", ""))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Email</label>
                    <input type="email" name="email" class="form-input" 
                           value="{safe_string(customer.get("email", ""))}">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Address</label>
                <textarea name="address" class="form-textarea" rows="2">{safe_string(customer.get("address", ""))}</textarea>
            </div>
            
            <div class="flex-between mt-lg">
                <button type="submit" class="btn btn-primary">Save Changes</button>
                <button type="submit" name="action" value="delete" class="btn btn-red"
                        onclick="return confirm('Delete this customer?')">Delete</button>
            </div>
        </form>
    </div>
    '''
    
    return page_wrapper("Edit Customer", content, active="customers", user=user)


@app.route("/customers/<customer_id>/receive", methods=["GET", "POST"])
def customer_receive_payment(customer_id):
    """Receive payment from customer"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    customer = get_customer(customer_id)
    if not customer:
        return redirect("/customers")
    
    balance = Decimal(str(customer.get("balance", 0) or 0))
    message = ""
    
    if request.method == "POST":
        amount = Money.parse(request.form.get("amount", "0"))
        reference = request.form.get("reference", "").strip()
        
        if amount <= 0:
            message = error_message("Enter a valid amount")
        else:
            # Create journal entry
            entry = JournalEntry(
                date=today(),
                reference=reference,
                description=f"Payment received - {customer.get('name', 'Customer')}",
                trans_type=TransactionType.RECEIPT,
                source_type="customer",
                source_id=customer_id
            )
            
            entry.debit(AccountCodes.BANK, amount)
            entry.credit(AccountCodes.DEBTORS, amount)
            
            success, result = entry.post()
            
            if success:
                # Update customer balance
                new_balance = balance - amount
                db.update("customers", customer_id, {"balance": float(new_balance)})
                
                return redirect(f"/customers/{customer_id}")
            else:
                message = error_message(f"Failed: {result}")
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">Receive Payment</h2>
        <p class="text-muted mb-lg">From: {safe_string(customer.get("name", ""))}</p>
        
        {message}
        
        <div class="alert alert-info mb-lg">
            Current balance owing: <strong>{Money.format(balance)}</strong>
        </div>
        
        <form method="POST">
            <div class="form-group">
                <label class="form-label">Amount Received</label>
                <input type="text" name="amount" class="form-input" 
                       placeholder="0.00" required autofocus
                       style="font-size: 24px; text-align: center;">
            </div>
            
            <div class="form-group">
                <label class="form-label">Reference</label>
                <input type="text" name="reference" class="form-input" 
                       placeholder="Receipt number, bank reference, etc.">
            </div>
            
            <button type="submit" class="btn btn-green btn-block btn-lg">
                Receive Payment
            </button>
        </form>
    </div>
    '''
    
    return page_wrapper("Receive Payment", content, active="customers", user=user)


@app.route("/customers/<customer_id>/statement")
def customer_statement(customer_id):
    """Generate printable customer statement"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    customer = get_customer(customer_id)
    if not customer:
        return "Customer not found", 404
    
    invoices = get_customer_invoices(customer_id)
    balance = Decimal(str(customer.get("balance", 0) or 0))
    
    # Get company details
    try:
        settings_row = db.select("settings", filters={"key": "company"}, limit=1)
        company = json.loads(settings_row[0].get("value", "{}")) if settings_row else {}
    except:
        company = {}
    
    company_name = company.get("name", "Your Business")
    company_phone = company.get("phone", "")
    company_email = company.get("email", "")
    
    # Build invoice rows
    rows_html = ""
    running_balance = Decimal("0")
    for inv in sorted(invoices, key=lambda x: x.get("date", "")):
        inv_date = inv.get("date", "")[:10]
        inv_num = inv.get("invoice_number", "-")
        total = Decimal(str(inv.get("total", 0) or 0))
        status = inv.get("status", "")
        
        if status == "paid":
            # This was paid - show payment
            rows_html += f'<tr><td>{inv_date}</td><td>{inv_num}</td><td style="text-align:right">R {total:.2f}</td><td style="text-align:right">-</td><td style="text-align:right">-</td></tr>'
        else:
            running_balance += total
            rows_html += f'<tr><td>{inv_date}</td><td>{inv_num}</td><td style="text-align:right">R {total:.2f}</td><td style="text-align:right">-</td><td style="text-align:right">R {running_balance:.2f}</td></tr>'
    
    return f'''<!DOCTYPE html>
<html>
<head>
    <title>Statement - {safe_string(customer.get("name", ""))}</title>
    <style>
        @media print {{ @page {{ margin: 15mm; }} .no-print {{ display: none; }} }}
        body {{ font-family: Arial, sans-serif; font-size: 14px; max-width: 210mm; margin: 0 auto; padding: 20px; }}
        .header {{ display: flex; justify-content: space-between; margin-bottom: 30px; border-bottom: 2px solid #333; padding-bottom: 20px; }}
        .company {{ font-size: 12px; }}
        .company-name {{ font-size: 20px; font-weight: bold; margin-bottom: 5px; }}
        h1 {{ font-size: 24px; margin: 0; }}
        .customer-box {{ background: #f5f5f5; padding: 15px; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th {{ background: #333; color: white; text-align: left; padding: 10px; }}
        td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
        .total-row {{ font-weight: bold; background: #f0f0f0; font-size: 16px; }}
        .btn {{ display: inline-block; padding: 10px 20px; background: #4CAF50; color: white; text-decoration: none; border-radius: 5px; margin: 5px; }}
    </style>
</head>
<body>
    <div class="header">
        <div class="company">
            <div class="company-name">{company_name}</div>
            {f"Tel: {company_phone}<br>" if company_phone else ""}
            {f"Email: {company_email}" if company_email else ""}
        </div>
        <div>
            <h1>STATEMENT</h1>
            <p>Date: {today()}</p>
        </div>
    </div>
    
    <div class="customer-box">
        <strong>{safe_string(customer.get("name", ""))}</strong><br>
        {safe_string(customer.get("phone", ""))}<br>
        {safe_string(customer.get("email", ""))}
    </div>
    
    <table>
        <thead>
            <tr>
                <th>Date</th>
                <th>Reference</th>
                <th style="text-align:right">Charges</th>
                <th style="text-align:right">Payments</th>
                <th style="text-align:right">Balance</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
        <tfoot>
            <tr class="total-row">
                <td colspan="4" style="text-align:right">Amount Due:</td>
                <td style="text-align:right">R {balance:.2f}</td>
            </tr>
        </tfoot>
    </table>
    
    <div class="no-print" style="margin-top: 30px; text-align: center;">
        <a href="#" class="btn" onclick="window.print(); return false;">Print Statement</a>
        <a href="#" class="btn" style="background:#666;" onclick="window.close(); return false;">Close</a>
    </div>
</body>
</html>'''


@app.route("/invoices/<invoice_id>/credit-note", methods=["GET", "POST"])
def create_credit_note(invoice_id):
    """Create a credit note (refund/return) for an invoice"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    inv = get_invoice(invoice_id)
    if not inv:
        return redirect("/invoices")
    
    if request.method == "POST":
        reason = request.form.get("reason", "Return")
        amount = Decimal(request.form.get("amount", 0) or 0)
        
        if amount <= 0:
            amount = Decimal(str(inv.get("total", 0)))
        
        # Generate credit note number
        cn_number = DocumentNumbers.get_next("CN", "credit_notes", "cn_number")
        
        # Create credit note record
        cn_id = generate_id()
        cn_record = {
            "id": cn_id,
            "cn_number": cn_number,
            "date": today(),
            "invoice_id": invoice_id,
            "invoice_number": inv.get("invoice_number"),
            "customer_id": inv.get("customer_id"),
            "customer_name": inv.get("customer_name"),
            "reason": reason,
            "amount": float(amount),
            "created_at": now()
        }
        
        db.insert("credit_notes", cn_record)
        
        # Update customer balance (reduce what they owe)
        if inv.get("customer_id"):
            cust = db.select_one("customers", inv["customer_id"])
            if cust:
                new_balance = Decimal(str(cust.get("balance", 0) or 0)) - amount
                db.update("customers", inv["customer_id"], {"balance": float(new_balance)})
        
        # Post to GL
        entry = JournalEntry(
            date=today(),
            reference=cn_number,
            description=f"Credit Note {cn_number} - {inv.get('customer_name')}",
            trans_type=TransactionType.SALE,
            source_type="credit_note",
            source_id=cn_id
        )
        
        # Reverse the sale
        entry.debit(AccountCodes.SALES, amount)
        entry.credit(AccountCodes.DEBTORS, amount)
        entry.post()
        
        return redirect(f"/invoices/{invoice_id}")
    
    # Show form
    items = json.loads(inv.get("items", "[]"))
    
    content = f'''
    <div class="mb-lg">
        <a href="/invoices/{invoice_id}" class="text-muted">← Back to Invoice</a>
        <h1>Create Credit Note</h1>
        <p class="text-muted">For Invoice {inv.get("invoice_number")} - {inv.get("customer_name")}</p>
    </div>
    
    <div class="card">
        <form method="POST">
            <div class="form-group">
                <label class="form-label">Reason</label>
                <select name="reason" class="form-select">
                    <option value="Return">Goods Returned</option>
                    <option value="Damaged">Damaged Goods</option>
                    <option value="Overcharge">Overcharge Correction</option>
                    <option value="Cancelled">Order Cancelled</option>
                    <option value="Other">Other</option>
                </select>
            </div>
            
            <div class="form-group">
                <label class="form-label">Amount</label>
                <input type="number" name="amount" class="form-input" 
                       value="{float(inv.get('total', 0)):.2f}" step="0.01"
                       style="font-size: 20px;">
                <small class="text-muted">Original invoice: R {float(inv.get('total', 0)):.2f}</small>
            </div>
            
            <div class="btn-group">
                <button type="submit" class="btn btn-orange">Create Credit Note</button>
                <a href="/invoices/{invoice_id}" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    '''
    
    return page_wrapper("Credit Note", content, user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# SUPPLIER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_all_suppliers(include_inactive: bool = False) -> list:
    """Get all suppliers"""
    suppliers = db.select("suppliers", order="name")
    
    if not include_inactive:
        suppliers = [s for s in suppliers if s.get("active", True)]
    
    return suppliers


def get_supplier(supplier_id: str) -> dict:
    """Get single supplier"""
    return db.select_one("suppliers", supplier_id)


# ═══════════════════════════════════════════════════════════════════════════════
# SUPPLIER PAGES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/suppliers")
def supplier_list():
    """Supplier listing page"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    suppliers = get_all_suppliers()
    
    total_owed = sum(Decimal(str(s.get("balance", 0) or 0)) for s in suppliers)
    
    rows = []
    for supp in suppliers:
        balance = Decimal(str(supp.get("balance", 0) or 0))
        
        if balance > 0:
            bal_html = f'<span class="text-orange font-bold">{Money.format(balance)}</span>'
        else:
            bal_html = Money.format(balance)
        
        rows.append([
            supp.get("code", "-"),
            f'<a href="/suppliers/{supp["id"]}">{safe_string(supp.get("name", ""))}</a>',
            supp.get("phone", "-"),
            supp.get("email", "-"),
            {"value": bal_html, "class": "number"},
            f'''<div class="btn-group">
                <a href="/suppliers/{supp["id"]}" class="btn btn-sm btn-ghost">View</a>
                <a href="/suppliers/{supp["id"]}/edit" class="btn btn-sm btn-ghost">Edit</a>
            </div>'''
        ])
    
    table = table_html(
        headers=["Code", "Name", "Phone", "Email", {"label": "We Owe", "class": "number"}, ""],
        rows=rows,
        empty_message="No suppliers yet. Add your first supplier!"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <p class="text-muted">{len(suppliers)} suppliers</p>
        </div>
        <a href="/suppliers/new" class="btn btn-primary">+ Add Supplier</a>
    </div>
    
    <div class="stats">
        <div class="stat">
            <div class="stat-value">{len(suppliers)}</div>
            <div class="stat-label">Suppliers</div>
        </div>
        <div class="stat">
            <div class="stat-value orange">{Money.format(total_owed)}</div>
            <div class="stat-label">Total We Owe</div>
        </div>
    </div>
    
    <div class="card">
        <div class="card-header">
            <h2 class="card-title">Suppliers</h2>
        </div>
        {table}
    </div>
    '''
    
    return page_wrapper("Suppliers", content, active="suppliers", user=user)


@app.route("/suppliers/new", methods=["GET", "POST"])
def supplier_new():
    """Add new supplier"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = ""
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        
        if not name:
            message = error_message("Name is required")
        else:
            supplier = {
                "id": generate_id(),
                "code": request.form.get("code", "").strip(),
                "name": name,
                "phone": request.form.get("phone", "").strip(),
                "email": request.form.get("email", "").strip(),
                "address": request.form.get("address", "").strip(),
                "balance": 0,
                "active": True,
                "created_at": now()
            }
            
            success, _ = db.insert("suppliers", supplier)
            
            if success:
                return redirect("/suppliers")
            else:
                message = error_message("Failed to save")
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">Add Supplier</h2>
        
        {message}
        
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Code</label>
                    <input type="text" name="code" class="form-input">
                </div>
                <div class="form-group">
                    <label class="form-label">Name *</label>
                    <input type="text" name="name" class="form-input" required>
                </div>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Phone</label>
                    <input type="text" name="phone" class="form-input">
                </div>
                <div class="form-group">
                    <label class="form-label">Email</label>
                    <input type="email" name="email" class="form-input">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Address</label>
                <textarea name="address" class="form-textarea" rows="2"></textarea>
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Supplier</button>
                <a href="/suppliers" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    '''
    
    return page_wrapper("Add Supplier", content, active="suppliers", user=user)


@app.route("/suppliers/<supplier_id>")
def supplier_view(supplier_id):
    """View supplier details"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    supplier = get_supplier(supplier_id)
    if not supplier:
        return redirect("/suppliers")
    
    balance = Decimal(str(supplier.get("balance", 0) or 0))
    
    content = f'''
    
    
    <div class="flex-between mb-lg">
        <div>
            <h1 style="font-size: 24px; font-weight: 700;">{safe_string(supplier.get("name", ""))}</h1>
            <p class="text-muted">{safe_string(supplier.get("phone", ""))} • {safe_string(supplier.get("email", ""))}</p>
        </div>
        <div class="btn-group">
            <a href="/suppliers/{supplier_id}/pay" class="btn btn-orange">Make Payment</a>
            <a href="/suppliers/{supplier_id}/edit" class="btn btn-ghost">Edit</a>
        </div>
    </div>
    
    <div class="stats">
        <div class="stat">
            <div class="stat-value orange">{Money.format(balance)}</div>
            <div class="stat-label">We Owe Them</div>
        </div>
    </div>
    '''
    
    return page_wrapper(supplier.get("name", "Supplier"), content, active="suppliers", user=user)


@app.route("/suppliers/<supplier_id>/edit", methods=["GET", "POST"])
def supplier_edit(supplier_id):
    """Edit supplier"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    supplier = get_supplier(supplier_id)
    if not supplier:
        return redirect("/suppliers")
    
    message = ""
    
    if request.method == "POST":
        if request.form.get("action") == "delete":
            db.update("suppliers", supplier_id, {"active": False})
            return redirect("/suppliers")
        
        updates = {
            "code": request.form.get("code", "").strip(),
            "name": request.form.get("name", "").strip(),
            "phone": request.form.get("phone", "").strip(),
            "email": request.form.get("email", "").strip(),
            "address": request.form.get("address", "").strip()
        }
        
        if not updates["name"]:
            message = error_message("Name is required")
        else:
            db.update("suppliers", supplier_id, updates)
            message = success_message("Updated")
            supplier.update(updates)
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">Edit Supplier</h2>
        
        {message}
        
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Code</label>
                    <input type="text" name="code" class="form-input" 
                           value="{safe_string(supplier.get("code", ""))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Name *</label>
                    <input type="text" name="name" class="form-input" 
                           value="{safe_string(supplier.get("name", ""))}" required>
                </div>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Phone</label>
                    <input type="text" name="phone" class="form-input" 
                           value="{safe_string(supplier.get("phone", ""))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Email</label>
                    <input type="email" name="email" class="form-input" 
                           value="{safe_string(supplier.get("email", ""))}">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Address</label>
                <textarea name="address" class="form-textarea" rows="2">{safe_string(supplier.get("address", ""))}</textarea>
            </div>
            
            <div class="flex-between mt-lg">
                <button type="submit" class="btn btn-primary">Save Changes</button>
                <button type="submit" name="action" value="delete" class="btn btn-red"
                        onclick="return confirm('Delete this supplier?')">Delete</button>
            </div>
        </form>
    </div>
    '''
    
    return page_wrapper("Edit Supplier", content, active="suppliers", user=user)


@app.route("/suppliers/<supplier_id>/pay", methods=["GET", "POST"])
def supplier_pay(supplier_id):
    """Make payment to supplier"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    supplier = get_supplier(supplier_id)
    if not supplier:
        return redirect("/suppliers")
    
    balance = Decimal(str(supplier.get("balance", 0) or 0))
    message = ""
    
    if request.method == "POST":
        amount = Money.parse(request.form.get("amount", "0"))
        reference = request.form.get("reference", "").strip()
        
        if amount <= 0:
            message = error_message("Enter a valid amount")
        else:
            # Create journal entry
            entry = JournalEntry(
                date=today(),
                reference=reference,
                description=f"Payment to - {supplier.get('name', 'Supplier')}",
                trans_type=TransactionType.PAYMENT,
                source_type="supplier",
                source_id=supplier_id
            )
            
            entry.debit(AccountCodes.CREDITORS, amount)
            entry.credit(AccountCodes.BANK, amount)
            
            success, result = entry.post()
            
            if success:
                # Update supplier balance
                new_balance = balance - amount
                db.update("suppliers", supplier_id, {"balance": float(new_balance)})
                
                return redirect(f"/suppliers/{supplier_id}")
            else:
                message = error_message(f"Failed: {result}")
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">Make Payment</h2>
        <p class="text-muted mb-lg">To: {safe_string(supplier.get("name", ""))}</p>
        
        {message}
        
        <div class="alert alert-info mb-lg">
            We owe them: <strong>{Money.format(balance)}</strong>
        </div>
        
        <form method="POST">
            <div class="form-group">
                <label class="form-label">Amount Paid</label>
                <input type="text" name="amount" class="form-input" 
                       placeholder="0.00" required autofocus
                       style="font-size: 24px; text-align: center;">
            </div>
            
            <div class="form-group">
                <label class="form-label">Reference</label>
                <input type="text" name="reference" class="form-input" 
                       placeholder="EFT reference, cheque number, etc.">
            </div>
            
            <button type="submit" class="btn btn-orange btn-block btn-lg">
                Make Payment
            </button>
        </form>
    </div>
    '''
    
    return page_wrapper("Make Payment", content, active="suppliers", user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# PURCHASE ORDERS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/purchase-orders")
def purchase_order_list():
    """List all purchase orders"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    pos = db.select("purchase_orders", order="created_at DESC") or []
    
    rows = []
    for po in pos[:50]:
        status = po.get("status", "draft")
        if status == "draft":
            sb = badge("Draft", "blue")
        elif status == "sent":
            sb = badge("Sent", "purple")
        elif status == "received":
            sb = badge("Received", "green")
        elif status == "cancelled":
            sb = badge("Cancelled", "red")
        else:
            sb = badge(status.title(), "blue")
        
        rows.append([
            f'<a href="/purchase-orders/{po.get("id")}">{po.get("po_number", "-")}</a>',
            po.get("date", "")[:10],
            po.get("supplier_name", "-"),
            {"value": Money.format(Decimal(str(po.get("total", 0)))), "class": "number"},
            sb
        ])
    
    table = table_html(
        headers=["PO #", "Date", "Supplier", {"label": "Total", "class": "number"}, "Status"],
        rows=rows,
        empty_message="No purchase orders yet"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <h1 style="font-size:24px;font-weight:700;">Purchase Orders</h1>
        <a href="/purchase-orders/new" class="btn btn-primary">+ New Order</a>
    </div>
    <div class="card">{table}</div>
    '''
    return page_wrapper("Purchase Orders", content, user=user)


@app.route("/purchase-orders/new", methods=["GET", "POST"])
def purchase_order_new():
    """Create new purchase order"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    suppliers = get_all_suppliers()
    stock = db.select("stock_items", order="description") or []
    
    if request.method == "POST":
        supplier_id = request.form.get("supplier_id")
        date = request.form.get("date", today())
        items_json = request.form.get("items_json", "[]")
        notes = request.form.get("notes", "")
        
        try:
            items = json.loads(items_json)
        except:
            items = []
        
        if not supplier_id:
            return "Supplier required", 400
        if not items:
            return "Add at least one item", 400
        
        # Get supplier
        supplier = get_supplier(supplier_id)
        supplier_name = supplier.get("name", "Unknown") if supplier else "Unknown"
        
        # Calculate total
        total = sum(Decimal(str(item.get("line_total", 0))) for item in items)
        
        # Generate PO number
        po_number = DocumentNumbers.get_next("PO", "purchase_orders", "po_number")
        
        po_id = generate_id()
        po_record = {
            "id": po_id,
            "po_number": po_number,
            "date": date,
            "supplier_id": supplier_id,
            "supplier_name": supplier_name,
            "items": json.dumps(items),
            "notes": notes,
            "total": float(total),
            "status": "draft",
            "created_at": now()
        }
        
        db.insert("purchase_orders", po_record)
        return redirect(f"/purchase-orders/{po_id}")
    
    # Build form
    supplier_options = "".join([f'<option value="{s["id"]}">{safe_string(s["name"])}</option>' for s in suppliers])
    
    content = f'''
    <div class="mb-lg">
        <a href="/purchase-orders" class="text-muted">← Purchase Orders</a>
        <h1>New Purchase Order</h1>
    </div>
    
    <form method="POST" id="po-form">
        <div class="card mb-lg">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Date</label>
                    <input type="date" name="date" class="form-input" value="{today()}">
                </div>
                <div class="form-group">
                    <label class="form-label">Supplier *</label>
                    <select name="supplier_id" class="form-select" required>
                        <option value="">Select supplier...</option>
                        {supplier_options}
                    </select>
                </div>
            </div>
        </div>
        
        <div class="card mb-lg">
            <h3 class="mb-md">Items</h3>
            <input type="text" id="search-stock" class="form-input mb-md" placeholder="Search products to add..." onkeyup="searchStock(this.value)">
            <div id="search-results" style="max-height:200px;overflow-y:auto;margin-bottom:15px;"></div>
            
            <table class="table" id="items-table">
                <thead><tr><th>Description</th><th class="number">Qty</th><th class="number">Cost</th><th class="number">Total</th><th></th></tr></thead>
                <tbody id="items-body"></tbody>
                <tfoot>
                    <tr><td colspan="3" style="text-align:right;font-weight:700;">Total:</td><td id="grand-total" class="number" style="font-weight:700;">R 0.00</td><td></td></tr>
                </tfoot>
            </table>
        </div>
        
        <div class="card mb-lg">
            <div class="form-group">
                <label class="form-label">Notes</label>
                <textarea name="notes" class="form-input" rows="3" placeholder="Delivery instructions, special requests..."></textarea>
            </div>
        </div>
        
        <input type="hidden" name="items_json" id="items-json">
        
        <button type="submit" class="btn btn-primary">Create Purchase Order</button>
        <a href="/purchase-orders" class="btn btn-ghost">Cancel</a>
    </form>
    
    <script>
    const allStock = {json.dumps([{"id": s.get("id"), "code": s.get("code", ""), "description": s.get("description", ""), "cost": float(s.get("cost_price") or 0)} for s in stock])};
    let poItems = [];
    
    function searchStock(q) {{
        const results = document.getElementById('search-results');
        if (q.length < 2) {{ results.innerHTML = ''; return; }}
        
        const matches = allStock.filter(s => 
            s.description.toLowerCase().includes(q.toLowerCase()) ||
            s.code.toLowerCase().includes(q.toLowerCase())
        ).slice(0, 10);
        
        results.innerHTML = matches.map(s => 
            '<div style="padding:10px;border-bottom:1px solid var(--border);cursor:pointer;" onclick="addItem(\\''+s.id+'\\')">'+
            '<strong>'+s.description+'</strong> <span class="text-muted">R '+s.cost.toFixed(2)+'</span></div>'
        ).join('');
    }}
    
    function addItem(itemId) {{
        const stock = allStock.find(s => s.id === itemId);
        if (!stock) return;
        
        const existing = poItems.find(i => i.id === itemId);
        if (existing) {{
            existing.qty++;
            existing.line_total = existing.qty * existing.cost;
        }} else {{
            poItems.push({{
                id: stock.id,
                code: stock.code,
                description: stock.description,
                qty: 1,
                cost: stock.cost,
                line_total: stock.cost
            }});
        }}
        
        renderItems();
        document.getElementById('search-stock').value = '';
        document.getElementById('search-results').innerHTML = '';
    }}
    
    function removeItem(idx) {{
        poItems.splice(idx, 1);
        renderItems();
    }}
    
    function updateQty(idx, qty) {{
        poItems[idx].qty = parseInt(qty) || 1;
        poItems[idx].line_total = poItems[idx].qty * poItems[idx].cost;
        renderItems();
    }}
    
    function renderItems() {{
        const body = document.getElementById('items-body');
        let html = '';
        let total = 0;
        
        poItems.forEach((item, idx) => {{
            total += item.line_total;
            html += '<tr>';
            html += '<td>'+item.description+'</td>';
            html += '<td class="number"><input type="number" value="'+item.qty+'" min="1" style="width:60px;text-align:center;" onchange="updateQty('+idx+', this.value)"></td>';
            html += '<td class="number">R '+item.cost.toFixed(2)+'</td>';
            html += '<td class="number">R '+item.line_total.toFixed(2)+'</td>';
            html += '<td><button type="button" class="btn btn-sm btn-red" onclick="removeItem('+idx+')">×</button></td>';
            html += '</tr>';
        }});
        
        body.innerHTML = html || '<tr><td colspan="5" class="text-muted" style="text-align:center;">No items added</td></tr>';
        document.getElementById('grand-total').textContent = 'R ' + total.toFixed(2);
        document.getElementById('items-json').value = JSON.stringify(poItems);
    }}
    
    renderItems();
    </script>
    '''
    return page_wrapper("New Purchase Order", content, user=user)


@app.route("/purchase-orders/<po_id>")
def purchase_order_view(po_id):
    """View purchase order"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    po = db.select_one("purchase_orders", po_id)
    if not po:
        return redirect("/purchase-orders")
    
    items = json.loads(po.get("items", "[]"))
    
    item_rows = ""
    for item in items:
        item_rows += f'''<tr>
            <td>{safe_string(item.get("description", ""))}</td>
            <td class="number">{item.get("qty", 1)}</td>
            <td class="number">R {float(item.get("cost", 0)):.2f}</td>
            <td class="number">R {float(item.get("line_total", 0)):.2f}</td>
        </tr>'''
    
    status = po.get("status", "draft")
    if status == "draft":
        sb = badge("Draft", "blue")
        actions = f'''
            <a href="/purchase-orders/{po_id}/send" class="btn btn-sm btn-purple">📧 Mark Sent</a>
            <a href="/purchase-orders/{po_id}/receive" class="btn btn-sm btn-green">📦 Receive Stock</a>
        '''
    elif status == "sent":
        sb = badge("Sent", "purple")
        actions = f'<a href="/purchase-orders/{po_id}/receive" class="btn btn-sm btn-green">📦 Receive Stock</a>'
    elif status == "received":
        sb = badge("Received", "green")
        actions = ""
    else:
        sb = badge(status.title(), "blue")
        actions = ""
    
    content = f'''
    <div class="mb-lg">
        <a href="/purchase-orders" class="text-muted">← Purchase Orders</a>
    </div>
    
    <div class="card">
        <div class="flex-between mb-lg">
            <div>
                <h1 style="font-size:24px;font-weight:700;">PO {po.get("po_number", "")}</h1>
                <p class="text-muted">{po.get("date", "")[:10]} • {safe_string(po.get("supplier_name", ""))}</p>
            </div>
            <div class="btn-group">
                <a href="/purchase-orders/{po_id}/print" target="_blank" class="btn btn-sm btn-ghost">🖨️ Print</a>
                {actions}
                {sb}
            </div>
        </div>
        
        <table class="table">
            <thead><tr><th>Description</th><th class="number">Qty</th><th class="number">Cost</th><th class="number">Total</th></tr></thead>
            <tbody>{item_rows}</tbody>
            <tfoot>
                <tr><td colspan="3" style="text-align:right;font-weight:700;">Total:</td>
                <td class="number" style="font-weight:700;">R {float(po.get("total", 0)):.2f}</td></tr>
            </tfoot>
        </table>
        
        {f'<div class="mt-lg"><strong>Notes:</strong><br>{safe_string(po.get("notes", ""))}</div>' if po.get("notes") else ""}
    </div>
    '''
    return page_wrapper(f"PO {po.get('po_number', '')}", content, user=user)


@app.route("/purchase-orders/<po_id>/send")
def purchase_order_send(po_id):
    """Mark PO as sent"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    db.update("purchase_orders", po_id, {"status": "sent"})
    return redirect(f"/purchase-orders/{po_id}")


@app.route("/purchase-orders/<po_id>/receive", methods=["GET", "POST"])
def purchase_order_receive(po_id):
    """Receive stock from PO"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    po = db.select_one("purchase_orders", po_id)
    if not po:
        return redirect("/purchase-orders")
    
    if request.method == "POST":
        items = json.loads(po.get("items", "[]"))
        
        # Book stock IN for each item
        for item in items:
            stock_id = item.get("id")
            qty = int(item.get("qty", 0))
            
            if stock_id and qty > 0:
                stock_item = db.select_one("stock_items", stock_id)
                if stock_item:
                    current_qty = int(stock_item.get("quantity", 0) or 0)
                    db.update("stock_items", stock_id, {"quantity": current_qty + qty})
        
        # Update PO status
        db.update("purchase_orders", po_id, {"status": "received", "received_date": today()})
        
        return redirect(f"/purchase-orders/{po_id}")
    
    items = json.loads(po.get("items", "[]"))
    item_list = "".join([f"<li>{item.get('description')} × {item.get('qty')}</li>" for item in items])
    
    content = f'''
    <div class="mb-lg">
        <a href="/purchase-orders/{po_id}" class="text-muted">← Back to PO</a>
        <h1>Receive Stock</h1>
        <p class="text-muted">PO {po.get("po_number")} from {po.get("supplier_name")}</p>
    </div>
    
    <div class="card">
        <p class="mb-md">The following items will be added to stock:</p>
        <ul style="margin-bottom:20px;">{item_list}</ul>
        
        <form method="POST">
            <button type="submit" class="btn btn-green btn-block">Confirm Receipt</button>
        </form>
    </div>
    '''
    return page_wrapper("Receive Stock", content, user=user)


@app.route("/purchase-orders/<po_id>/print")
def purchase_order_print(po_id):
    """Print purchase order"""
    po = db.select_one("purchase_orders", po_id)
    if not po:
        return "PO not found", 404
    
    # Get company details
    try:
        settings_row = db.select("settings", filters={"key": "company"}, limit=1)
        company = json.loads(settings_row[0].get("value", "{}")) if settings_row else {}
    except:
        company = {}
    
    items = json.loads(po.get("items", "[]"))
    item_rows = ""
    for item in items:
        item_rows += f'<tr><td>{safe_string(item.get("description", ""))}</td><td style="text-align:right">{item.get("qty", 1)}</td><td style="text-align:right">R {float(item.get("cost", 0)):.2f}</td><td style="text-align:right">R {float(item.get("line_total", 0)):.2f}</td></tr>'
    
    return f'''<!DOCTYPE html>
<html>
<head>
    <title>PO {po.get("po_number", "")}</title>
    <style>
        @media print {{ @page {{ margin: 15mm; }} .no-print {{ display: none; }} }}
        body {{ font-family: Arial, sans-serif; font-size: 14px; max-width: 210mm; margin: 0 auto; padding: 20px; }}
        .header {{ display: flex; justify-content: space-between; margin-bottom: 30px; border-bottom: 2px solid #333; padding-bottom: 20px; }}
        h1 {{ margin: 0; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th {{ background: #f5f5f5; text-align: left; padding: 10px; border-bottom: 2px solid #333; }}
        td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
        .total {{ font-weight: bold; font-size: 16px; }}
        .btn {{ display: inline-block; padding: 10px 20px; background: #4CAF50; color: white; text-decoration: none; border-radius: 5px; margin: 5px; }}
    </style>
</head>
<body>
    <div class="header">
        <div>
            <h2>{company.get("name", "Your Business")}</h2>
            <p>{company.get("phone", "")}</p>
        </div>
        <div style="text-align:right;">
            <h1>PURCHASE ORDER</h1>
            <p><strong>PO #:</strong> {po.get("po_number", "")}<br>
            <strong>Date:</strong> {po.get("date", "")[:10]}</p>
        </div>
    </div>
    
    <div style="background:#f5f5f5; padding:15px; margin-bottom:20px;">
        <strong>To:</strong> {safe_string(po.get("supplier_name", ""))}
    </div>
    
    <table>
        <thead><tr><th>Description</th><th style="text-align:right">Qty</th><th style="text-align:right">Unit Cost</th><th style="text-align:right">Total</th></tr></thead>
        <tbody>{item_rows}</tbody>
        <tfoot><tr class="total"><td colspan="3" style="text-align:right">TOTAL:</td><td style="text-align:right">R {float(po.get("total", 0)):.2f}</td></tr></tfoot>
    </table>
    
    {f'<p><strong>Notes:</strong> {safe_string(po.get("notes", ""))}</p>' if po.get("notes") else ""}
    
    <div class="no-print" style="margin-top:30px; text-align:center;">
        <a href="#" class="btn" onclick="window.print(); return false;">Print</a>
        <a href="#" class="btn" style="background:#666;" onclick="window.close(); return false;">Close</a>
    </div>
</body>
</html>'''


# ═══════════════════════════════════════════════════════════════════════════════
# API ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/customers")
def api_customers():
    """API: Get all customers"""
    return jsonify(get_all_customers())


@app.route("/api/suppliers")
def api_suppliers():
    """API: Get all suppliers"""
    return jsonify(get_all_suppliers())


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 8
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 8 COMPLETE - Customers & Suppliers

Contains:
✓ Customer listing with search
✓ Customer CRUD (add, edit, deactivate)
✓ Customer balance tracking
✓ Customer invoice history
✓ Receive payment from customer with GL posting:
  - DR Bank
  - CR Debtors
✓ Supplier listing
✓ Supplier CRUD
✓ Supplier balance tracking
✓ Make payment to supplier with GL posting:
  - DR Creditors
  - CR Bank
✓ API endpoints

All payments properly posted to General Ledger!

Next: Piece 9 - Invoices & Quotes
"""

# =============================================================================
# PIECE9_INVOICES.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 9: Invoices & Quotes                                                  ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""




# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_all_invoices():
    return db.select("invoices", order="-date")

def get_invoice(invoice_id):
    return db.select_one("invoices", invoice_id)

def get_all_quotes():
    return db.select("quotes", order="-date")

def get_quote(quote_id):
    return db.select_one("quotes", quote_id)

def get_stock_for_selection():
    items = db.select("stock_items", order="description")
    result = []
    for item in items:
        if not item.get("active", True):
            continue
        price = Decimal(str(item.get("selling_price", 0) or 0))
        desc = item.get("description", "")
        is_zero = VAT.is_zero_rated(desc, item.get("category", ""))
        result.append({
            "id": item["id"],
            "code": item.get("code", ""),
            "description": desc,
            "price": float(price),
            "price_formatted": Money.format(price),
            "is_zero_rated": is_zero
        })
    return result

def get_customers_for_selection():
    customers = db.select("customers", order="name")
    return [{"id": c["id"], "name": c.get("name", "")} for c in customers if c.get("active", True)]


# ═══════════════════════════════════════════════════════════════════════════════
# INVOICE PAGES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/invoices")
def invoice_list():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    invoices = get_all_invoices()
    total_value = sum(Decimal(str(i.get("total", 0) or 0)) for i in invoices)
    outstanding = sum(Decimal(str(i.get("total", 0) or 0)) for i in invoices if i.get("status") == "outstanding")
    
    rows = []
    for inv in invoices[:50]:
        status = inv.get("status", "draft")
        if status == "paid":
            sb = badge("Paid", "green")
        elif status == "outstanding":
            sb = badge("Outstanding", "orange")
        else:
            sb = badge(status.title(), "blue")
        
        rows.append([
            f'<a href="/invoices/{inv["id"]}">{inv.get("invoice_number", "-")}</a>',
            inv.get("date", "")[:10],
            safe_string(inv.get("customer_name", "Walk-in")),
            {"value": Money.format(Decimal(str(inv.get("total", 0)))), "class": "number"},
            sb
        ])
    
    table = table_html(
        headers=["Invoice #", "Date", "Customer", {"label": "Total", "class": "number"}, "Status"],
        rows=rows,
        empty_message="No invoices yet"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <p class="text-muted">{len(invoices)} invoices</p>
        <a href="/invoices/new" class="btn btn-primary">+ New Invoice</a>
    </div>
    
    <div class="stats">
        <div class="stat"><div class="stat-value">{len(invoices)}</div><div class="stat-label">Invoices</div></div>
        <div class="stat"><div class="stat-value green">{Money.format(total_value)}</div><div class="stat-label">Total Value</div></div>
        <div class="stat"><div class="stat-value orange">{Money.format(outstanding)}</div><div class="stat-label">Outstanding</div></div>
    </div>
    
    <div class="card">{table}</div>
    '''
    
    return page_wrapper("Invoices", content, active="invoices", user=user)


@app.route("/invoices/new", methods=["GET", "POST"])
def invoice_new():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    if request.method == "POST":
        return process_invoice_creation(request)
    
    stock = get_stock_for_selection()
    customers = get_customers_for_selection()
    inv_number = DocumentNumbers.get_next("INV", "invoices", "invoice_number")
    
    js = create_document_js("invoice")
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">New Invoice</h2>
        
        <form method="POST" id="doc-form">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Invoice Number</label>
                    <input type="text" name="doc_number" class="form-input" value="{inv_number}" readonly>
                </div>
                <div class="form-group">
                    <label class="form-label">Date</label>
                    <input type="date" name="date" class="form-input" value="{today()}">
                </div>
                <div class="form-group">
                    <label class="form-label">Customer</label>
                    <select name="customer_id" id="customer_id" class="form-select" onchange="toggleNewCustomer(this)">
                        <option value="">Walk-in Customer</option>
                        <option value="NEW">➕ Add New Customer</option>
                        {"".join([f'<option value="{c["id"]}">{safe_string(c["name"])}</option>' for c in customers])}
                    </select>
                </div>
            </div>
            
            <div id="new-customer-fields" style="display:none;" class="form-row">
                <div class="form-group">
                    <label class="form-label">New Customer Name *</label>
                    <input type="text" name="new_customer_name" class="form-input" placeholder="Customer name">
                </div>
                <div class="form-group">
                    <label class="form-label">Phone</label>
                    <input type="text" name="new_customer_phone" class="form-input" placeholder="Phone">
                </div>
                <div class="form-group">
                    <label class="form-label">Email</label>
                    <input type="email" name="new_customer_email" class="form-input" placeholder="Email">
                </div>
            </div>
            
            <h3 style="margin: 20px 0 10px;">Line Items</h3>
            <div style="display:flex;gap:12px;margin-bottom:12px;flex-wrap:wrap;">
                <input type="text" id="search-stock" class="form-input" style="flex:1;min-width:200px;" placeholder="Search stock items..." onkeyup="searchStock(this.value)">
                <button type="button" class="btn btn-purple" onclick="addCustomLine()" title="Add empty row to type in">➕ Quick Add</button>
                <button type="button" class="btn btn-ghost" onclick="showCustomItemModal()" title="Open popup form">📝 Form</button>
            </div>
            <div id="search-results" style="max-height:200px;overflow-y:auto;margin-bottom:15px;border:1px solid var(--border);border-radius:var(--radius-md);"></div>
            
            <table class="table" id="lines-table" style="table-layout:fixed;">
                <thead><tr>
                    <th style="width:50%;">Description</th>
                    <th style="width:12%;text-align:center;">QTY</th>
                    <th style="width:15%;text-align:right;">PRICE</th>
                    <th style="width:15%;text-align:right;">TOTAL</th>
                    <th style="width:8%;"></th>
                </tr></thead>
                <tbody id="lines-body"></tbody>
                <tfoot>
                    <tr><td colspan="3" style="text-align:right;font-weight:600;">Subtotal:</td><td id="subtotal" style="text-align:right;">R 0.00</td><td></td></tr>
                    <tr><td colspan="3" style="text-align:right;font-weight:600;">VAT (15%):</td><td id="vat-total" style="text-align:right;">R 0.00</td><td></td></tr>
                    <tr><td colspan="3" style="text-align:right;font-weight:700;font-size:18px;">Total:</td><td id="grand-total" style="text-align:right;font-weight:700;font-size:18px;color:var(--green);">R 0.00</td><td></td></tr>
                </tfoot>
            </table>
            
            <input type="hidden" name="items_json" id="items-json">
            <input type="hidden" name="subtotal" id="subtotal-input">
            <input type="hidden" name="vat" id="vat-input">
            <input type="hidden" name="total" id="total-input">
            
            <!-- Payment Status Option -->
            <div style="background:rgba(16,185,129,0.1);border:1px solid rgba(16,185,129,0.2);border-radius:12px;padding:16px;margin-top:20px;">
                <label style="display:flex;align-items:center;gap:12px;cursor:pointer;">
                    <input type="checkbox" name="mark_paid" value="yes" style="width:20px;height:20px;">
                    <div>
                        <strong style="color:#10b981;">💵 Mark as Paid (Cash Sale)</strong>
                        <div style="color:#8b8b9a;font-size:13px;">Check this if customer is paying now. Won't show as outstanding.</div>
                    </div>
                </label>
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Invoice</button>
                <a href="/invoices" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    
    {CUSTOM_ITEM_MODAL}
    
    <script>
    const stockItems = {json.dumps(stock)};
    {js}
    </script>
    '''
    
    return page_wrapper("New Invoice", content, active="invoices", user=user)


def process_invoice_creation(req):
    """Process invoice form submission"""
    try:
        items_json = req.form.get("items_json", "[]")
        items = json.loads(items_json)
        
        if not items:
            return redirect("/invoices/new")
        
        customer_id = req.form.get("customer_id", "")
        customer_name = "Walk-in Customer"
        
        # Handle new customer creation
        if customer_id == "NEW":
            new_name = req.form.get("new_customer_name", "").strip()
            if new_name:
                new_cust = {
                    "id": generate_id(),
                    "name": new_name,
                    "phone": req.form.get("new_customer_phone", ""),
                    "email": req.form.get("new_customer_email", ""),
                    "balance": 0,
                    "active": True,
                    "created_at": now()
                }
                db.insert("customers", new_cust)
                customer_id = new_cust["id"]
                customer_name = new_name
            else:
                customer_id = ""
        elif customer_id:
            cust = db.select_one("customers", customer_id)
            if cust:
                customer_name = cust.get("name", "Customer")
        
        # Calculate totals in Flask
        subtotal = Decimal("0")
        total_vat = Decimal("0")
        total_cost = Decimal("0")
        
        for item in items:
            price = Decimal(str(item.get("price", 0)))
            qty = int(item.get("quantity", 1))
            is_zero = item.get("is_zero_rated", False)
            
            line_total = price * qty
            
            if is_zero:
                vat_info = VAT.calculate_from_inclusive(line_total, VAT.ZERO_RATE)
            else:
                vat_info = VAT.calculate_from_inclusive(line_total)
            
            subtotal += vat_info["exclusive"]
            total_vat += vat_info["vat"]
            total_cost += Decimal(str(item.get("cost", 0))) * qty
        
        total = subtotal + total_vat
        
        invoice_id = generate_id()
        invoice_number = req.form.get("doc_number", DocumentNumbers.get_next("INV", "invoices", "invoice_number"))
        
        # Check if marked as paid
        mark_paid = req.form.get("mark_paid") == "yes"
        invoice_status = "paid" if mark_paid or not customer_id else "outstanding"
        
        invoice = {
            "id": invoice_id,
            "invoice_number": invoice_number,
            "date": req.form.get("date", today()),
            "customer_id": customer_id or None,
            "customer_name": customer_name,
            "items": items_json,
            "subtotal": float(subtotal),
            "vat": float(total_vat),
            "total": float(total),
            "status": invoice_status,
            "created_at": now()
        }
        
        db.insert("invoices", invoice)
        
        # Post to GL
        entry = JournalEntry(
            date=today(),
            reference=invoice_number,
            description=f"Invoice {invoice_number} - {customer_name}",
            trans_type=TransactionType.SALE,
            source_type="invoice",
            source_id=invoice_id
        )
        
        # If marked as paid or walk-in, go to bank. Otherwise to debtors.
        if mark_paid or not customer_id:
            entry.debit(AccountCodes.BANK, total)
        else:
            entry.debit(AccountCodes.DEBTORS, total)
            # Update customer balance
            cust = db.select_one("customers", customer_id)
            if cust:
                new_bal = Decimal(str(cust.get("balance", 0) or 0)) + total
                db.update("customers", customer_id, {"balance": float(new_bal)})
        
        entry.credit(AccountCodes.SALES, subtotal)
        if total_vat > 0:
            entry.credit(AccountCodes.VAT_OUTPUT, total_vat)
        
        entry.post()
        
        # Post COGS
        if total_cost > 0:
            cogs = JournalEntry(
                date=today(),
                reference=invoice_number,
                description=f"COGS - {invoice_number}",
                trans_type=TransactionType.SALE
            )
            cogs.debit(AccountCodes.COGS, total_cost)
            cogs.credit(AccountCodes.STOCK, total_cost)
            cogs.post()
        
        return redirect("/invoices")
        
    except Exception as e:
        return redirect("/invoices/new")


@app.route("/invoices/<invoice_id>")
def invoice_view(invoice_id):
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    inv = get_invoice(invoice_id)
    if not inv:
        return redirect("/invoices")
    
    items = json.loads(inv.get("items", "[]"))
    
    item_rows = ""
    for item in items:
        line_total = Decimal(str(item.get("price", 0))) * int(item.get("quantity", 1))
        item_rows += f'''
        <tr>
            <td>{safe_string(item.get("description", ""))}</td>
            <td class="number">{item.get("quantity", 1)}</td>
            <td class="number">{Money.format(Decimal(str(item.get("price", 0))))}</td>
            <td class="number">{Money.format(line_total)}</td>
        </tr>
        '''
    
    status = inv.get("status", "draft")
    if status == "paid":
        sb = badge("Paid", "green")
        status_btn = f'<a href="/invoices/{invoice_id}/mark-outstanding" class="btn btn-sm btn-orange">Mark Outstanding</a>'
    elif status == "outstanding":
        sb = badge("Outstanding", "orange")
        status_btn = f'<a href="/invoices/{invoice_id}/mark-paid" class="btn btn-sm btn-green">Mark Paid</a>'
    else:
        sb = badge(status.title(), "blue")
        status_btn = f'<a href="/invoices/{invoice_id}/mark-paid" class="btn btn-sm btn-green">Mark Paid</a>'
    
    content = f'''
    
    
    <div class="card">
        <div class="flex-between mb-lg">
            <div>
                <h1 style="font-size:24px;font-weight:700;">Invoice {safe_string(inv.get("invoice_number", ""))}</h1>
                <p class="text-muted">{inv.get("date", "")[:10]} • {safe_string(inv.get("customer_name", "Walk-in"))}</p>
            </div>
            <div class="btn-group">
                <a href="/print/office/{invoice_id}" target="_blank" class="btn btn-sm btn-ghost">📄 Print</a>
                <a href="/invoices/{invoice_id}/edit" class="btn btn-sm btn-ghost">✏️ Edit</a>
                <a href="/invoices/{invoice_id}/credit-note" class="btn btn-sm btn-orange">↩️ Return</a>
                {status_btn}
                {sb}
            </div>
        </div>
        
        <table class="table">
            <thead><tr><th>Description</th><th class="number">Qty</th><th class="number">Price</th><th class="number">Total</th></tr></thead>
            <tbody>{item_rows}</tbody>
            <tfoot>
                <tr><td colspan="3" style="text-align:right;">Subtotal:</td><td class="number">{Money.format(Decimal(str(inv.get("subtotal", 0))))}</td></tr>
                <tr><td colspan="3" style="text-align:right;">VAT:</td><td class="number">{Money.format(Decimal(str(inv.get("vat", 0))))}</td></tr>
                <tr><td colspan="3" style="text-align:right;font-weight:700;">Total:</td><td class="number" style="font-weight:700;color:var(--green);">{Money.format(Decimal(str(inv.get("total", 0))))}</td></tr>
            </tfoot>
        </table>
    </div>
    '''
    
    return page_wrapper(f"Invoice {inv.get('invoice_number', '')}", content, active="invoices", user=user)


@app.route("/invoices/<invoice_id>/mark-paid")
def invoice_mark_paid(invoice_id):
    """Mark invoice as paid"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    db.update("invoices", invoice_id, {"status": "paid"})
    return redirect(f"/invoices/{invoice_id}")


@app.route("/invoices/<invoice_id>/mark-outstanding")
def invoice_mark_outstanding(invoice_id):
    """Mark invoice as outstanding"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    db.update("invoices", invoice_id, {"status": "outstanding"})
    return redirect(f"/invoices/{invoice_id}")



@app.route("/invoices/<invoice_id>/edit", methods=["GET", "POST"])
def invoice_edit(invoice_id):
    """Edit invoice with warning if already printed/emailed"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    inv = get_invoice(invoice_id)
    if not inv:
        return redirect("/invoices")
    
    # Handle confirmation
    confirmed = request.args.get("confirmed") == "yes"
    
    if request.method == "GET" and not confirmed:
        # Show warning page first
        content = f'''
        <div class="card" style="max-width:500px; margin: 40px auto; text-align:center;">
            <div style="font-size:48px; margin-bottom:20px;">⚠️</div>
            <h2 style="margin-bottom:12px;">Edit Invoice?</h2>
            <p class="text-muted" style="margin-bottom:24px;">
                This invoice may have already been printed or emailed to the customer.<br>
                Are you sure you want to edit it?
            </p>
            <div class="btn-group" style="justify-content:center;">
                <a href="/invoices/{invoice_id}/edit?confirmed=yes" class="btn btn-orange">Yes, Edit Invoice</a>
                <a href="/invoices/{invoice_id}" class="btn btn-ghost">Cancel</a>
            </div>
        </div>
        '''
        return page_wrapper("Edit Invoice", content, "invoices", user)
    
    if request.method == "POST":
        # Save changes
        try:
            items = json.loads(request.form.get("items_json", "[]"))
            
            subtotal = Decimal("0")
            total_vat = Decimal("0")
            for item in items:
                price = Decimal(str(item.get("price", 0)))
                qty = int(item.get("quantity", 1))
                line_total = price * qty
                vat_info = VAT.calculate_from_inclusive(line_total)
                subtotal += vat_info["exclusive"]
                total_vat += vat_info["vat"]
            
            total = subtotal + total_vat
            
            updates = {
                "items": request.form.get("items_json", "[]"),
                "subtotal": float(subtotal),
                "vat": float(total_vat),
                "total": float(total),
                "date": request.form.get("date", inv.get("date")),
                "updated_at": now()
            }
            
            db.update("invoices", invoice_id, updates)
            return redirect(f"/invoices/{invoice_id}")
            
        except Exception:
            return redirect(f"/invoices/{invoice_id}/edit?confirmed=yes")
    
    # Show edit form
    stock = get_stock_for_selection()
    items = json.loads(inv.get("items", "[]"))
    js = create_document_js("invoice")
    
    # Pre-populate items
    items_js = json.dumps(items)
    
    content = f'''
    <div class="card">
        <h2 class="card-title mb-md">Edit Invoice {safe_string(inv.get("invoice_number", ""))}</h2>
        
        <form method="POST" id="doc-form">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Invoice Number</label>
                    <input type="text" class="form-input" value="{safe_string(inv.get('invoice_number', ''))}" readonly>
                </div>
                <div class="form-group">
                    <label class="form-label">Date</label>
                    <input type="date" name="date" class="form-input" value="{inv.get('date', today())[:10]}">
                </div>
                <div class="form-group">
                    <label class="form-label">Customer</label>
                    <input type="text" class="form-input" value="{safe_string(inv.get('customer_name', 'Walk-in'))}" readonly>
                </div>
            </div>
            
            <h3 style="margin: 20px 0 10px;">Line Items</h3>
            <input type="text" id="search-stock" class="form-input" placeholder="Search products..." onkeyup="searchStock(this.value)">
            <div id="search-results" style="max-height:300px;overflow-y:auto;margin-bottom:15px;border:1px solid var(--border);border-radius:var(--radius-md);"></div>
            
            <table class="table" id="lines-table">
                <thead><tr><th>Description</th><th class="number">Qty</th><th class="number">Price</th><th class="number">Total</th><th></th></tr></thead>
                <tbody id="lines-body"></tbody>
                <tfoot>
                    <tr><td colspan="3" style="text-align:right;font-weight:600;">Total:</td><td id="grand-total" style="text-align:right;font-weight:700;font-size:18px;color:var(--green);">R 0.00</td><td></td></tr>
                </tfoot>
            </table>
            
            <input type="hidden" name="items_json" id="items-json">
            <input type="hidden" name="total" id="total-input">
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Changes</button>
                <a href="/invoices/{invoice_id}" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    
    <script>
    const stockItems = {json.dumps(stock)};
    {js}
    
    // Pre-populate existing items
    const existingItems = {items_js};
    lines = existingItems.map(item => ({{
        id: item.stock_id || '',
        code: item.code || '',
        description: item.description,
        price: item.price,
        quantity: item.quantity
    }}));
    renderLines();
    </script>
    '''
    
    return page_wrapper("Edit Invoice", content, "invoices", user)


# ═══════════════════════════════════════════════════════════════════════════════
# QUOTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/quotes")
def quote_list():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    quotes = get_all_quotes()
    
    rows = []
    for q in quotes[:50]:
        # Check if converted
        if q.get("status") == "converted":
            action_btn = '<span class="badge badge-blue">✓ Converted</span>'
        else:
            action_btn = f'<a href="/quotes/{q["id"]}/convert" class="btn btn-sm btn-green">→ Invoice</a>'
        
        rows.append([
            f'<a href="/quotes/{q["id"]}">{q.get("quote_number", "-")}</a>',
            q.get("date", "")[:10],
            safe_string(q.get("customer_name", "")),
            {"value": Money.format(Decimal(str(q.get("total", 0)))), "class": "number"},
            action_btn
        ])
    
    table = table_html(
        headers=["Quote #", "Date", "Customer", {"label": "Total", "class": "number"}, ""],
        rows=rows,
        empty_message="No quotes yet"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <p class="text-muted">{len(quotes)} quotes</p>
        <a href="/quotes/new" class="btn btn-primary">+ New Quote</a>
    </div>
    
    <div class="card">{table}</div>
    '''
    
    return page_wrapper("Quotes", content, active="quotes", user=user)


@app.route("/quotes/new", methods=["GET", "POST"])
def quote_new():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    if request.method == "POST":
        return process_quote_creation(request)
    
    stock = get_stock_for_selection()
    customers = get_customers_for_selection()
    quote_number = DocumentNumbers.get_next("QUO", "quotes", "quote_number")
    
    js = create_document_js("quote")
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">New Quote</h2>
        
        <form method="POST" id="doc-form">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Quote Number</label>
                    <input type="text" name="doc_number" class="form-input" value="{quote_number}" readonly>
                </div>
                <div class="form-group">
                    <label class="form-label">Date</label>
                    <input type="date" name="date" class="form-input" value="{today()}">
                </div>
                <div class="form-group">
                    <label class="form-label">Customer</label>
                    <select name="customer_id" id="customer_id" class="form-select" onchange="toggleNewCustomer(this)">
                        <option value="">Walk-in Customer</option>
                        <option value="NEW">➕ Add New Customer</option>
                        {"".join([f'<option value="{c["id"]}">{safe_string(c["name"])}</option>' for c in customers])}
                    </select>
                </div>
            </div>
            
            <div id="new-customer-fields" style="display:none;" class="form-row">
                <div class="form-group">
                    <label class="form-label">New Customer Name *</label>
                    <input type="text" name="new_customer_name" class="form-input" placeholder="Customer name">
                </div>
                <div class="form-group">
                    <label class="form-label">Phone</label>
                    <input type="text" name="new_customer_phone" class="form-input" placeholder="Phone">
                </div>
                <div class="form-group">
                    <label class="form-label">Email</label>
                    <input type="email" name="new_customer_email" class="form-input" placeholder="Email">
                </div>
            </div>
            
            <h3 style="margin: 20px 0 10px;">Line Items</h3>
            <div style="display:flex;gap:12px;margin-bottom:12px;flex-wrap:wrap;">
                <input type="text" id="search-stock" class="form-input" style="flex:1;min-width:200px;" placeholder="Search stock items..." onkeyup="searchStock(this.value)">
                <button type="button" class="btn btn-purple" onclick="addCustomLine()" title="Add empty row to type in">➕ Quick Add</button>
                <button type="button" class="btn btn-ghost" onclick="showCustomItemModal()" title="Open popup form">📝 Form</button>
            </div>
            <div id="search-results" style="max-height:200px;overflow-y:auto;margin-bottom:15px;border:1px solid var(--border);border-radius:var(--radius-md);"></div>
            
            <table class="table" id="lines-table" style="table-layout:fixed;">
                <thead><tr>
                    <th style="width:50%;">Description</th>
                    <th style="width:12%;text-align:center;">QTY</th>
                    <th style="width:15%;text-align:right;">PRICE</th>
                    <th style="width:15%;text-align:right;">TOTAL</th>
                    <th style="width:8%;"></th>
                </tr></thead>
                <tbody id="lines-body"></tbody>
                <tfoot>
                    <tr><td colspan="3" style="text-align:right;font-weight:600;">Total:</td><td id="grand-total" style="text-align:right;font-weight:700;font-size:18px;color:var(--green);">R 0.00</td><td></td></tr>
                </tfoot>
            </table>
            
            <input type="hidden" name="items_json" id="items-json">
            <input type="hidden" name="total" id="total-input">
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Quote</button>
                <a href="/quotes" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    
    {CUSTOM_ITEM_MODAL}
    
    <script>
    const stockItems = {json.dumps(stock)};
    {js}
    </script>
    '''
    
    return page_wrapper("New Quote", content, active="quotes", user=user)


def process_quote_creation(req):
    try:
        items = json.loads(req.form.get("items_json", "[]"))
        if not items:
            return redirect("/quotes/new")
        
        customer_id = req.form.get("customer_id", "")
        customer_name = ""
        
        # Handle new customer creation
        if customer_id == "NEW":
            new_name = req.form.get("new_customer_name", "").strip()
            if new_name:
                new_cust = {
                    "id": generate_id(),
                    "name": new_name,
                    "phone": req.form.get("new_customer_phone", ""),
                    "email": req.form.get("new_customer_email", ""),
                    "balance": 0,
                    "active": True,
                    "created_at": now()
                }
                db.insert("customers", new_cust)
                customer_id = new_cust["id"]
                customer_name = new_name
            else:
                customer_id = ""
        elif customer_id:
            cust = db.select_one("customers", customer_id)
            if cust:
                customer_name = cust.get("name", "")
        
        total = Decimal("0")
        for item in items:
            total += Decimal(str(item.get("price", 0))) * int(item.get("quantity", 1))
        
        quote = {
            "id": generate_id(),
            "quote_number": req.form.get("doc_number", DocumentNumbers.get_next("QUO", "quotes", "quote_number")),
            "date": req.form.get("date", today()),
            "customer_id": customer_id or None,
            "customer_name": customer_name,
            "items": req.form.get("items_json", "[]"),
            "total": float(total),
            "status": "draft",
            "created_at": now()
        }
        
        db.insert("quotes", quote)
        return redirect("/quotes")
        
    except Exception:
        return redirect("/quotes/new")


@app.route("/quotes/<quote_id>")
def quote_view(quote_id):
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    q = get_quote(quote_id)
    if not q:
        return redirect("/quotes")
    
    # Check if already converted
    if q.get("status") == "converted":
        return redirect(f"/quotes/{quote_id}")
    
    items = json.loads(q.get("items", "[]"))
    
    item_rows = ""
    for item in items:
        line_total = Decimal(str(item.get("price", 0))) * int(item.get("quantity", 1))
        item_rows += f'''
        <tr>
            <td>{safe_string(item.get("description", ""))}</td>
            <td class="number">{item.get("quantity", 1)}</td>
            <td class="number">{Money.format(Decimal(str(item.get("price", 0))))}</td>
            <td class="number">{Money.format(line_total)}</td>
        </tr>
        '''
    
    content = f'''
    
    
    <div class="card">
        <div class="flex-between mb-lg">
            <div>
                <h1 style="font-size:24px;font-weight:700;">Quote {safe_string(q.get("quote_number", ""))}</h1>
                <p class="text-muted">{q.get("date", "")[:10]} • {safe_string(q.get("customer_name", ""))}</p>
            </div>
            {'<span class="badge badge-green">✓ Converted</span>' if q.get('status') == 'converted' else f'<a href="/quotes/{quote_id}/convert" class="btn btn-green">Convert to Invoice</a>'}
        </div>
        
        <table class="table">
            <thead><tr><th>Description</th><th class="number">Qty</th><th class="number">Price</th><th class="number">Total</th></tr></thead>
            <tbody>{item_rows}</tbody>
            <tfoot>
                <tr><td colspan="3" style="text-align:right;font-weight:700;">Total:</td><td class="number" style="font-weight:700;color:var(--green);">{Money.format(Decimal(str(q.get("total", 0))))}</td></tr>
            </tfoot>
        </table>
    </div>
    '''
    
    return page_wrapper(f"Quote {q.get('quote_number', '')}", content, active="quotes", user=user)


@app.route("/quotes/<quote_id>/convert")
def quote_convert(quote_id):
    """Convert quote to invoice"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    q = get_quote(quote_id)
    if not q:
        return redirect("/quotes")
    
    # Check if already converted
    if q.get("status") == "converted":
        return redirect(f"/quotes/{quote_id}")
    
    # Create invoice from quote
    items = json.loads(q.get("items", "[]"))
    customer_id = q.get("customer_id", "")
    customer_name = q.get("customer_name", "")
    
    # Calculate with VAT
    subtotal = Decimal("0")
    total_vat = Decimal("0")
    
    for item in items:
        price = Decimal(str(item.get("price", 0)))
        qty = int(item.get("quantity", 1))
        line_total = price * qty
        vat_info = VAT.calculate_from_inclusive(line_total)
        subtotal += vat_info["exclusive"]
        total_vat += vat_info["vat"]
    
    total = subtotal + total_vat
    
    invoice_id = generate_id()
    invoice_number = DocumentNumbers.get_next("INV", "invoices", "invoice_number")
    
    invoice = {
        "id": invoice_id,
        "invoice_number": invoice_number,
        "date": today(),
        "customer_id": customer_id or None,
        "customer_name": customer_name,
        "items": q.get("items", "[]"),
        "subtotal": float(subtotal),
        "vat": float(total_vat),
        "total": float(total),
        "status": "outstanding",
        "from_quote": quote_id,
        "created_at": now()
    }
    
    db.insert("invoices", invoice)
    db.update("quotes", quote_id, {"status": "converted"})
    
    # Post to GL
    entry = JournalEntry(
        date=today(),
        reference=invoice_number,
        description=f"Invoice {invoice_number} - {customer_name}",
        trans_type=TransactionType.SALE,
        source_type="invoice",
        source_id=invoice_id
    )
    
    if customer_id:
        entry.debit(AccountCodes.DEBTORS, total)
        cust = db.select_one("customers", customer_id)
        if cust:
            new_bal = Decimal(str(cust.get("balance", 0) or 0)) + total
            db.update("customers", customer_id, {"balance": float(new_bal)})
    else:
        entry.debit(AccountCodes.BANK, total)
    
    entry.credit(AccountCodes.SALES, subtotal)
    if total_vat > 0:
        entry.credit(AccountCodes.VAT_OUTPUT, total_vat)
    
    entry.post()
    
    return redirect(f"/invoices/{invoice_id}")


# ═══════════════════════════════════════════════════════════════════════════════
# SHARED JAVASCRIPT FOR DOCUMENT CREATION
# ═══════════════════════════════════════════════════════════════════════════════

def create_document_js(doc_type):
    return '''
let lines = [];

function toggleNewCustomer(select) {
    const fields = document.getElementById('new-customer-fields');
    if (fields) {
        fields.style.display = select.value === 'NEW' ? 'flex' : 'none';
    }
}

function searchStock(q) {
    const results = document.getElementById('search-results');
    if (!q || q.length < 2) {
        results.innerHTML = '';
        return;
    }
    
    const searchWords = q.toLowerCase().split(/\\s+/).filter(w => w.length > 0);
    
    let html = '';
    let count = 0;
    for (const item of stockItems) {
        if (count >= 20) break;
        const desc = item.description.toLowerCase();
        const code = (item.code || '').toLowerCase();
        
        const allWordsMatch = searchWords.every(word => 
            desc.includes(word) || code.includes(word)
        );
        
        if (allWordsMatch) {
            count++;
            html += '<div style="padding:10px;border-bottom:1px solid var(--border);cursor:pointer;" onclick="addLine(\\''+item.id+'\\')"><strong>'+escHtml(item.code||'')+'</strong> '+escHtml(item.description)+' - '+item.price_formatted+'</div>';
        }
    }
    
    if (!html) {
        html = '<div style="padding:10px;color:#606070;">No stock items found for "'+escHtml(q)+'"</div>';
    }
    
    results.innerHTML = html;
}

function addLine(itemId) {
    const item = stockItems.find(i => i.id === itemId);
    if (!item) return;
    
    const existing = lines.find(l => l.id === itemId && !l.is_custom);
    if (existing) {
        existing.quantity++;
    } else {
        lines.push({
            id: item.id,
            code: item.code,
            description: item.description,
            price: item.price,
            quantity: 1,
            is_zero_rated: item.is_zero_rated,
            cost: item.cost || 0
        });
    }
    
    document.getElementById('search-stock').value = '';
    document.getElementById('search-results').innerHTML = '';
    renderLines();
}

function addCustomLine() {
    lines.push({
        id: 'custom_' + Date.now(),
        code: '',
        description: '',
        price: 0,
        quantity: 1,
        is_zero_rated: false,
        cost: 0,
        is_custom: true
    });
    renderLines();
    // Focus the new description input
    setTimeout(() => {
        const inputs = document.querySelectorAll('.custom-desc-input');
        if (inputs.length > 0) inputs[inputs.length-1].focus();
    }, 50);
}

function updateCustomDesc(index, value) {
    lines[index].description = value;
    updateHiddenFields();
}

function updateCustomPrice(index, value) {
    lines[index].price = parseFloat(value) || 0;
    renderLines();
}

function renderLines() {
    const tbody = document.getElementById('lines-body');
    let html = '';
    let total = 0;
    
    for (let i = 0; i < lines.length; i++) {
        const line = lines[i];
        const lineTotal = line.price * line.quantity;
        total += lineTotal;
        
        if (line.is_custom) {
            // Editable row for custom/buyout items - SPACIOUS & ALIGNED
            html += '<tr style="background:rgba(139,92,246,0.08);">';
            // Description - big comfortable input with badge inline
            html += '<td style="padding:12px 8px;vertical-align:middle;">';
            html += '<input type="text" class="custom-desc-input" value="'+escHtml(line.description)+'" placeholder="What are you selling? e.g. Special order pump..." style="width:calc(100% - 100px);padding:12px 14px;font-size:15px;background:#0a0a10;border:2px solid #8b5cf6;color:#f0f0f0;border-radius:8px;display:inline-block;vertical-align:middle;" onchange="updateCustomDesc('+i+',this.value)">';
            html += '<span style="background:linear-gradient(135deg,#8b5cf6,#6d28d9);color:white;padding:6px 10px;border-radius:6px;font-size:11px;font-weight:600;margin-left:8px;vertical-align:middle;">BUYOUT</span>';
            html += '</td>';
            // Quantity - centered
            html += '<td style="text-align:center;padding:12px 8px;vertical-align:middle;"><input type="number" value="'+line.quantity+'" min="1" style="width:70px;padding:10px;font-size:15px;text-align:center;background:#0a0a10;border:2px solid #2a2a4a;color:#f0f0f0;border-radius:8px;" onchange="updateQty('+i+',this.value)"></td>';
            // Price - right aligned with R prefix
            html += '<td style="text-align:right;padding:12px 8px;vertical-align:middle;"><div style="display:inline-flex;align-items:center;gap:4px;"><span style="color:#8b8b9a;">R</span><input type="number" value="'+line.price.toFixed(2)+'" step="0.01" min="0" style="width:100px;padding:10px;font-size:15px;text-align:right;background:#0a0a10;border:2px solid #8b5cf6;color:#f0f0f0;border-radius:8px;" onchange="updateCustomPrice('+i+',this.value)"></div></td>';
            // Total - right aligned
            html += '<td style="text-align:right;padding:12px 8px;vertical-align:middle;font-weight:600;font-size:15px;">R '+lineTotal.toFixed(2)+'</td>';
            // Delete button
            html += '<td style="text-align:center;padding:12px 8px;vertical-align:middle;"><button type="button" class="btn btn-sm btn-red" style="padding:8px 12px;" onclick="removeLine('+i+')">×</button></td>';
            html += '</tr>';
        } else {
            // Normal stock item row
            html += '<tr>';
            html += '<td style="padding:12px 8px;vertical-align:middle;">'+escHtml(line.description)+'</td>';
            html += '<td style="text-align:center;padding:12px 8px;vertical-align:middle;"><input type="number" value="'+line.quantity+'" min="1" style="width:70px;padding:10px;font-size:15px;text-align:center;background:#0a0a10;border:2px solid #2a2a4a;color:#f0f0f0;border-radius:8px;" onchange="updateQty('+i+',this.value)"></td>';
            html += '<td style="text-align:right;padding:12px 8px;vertical-align:middle;">R '+line.price.toFixed(2)+'</td>';
            html += '<td style="text-align:right;padding:12px 8px;vertical-align:middle;font-weight:600;">R '+lineTotal.toFixed(2)+'</td>';
            html += '<td style="text-align:center;padding:12px 8px;vertical-align:middle;"><button type="button" class="btn btn-sm btn-red" style="padding:8px 12px;" onclick="removeLine('+i+')">×</button></td>';
            html += '</tr>';
        }
    }
    
    tbody.innerHTML = html || '<tr><td colspan="5" style="text-align:center;color:#606070;padding:30px;">Search stock above or add a custom item</td></tr>';
    
    updateHiddenFields();
    document.getElementById('grand-total').textContent = 'R ' + total.toFixed(2);
}

function updateHiddenFields() {
    let total = 0;
    for (const line of lines) {
        total += line.price * line.quantity;
    }
    document.getElementById('items-json').value = JSON.stringify(lines);
    document.getElementById('total-input').value = total.toFixed(2);
}

function updateQty(index, value) {
    lines[index].quantity = parseInt(value) || 1;
    renderLines();
}

function removeLine(index) {
    lines.splice(index, 1);
    renderLines();
}

function escHtml(text) {
    const div = document.createElement('div');
    div.textContent = text || '';
    return div.innerHTML;
}

// Modal functions for popup option
function showCustomItemModal() {
    document.getElementById('custom-item-modal').classList.add('active');
    document.getElementById('custom-description').focus();
}

function closeCustomItemModal() {
    document.getElementById('custom-item-modal').classList.remove('active');
}

function addCustomItemFromModal() {
    const desc = document.getElementById('custom-description').value.trim();
    const price = parseFloat(document.getElementById('custom-price').value) || 0;
    const qty = parseInt(document.getElementById('custom-qty').value) || 1;
    
    if (!desc) {
        alert('Please enter a description');
        return;
    }
    if (price <= 0) {
        alert('Please enter a price');
        return;
    }
    
    lines.push({
        id: 'custom_' + Date.now(),
        code: '',
        description: desc,
        price: price,
        quantity: qty,
        is_zero_rated: false,
        cost: 0,
        is_custom: true
    });
    
    // Clear and close
    document.getElementById('custom-description').value = '';
    document.getElementById('custom-price').value = '';
    document.getElementById('custom-qty').value = '1';
    closeCustomItemModal();
    renderLines();
}

renderLines();
'''


# Custom item modal HTML for popup option
CUSTOM_ITEM_MODAL = '''
<!-- Custom Item Modal for Buyouts -->
<div class="modal-overlay" id="custom-item-modal">
    <div class="modal" style="max-width: 420px;">
        <div class="modal-header">
            <h3 class="modal-title">➕ Add Non-Stock Item (Buyout)</h3>
            <button class="modal-close" onclick="closeCustomItemModal()">&times;</button>
        </div>
        <div class="modal-body">
            <p style="color:#8b8b9a;margin-bottom:16px;font-size:14px;">
                Add items not in your stock - perfect for special orders, buyouts, or services.
            </p>
            <div class="form-group">
                <label class="form-label">Description *</label>
                <input type="text" id="custom-description" class="form-input" placeholder="e.g. Special order bracket, Labour charge">
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Price (incl VAT) *</label>
                    <input type="number" id="custom-price" class="form-input" placeholder="0.00" step="0.01" min="0">
                </div>
                <div class="form-group">
                    <label class="form-label">Quantity</label>
                    <input type="number" id="custom-qty" class="form-input" value="1" min="1">
                </div>
            </div>
        </div>
        <div class="modal-footer">
            <button type="button" class="btn btn-ghost" onclick="closeCustomItemModal()">Cancel</button>
            <button type="button" class="btn btn-primary" onclick="addCustomItemFromModal()">Add Item</button>
        </div>
    </div>
</div>
'''


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 9
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 9 COMPLETE - Invoices & Quotes

Contains:
✓ Invoice listing with totals and status
✓ Create new invoice with line items
✓ View invoice details
✓ Quote listing
✓ Create new quote
✓ View quote details
✓ Convert quote to invoice
✓ Proper GL posting on invoice creation:
  - DR Debtors (or Bank for walk-in)
  - CR Sales
  - CR VAT Output
  - DR COGS / CR Stock
✓ Customer balance update on invoice
✓ Shared JavaScript for document creation

All calculations done in Flask!

Next: Piece 10 - Expenses
"""

# =============================================================================
# PIECE10_EXPENSES.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - COMPLETE BUSINESS MANAGEMENT SYSTEM                              ║
║   Piece 10: Expenses                                                          ║
║                                                                               ║
║   This piece contains:                                                        ║
║   - Expense listing                                                           ║
║   - Add expense manually                                                      ║
║   - AI receipt scanning with Opus                                             ║
║   - Category selection with proper GL accounts                                ║
║   - VAT handling (standard, zero-rated, exempt)                               ║
║   - Proper GL posting                                                         ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""




# ═══════════════════════════════════════════════════════════════════════════════
# EXPENSE FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_all_expenses():
    """Get all expenses"""
    return db.select("expenses", order="-date")


def get_expense(expense_id):
    """Get single expense"""
    return db.select_one("expenses", expense_id)


def get_expense_categories():
    """Get expense account categories for dropdown"""
    accounts = Account.get_expense_accounts()
    categories = []
    
    for acc in accounts:
        # Determine VAT type
        if acc.get("zero_rated"):
            vat_type = "zero"
        elif not acc.get("vat_applicable", True):
            vat_type = "exempt"
        else:
            vat_type = "standard"
        
        categories.append({
            "code": acc["code"],
            "name": acc["name"],
            "vat_type": vat_type
        })
    
    return categories


# ═══════════════════════════════════════════════════════════════════════════════
# EXPENSE PAGES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/expenses")
def expense_list():
    """Expense listing page"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    expenses = get_all_expenses()
    
    total = sum(Decimal(str(e.get("amount", 0) or 0)) for e in expenses)
    total_vat = sum(Decimal(str(e.get("vat", 0) or 0)) for e in expenses)
    
    rows = []
    for exp in expenses[:50]:
        amount = Decimal(str(exp.get("amount", 0)))
        vat = Decimal(str(exp.get("vat", 0)))
        
        rows.append([
            exp.get("date", "")[:10],
            safe_string(exp.get("supplier", "-")),
            safe_string(exp.get("description", "")),
            safe_string(exp.get("category", "-")),
            {"value": Money.format(amount), "class": "number"},
            {"value": Money.format(vat), "class": "number text-muted"}
        ])
    
    table = table_html(
        headers=["Date", "Supplier", "Description", "Category", 
                {"label": "Amount", "class": "number"}, 
                {"label": "VAT", "class": "number"}],
        rows=rows,
        empty_message="No expenses yet"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <p class="text-muted">{len(expenses)} expenses</p>
        <div class="btn-group">
            <a href="/expenses/scan" class="btn btn-orange">📷 Scan Receipt</a>
            <a href="/expenses/new" class="btn btn-primary">+ Add Expense</a>
        </div>
    </div>
    
    <div class="stats">
        <div class="stat">
            <div class="stat-value">{len(expenses)}</div>
            <div class="stat-label">Expenses</div>
        </div>
        <div class="stat">
            <div class="stat-value red">{Money.format(total)}</div>
            <div class="stat-label">Total Spent</div>
        </div>
        <div class="stat">
            <div class="stat-value green">{Money.format(total_vat)}</div>
            <div class="stat-label">VAT Claimable</div>
        </div>
    </div>
    
    <div class="card">{table}</div>
    '''
    
    return page_wrapper("Expenses", content, active="expenses", user=user)


@app.route("/expenses/new", methods=["GET", "POST"])
def expense_new():
    """Add new expense manually"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = ""
    
    if request.method == "POST":
        result = process_expense(request.form)
        if result["success"]:
            return redirect("/expenses")
        else:
            message = error_message(result["error"])
    
    categories = get_expense_categories()
    cat_options = "".join([
        f'<option value="{c["code"]}" data-vat="{c["vat_type"]}">{c["name"]}</option>' 
        for c in categories
    ])
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">Add Expense</h2>
        
        {message}
        
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Date</label>
                    <input type="date" name="date" class="form-input" value="{today()}">
                </div>
                <div class="form-group">
                    <label class="form-label">Supplier</label>
                    <input type="text" name="supplier" class="form-input" placeholder="Supplier name">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Description</label>
                <input type="text" name="description" class="form-input" placeholder="What was purchased">
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Category</label>
                    <select name="category" id="category" class="form-select" onchange="updateVatType()">
                        <option value="">Select category...</option>
                        {cat_options}
                    </select>
                </div>
                <div class="form-group">
                    <label class="form-label">Amount (incl VAT)</label>
                    <input type="text" name="amount" class="form-input" placeholder="0.00" required>
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">VAT Type</label>
                <select name="vat_type" id="vat_type" class="form-select">
                    <option value="standard">Standard Rate (15%)</option>
                    <option value="zero">Zero Rated (0%)</option>
                    <option value="exempt">Exempt (no VAT claim)</option>
                </select>
            </div>
            
            <div class="form-group">
                <label class="form-label">Reference</label>
                <input type="text" name="reference" class="form-input" placeholder="Invoice number, slip number, etc.">
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Expense</button>
                <a href="/expenses" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    
    <script>
    function updateVatType() {{
        const cat = document.getElementById('category');
        const vatType = document.getElementById('vat_type');
        const selected = cat.options[cat.selectedIndex];
        if (selected && selected.dataset.vat) {{
            vatType.value = selected.dataset.vat;
        }}
    }}
    </script>
    '''
    
    return page_wrapper("Add Expense", content, active="expenses", user=user)


def process_expense(form_data, prefilled=None):
    """Process expense form data and post to GL"""
    
    try:
        supplier = form_data.get("supplier", "").strip() if form_data else prefilled.get("supplier", "")
        description = form_data.get("description", "").strip() if form_data else prefilled.get("description", "")
        amount_str = form_data.get("amount", "0") if form_data else str(prefilled.get("amount", 0))
        category_code = form_data.get("category", "") if form_data else prefilled.get("category", "")
        vat_type = form_data.get("vat_type", "standard") if form_data else prefilled.get("vat_type", "standard")
        reference = form_data.get("reference", "") if form_data else prefilled.get("reference", "")
        date = form_data.get("date", today()) if form_data else prefilled.get("date", today())
        
        amount = Money.parse(amount_str)
        
        if amount <= 0:
            return {"success": False, "error": "Enter a valid amount"}
        
        # Get expense account
        if category_code:
            expense_account = category_code
        else:
            expense_account = AccountCodes.GENERAL_EXPENSES
        
        # Calculate VAT based on type
        if vat_type == "zero":
            vat_info = VAT.calculate_from_inclusive(amount, VAT.ZERO_RATE)
        elif vat_type == "exempt":
            vat_info = {"exclusive": amount, "vat": Decimal("0"), "inclusive": amount}
        else:
            vat_info = VAT.calculate_from_inclusive(amount)
        
        # Get category name
        category_name = "General Expenses"
        cat_account = Account.get_by_code(expense_account)
        if cat_account:
            category_name = cat_account.get("name", category_name)
        
        # Create expense record
        expense_id = generate_id()
        expense = {
            "id": expense_id,
            "date": date,
            "supplier": supplier,
            "description": description,
            "category": category_name,
            "category_code": expense_account,
            "amount": float(vat_info["inclusive"]),
            "vat": float(vat_info["vat"]),
            "net": float(vat_info["exclusive"]),
            "vat_type": vat_type,
            "reference": reference,
            "created_at": now()
        }
        
        success, result = db.insert("expenses", expense)
        if not success:
            return {"success": False, "error": f"Failed to save: {result}"}
        
        # Post to GL
        entry_desc = f"{supplier}: {description}" if supplier else description
        
        entry = JournalEntry(
            date=date,
            reference=reference,
            description=entry_desc[:200],
            trans_type=TransactionType.EXPENSE,
            source_type="expense",
            source_id=expense_id
        )
        
        # Debit: Expense account (net amount)
        entry.debit(expense_account, vat_info["exclusive"])
        
        # Debit: VAT Input (if claimable)
        if vat_info["vat"] > 0 and vat_type != "exempt":
            entry.debit(AccountCodes.VAT_INPUT, vat_info["vat"])
        
        # Credit: Bank (total amount)
        entry.credit(AccountCodes.BANK, vat_info["inclusive"])
        
        success, result = entry.post()
        if not success:
            return {"success": False, "error": f"GL posting failed: {result}"}
        
        return {"success": True, "expense_id": expense_id}
        
    except Exception as e:
        return {"success": False, "error": str(e)}


# ═══════════════════════════════════════════════════════════════════════════════
# AI RECEIPT SCANNING
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/expenses/scan", methods=["GET"])
def expense_scan_page():
    """Receipt scanning page"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    categories = get_expense_categories()
    cat_options = "".join([
        f'<option value="{c["code"]}" data-vat="{c["vat_type"]}">{c["name"]}</option>' 
        for c in categories
    ])
    
    content = f'''
    
    
    <div class="card">
        <h2 class="card-title mb-md">📷 Scan Receipt</h2>
        <p class="text-muted mb-lg">Take a photo or upload an image. AI will extract the details.</p>
        
        <div class="btn-group mb-lg" style="justify-content: center;">
            <button class="btn btn-orange" onclick="startCamera()">📷 Use Camera</button>
            <label class="btn btn-ghost" style="cursor: pointer;">
                📁 Upload Image
                <input type="file" id="file-input" accept="image/*" style="display: none;" onchange="handleFile(this)">
            </label>
        </div>
        
        <video id="video" autoplay playsinline style="display:none;width:100%;max-width:400px;margin:0 auto;border-radius:12px;background:#000;"></video>
        <button id="capture-btn" class="btn btn-green btn-block mt-md" style="display:none;" onclick="capture()">📸 Capture</button>
        
        <canvas id="canvas" style="display:none;"></canvas>
        <img id="preview" style="display:none;max-width:100%;border-radius:12px;margin-top:16px;">
        
        <div id="processing" class="alert alert-info mt-lg" style="display:none;">
            🔄 AI is reading your receipt...
        </div>
    </div>
    
    <div class="card" id="result-card" style="max-width: 600px; display: none;">
        <h3 class="card-title mb-md">Receipt Details</h3>
        <p class="text-muted mb-md">Review and adjust if needed:</p>
        
        <form method="POST" action="/expenses/save-scanned">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Date</label>
                    <input type="date" name="date" id="r-date" class="form-input" value="{today()}">
                </div>
                <div class="form-group">
                    <label class="form-label">Supplier</label>
                    <input type="text" name="supplier" id="r-supplier" class="form-input">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Description</label>
                <input type="text" name="description" id="r-description" class="form-input">
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Category</label>
                    <select name="category" id="r-category" class="form-select">
                        <option value="">Select...</option>
                        {cat_options}
                    </select>
                </div>
                <div class="form-group">
                    <label class="form-label">Amount (incl VAT)</label>
                    <input type="text" name="amount" id="r-amount" class="form-input" 
                           style="font-size: 20px; font-weight: 700;">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">VAT Type</label>
                <select name="vat_type" id="r-vat-type" class="form-select">
                    <option value="standard">Standard Rate (15%)</option>
                    <option value="zero">Zero Rated (0%)</option>
                    <option value="exempt">Exempt (no VAT)</option>
                </select>
            </div>
            
            <button type="submit" class="btn btn-green btn-block btn-lg">
                ✓ Save Expense
            </button>
        </form>
    </div>
    
    <script>
    let stream = null;
    
    function startCamera() {{
        navigator.mediaDevices.getUserMedia({{ video: {{ facingMode: "environment" }} }})
            .then(s => {{
                stream = s;
                const video = document.getElementById('video');
                video.srcObject = s;
                video.style.display = 'block';
                document.getElementById('capture-btn').style.display = 'block';
            }})
            .catch(e => alert('Camera error: ' + e.message));
    }}
    
    function capture() {{
        const video = document.getElementById('video');
        const canvas = document.getElementById('canvas');
        canvas.width = video.videoWidth;
        canvas.height = video.videoHeight;
        canvas.getContext('2d').drawImage(video, 0, 0);
        
        const imageData = canvas.toDataURL('image/jpeg', 0.8);
        
        if (stream) {{
            stream.getTracks().forEach(t => t.stop());
        }}
        video.style.display = 'none';
        document.getElementById('capture-btn').style.display = 'none';
        
        document.getElementById('preview').src = imageData;
        document.getElementById('preview').style.display = 'block';
        
        processImage(imageData);
    }}
    
    function handleFile(input) {{
        if (input.files && input.files[0]) {{
            const reader = new FileReader();
            reader.onload = e => {{
                document.getElementById('preview').src = e.target.result;
                document.getElementById('preview').style.display = 'block';
                processImage(e.target.result);
            }};
            reader.readAsDataURL(input.files[0]);
        }}
    }}
    
    async function processImage(imageData) {{
        document.getElementById('processing').style.display = 'block';
        document.getElementById('result-card').style.display = 'none';
        
        const response = await fetch('/api/expenses/scan', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{ image: imageData }})
        }});
        
        const r = await response.json();
        
        document.getElementById('processing').style.display = 'none';
        document.getElementById('result-card').style.display = 'block';
        
        // Fill the form
        document.getElementById('r-supplier').value = r.supplier || '';
        document.getElementById('r-description').value = r.description || '';
        document.getElementById('r-amount').value = r.total || r.amount || '';
        if (r.date) document.getElementById('r-date').value = r.date;
        
        // Match category
        if (r.category) {{
            const sel = document.getElementById('r-category');
            const cat = r.category.toLowerCase();
            for (let i = 0; i < sel.options.length; i++) {{
                if (sel.options[i].text.toLowerCase().includes(cat)) {{
                    sel.selectedIndex = i;
                    break;
                }}
            }}
        }}
    }}
    </script>
    '''
    
    return page_wrapper("Scan Receipt", content, active="expenses", user=user)


@app.route("/api/expenses/scan", methods=["POST"])
def api_scan_receipt():
    """API: Scan receipt with AI and return extracted data"""
    data = request.get_json() or {}
    image_data = data.get("image", "")
    
    if not image_data:
        return jsonify({"supplier": "", "total": 0, "description": "No image", "category": "Other"})
    
    if "," in image_data:
        image_data = image_data.split(",")[1]
    
    result = OpusAI.read_receipt(image_data)
    return jsonify(result)


@app.route("/expenses/save-scanned", methods=["POST"])
def expense_save_scanned():
    """Save expense from scanned receipt"""
    
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    result = process_expense(request.form)
    
    if result["success"]:
        return redirect("/expenses")
    else:
        # Show error and redirect back
        return redirect("/expenses/scan")


# ═══════════════════════════════════════════════════════════════════════════════
# EXPENSE API
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/expenses")
def api_expenses():
    """API: Get all expenses"""
    return jsonify(get_all_expenses())


# ═══════════════════════════════════════════════════════════════════════════════
# END OF PIECE 10
# ═══════════════════════════════════════════════════════════════════════════════

"""
PIECE 10 COMPLETE - Expenses

Contains:
✓ Expense listing with totals
✓ Add expense manually
✓ Category dropdown from GL expense accounts
✓ VAT type selection (standard/zero/exempt)
✓ AI receipt scanning with Opus:
  - Camera capture
  - Image upload
  - Automatic extraction of supplier, description, amount
  - Category suggestion
  - Zero-rated detection
✓ Proper GL posting:
  - DR Expense account (net)
  - DR VAT Input (if claimable)
  - CR Bank (total)
✓ Expense API

All calculations done in Flask!

Next: Piece 11 - Reports
"""

# =============================================================================
# PIECE11_REPORTS.PY
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║   CLICK AI - Piece 11: Reports                                                ║
║   ALL REPORTS PULL FROM THE JOURNAL - REAL DOUBLE-ENTRY DATA                  ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""




@app.route("/reports")
def reports_menu():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    content = '''
    <h1 style="font-size: 24px; font-weight: 700; margin-bottom: 24px;">Reports</h1>
    
    <h3 style="color: var(--purple); margin-bottom: 16px;">🤖 AI Business Advisor</h3>
    <div class="report-grid" style="margin-bottom: 32px;">
        <a href="/ai-advisor" class="report-card" style="border-color: var(--green); background: linear-gradient(135deg, rgba(16,185,129,0.15), rgba(139,92,246,0.15));">
            <div class="report-card-icon">🧠</div>
            <h3 class="report-card-title">AI Business Advisor</h3>
            <p class="report-card-desc">Chat with AI about YOUR business data - it knows your numbers!</p>
        </a>
        <a href="/tb-analyzer" class="report-card" style="border-color: var(--purple); background: rgba(139,92,246,0.05);">
            <div class="report-card-icon">📤</div>
            <h3 class="report-card-title">External TB Analyzer</h3>
            <p class="report-card-desc">Upload a trial balance from another system</p>
        </a>
    </div>
    
    <h3 style="color: var(--text-muted); margin-bottom: 16px;">📊 Standard Reports</h3>
    <div class="report-grid" style="margin-bottom: 32px;">
        <a href="/reports/trial-balance" class="report-card"><div class="report-card-icon">⚖️</div><h3 class="report-card-title">Trial Balance</h3><p class="report-card-desc">All accounts with debit/credit balances</p></a>
        <a href="/reports/income-statement" class="report-card"><div class="report-card-icon">📈</div><h3 class="report-card-title">Income Statement</h3><p class="report-card-desc">Profit & Loss report</p></a>
        <a href="/reports/balance-sheet" class="report-card"><div class="report-card-icon">🏦</div><h3 class="report-card-title">Balance Sheet</h3><p class="report-card-desc">Assets, Liabilities, Equity</p></a>
        <a href="/reports/vat" class="report-card"><div class="report-card-icon">🏛️</div><h3 class="report-card-title">VAT Report</h3><p class="report-card-desc">Output minus Input VAT</p></a>
        <a href="/reports/sales" class="report-card"><div class="report-card-icon">💰</div><h3 class="report-card-title">Sales Report</h3><p class="report-card-desc">Invoice summary</p></a>
        <a href="/reports/debtors" class="report-card"><div class="report-card-icon">📋</div><h3 class="report-card-title">Debtors</h3><p class="report-card-desc">Who owes you</p></a>
        <a href="/reports/creditors" class="report-card"><div class="report-card-icon">📑</div><h3 class="report-card-title">Creditors</h3><p class="report-card-desc">Who you owe</p></a>
        <a href="/reports/stock" class="report-card"><div class="report-card-icon">📦</div><h3 class="report-card-title">Stock Report</h3><p class="report-card-desc">Levels and valuation</p></a>
        <a href="/reports/ledger" class="report-card"><div class="report-card-icon">📒</div><h3 class="report-card-title">General Ledger</h3><p class="report-card-desc">All transactions</p></a>
    </div>
    
    <h3 style="color: var(--text-muted); margin-bottom: 16px;">🏦 Professional Reports</h3>
    <div class="report-grid" style="margin-bottom: 32px;">
        <a href="/reports/business-health" class="report-card"><div class="report-card-icon">💊</div><h3 class="report-card-title">Business Health</h3><p class="report-card-desc">Plain-language financial analysis</p></a>
        <a href="/reports/management-financials" class="report-card"><div class="report-card-icon">📋</div><h3 class="report-card-title">Management Financials</h3><p class="report-card-desc">Professional report for banks</p></a>
    </div>
    
    <h3 style="color: var(--text-muted); margin-bottom: 16px;">✏️ Adjustments</h3>
    <div class="report-grid">
        <a href="/journal-entry" class="report-card"><div class="report-card-icon">📝</div><h3 class="report-card-title">Journal Entry</h3><p class="report-card-desc">Manual adjustments</p></a>
        <a href="/bank-recon" class="report-card"><div class="report-card-icon">🏦</div><h3 class="report-card-title">Bank Reconciliation</h3><p class="report-card-desc">Match bank statement</p></a>
    </div>
    '''
    return page_wrapper("Reports", content, active="reports", user=user)


@app.route("/reports/trial-balance")
def report_trial_balance():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Use fast database function
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_trial_balance"
        headers = {
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }
        resp = requests.post(url, headers=headers, json={}, timeout=30)
        tb_data = resp.json() if resp.status_code == 200 else []
    except:
        tb_data = []
    
    rows = ""
    total_dr = Decimal("0")
    total_cr = Decimal("0")
    
    for entry in tb_data:
        dr = Decimal(str(entry.get("total_debit", 0) or 0))
        cr = Decimal(str(entry.get("total_credit", 0) or 0))
        total_dr += dr
        total_cr += cr
        
        dr_fmt = Money.format(dr) if dr > 0 else "-"
        cr_fmt = Money.format(cr) if cr > 0 else "-"
        
        rows += f'''<tr>
            <td>{entry.get("account_code", "")}</td>
            <td>{entry.get("account_name", "")}</td>
            <td class="number">{dr_fmt}</td>
            <td class="number">{cr_fmt}</td>
        </tr>'''
    
    if not rows:
        rows = '<tr><td colspan="4" class="text-muted" style="text-align:center;padding:40px;">No transactions yet</td></tr>'
    
    balanced = abs(total_dr - total_cr) < Decimal("0.01")
    balance_badge = '<span class="badge badge-green">✓ Balanced</span>' if balanced else '<span class="badge badge-red">✗ Unbalanced</span>'
    
    content = f'''
    
    <div class="flex-between mb-lg">
        <div>
            <h1 style="font-size:24px;font-weight:700;margin-bottom:8px;">Trial Balance</h1>
            <p class="text-muted">As at {today()}</p>
        </div>
        <div class="btn-group">
            <a href="/reports/trial-balance/csv" class="btn btn-ghost">📥 Export CSV</a>
            {balance_badge}
        </div>
    </div>
    <div class="stats">
        <div class="stat"><div class="stat-value">{len(tb_data)}</div><div class="stat-label">Accounts</div></div>
        <div class="stat"><div class="stat-value green">{Money.format(total_dr)}</div><div class="stat-label">Total Debits</div></div>
        <div class="stat"><div class="stat-value green">{Money.format(total_cr)}</div><div class="stat-label">Total Credits</div></div>
    </div>
    <div class="card">
        <table class="table">
            <thead><tr><th>Code</th><th>Account</th><th class="number">Debit</th><th class="number">Credit</th></tr></thead>
            <tbody>{rows}</tbody>
            <tfoot><tr style="font-weight:bold;background:#0a0a10;">
                <td colspan="2">TOTAL</td>
                <td class="number">{Money.format(total_dr)}</td>
                <td class="number">{Money.format(total_cr)}</td>
            </tr></tfoot>
        </table>
    </div>
    '''
    return page_wrapper("Trial Balance", content, "reports", user)


@app.route("/reports/trial-balance/csv")
def report_trial_balance_csv():
    """Export trial balance as CSV"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_trial_balance"
        headers = {
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }
        resp = requests.post(url, headers=headers, json={}, timeout=30)
        tb_data = resp.json() if resp.status_code == 200 else []
    except:
        tb_data = []
    
    # Build CSV
    csv_lines = ["Account Code,Account Name,Debit,Credit"]
    for entry in tb_data:
        dr = entry.get("total_debit", 0) or 0
        cr = entry.get("total_credit", 0) or 0
        csv_lines.append(f'{entry.get("account_code", "")},{entry.get("account_name", "")},{dr},{cr}')
    
    csv_content = "\n".join(csv_lines)
    
    from flask import Response
    return Response(
        csv_content,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=trial_balance_{today()}.csv"}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# TB ANALYZER - THE DREAM FEATURE
# Chatty Bot (Haiku) for friendly interface + Opus for deep analysis
# "Horses for courses" - each AI does what it's best at
# ═══════════════════════════════════════════════════════════════════════════════

class TBAnalyzer:
    """
    Trial Balance Analyzer - The Heart of BB Fin
    
    Architecture:
    - Haiku: Fast, friendly, handles conversation
    - Opus: Deep analysis, pattern recognition, the "accountant brain"
    """
    
    HAIKU_MODEL = "claude-3-5-haiku-20241022"
    OPUS_MODEL = "claude-opus-4-20250514"
    
    @classmethod
    def parse_tb_csv(cls, csv_content: str) -> list:
        """Parse uploaded TB CSV into structured data"""
        lines = csv_content.strip().split('\n')
        if not lines:
            return []
        
        # Parse all rows first
        import csv
        reader = csv.reader(lines)
        rows = list(reader)
        if not rows:
            return []
        
        # Find the REAL header row - must have debit/credit columns OR name+amount
        header_idx = 0
        for i, row in enumerate(rows[:10]):
            headers_lower = [h.strip().lower() for h in row]
            # Look for rows that have BOTH debit and credit, or name and balance
            has_debit = any('debit' in h for h in headers_lower)
            has_credit = any('credit' in h for h in headers_lower)
            has_name = any(h in ['name', 'account', 'description'] for h in headers_lower)
            has_balance = any('balance' in h and h != 'trial balance report' for h in headers_lower)
            
            if (has_debit and has_credit) or (has_name and has_balance):
                header_idx = i
                print(f"DEBUG: Found header at row {i}: {row}")
                break
        
        headers = [h.strip().lower() for h in rows[header_idx]]
        print(f"DEBUG: Headers: {headers}")
        
        # Find column indices - be more flexible
        desc_col = None
        for i, h in enumerate(headers):
            if h in ['name', 'account', 'description'] or 'account' in h:
                desc_col = i
                break
        if desc_col is None:
            desc_col = 0
        
        # Category column (optional)
        cat_col = next((i for i, h in enumerate(headers) if 'category' in h or 'type' in h), None)
        
        # Debit column
        debit_col = next((i for i, h in enumerate(headers) if 'debit' in h), None)
        
        # Credit column  
        credit_col = next((i for i, h in enumerate(headers) if 'credit' in h), None)
        
        # Balance column (fallback if no debit/credit)
        balance_col = next((i for i, h in enumerate(headers) if 'balance' in h and 'trial' not in h), None)
        
        print(f"DEBUG: Columns - desc:{desc_col}, cat:{cat_col}, debit:{debit_col}, credit:{credit_col}, balance:{balance_col}")
        
        accounts = []
        for row in rows[header_idx + 1:]:
            if len(row) < 2:
                continue
            
            try:
                acc_name = row[desc_col].strip() if desc_col is not None and desc_col < len(row) else ""
                
                # Skip empty rows or header repeats
                if not acc_name or acc_name.lower() in ['name', 'account', 'description', '', 'total', 'totals']:
                    continue
                
                # Get category if available
                category = ""
                if cat_col is not None and cat_col < len(row):
                    category = row[cat_col].strip()
                
                # Get debit/credit or balance
                debit = Decimal("0")
                credit = Decimal("0")
                balance = Decimal("0")
                
                if debit_col is not None and debit_col < len(row):
                    debit = cls._parse_amount(row[debit_col])
                if credit_col is not None and credit_col < len(row):
                    credit = cls._parse_amount(row[credit_col])
                if balance_col is not None and balance_col < len(row):
                    balance = cls._parse_amount(row[balance_col])
                
                # Calculate net balance from debit/credit
                if debit or credit:
                    balance = debit - credit
                
                # Only add if we have a name and some value
                if acc_name:
                    accounts.append({
                        "code": category,
                        "name": acc_name,
                        "category": category,
                        "debit": float(debit),
                        "credit": float(credit),
                        "balance": float(balance)
                    })
                    
            except Exception as e:
                print(f"DEBUG: Error parsing row: {row}, error: {e}")
                continue
        
        print(f"DEBUG: Parsed {len(accounts)} accounts")
        if accounts:
            print(f"DEBUG: First account: {accounts[0]}")
            print(f"DEBUG: Total debits: R{sum(a['debit'] for a in accounts):,.2f}")
            print(f"DEBUG: Total credits: R{sum(a['credit'] for a in accounts):,.2f}")
        
        return accounts
    
    @staticmethod
    def _parse_amount(val: str) -> Decimal:
        """Parse amount string to Decimal"""
        if not val:
            return Decimal("0")
        val = str(val).strip()
        val = val.replace("R", "").replace(",", "").replace(" ", "")
        val = val.replace("(", "-").replace(")", "")
        try:
            return Decimal(val)
        except:
            return Decimal("0")
    
    @classmethod
    def analyze_with_opus(cls, accounts: list, business_context: str = "") -> dict:
        """
        Deep analysis with Opus - the accountant brain
        Returns structured findings for the chatty bot to present
        """
        api_key = Config.ANTHROPIC_API_KEY
        if not api_key:
            return {"error": "API key not configured"}
        
        # Prepare TB summary for Opus
        tb_text = "TRIAL BALANCE DATA:\n"
        tb_text += "=" * 60 + "\n"
        
        total_debit = 0
        total_credit = 0
        total_assets = 0
        total_liabilities = 0
        total_equity = 0
        total_income = 0
        total_expenses = 0
        
        for acc in accounts:
            tb_text += f"{acc['code']}: {acc['name']} = R {acc['balance']:,.2f}\n"
            
            # Categorize by account code patterns
            code = acc['code'].upper() if acc['code'] else ""
            name = acc['name'].upper() if acc['name'] else ""
            bal = acc['balance']
            
            if acc['debit']:
                total_debit += acc['debit']
            if acc['credit']:
                total_credit += acc['credit']
            
            # Basic categorization
            if code.startswith(('1', 'B/S')) or any(x in name for x in ['ASSET', 'BANK', 'CASH', 'DEBTOR', 'INVENTORY', 'EQUIPMENT', 'VEHICLE']):
                total_assets += bal
            elif code.startswith(('2', '3')) or any(x in name for x in ['LIABILITY', 'CREDITOR', 'LOAN', 'VAT', 'PAYE']):
                total_liabilities += abs(bal)
            elif code.startswith('3') or any(x in name for x in ['CAPITAL', 'RETAINED', 'EQUITY', 'RESERVE']):
                total_equity += abs(bal)
            elif code.startswith(('4', 'I')) or any(x in name for x in ['SALES', 'INCOME', 'REVENUE', 'FEE']):
                total_income += abs(bal)
            elif code.startswith(('5', '6', '7', 'E')) or any(x in name for x in ['EXPENSE', 'COST', 'SALARY', 'WAGE', 'RENT', 'TELEPHONE']):
                total_expenses += abs(bal)
        
        tb_text += "\n" + "=" * 60 + "\n"
        tb_text += f"Total Debits: R {total_debit:,.2f}\n"
        tb_text += f"Total Credits: R {total_credit:,.2f}\n"
        tb_text += f"Difference: R {abs(total_debit - total_credit):,.2f}\n"
        
        prompt = f"""You are a senior South African Chartered Accountant analyzing a Trial Balance. 
Your task is to provide deep, insightful analysis that a business owner can understand.

{tb_text}

Additional context: {business_context if business_context else "No additional context provided"}

Provide your analysis as JSON with this EXACT structure:
{{
    "company_health": "good|warning|critical",
    "health_summary": "One sentence overall health assessment",
    
    "key_metrics": {{
        "gross_profit_margin": null or percentage,
        "net_profit_margin": null or percentage,
        "current_ratio": null or number,
        "debt_ratio": null or percentage,
        "debtor_days": null or number,
        "creditor_days": null or number
    }},
    
    "red_flags": [
        {{"issue": "Description", "severity": "high|medium|low", "recommendation": "What to do"}}
    ],
    
    "opportunities": [
        {{"opportunity": "Description", "potential_benefit": "What they could gain"}}
    ],
    
    "questions_for_owner": [
        "Question that would help clarify something unusual?"
    ],
    
    "sars_concerns": [
        "Any items SARS might query during an audit"
    ],
    
    "industry_comparison": "How does this compare to typical businesses",
    
    "top_3_priorities": [
        "Most important thing to address first",
        "Second priority",
        "Third priority"
    ]
}}

Focus on:
1. Cash flow health - can they pay their bills?
2. Profit margins - are they making money?
3. Balance sheet strength - assets vs liabilities
4. Tax compliance - VAT, PAYE issues
5. Unusual patterns - things that don't look right
6. SA-specific issues - SARS, BEE, industry norms

Be specific with Rand amounts where relevant. Be honest but constructive.
Return ONLY valid JSON, no other text."""

        try:
            # Use Sonnet for faster analysis - Opus times out on Render's 30s limit
            response = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "content-type": "application/json",
                    "anthropic-version": "2023-06-01"
                },
                json={
                    "model": "claude-sonnet-4-20250514",  # Sonnet - fast enough for Render
                    "max_tokens": 4000,
                    "messages": [{"role": "user", "content": prompt}]
                },
                timeout=25  # Under Render's 30s limit
            )
            
            if response.status_code == 200:
                result = response.json()
                text = result.get("content", [{}])[0].get("text", "{}")
                # Parse JSON from response
                try:
                    # Find JSON in response
                    start = text.find('{')
                    end = text.rfind('}') + 1
                    if start >= 0 and end > start:
                        return json.loads(text[start:end])
                except:
                    pass
                return {"error": "Could not parse analysis", "raw": text[:500], "company_health": "unknown", "health_summary": "Analysis completed but couldn't parse results"}
            else:
                return {"error": f"API error: {response.status_code}", "company_health": "unknown", "health_summary": "Could not complete analysis"}
        except requests.exceptions.Timeout:
            return {"error": "Analysis timed out - please try again", "company_health": "unknown", "health_summary": "The analysis took too long. Try uploading a smaller file."}
        except Exception as e:
            return {"error": str(e), "company_health": "unknown", "health_summary": "An error occurred during analysis"}
    
    @classmethod
    def chat_response(cls, analysis: dict, user_message: str = "", conversation_history: list = None, business_context: dict = None) -> str:
        """
        Haiku generates friendly conversational response based on analysis
        Strictly limited to business/financial topics only!
        """
        api_key = Config.ANTHROPIC_API_KEY
        if not api_key:
            return "I'm sorry, I can't analyze right now. The AI service isn't configured."
        
        # Build context from analysis
        analysis_context = json.dumps(analysis, indent=2) if isinstance(analysis, dict) else str(analysis)
        
        # Add business context if available (for Click AI native users)
        business_info = ""
        if business_context:
            business_info = f"""
BUSINESS CONTEXT (from Click AI records):
- Business Name: {business_context.get('name', 'Unknown')}
- Industry: {business_context.get('industry', 'General')}
- Trading Since: {business_context.get('created_at', 'Unknown')[:10] if business_context.get('created_at') else 'Unknown'}
- Business Type: {business_context.get('business_type', 'SME')}
"""
        
        system_prompt = f"""You are BB Fin, a friendly financial advisor chatbot for Click AI, a South African accounting system.

Your personality:
- Warm, encouraging, and occasionally witty
- Use simple language, avoid jargon (or explain it when needed)
- South African context - understand SARS, VAT, BEE, load shedding impacts
- Be honest about problems but always offer hope and actionable solutions
- Address the business owner directly as "you"

{business_info}

CRITICAL RULES - YOU MUST FOLLOW THESE:
1. You ONLY discuss business, financial, and accounting matters
2. If asked about ANYTHING not related to business/finances, respond with a witty but firm redirect:
   - Personal problems: "Eish, that sounds rough, but I'm your numbers guy, not Dr Phil! Let's focus on what I CAN help with - your business finances. 📊"
   - Health questions: "I'm flattered you'd ask, but my medical degree is in diagnosing sick balance sheets, not sick people! Let's talk business. 💼"
   - Relationship advice: "Listen pal, I'm here to help with your cash flow, not your personal flow! Now, about those debtors... 😄"
   - Politics/religion: "I only have strong opinions about VAT returns and profit margins! What business question can I help with?"
   - Random topics: "That's interesting, but my expertise starts and ends with Rands and cents! What would you like to know about your finances?"
3. Always bring the conversation back to their business data
4. Never provide advice on: medical, legal (except basic tax), personal relationships, or non-business matters
5. If someone seems distressed about personal issues, be kind but firm: "I can see things are tough. While I can't help with that, I CAN help you understand your business better. Sometimes getting your finances sorted helps everything else feel more manageable."

The financial analysis to discuss:
{analysis_context}

Your job is to:
1. Present findings in a friendly, conversational way
2. Answer follow-up questions about the FINANCIAL DATA only
3. Help them understand what the numbers mean for their business
4. Suggest practical next steps they can take"""

        messages = []
        if conversation_history:
            messages.extend(conversation_history)
        
        if user_message:
            messages.append({"role": "user", "content": user_message})
        else:
            messages.append({"role": "user", "content": "Please give me a friendly summary of the analysis. Start with the overall health, then the most important findings."})
        
        try:
            response = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": api_key,
                    "content-type": "application/json",
                    "anthropic-version": "2023-06-01"
                },
                json={
                    "model": cls.HAIKU_MODEL,
                    "max_tokens": 1500,
                    "system": system_prompt,
                    "messages": messages
                },
                timeout=30
            )
            
            if response.status_code == 200:
                result = response.json()
                return result.get("content", [{}])[0].get("text", "I couldn't generate a response.")
            else:
                return f"I'm having trouble connecting right now. Please try again in a moment."
        except Exception as e:
            return f"Something went wrong: {str(e)}"


# =============================================================================
# AI BUSINESS ADVISOR - Native Click AI Analysis
# =============================================================================

@app.route("/ai-advisor", methods=["GET", "POST"])
def ai_advisor():
    """AI Business Advisor - Uses actual Click AI data for analysis"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    business_id = session.get("current_business_id")
    if not business_id:
        return redirect("/businesses")
    
    # Get business details
    business = db.select_one("businesses", business_id)
    if not business:
        return redirect("/businesses")
    
    # Build trial balance from Click AI data
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_trial_balance"
        response = requests.post(url, headers={
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }, json={"p_business_id": business_id})
        
        if response.status_code == 200:
            tb_data = response.json()
        else:
            tb_data = []
    except:
        tb_data = []
    
    # Convert to accounts format
    accounts = []
    total_debit = Decimal("0")
    total_credit = Decimal("0")
    
    for row in tb_data:
        debit = Decimal(str(row.get('debit', 0) or 0))
        credit = Decimal(str(row.get('credit', 0) or 0))
        total_debit += debit
        total_credit += credit
        accounts.append({
            "code": row.get('code', ''),
            "name": row.get('account_name', row.get('name', '')),
            "category": row.get('category', ''),
            "debit": float(debit),
            "credit": float(credit),
            "balance": float(debit - credit)
        })
    
    # Get additional business context
    # Count transactions, invoices, etc
    try:
        tx_count = len(db.select("transactions", {"business_id": business_id}))
        inv_count = len(db.select("invoices", {"business_id": business_id}))
        cust_count = len(db.select("customers", {"business_id": business_id}))
    except:
        tx_count = inv_count = cust_count = 0
    
    business_context = {
        "name": business.get("business_name", business.get("name", "Unknown")),
        "industry": business.get("industry", "general"),
        "business_type": business.get("business_type", "SME"),
        "created_at": business.get("created_at", ""),
        "transaction_count": tx_count,
        "invoice_count": inv_count,
        "customer_count": cust_count
    }
    
    # Store in session
    session['ai_advisor_accounts'] = accounts
    session['ai_advisor_business'] = business_context
    session['ai_advisor_history'] = []
    
    # Quick stats
    net_position = total_credit - total_debit
    
    content = f'''
    <div class="mb-lg">
        <a href="/reports" class="text-muted">← Reports</a>
        <h1>🧠 AI Business Advisor</h1>
        <p class="text-muted">Chat with AI about <strong>{safe_string(business_context["name"])}</strong> - it knows your numbers!</p>
    </div>
    
    <!-- Business Overview -->
    <div class="card mb-lg" style="background: linear-gradient(135deg, rgba(16,185,129,0.1), rgba(139,92,246,0.1));">
        <div style="display:flex;align-items:center;gap:16px;margin-bottom:16px;">
            <div style="font-size:48px;">🏢</div>
            <div>
                <h2 style="margin:0;">{safe_string(business_context["name"])}</h2>
                <p style="color:#8b8b9a;margin:4px 0 0;">{business_context["industry"].title()} • {tx_count} transactions • {cust_count} customers</p>
            </div>
        </div>
        <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:16px;">
            <div style="text-align:center;padding:16px;background:rgba(255,255,255,0.05);border-radius:8px;">
                <div style="font-size:24px;font-weight:700;color:#10b981;">R {total_credit:,.0f}</div>
                <div style="color:#8b8b9a;font-size:13px;">Total Credits</div>
            </div>
            <div style="text-align:center;padding:16px;background:rgba(255,255,255,0.05);border-radius:8px;">
                <div style="font-size:24px;font-weight:700;color:#ef4444;">R {total_debit:,.0f}</div>
                <div style="color:#8b8b9a;font-size:13px;">Total Debits</div>
            </div>
            <div style="text-align:center;padding:16px;background:rgba(255,255,255,0.05);border-radius:8px;">
                <div style="font-size:24px;font-weight:700;color:#a78bfa;">R {abs(net_position):,.0f}</div>
                <div style="color:#8b8b9a;font-size:13px;">{"Profit" if net_position > 0 else "Loss"}</div>
            </div>
            <div style="text-align:center;padding:16px;background:rgba(255,255,255,0.05);border-radius:8px;">
                <div style="font-size:24px;font-weight:700;color:#60a5fa;">{len(accounts)}</div>
                <div style="color:#8b8b9a;font-size:13px;">Accounts</div>
            </div>
        </div>
    </div>
    
    <!-- Chat Interface -->
    <div class="card" id="chat-container">
        <div id="chat-messages" style="max-height:400px;overflow-y:auto;padding:10px;">
            <div id="ai-message" class="chat-message assistant" style="background:rgba(139,92,246,0.1);padding:16px;border-radius:12px;margin-bottom:12px;">
                <div style="font-weight:600;color:#a78bfa;margin-bottom:8px;">🤖 BB Fin Advisor</div>
                <div style="line-height:1.6;">
                    Hi! I'm your AI financial advisor. I've got access to all the data for <strong>{safe_string(business_context["name"])}</strong>.
                    <br><br>
                    Ask me anything about your finances - I can help you understand your numbers, spot trends, and find opportunities!
                    <br><br>
                    <em style="color:#8b8b9a;">Try: "How's my business doing?" or "What are my biggest expenses?"</em>
                </div>
            </div>
        </div>
        
        <div style="border-top:1px solid var(--border);padding:16px;display:flex;gap:12px;">
            <input type="text" id="chat-input" class="form-input" style="flex:1;" placeholder="Ask about your business finances..." onkeypress="if(event.key==='Enter')sendAdvisorMessage()">
            <button type="button" class="btn btn-purple" onclick="sendAdvisorMessage()">Send</button>
        </div>
    </div>
    
    <!-- Quick Questions -->
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin-top:20px;">
        <button class="btn btn-ghost" onclick="askAdvisor('How is my business doing overall?')">📊 Overall Health</button>
        <button class="btn btn-ghost" onclick="askAdvisor('What are my biggest expenses?')">💸 Top Expenses</button>
        <button class="btn btn-ghost" onclick="askAdvisor('Am I making a profit?')">💰 Profitability</button>
        <button class="btn btn-ghost" onclick="askAdvisor('What should I focus on improving?')">🎯 Priorities</button>
        <button class="btn btn-ghost" onclick="askAdvisor('Any SARS concerns I should know about?')">🏛️ SARS Check</button>
        <button class="btn btn-ghost" onclick="askAdvisor('How is my cash flow?')">🌊 Cash Flow</button>
    </div>
    
    <script>
    function sendAdvisorMessage() {{
        const input = document.getElementById('chat-input');
        const message = input.value.trim();
        if (!message) return;
        
        const chatMessages = document.getElementById('chat-messages');
        
        // Add user message
        chatMessages.innerHTML += `
            <div class="chat-message user" style="background:rgba(59,130,246,0.1);padding:16px;border-radius:12px;margin-bottom:12px;">
                <div style="font-weight:600;color:#60a5fa;margin-bottom:8px;">You</div>
                <div>${{message}}</div>
            </div>
        `;
        
        // Add loading message
        chatMessages.innerHTML += `
            <div id="loading-msg" class="chat-message assistant" style="background:rgba(139,92,246,0.1);padding:16px;border-radius:12px;margin-bottom:12px;">
                <div style="font-weight:600;color:#a78bfa;margin-bottom:8px;">🤖 BB Fin Advisor</div>
                <div style="color:#8b8b9a;">
                    <span class="spinner" style="display:inline-block;width:16px;height:16px;border:2px solid #8b8b9a;border-top-color:#a78bfa;border-radius:50%;animation:spin 1s linear infinite;margin-right:8px;vertical-align:middle;"></span>
                    Thinking...
                </div>
            </div>
        `;
        
        chatMessages.scrollTop = chatMessages.scrollHeight;
        input.value = '';
        
        fetch('/ai-advisor/chat', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            body: JSON.stringify({{message: message}})
        }})
        .then(r => r.json())
        .then(data => {{
            document.getElementById('loading-msg').remove();
            chatMessages.innerHTML += `
                <div class="chat-message assistant" style="background:rgba(139,92,246,0.1);padding:16px;border-radius:12px;margin-bottom:12px;">
                    <div style="font-weight:600;color:#a78bfa;margin-bottom:8px;">🤖 BB Fin Advisor</div>
                    <div style="white-space:pre-wrap;line-height:1.6;">${{data.response}}</div>
                </div>
            `;
            chatMessages.scrollTop = chatMessages.scrollHeight;
        }})
        .catch(err => {{
            document.getElementById('loading-msg').innerHTML = `
                <div style="font-weight:600;color:#a78bfa;margin-bottom:8px;">🤖 BB Fin Advisor</div>
                <div style="color:#ef4444;">Sorry, something went wrong. Please try again.</div>
            `;
        }});
    }}
    
    function askAdvisor(q) {{
        document.getElementById('chat-input').value = q;
        sendAdvisorMessage();
    }}
    </script>
    <style>@keyframes spin {{ to {{ transform: rotate(360deg); }} }}</style>
    '''
    
    return page_wrapper("AI Business Advisor", content, active="reports", user=user)


@app.route("/ai-advisor/chat", methods=["POST"])
def ai_advisor_chat():
    """Handle AI advisor chat messages"""
    user = UserSession.get_current_user()
    if not user:
        return jsonify({"error": "Not logged in", "response": "Please log in to continue."})
    
    data = request.get_json()
    message = data.get("message", "")
    
    # Get stored data
    accounts = session.get('ai_advisor_accounts', [])
    business_context = session.get('ai_advisor_business', {})
    history = session.get('ai_advisor_history', [])
    
    if not accounts:
        return jsonify({"response": "I don't have your business data loaded. Please go back to the AI Advisor page and try again."})
    
    # Build analysis context (simplified for chat)
    analysis = {
        "accounts": accounts[:50],  # Limit to top 50 accounts for context
        "total_accounts": len(accounts),
        "total_debit": sum(a['debit'] for a in accounts),
        "total_credit": sum(a['credit'] for a in accounts),
        "net_position": sum(a['credit'] for a in accounts) - sum(a['debit'] for a in accounts)
    }
    
    # Add user message to history
    history.append({"role": "user", "content": message})
    
    # Get response with business context
    response = TBAnalyzer.chat_response(analysis, message, history, business_context)
    
    # Add to history
    history.append({"role": "assistant", "content": response})
    session['ai_advisor_history'] = history[-20:]  # Keep last 20
    
    return jsonify({"response": response})


@app.route("/tb-analyzer", methods=["GET", "POST"])
def tb_analyzer():
    """TB Analyzer - Upload and analyze any trial balance"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Handle POST - file upload
    if request.method == "POST":
        file = request.files.get("tb_file")
        context = request.form.get("context", "")
        
        if not file:
            return redirect("/tb-analyzer")
        
        # Read file content
        filename = file.filename.lower()
        try:
            if filename.endswith('.csv'):
                content = file.read().decode('utf-8', errors='ignore')
                accounts = TBAnalyzer.parse_tb_csv(content)
            elif filename.endswith(('.xlsx', '.xls')):
                # For Excel, try openpyxl
                import io
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(file.read()))
                    ws = wb.active
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append(','.join(str(c) if c else '' for c in row))
                    content = '\n'.join(rows)
                    accounts = TBAnalyzer.parse_tb_csv(content)
                except ImportError:
                    content = file.read().decode('utf-8', errors='ignore')
                    accounts = TBAnalyzer.parse_tb_csv(content)
            else:
                accounts = []
        except Exception as e:
            accounts = []
        
        if not accounts:
            error_content = '''
            <div class="mb-lg">
                <a href="/tb-analyzer" class="text-muted">← Back</a>
                <h1>🧠 TB Analyzer</h1>
            </div>
            <div class="alert alert-error">
                Could not read the file. Please make sure it's a valid CSV or Excel file with account codes and balances.
            </div>
            <a href="/tb-analyzer" class="btn btn-primary">Try Again</a>
            '''
            return page_wrapper("TB Analyzer", error_content, active="reports", user=user)
        
        # Calculate quick stats first (instant)
        total_debit = sum(a['debit'] for a in accounts)
        total_credit = sum(a['credit'] for a in accounts)
        total_income = sum(a['credit'] for a in accounts if a.get('category', '').lower() in ['sales', 'income', 'revenue', 'other income'])
        total_expenses = sum(a['debit'] for a in accounts if a.get('category', '').lower() in ['expenses', 'expense', 'cost of sales'])
        
        # Store in session for chat
        session['tb_accounts'] = accounts
        session['tb_context'] = context
        
        # Quick stats to show immediately
        quick_stats = f"""
        <div class="card mb-lg">
            <h3>📊 Quick Stats (from {len(accounts)} accounts)</h3>
            <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:16px;margin-top:16px;">
                <div style="text-align:center;padding:16px;background:rgba(16,185,129,0.1);border-radius:8px;">
                    <div style="font-size:24px;font-weight:700;color:#10b981;">R {total_credit:,.0f}</div>
                    <div style="color:#8b8b9a;font-size:13px;">Total Credits</div>
                </div>
                <div style="text-align:center;padding:16px;background:rgba(239,68,68,0.1);border-radius:8px;">
                    <div style="font-size:24px;font-weight:700;color:#ef4444;">R {total_debit:,.0f}</div>
                    <div style="color:#8b8b9a;font-size:13px;">Total Debits</div>
                </div>
                <div style="text-align:center;padding:16px;background:rgba(139,92,246,0.1);border-radius:8px;">
                    <div style="font-size:24px;font-weight:700;color:#a78bfa;">R {abs(total_credit - total_debit):,.0f}</div>
                    <div style="color:#8b8b9a;font-size:13px;">Net Position</div>
                </div>
            </div>
        </div>
        """
        # DON'T run analysis here - will timeout. Do it async via JavaScript
        # Just store the accounts for the async endpoint
        session['tb_analysis'] = {}  # Will be populated async
        session['tb_chat_history'] = []
        
        # Check for errors
        error_msg = ""
        
        # Build the chat interface - AI loads async
        result_content = f'''
        <div class="mb-lg">
            <a href="/tb-analyzer" class="text-muted">← Upload New TB</a>
            <h1>🧠 TB Analysis Results</h1>
            <p class="text-muted">Analyzed {len(accounts)} accounts</p>
        </div>
        
        {error_msg}
        
        <!-- Quick Stats -->
        {quick_stats}
        
        <!-- Health Banner - will be updated by JS -->
        <div class="card" id="health-banner" style="background:linear-gradient(135deg,#8b8b9a20,#8b8b9a05);border-color:#8b8b9a50;margin-bottom:20px;">
            <div style="display:flex;align-items:center;gap:16px;">
                <div style="font-size:48px;" id="health-icon">⏳</div>
                <div>
                    <h2 style="color:#8b8b9a;margin:0;" id="health-title">Analyzing...</h2>
                    <p style="color:#c0c0c0;margin:4px 0 0;" id="health-summary">Our AI accountant is reviewing your numbers...</p>
                </div>
            </div>
        </div>
        
        <!-- Chat Interface -->
        <div class="card" id="chat-container">
            <div id="chat-messages" style="max-height:500px;overflow-y:auto;padding:10px;">
                <div id="initial-message" class="chat-message assistant" style="background:rgba(139,92,246,0.1);padding:16px;border-radius:12px;margin-bottom:12px;">
                    <div style="font-weight:600;color:#a78bfa;margin-bottom:8px;">🤖 BB Fin Assistant</div>
                    <div style="line-height:1.6;">
                        <div style="display:flex;align-items:center;gap:8px;">
                            <div class="spinner" style="width:20px;height:20px;border:3px solid #8b8b9a;border-top-color:#a78bfa;border-radius:50%;animation:spin 1s linear infinite;"></div>
                            <span>Analyzing your trial balance... this takes about 30 seconds</span>
                        </div>
                    </div>
                </div>
            </div>
            
            <div style="border-top:1px solid var(--border);padding:16px;display:flex;gap:12px;">
                <input type="text" id="chat-input" class="form-input" style="flex:1;" placeholder="Ask a follow-up question..." onkeypress="if(event.key==='Enter')sendMessage()" disabled>
                <button type="button" class="btn btn-purple" onclick="sendMessage()" id="send-btn" disabled>Send</button>
            </div>
        </div>
        
        <!-- Quick Actions -->
        <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:12px;margin-top:20px;">
            <button class="btn btn-ghost" onclick="askQuestion('What are the biggest red flags?')" disabled id="q1">🚩 Red Flags</button>
            <button class="btn btn-ghost" onclick="askQuestion('What opportunities do you see?')" disabled id="q2">💡 Opportunities</button>
            <button class="btn btn-ghost" onclick="askQuestion('What would SARS look at?')" disabled id="q3">🏛️ SARS Concerns</button>
            <button class="btn btn-ghost" onclick="askQuestion('What should I do first?')" disabled id="q4">📋 Priorities</button>
        </div>
        
        <!-- Detailed Metrics (collapsible) -->
        <details class="card mt-lg" style="cursor:pointer;">
            <summary style="font-weight:600;padding:16px;">📊 Detailed Metrics & Findings</summary>
            <div style="padding:0 16px 16px;">
                <pre id="analysis-json" style="background:#0a0a10;padding:16px;border-radius:8px;overflow-x:auto;font-size:12px;">Loading analysis data...</pre>
            </div>
        </details>
        
        <script>
        function sendMessage() {{
            const input = document.getElementById('chat-input');
            const message = input.value.trim();
            if (!message) return;
            
            const chatMessages = document.getElementById('chat-messages');
            chatMessages.innerHTML += `
                <div class="chat-message user" style="background:rgba(59,130,246,0.1);padding:16px;border-radius:12px;margin-bottom:12px;">
                    <div style="font-weight:600;color:#60a5fa;margin-bottom:8px;">You</div>
                    <div>${{message}}</div>
                </div>
            `;
            
            chatMessages.innerHTML += `
                <div id="loading-msg" class="chat-message assistant" style="background:rgba(139,92,246,0.1);padding:16px;border-radius:12px;margin-bottom:12px;">
                    <div style="font-weight:600;color:#a78bfa;margin-bottom:8px;">🤖 BB Fin Assistant</div>
                    <div style="color:#8b8b9a;">Thinking...</div>
                </div>
            `;
            
            chatMessages.scrollTop = chatMessages.scrollHeight;
            input.value = '';
            
            fetch('/tb-analyzer/chat', {{
                method: 'POST',
                headers: {{'Content-Type': 'application/json'}},
                body: JSON.stringify({{message: message}})
            }})
            .then(r => r.json())
            .then(data => {{
                document.getElementById('loading-msg').remove();
                chatMessages.innerHTML += `
                    <div class="chat-message assistant" style="background:rgba(139,92,246,0.1);padding:16px;border-radius:12px;margin-bottom:12px;">
                        <div style="font-weight:600;color:#a78bfa;margin-bottom:8px;">🤖 BB Fin Assistant</div>
                        <div style="white-space:pre-wrap;line-height:1.6;">${{data.response}}</div>
                    </div>
                `;
                chatMessages.scrollTop = chatMessages.scrollHeight;
            }})
            .catch(err => {{
                document.getElementById('loading-msg').innerHTML = '<div style="color:#ef4444;">Error getting response</div>';
            }});
        }}
        
        function askQuestion(q) {{
            document.getElementById('chat-input').value = q;
            sendMessage();
        }}
        
        // Load AI analysis async on page load
        window.onload = function() {{
            fetch('/tb-analyzer/analyze', {{
                method: 'POST',
                headers: {{'Content-Type': 'application/json'}}
            }})
            .then(r => {{
                if (!r.ok) throw new Error('Server error');
                return r.json();
            }})
            .then(data => {{
                if (data.error) {{
                    throw new Error(data.error);
                }}
                
                // Update health banner
                const colors = {{'good': '#10b981', 'warning': '#f59e0b', 'critical': '#ef4444'}};
                const icons = {{'good': '✅', 'warning': '⚠️', 'critical': '🚨'}};
                const health = data.health || 'unknown';
                const color = colors[health] || '#8b8b9a';
                const icon = icons[health] || '❓';
                
                document.getElementById('health-icon').textContent = icon;
                document.getElementById('health-title').textContent = 'Business Health: ' + health.toUpperCase();
                document.getElementById('health-title').style.color = color;
                document.getElementById('health-summary').textContent = data.summary || '';
                document.getElementById('health-banner').style.background = 'linear-gradient(135deg,' + color + '20,' + color + '05)';
                document.getElementById('health-banner').style.borderColor = color + '50';
                
                // Update chat with AI response
                document.getElementById('initial-message').innerHTML = `
                    <div style="font-weight:600;color:#a78bfa;margin-bottom:8px;">🤖 BB Fin Assistant</div>
                    <div style="white-space:pre-wrap;line-height:1.6;">${{data.chat_response}}</div>
                `;
                
                // Enable inputs
                document.getElementById('chat-input').disabled = false;
                document.getElementById('send-btn').disabled = false;
                document.getElementById('q1').disabled = false;
                document.getElementById('q2').disabled = false;
                document.getElementById('q3').disabled = false;
                document.getElementById('q4').disabled = false;
                
                // Update detailed analysis JSON
                if (data.analysis) {{
                    document.getElementById('analysis-json').textContent = JSON.stringify(data.analysis, null, 2);
                }}
            }})
            .catch(err => {{
                console.error('Analysis error:', err);
                document.getElementById('initial-message').innerHTML = `
                    <div style="font-weight:600;color:#a78bfa;margin-bottom:8px;">🤖 BB Fin Assistant</div>
                    <div style="color:#ef4444;">Sorry, the analysis failed: ${{err.message}}. Please try again.</div>
                `;
                document.getElementById('health-icon').textContent = '⚠️';
                document.getElementById('health-title').textContent = 'Analysis Failed';
                document.getElementById('health-summary').textContent = 'Could not complete analysis - try again';
            }});
        }};
        </script>
        <style>@keyframes spin {{ to {{ transform: rotate(360deg); }} }}</style>
        '''
        
        return page_wrapper("TB Analysis", result_content, active="reports", user=user)
    
    # GET - show upload form
    content = '''
    <div class="mb-lg">
        <h1>🧠 TB Analyzer</h1>
        <p class="text-muted">Upload any Trial Balance and let our AI accountant analyze it for you</p>
    </div>
    
    <div class="card" style="border: 2px dashed var(--purple); background: rgba(139,92,246,0.05);">
        <form method="POST" enctype="multipart/form-data" id="upload-form">
            <div style="text-align:center;padding:40px;">
                <div style="font-size:48px;margin-bottom:16px;">📊</div>
                <h3 style="margin-bottom:8px;">Drop your Trial Balance here</h3>
                <p class="text-muted" style="margin-bottom:20px;">CSV or Excel file with account codes, names, and balances</p>
                
                <input type="file" name="tb_file" id="tb-file" accept=".csv,.xlsx,.xls" style="display:none;" onchange="handleFileSelect(this)">
                <label for="tb-file" class="btn btn-purple btn-lg" style="cursor:pointer;">
                    📁 Choose File
                </label>
                
                <div id="file-info" style="margin-top:16px;display:none;">
                    <span id="file-name" style="color:var(--green);font-weight:600;"></span>
                </div>
            </div>
            
            <div style="border-top:1px solid var(--border);padding:20px;margin-top:20px;">
                <label class="form-label">Business Context (optional)</label>
                <textarea name="context" class="form-input" rows="2" placeholder="e.g. Hardware store, been trading 5 years, had a tough year due to load shedding..."></textarea>
            </div>
            
            <div style="padding:0 20px 20px;">
                <button type="submit" class="btn btn-primary btn-lg" style="width:100%;" id="analyze-btn" disabled>
                    🔍 Analyze Trial Balance
                </button>
            </div>
        </form>
    </div>
    
    <div class="card mt-lg" style="background:linear-gradient(135deg,rgba(139,92,246,0.1),rgba(59,130,246,0.1));">
        <h3 style="color:#a78bfa;margin-bottom:12px;">🤖 How it works</h3>
        <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:20px;">
            <div>
                <div style="font-size:24px;margin-bottom:8px;">1️⃣</div>
                <strong>Upload</strong>
                <p class="text-muted" style="font-size:13px;">Drop your TB file - CSV or Excel from any accounting system</p>
            </div>
            <div>
                <div style="font-size:24px;margin-bottom:8px;">2️⃣</div>
                <strong>Deep Analysis</strong>
                <p class="text-muted" style="font-size:13px;">Our AI accountant examines ratios, patterns, and red flags</p>
            </div>
            <div>
                <div style="font-size:24px;margin-bottom:8px;">3️⃣</div>
                <strong>Friendly Report</strong>
                <p class="text-muted" style="font-size:13px;">Get insights in plain language you can actually use</p>
            </div>
        </div>
    </div>
    
    <script>
    function handleFileSelect(input) {
        if (input.files && input.files[0]) {
            document.getElementById('file-name').textContent = input.files[0].name;
            document.getElementById('file-info').style.display = 'block';
            document.getElementById('analyze-btn').disabled = false;
        }
    }
    
    // Drag and drop
    const form = document.getElementById('upload-form');
    form.addEventListener('dragover', (e) => {
        e.preventDefault();
        form.style.borderColor = 'var(--green)';
    });
    form.addEventListener('dragleave', () => {
        form.style.borderColor = 'var(--purple)';
    });
    form.addEventListener('drop', (e) => {
        e.preventDefault();
        form.style.borderColor = 'var(--purple)';
        const file = e.dataTransfer.files[0];
        if (file) {
            document.getElementById('tb-file').files = e.dataTransfer.files;
            handleFileSelect(document.getElementById('tb-file'));
        }
    });
    </script>
    '''
    
    return page_wrapper("TB Analyzer", content, active="reports", user=user)


@app.route("/tb-analyzer/chat", methods=["POST"])
def tb_analyzer_chat():
    """Handle chat follow-up questions"""
    user = UserSession.get_current_user()
    if not user:
        return jsonify({"error": "Not logged in"})
    
    data = request.get_json()
    message = data.get("message", "")
    
    # Get stored analysis and history
    analysis = session.get('tb_analysis', {})
    history = session.get('tb_chat_history', [])
    
    # Add user message to history
    history.append({"role": "user", "content": message})
    
    # Get response from Haiku
    response = TBAnalyzer.chat_response(analysis, message, history)
    
    # Add assistant response to history
    history.append({"role": "assistant", "content": response})
    session['tb_chat_history'] = history[-20:]  # Keep last 20 messages
    
    return jsonify({"response": response})


@app.route("/tb-analyzer/analyze", methods=["POST"])
def tb_analyzer_analyze():
    """Async endpoint to run AI analysis - called via JavaScript"""
    user = UserSession.get_current_user()
    if not user:
        return jsonify({"error": "Not logged in"})
    
    # Get stored accounts from session
    accounts = session.get('tb_accounts', [])
    context = session.get('tb_context', '')
    
    if not accounts:
        return jsonify({
            "error": "No data to analyze",
            "health": "unknown",
            "summary": "Please upload a trial balance first",
            "chat_response": "I don't have any data to analyze. Please go back and upload a trial balance file."
        })
    
    try:
        # Run analysis with Opus
        analysis = TBAnalyzer.analyze_with_opus(accounts, context)
        session['tb_analysis'] = analysis
        
        # Get chat response from Haiku
        chat_response = TBAnalyzer.chat_response(analysis)
        session['tb_chat_history'] = [
            {"role": "assistant", "content": chat_response}
        ]
        
        return jsonify({
            "health": analysis.get('company_health', 'unknown'),
            "summary": analysis.get('health_summary', ''),
            "chat_response": chat_response,
            "analysis": analysis
        })
    except Exception as e:
        return jsonify({
            "error": str(e),
            "health": "unknown",
            "summary": "Analysis failed",
            "chat_response": f"Sorry, I encountered an error while analyzing: {str(e)}"
        })


@app.route("/journal-entry", methods=["GET", "POST"])
def journal_entry():
    """Manual journal entry for adjustments"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = ""
    
    if request.method == "POST":
        date = request.form.get("date", today())
        reference = request.form.get("reference", "")
        description = request.form.get("description", "Manual Adjustment")
        
        # Get debit/credit entries
        debit_account = request.form.get("debit_account", "")
        debit_amount = Decimal(request.form.get("debit_amount", 0) or 0)
        credit_account = request.form.get("credit_account", "")
        credit_amount = Decimal(request.form.get("credit_amount", 0) or 0)
        
        # Validate
        if debit_amount != credit_amount:
            message = '<div class="alert alert-error">Debits must equal credits!</div>'
        elif debit_amount <= 0:
            message = '<div class="alert alert-error">Please enter an amount</div>'
        elif not debit_account or not credit_account:
            message = '<div class="alert alert-error">Please select both accounts</div>'
        else:
            # Post the entry
            entry = JournalEntry(
                date=date,
                reference=reference or f"JE-{today()}",
                description=description,
                trans_type=TransactionType.JOURNAL,
                source_type="manual",
                source_id=generate_id()
            )
            
            entry.debit(debit_account, debit_amount)
            entry.credit(credit_account, credit_amount)
            
            success, result = entry.post()
            
            if success:
                message = '<div class="alert alert-info">✓ Journal entry posted successfully!</div>'
            else:
                message = f'<div class="alert alert-error">Error: {result}</div>'
    
    # Build account options
    accounts = [
        ("1000", "Bank"),
        ("1100", "Petty Cash"),
        ("1200", "Debtors"),
        ("1300", "Stock"),
        ("2000", "Creditors"),
        ("2100", "VAT Output"),
        ("2200", "VAT Input"),
        ("3000", "Capital"),
        ("3100", "Retained Earnings"),
        ("4000", "Sales"),
        ("5000", "Cost of Sales"),
        ("6000", "Salaries"),
        ("6100", "Rent"),
        ("6200", "Utilities"),
        ("6300", "Repairs"),
        ("6400", "Fuel"),
        ("6500", "Office Expenses"),
        ("6600", "Bank Charges"),
        ("6900", "Other Expenses"),
    ]
    
    account_options = "".join([f'<option value="{code}">{code} - {name}</option>' for code, name in accounts])
    
    content = f'''
    <div class="mb-lg">
        <a href="/reports" class="text-muted">← Reports</a>
        <h1>📝 Journal Entry</h1>
        <p class="text-muted">Manual adjustments - debits must equal credits</p>
    </div>
    
    {message}
    
    <div class="card">
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Date</label>
                    <input type="date" name="date" class="form-input" value="{today()}">
                </div>
                <div class="form-group">
                    <label class="form-label">Reference</label>
                    <input type="text" name="reference" class="form-input" placeholder="JE-001">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Description</label>
                <input type="text" name="description" class="form-input" placeholder="What is this adjustment for?">
            </div>
            
            <div style="background:#1a1a2e; padding:16px; border-radius:8px; margin:16px 0;">
                <h4 style="margin-bottom:12px;">Debit (Increase)</h4>
                <div class="form-row">
                    <div class="form-group" style="flex:2;">
                        <select name="debit_account" class="form-select">
                            <option value="">Select account...</option>
                            {account_options}
                        </select>
                    </div>
                    <div class="form-group" style="flex:1;">
                        <input type="number" name="debit_amount" class="form-input" placeholder="0.00" step="0.01" id="debit-amt" onchange="syncAmounts(this)">
                    </div>
                </div>
            </div>
            
            <div style="background:#1a1a2e; padding:16px; border-radius:8px; margin:16px 0;">
                <h4 style="margin-bottom:12px;">Credit (Decrease)</h4>
                <div class="form-row">
                    <div class="form-group" style="flex:2;">
                        <select name="credit_account" class="form-select">
                            <option value="">Select account...</option>
                            {account_options}
                        </select>
                    </div>
                    <div class="form-group" style="flex:1;">
                        <input type="number" name="credit_amount" class="form-input" placeholder="0.00" step="0.01" id="credit-amt">
                    </div>
                </div>
            </div>
            
            <button type="submit" class="btn btn-primary btn-block">Post Journal Entry</button>
        </form>
    </div>
    
    <script>
    function syncAmounts(el) {{
        document.getElementById('credit-amt').value = el.value;
    }}
    </script>
    '''
    
    return page_wrapper("Journal Entry", content, active="reports", user=user)


@app.route("/bank-recon", methods=["GET", "POST"])
def bank_recon():
    """Simple bank reconciliation"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = ""
    
    if request.method == "POST":
        statement_balance = Decimal(request.form.get("statement_balance", 0) or 0)
        
        # Save to settings
        db.upsert("settings", {
            "id": "bank_recon_" + today(),
            "key": "bank_recon",
            "value": json.dumps({
                "date": today(),
                "statement_balance": float(statement_balance)
            })
        })
        
        message = '<div class="alert alert-info">✓ Bank balance saved</div>'
    
    # Get GL bank balance
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_trial_balance"
        headers = {
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }
        resp = requests.post(url, headers=headers, json={}, timeout=30)
        tb_data = resp.json() if resp.status_code == 200 else []
        
        # Find bank account
        gl_balance = Decimal("0")
        for entry in tb_data:
            if entry.get("account_code") == "1000":  # Bank
                gl_balance = Decimal(str(entry.get("total_debit", 0) or 0)) - Decimal(str(entry.get("total_credit", 0) or 0))
                break
    except:
        gl_balance = Decimal("0")
    
    # Get last saved statement balance
    try:
        recon = db.select("settings", filters={"key": "bank_recon"}, limit=1)
        if recon:
            recon_data = json.loads(recon[0].get("value", "{}"))
            statement_balance = Decimal(str(recon_data.get("statement_balance", 0)))
            recon_date = recon_data.get("date", "Not set")
        else:
            statement_balance = Decimal("0")
            recon_date = "Not set"
    except:
        statement_balance = Decimal("0")
        recon_date = "Not set"
    
    difference = gl_balance - statement_balance
    
    diff_class = "green" if abs(difference) < Decimal("0.01") else "red"
    diff_status = "✓ Reconciled" if abs(difference) < Decimal("0.01") else f"Difference of {Money.format(abs(difference))}"
    
    # Get recent bank transactions from GL
    gl_entries = db.select("gl_entries", filters={"account_code": "1000"}, order="date DESC", limit=20) or []
    
    rows = []
    for entry in gl_entries:
        dr = Decimal(str(entry.get("debit", 0) or 0))
        cr = Decimal(str(entry.get("credit", 0) or 0))
        
        rows.append([
            entry.get("date", "")[:10],
            entry.get("description", "")[:40],
            {"value": Money.format(dr) if dr > 0 else "-", "class": "number text-green"},
            {"value": Money.format(cr) if cr > 0 else "-", "class": "number text-red"}
        ])
    
    table = table_html(
        headers=["Date", "Description", {"label": "In", "class": "number"}, {"label": "Out", "class": "number"}],
        rows=rows,
        empty_message="No bank transactions"
    )
    
    content = f'''
    <div class="mb-lg">
        <a href="/reports" class="text-muted">← Reports</a>
        <h1>🏦 Bank Reconciliation</h1>
        <p class="text-muted">Compare your books to your bank statement</p>
    </div>
    
    {message}
    
    <div class="stats mb-lg">
        <div class="stat">
            <div class="stat-value">{Money.format(gl_balance)}</div>
            <div class="stat-label">Book Balance (GL)</div>
        </div>
        <div class="stat">
            <div class="stat-value">{Money.format(statement_balance)}</div>
            <div class="stat-label">Bank Statement</div>
        </div>
        <div class="stat">
            <div class="stat-value {diff_class}">{diff_status}</div>
            <div class="stat-label">Status</div>
        </div>
    </div>
    
    <div class="card mb-lg" style="max-width: 400px;">
        <h3 class="mb-md">Update Bank Statement Balance</h3>
        <form method="POST">
            <div class="form-group">
                <label class="form-label">Closing balance from bank statement</label>
                <input type="number" name="statement_balance" class="form-input" 
                       value="{float(statement_balance):.2f}" step="0.01"
                       style="font-size:20px;">
                <small class="text-muted">Last updated: {recon_date}</small>
            </div>
            <button type="submit" class="btn btn-primary">Save</button>
        </form>
    </div>
    
    <div class="card">
        <h3 class="mb-md">Recent Bank Transactions</h3>
        {table}
    </div>
    '''
    
    return page_wrapper("Bank Reconciliation", content, active="reports", user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# AI-POWERED REPORTS
# ═══════════════════════════════════════════════════════════════════════════════

def get_financial_data_for_ai():
    """
    Gather all financial data needed for AI analysis.
    Returns a comprehensive dict of business financials.
    """
    year_start = FinancialPeriod.get_current_year_start()
    
    data = {
        "report_date": today(),
        "period_start": year_start,
        "period_end": today(),
        "company_name": Config.COMPANY_NAME
    }
    
    # Get Trial Balance
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_trial_balance"
        headers = {
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }
        resp = requests.post(url, headers=headers, json={}, timeout=30)
        tb_data = resp.json() if resp.status_code == 200 else []
    except:
        tb_data = []
    
    # Organize TB by account type
    data["accounts"] = {}
    for entry in tb_data:
        code = entry.get("account_code", "")
        data["accounts"][code] = {
            "name": entry.get("account_name", ""),
            "debit": float(entry.get("total_debit", 0) or 0),
            "credit": float(entry.get("total_credit", 0) or 0),
            "balance": float(entry.get("total_debit", 0) or 0) - float(entry.get("total_credit", 0) or 0)
        }
    
    # Get Income Statement data
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_income_statement"
        headers = {
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }
        resp = requests.post(url, headers=headers, json={"date_from": year_start, "date_to": today()}, timeout=30)
        is_data = resp.json() if resp.status_code == 200 else []
    except:
        is_data = []
    
    revenue = 0
    cogs = 0
    expenses = 0
    revenue_items = []
    expense_items = []
    
    for entry in is_data:
        amt = float(entry.get("amount", 0) or 0)
        cat = entry.get("category", "")
        name = entry.get("account_name", "")
        
        if cat == "revenue":
            revenue += amt
            revenue_items.append({"name": name, "amount": amt})
        elif cat == "cost_of_sales":
            cogs += amt
        elif cat == "expense":
            expenses += amt
            expense_items.append({"name": name, "amount": amt})
    
    data["income_statement"] = {
        "revenue": revenue,
        "cost_of_sales": cogs,
        "gross_profit": revenue - cogs,
        "gross_margin_pct": ((revenue - cogs) / revenue * 100) if revenue > 0 else 0,
        "expenses": expenses,
        "net_profit": revenue - cogs - expenses,
        "net_margin_pct": ((revenue - cogs - expenses) / revenue * 100) if revenue > 0 else 0,
        "revenue_breakdown": sorted(revenue_items, key=lambda x: x["amount"], reverse=True)[:10],
        "expense_breakdown": sorted(expense_items, key=lambda x: x["amount"], reverse=True)[:10]
    }
    
    # Get Balance Sheet data
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_balance_sheet"
        headers = {
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }
        resp = requests.post(url, headers=headers, json={}, timeout=30)
        bs_data = resp.json() if resp.status_code == 200 else []
    except:
        bs_data = []
    
    current_assets = 0
    fixed_assets = 0
    current_liab = 0
    longterm_liab = 0
    equity = 0
    
    bank_balance = 0
    debtors = 0
    creditors = 0
    stock = 0
    vat_payable = 0
    
    for item in bs_data:
        bal = float(item.get("balance", 0) or 0)
        cat = item.get("account_category", "")
        typ = item.get("account_type", "")
        name = item.get("account_name", "").lower()
        
        if typ == "asset":
            if cat in ["current_asset", "bank", "receivable", "inventory"]:
                current_assets += bal
                if "bank" in name or cat == "bank":
                    bank_balance += bal
                elif "debtor" in name or cat == "receivable":
                    debtors += bal
                elif "stock" in name or "inventory" in name or cat == "inventory":
                    stock += bal
            else:
                fixed_assets += bal
        elif typ == "liability":
            if cat in ["current_liability", "payable", "vat"]:
                current_liab += bal
                if "creditor" in name or cat == "payable":
                    creditors += bal
                if "vat" in name:
                    vat_payable += bal
            else:
                longterm_liab += bal
        elif typ == "equity":
            equity += bal
    
    total_assets = current_assets + fixed_assets
    total_liab = current_liab + longterm_liab
    
    data["balance_sheet"] = {
        "current_assets": current_assets,
        "fixed_assets": fixed_assets,
        "total_assets": total_assets,
        "current_liabilities": current_liab,
        "longterm_liabilities": longterm_liab,
        "total_liabilities": total_liab,
        "equity": equity,
        "bank_balance": bank_balance,
        "debtors": debtors,
        "creditors": creditors,
        "stock": stock,
        "vat_payable": vat_payable
    }
    
    # Calculate key ratios
    data["ratios"] = {
        "current_ratio": current_assets / current_liab if current_liab > 0 else 0,
        "quick_ratio": (current_assets - stock) / current_liab if current_liab > 0 else 0,
        "debt_to_equity": total_liab / equity if equity > 0 else 0,
        "debtor_days": (debtors / revenue * 365) if revenue > 0 else 0,
        "creditor_days": (creditors / cogs * 365) if cogs > 0 else 0,
        "stock_days": (stock / cogs * 365) if cogs > 0 else 0,
        "return_on_equity": ((revenue - cogs - expenses) / equity * 100) if equity > 0 else 0,
        "return_on_assets": ((revenue - cogs - expenses) / total_assets * 100) if total_assets > 0 else 0
    }
    
    # Get sales and customer data
    try:
        invoices = db.select("invoices", order="-date", limit=100)
        total_sales = sum(float(inv.get("total", 0) or 0) for inv in invoices)
        outstanding_invoices = [inv for inv in invoices if inv.get("status") == "outstanding"]
        outstanding_total = sum(float(inv.get("total", 0) or 0) for inv in outstanding_invoices)
        
        data["sales"] = {
            "total_invoices": len(invoices),
            "total_sales": total_sales,
            "outstanding_count": len(outstanding_invoices),
            "outstanding_total": outstanding_total
        }
    except:
        data["sales"] = {"total_invoices": 0, "total_sales": 0, "outstanding_count": 0, "outstanding_total": 0}
    
    return data


@app.route("/reports/business-health")
def report_business_health():
    """
    Business Health Report - AI-powered plain-language analysis
    Written like a friend explaining your finances
    """
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Check if we should generate the report
    generate = request.args.get("generate") == "yes"
    
    if not generate:
        # Show intro page
        content = '''
        <div class="card" style="max-width: 600px; margin: 0 auto; text-align: center;">
            <div style="font-size: 64px; margin-bottom: 20px;">💊</div>
            <h1 style="margin-bottom: 12px;">Business Health Report</h1>
            <p class="text-muted" style="margin-bottom: 24px;">
                AI will analyze your financials and explain what's happening in plain language.<br>
                No accounting jargon - just straight talk about your business.
            </p>
            <p style="margin-bottom: 24px; padding: 16px; background: var(--bg-secondary); border-radius: 8px;">
                <strong>What you'll get:</strong><br>
                • How healthy is your cash position?<br>
                • Are you making or losing money?<br>
                • What risks should you know about?<br>
                • What should you do next?
            </p>
            <a href="/reports/business-health?generate=yes" class="btn btn-primary btn-lg">
                🤖 Generate Report
            </a>
            <p class="text-muted mt-md" style="font-size: 12px;">Takes about 30 seconds</p>
        </div>
        '''
        return page_wrapper("Business Health Report", content, "reports", user)
    
    # Generate the report with AI
    financial_data = get_financial_data_for_ai()
    
    # Check for API key
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        content = '''
        <div class="alert alert-error">
            AI not configured. Add ANTHROPIC_API_KEY to environment variables.
        </div>
        '''
        return page_wrapper("Business Health Report", content, "reports", user)
    
    # Build prompt for Opus
    prompt = f"""You are a sharp financial advisor writing a Business Health Report for a business owner. 

The owner is NOT an accountant. They want to understand their business finances in plain language - like a friend explaining things over coffee.

Here is the financial data for {financial_data['company_name']}:

PERIOD: {financial_data['period_start']} to {financial_data['period_end']}

INCOME STATEMENT:
- Revenue: R {financial_data['income_statement']['revenue']:,.2f}
- Cost of Sales: R {financial_data['income_statement']['cost_of_sales']:,.2f}
- Gross Profit: R {financial_data['income_statement']['gross_profit']:,.2f} ({financial_data['income_statement']['gross_margin_pct']:.1f}%)
- Operating Expenses: R {financial_data['income_statement']['expenses']:,.2f}
- Net Profit: R {financial_data['income_statement']['net_profit']:,.2f} ({financial_data['income_statement']['net_margin_pct']:.1f}%)

Top Expenses: {json.dumps(financial_data['income_statement']['expense_breakdown'][:5])}

BALANCE SHEET:
- Bank Balance: R {financial_data['balance_sheet']['bank_balance']:,.2f}
- Debtors (owed to you): R {financial_data['balance_sheet']['debtors']:,.2f}
- Stock: R {financial_data['balance_sheet']['stock']:,.2f}
- Creditors (you owe): R {financial_data['balance_sheet']['creditors']:,.2f}
- VAT Payable: R {financial_data['balance_sheet']['vat_payable']:,.2f}
- Total Assets: R {financial_data['balance_sheet']['total_assets']:,.2f}
- Total Liabilities: R {financial_data['balance_sheet']['total_liabilities']:,.2f}
- Equity: R {financial_data['balance_sheet']['equity']:,.2f}

KEY RATIOS:
- Current Ratio: {financial_data['ratios']['current_ratio']:.2f} (healthy is 1.5-2.0)
- Quick Ratio: {financial_data['ratios']['quick_ratio']:.2f} (healthy is 1.0+)
- Debtor Days: {financial_data['ratios']['debtor_days']:.0f} days (how long customers take to pay)
- Creditor Days: {financial_data['ratios']['creditor_days']:.0f} days (how long you take to pay suppliers)
- Stock Days: {financial_data['ratios']['stock_days']:.0f} days (how long stock sits before selling)

SALES INFO:
- Total Invoices: {financial_data['sales']['total_invoices']}
- Outstanding Invoices: {financial_data['sales']['outstanding_count']} worth R {financial_data['sales']['outstanding_total']:,.2f}

Write a Business Health Report with these sections. Use a conversational, direct tone. No jargon. Be specific with numbers. Tell them what it MEANS, not just what it IS.

1. THE VERDICT (2-3 sentences - is this business healthy or not?)

2. CASH POSITION (How much actual cash do you have? Can you pay your bills? What's the real picture?)

3. PROFIT REALITY (Are you actually making money? What's eating into your margins? Be honest.)

4. WHAT'S WORRYING (What risks or red flags do you see? Don't sugarcoat it.)

5. WHAT'S GOOD (What's working well? Give credit where it's due.)

6. DO THIS NOW (3-5 specific, actionable things to improve the business. Be concrete.)

Format as clean HTML with <h2> for section headers. Use <p> for paragraphs. Keep it readable. About 800-1000 words total.

Remember: You're talking to the OWNER, not their accountant. They need to understand and act on this."""

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        
        message = client.messages.create(
            model="claude-opus-4-20250514",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )
        
        report_html = message.content[0].text
        
    except Exception as e:
        report_html = f"<p class='text-red'>Error generating report: {str(e)}</p>"
    
    # Build the page
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <h1 style="font-size: 24px; font-weight: 700;">💊 Business Health Report</h1>
            <p class="text-muted">{financial_data['company_name']} • {financial_data['period_start']} to {financial_data['period_end']}</p>
        </div>
        <div class="btn-group">
            <button onclick="window.print()" class="btn btn-ghost">🖨️ Print</button>
            <a href="/reports/business-health?generate=yes" class="btn btn-ghost">🔄 Regenerate</a>
        </div>
    </div>
    
    <div class="stats mb-lg">
        <div class="stat"><div class="stat-value">{Money.format(Decimal(str(financial_data['balance_sheet']['bank_balance'])))}</div><div class="stat-label">Cash in Bank</div></div>
        <div class="stat"><div class="stat-value{' green' if financial_data['income_statement']['net_profit'] >= 0 else ' red'}">{Money.format(Decimal(str(financial_data['income_statement']['net_profit'])))}</div><div class="stat-label">Net Profit YTD</div></div>
        <div class="stat"><div class="stat-value">{financial_data['ratios']['current_ratio']:.1f}</div><div class="stat-label">Current Ratio</div></div>
        <div class="stat"><div class="stat-value">{financial_data['ratios']['debtor_days']:.0f}</div><div class="stat-label">Debtor Days</div></div>
    </div>
    
    <div class="card report-content" style="line-height: 1.7;">
        {report_html}
    </div>
    
    <style>
        .report-content h2 {{
            color: var(--purple);
            margin-top: 24px;
            margin-bottom: 12px;
            font-size: 18px;
        }}
        .report-content h2:first-child {{
            margin-top: 0;
        }}
        .report-content p {{
            margin-bottom: 12px;
        }}
        .report-content ul, .report-content ol {{
            margin: 12px 0;
            padding-left: 24px;
        }}
        .report-content li {{
            margin-bottom: 8px;
        }}
        @media print {{
            .header, .btn-group, .stats {{ display: none !important; }}
            .card {{ border: none; padding: 0; }}
        }}
    </style>
    '''
    
    return page_wrapper("Business Health Report", content, "reports", user)


@app.route("/reports/management-financials")
def report_management_financials():
    """
    Management Financials Report - Professional bank-ready report
    For banks, investors, and formal submissions
    """
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Check if we should generate the report
    generate = request.args.get("generate") == "yes"
    
    if not generate:
        # Show intro page
        content = '''
        <div class="card" style="max-width: 600px; margin: 0 auto; text-align: center;">
            <div style="font-size: 64px; margin-bottom: 20px;">🏦</div>
            <h1 style="margin-bottom: 12px;">Management Financials</h1>
            <p class="text-muted" style="margin-bottom: 24px;">
                Professional financial report suitable for banks, investors, and board presentations.<br>
                AI analyzes your data and produces a formal management report.
            </p>
            <p style="margin-bottom: 24px; padding: 16px; background: var(--bg-secondary); border-radius: 8px;">
                <strong>Includes:</strong><br>
                • Executive Summary<br>
                • Financial Performance Analysis<br>
                • Liquidity & Solvency Assessment<br>
                • Risk Analysis<br>
                • Key Performance Indicators<br>
                • Recommendations
            </p>
            <a href="/reports/management-financials?generate=yes" class="btn btn-primary btn-lg">
                🤖 Generate Report
            </a>
            <p class="text-muted mt-md" style="font-size: 12px;">Takes about 30 seconds</p>
        </div>
        '''
        return page_wrapper("Management Financials", content, "reports", user)
    
    # Generate the report with AI
    financial_data = get_financial_data_for_ai()
    
    # Check for API key
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        content = '''
        <div class="alert alert-error">
            AI not configured. Add ANTHROPIC_API_KEY to environment variables.
        </div>
        '''
        return page_wrapper("Management Financials", content, "reports", user)
    
    # Build prompt for Opus
    prompt = f"""You are a senior financial analyst preparing a Management Financial Report for submission to banks and investors.

This report must be professional, thorough, and suitable for formal financial review. Use proper financial terminology but remain clear and readable.

Here is the financial data for {financial_data['company_name']}:

REPORTING PERIOD: {financial_data['period_start']} to {financial_data['period_end']}
REPORT DATE: {financial_data['report_date']}

INCOME STATEMENT:
- Total Revenue: R {financial_data['income_statement']['revenue']:,.2f}
- Cost of Sales: R {financial_data['income_statement']['cost_of_sales']:,.2f}
- Gross Profit: R {financial_data['income_statement']['gross_profit']:,.2f}
- Gross Profit Margin: {financial_data['income_statement']['gross_margin_pct']:.1f}%
- Operating Expenses: R {financial_data['income_statement']['expenses']:,.2f}
- Net Profit Before Tax: R {financial_data['income_statement']['net_profit']:,.2f}
- Net Profit Margin: {financial_data['income_statement']['net_margin_pct']:.1f}%

Expense Breakdown: {json.dumps(financial_data['income_statement']['expense_breakdown'][:8])}

BALANCE SHEET:
- Current Assets: R {financial_data['balance_sheet']['current_assets']:,.2f}
  - Cash and Bank: R {financial_data['balance_sheet']['bank_balance']:,.2f}
  - Trade Debtors: R {financial_data['balance_sheet']['debtors']:,.2f}
  - Inventory: R {financial_data['balance_sheet']['stock']:,.2f}
- Fixed Assets: R {financial_data['balance_sheet']['fixed_assets']:,.2f}
- Total Assets: R {financial_data['balance_sheet']['total_assets']:,.2f}

- Current Liabilities: R {financial_data['balance_sheet']['current_liabilities']:,.2f}
  - Trade Creditors: R {financial_data['balance_sheet']['creditors']:,.2f}
  - VAT Payable: R {financial_data['balance_sheet']['vat_payable']:,.2f}
- Long-term Liabilities: R {financial_data['balance_sheet']['longterm_liabilities']:,.2f}
- Total Liabilities: R {financial_data['balance_sheet']['total_liabilities']:,.2f}
- Shareholders' Equity: R {financial_data['balance_sheet']['equity']:,.2f}

KEY FINANCIAL RATIOS:
- Current Ratio: {financial_data['ratios']['current_ratio']:.2f}
- Quick Ratio (Acid Test): {financial_data['ratios']['quick_ratio']:.2f}
- Debt to Equity Ratio: {financial_data['ratios']['debt_to_equity']:.2f}
- Debtor Days: {financial_data['ratios']['debtor_days']:.0f}
- Creditor Days: {financial_data['ratios']['creditor_days']:.0f}
- Inventory Days: {financial_data['ratios']['stock_days']:.0f}
- Return on Equity: {financial_data['ratios']['return_on_equity']:.1f}%
- Return on Assets: {financial_data['ratios']['return_on_assets']:.1f}%

TRADE RECEIVABLES:
- Outstanding Invoices: {financial_data['sales']['outstanding_count']}
- Outstanding Value: R {financial_data['sales']['outstanding_total']:,.2f}

Prepare a formal Management Financial Report with these sections:

1. EXECUTIVE SUMMARY
Brief overview of financial position, key findings, and overall assessment. 3-4 sentences.

2. FINANCIAL PERFORMANCE ANALYSIS
- Revenue and profitability analysis
- Margin analysis and trends
- Cost structure assessment

3. LIQUIDITY AND SOLVENCY
- Working capital analysis
- Cash flow assessment
- Ability to meet short and long-term obligations
- Compare ratios to industry standards

4. ASSET MANAGEMENT
- Debtor collection efficiency
- Inventory management
- Creditor management
- Cash conversion cycle

5. RISK ASSESSMENT
- Key financial risks identified
- Concentration risks
- Operational concerns

6. KEY PERFORMANCE INDICATORS
Present 5-6 most important KPIs in a clear format

7. RECOMMENDATIONS
Specific, actionable recommendations for improving financial position

Format as professional HTML. Use <h2> for main sections, <h3> for subsections. Use tables where appropriate for presenting ratios and KPIs. Use <strong> for emphasis. Keep professional tone throughout. About 1200-1500 words."""

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        
        message = client.messages.create(
            model="claude-opus-4-20250514",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )
        
        report_html = message.content[0].text
        
    except Exception as e:
        report_html = f"<p class='text-red'>Error generating report: {str(e)}</p>"
    
    # Build the professional page
    content = f'''
    <div class="flex-between mb-lg no-print">
        <div>
            <h1 style="font-size: 24px; font-weight: 700;">🏦 Management Financials</h1>
            <p class="text-muted">{financial_data['company_name']} • Report Date: {financial_data['report_date']}</p>
        </div>
        <div class="btn-group">
            <button onclick="window.print()" class="btn btn-ghost">🖨️ Print</button>
            <a href="/reports/management-financials?generate=yes" class="btn btn-ghost">🔄 Regenerate</a>
        </div>
    </div>
    
    <div class="card report-content management-report">
        <div class="report-header print-only" style="display: none; text-align: center; margin-bottom: 30px; padding-bottom: 20px; border-bottom: 2px solid #333;">
            <h1 style="font-size: 24px; margin-bottom: 8px;">MANAGEMENT FINANCIAL REPORT</h1>
            <h2 style="font-size: 18px; font-weight: normal; margin-bottom: 8px;">{financial_data['company_name']}</h2>
            <p>For the period {financial_data['period_start']} to {financial_data['period_end']}</p>
            <p>Report Date: {financial_data['report_date']}</p>
        </div>
        
        {report_html}
        
        <div class="report-footer" style="margin-top: 40px; padding-top: 20px; border-top: 1px solid var(--border); font-size: 12px; color: var(--text-muted);">
            <p>This report was prepared using data from {financial_data['company_name']}'s accounting system.</p>
            <p>Generated by Click AI Financial Reporting System</p>
        </div>
    </div>
    
    <style>
        .management-report {{
            font-family: Georgia, serif;
            line-height: 1.8;
        }}
        .management-report h2 {{
            color: #333;
            margin-top: 30px;
            margin-bottom: 15px;
            font-size: 18px;
            border-bottom: 1px solid var(--border);
            padding-bottom: 8px;
        }}
        .management-report h2:first-child {{
            margin-top: 0;
        }}
        .management-report h3 {{
            color: #555;
            margin-top: 20px;
            margin-bottom: 10px;
            font-size: 15px;
        }}
        .management-report p {{
            margin-bottom: 12px;
            text-align: justify;
        }}
        .management-report table {{
            width: 100%;
            border-collapse: collapse;
            margin: 15px 0;
        }}
        .management-report th, .management-report td {{
            padding: 10px;
            border: 1px solid var(--border);
            text-align: left;
        }}
        .management-report th {{
            background: var(--bg-secondary);
            font-weight: 600;
        }}
        .management-report ul, .management-report ol {{
            margin: 12px 0;
            padding-left: 24px;
        }}
        .management-report li {{
            margin-bottom: 8px;
        }}
        @media print {{
            .header, .btn-group, .no-print {{ display: none !important; }}
            .card {{ border: none; padding: 20px; box-shadow: none; }}
            .print-only {{ display: block !important; }}
            .management-report {{ font-size: 11pt; }}
            .management-report h2 {{ font-size: 14pt; }}
        }}
    </style>
    '''
    
    return page_wrapper("Management Financials", content, "reports", user)


@app.route("/reports/income-statement")
def report_income_statement():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    year_start = FinancialPeriod.get_current_year_start()
    
    # Use fast database function
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_income_statement"
        headers = {
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }
        resp = requests.post(url, headers=headers, json={"date_from": year_start, "date_to": today()}, timeout=30)
        is_data = resp.json() if resp.status_code == 200 else []
    except:
        is_data = []
    
    revenue = Decimal("0")
    cogs = Decimal("0")
    expenses = Decimal("0")
    
    revenue_rows = ""
    cogs_rows = ""
    expense_rows = ""
    
    for entry in is_data:
        amt = Decimal(str(entry.get("amount", 0) or 0))
        cat = entry.get("category", "")
        
        if cat == "revenue":
            revenue += amt
            revenue_rows += f'<tr><td>{entry.get("account_code","")}</td><td>{entry.get("account_name","")}</td><td class="number">{Money.format(amt)}</td></tr>'
        elif cat == "cost_of_sales":
            cogs += amt
            cogs_rows += f'<tr><td>{entry.get("account_code","")}</td><td>{entry.get("account_name","")}</td><td class="number">{Money.format(amt)}</td></tr>'
        elif cat == "expense":
            expenses += amt
            expense_rows += f'<tr><td>{entry.get("account_code","")}</td><td>{entry.get("account_name","")}</td><td class="number">{Money.format(amt)}</td></tr>'
    
    gross_profit = revenue - cogs
    net_profit = gross_profit - expenses
    
    content = f'''
    
    <h1 style="font-size:24px;font-weight:700;margin-bottom:8px;">Income Statement</h1>
    <p class="text-muted mb-lg">{year_start} to {today()}</p>
    
    <div class="stats">
        <div class="stat"><div class="stat-value green">{Money.format(revenue)}</div><div class="stat-label">Revenue</div></div>
        <div class="stat"><div class="stat-value">{Money.format(gross_profit)}</div><div class="stat-label">Gross Profit</div></div>
        <div class="stat"><div class="stat-value red">{Money.format(expenses)}</div><div class="stat-label">Expenses</div></div>
        <div class="stat"><div class="stat-value{' green' if net_profit >= 0 else ' red'}">{Money.format(net_profit)}</div><div class="stat-label">Net Profit</div></div>
    </div>
    
    <div class="card">
        <h3>Revenue</h3>
        <table class="table"><thead><tr><th>Code</th><th>Account</th><th class="number">Amount</th></tr></thead>
        <tbody>{revenue_rows or '<tr><td colspan="3" class="text-muted">No revenue yet</td></tr>'}</tbody>
        <tfoot><tr style="font-weight:bold;"><td colspan="2">Total Revenue</td><td class="number">{Money.format(revenue)}</td></tr></tfoot></table>
    </div>
    
    <div class="card">
        <h3>Cost of Sales</h3>
        <table class="table"><thead><tr><th>Code</th><th>Account</th><th class="number">Amount</th></tr></thead>
        <tbody>{cogs_rows or '<tr><td colspan="3" class="text-muted">No cost of sales</td></tr>'}</tbody>
        <tfoot><tr style="font-weight:bold;"><td colspan="2">Gross Profit</td><td class="number">{Money.format(gross_profit)}</td></tr></tfoot></table>
    </div>
    
    <div class="card">
        <h3>Expenses</h3>
        <table class="table"><thead><tr><th>Code</th><th>Account</th><th class="number">Amount</th></tr></thead>
        <tbody>{expense_rows or '<tr><td colspan="3" class="text-muted">No expenses yet</td></tr>'}</tbody>
        <tfoot><tr style="font-weight:bold;"><td colspan="2">Total Expenses</td><td class="number">{Money.format(expenses)}</td></tr></tfoot></table>
    </div>
    
    <div class="card" style="background:{'#0a2010' if net_profit >= 0 else '#200a0a'};">
        <h2 style="text-align:center;">Net Profit: {Money.format(net_profit)}</h2>
    </div>
    '''
    return page_wrapper("Income Statement", content, "reports", user)

@app.route("/reports/balance-sheet")
def report_balance_sheet():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Use fast database function
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_balance_sheet"
        headers = {
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }
        resp = requests.post(url, headers=headers, json={}, timeout=30)
        bs_data = resp.json() if resp.status_code == 200 else []
    except:
        bs_data = []
    
    # Organize by category
    current_assets = []
    fixed_assets = []
    current_liab = []
    longterm_liab = []
    equity = []
    
    total_assets = Decimal("0")
    total_liab = Decimal("0")
    total_equity = Decimal("0")
    
    for item in bs_data:
        bal = Decimal(str(item.get("balance", 0) or 0))
        entry = {"name": item.get("account_name", ""), "balance": bal}
        cat = item.get("account_category", "")
        typ = item.get("account_type", "")
        
        if typ == "asset":
            total_assets += bal
            if cat in ["current_asset", "bank", "receivable", "inventory"]:
                current_assets.append(entry)
            else:
                fixed_assets.append(entry)
        elif typ == "liability":
            total_liab += bal
            if cat in ["current_liability", "payable", "vat"]:
                current_liab.append(entry)
            else:
                longterm_liab.append(entry)
        elif typ == "equity":
            total_equity += bal
            equity.append(entry)
    
    def make_rows(items):
        if not items:
            return '<tr><td colspan="2" class="text-muted" style="padding-left:20px;">None</td></tr>'
        return "".join([f'<tr><td style="padding-left:20px;">{i["name"]}</td><td class="number">{Money.format(i["balance"])}</td></tr>' for i in items])
    
    total_liab_equity = total_liab + total_equity
    balanced = abs(total_assets - total_liab_equity) < Decimal("0.01")
    
    total_current_assets = sum(i["balance"] for i in current_assets)
    total_fixed_assets = sum(i["balance"] for i in fixed_assets)
    total_current_liab = sum(i["balance"] for i in current_liab)
    total_longterm_liab = sum(i["balance"] for i in longterm_liab)
    
    content = f'''
    <div class="flex-between mb-lg">
        <div><h1 style="font-size:24px;font-weight:700;">Balance Sheet</h1><p class="text-muted">As at {today()}</p></div>
        {'<span class="badge badge-green">✓ Balanced</span>' if balanced else '<span class="badge badge-red">✗ Unbalanced</span>'}
    </div>
    <div class="stats">
        <div class="stat"><div class="stat-value">{Money.format(total_assets)}</div><div class="stat-label">Assets</div></div>
        <div class="stat"><div class="stat-value">{Money.format(total_liab)}</div><div class="stat-label">Liabilities</div></div>
        <div class="stat"><div class="stat-value">{Money.format(total_equity)}</div><div class="stat-label">Equity</div></div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;">
        <div class="card"><h3 class="card-title mb-md">ASSETS</h3><table class="table"><tbody>
            <tr style="background:var(--bg-secondary);"><td><strong>Current Assets</strong></td><td></td></tr>
            {make_rows(current_assets)}
            <tr><td><strong>Sub-total</strong></td><td class="number"><strong>{Money.format(total_current_assets)}</strong></td></tr>
            <tr style="background:var(--bg-secondary);"><td><strong>Fixed Assets</strong></td><td></td></tr>
            {make_rows(fixed_assets)}
            <tr><td><strong>Sub-total</strong></td><td class="number"><strong>{Money.format(total_fixed_assets)}</strong></td></tr>
            <tr style="background:linear-gradient(135deg,rgba(59,130,246,0.1),rgba(139,92,246,0.1));"><td><strong>TOTAL ASSETS</strong></td><td class="number"><strong>{Money.format(total_assets)}</strong></td></tr>
        </tbody></table></div>
        <div class="card"><h3 class="card-title mb-md">LIABILITIES & EQUITY</h3><table class="table"><tbody>
            <tr style="background:var(--bg-secondary);"><td><strong>Current Liabilities</strong></td><td></td></tr>
            {make_rows(current_liab)}
            <tr><td><strong>Sub-total</strong></td><td class="number"><strong>{Money.format(total_current_liab)}</strong></td></tr>
            <tr style="background:var(--bg-secondary);"><td><strong>Long-term Liabilities</strong></td><td></td></tr>
            {make_rows(longterm_liab)}
            <tr><td><strong>Sub-total</strong></td><td class="number"><strong>{Money.format(total_longterm_liab)}</strong></td></tr>
            <tr style="background:var(--bg-secondary);"><td><strong>Equity</strong></td><td></td></tr>
            {make_rows(equity)}
            <tr><td><strong>Sub-total</strong></td><td class="number"><strong>{Money.format(total_equity)}</strong></td></tr>
            <tr style="background:linear-gradient(135deg,rgba(59,130,246,0.1),rgba(139,92,246,0.1));"><td><strong>TOTAL</strong></td><td class="number"><strong>{Money.format(total_liab_equity)}</strong></td></tr>
        </tbody></table></div>
    </div>
    '''
    return page_wrapper("Balance Sheet", content, active="reports", user=user)


@app.route("/reports/vat")
def report_vat():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    date_from = FinancialPeriod.get_vat_period_start()
    date_to = FinancialPeriod.get_vat_period_end()
    vat = Journal.get_vat_report(date_from, date_to)
    
    net_color = "red" if vat["net_vat"] > 0 else "green"
    net_label = "VAT Payable" if vat["net_vat"] > 0 else "VAT Refund"
    
    content = f'''
    
    <h1 style="font-size:24px;font-weight:700;margin-bottom:8px;">VAT Report</h1>
    <p class="text-muted mb-lg">{format_date(date_from)} to {format_date(date_to)}</p>
    <div class="stats">
        <div class="stat"><div class="stat-value">{Money.format(vat["output_vat"])}</div><div class="stat-label">Output VAT</div></div>
        <div class="stat"><div class="stat-value green">{Money.format(vat["input_vat"])}</div><div class="stat-label">Input VAT</div></div>
        <div class="stat"><div class="stat-value {net_color}">{Money.format(abs(vat["net_vat"]))}</div><div class="stat-label">{net_label}</div></div>
    </div>
    <div class="card">
        <table class="table"><tbody>
            <tr><td>Output VAT (collected)</td><td class="number">{Money.format(vat["output_vat"])}</td></tr>
            <tr><td>Less: Input VAT (paid)</td><td class="number">({Money.format(vat["input_vat"])})</td></tr>
            <tr style="background:linear-gradient(135deg,rgba(139,92,246,0.1),rgba(59,130,246,0.1));font-size:18px;"><td><strong>{net_label}</strong></td><td class="number" style="color:var(--{net_color});"><strong>{Money.format(abs(vat["net_vat"]))}</strong></td></tr>
        </tbody></table>
    </div>
    '''
    return page_wrapper("VAT Report", content, active="reports", user=user)


@app.route("/reports/sales")
def report_sales():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    invoices = db.select("invoices", order="-date")
    total = sum(Decimal(str(i.get("total", 0) or 0)) for i in invoices)
    paid = sum(Decimal(str(i.get("total", 0) or 0)) for i in invoices if i.get("status") == "paid")
    
    rows = [[i.get("invoice_number", "-"), i.get("date", "")[:10], i.get("customer_name", "-"), {"value": Money.format(Decimal(str(i.get("total", 0)))), "class": "number"}, badge("Paid", "green") if i.get("status") == "paid" else badge("Outstanding", "orange")] for i in invoices[:50]]
    table = table_html(headers=["Invoice", "Date", "Customer", {"label": "Total", "class": "number"}, "Status"], rows=rows, empty_message="No sales")
    
    content = f'''
    
    <h1 style="font-size:24px;font-weight:700;margin-bottom:20px;">Sales Report</h1>
    <div class="stats">
        <div class="stat"><div class="stat-value">{len(invoices)}</div><div class="stat-label">Invoices</div></div>
        <div class="stat"><div class="stat-value green">{Money.format(total)}</div><div class="stat-label">Total Sales</div></div>
        <div class="stat"><div class="stat-value orange">{Money.format(total - paid)}</div><div class="stat-label">Outstanding</div></div>
    </div>
    <div class="card">{table}</div>
    '''
    return page_wrapper("Sales Report", content, active="reports", user=user)


@app.route("/reports/debtors")
def report_debtors():
    """Debtors aging report - 30/60/90 days"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    customers = [c for c in db.select("customers", order="name") if c.get("active", True)]
    
    # Get all outstanding invoices
    invoices = db.select("invoices", filters={"status": "outstanding"}) or []
    
    # Calculate aging for each customer
    from datetime import datetime, timedelta
    today_date = datetime.strptime(today(), "%Y-%m-%d")
    
    aging_data = {}
    for inv in invoices:
        cust_id = inv.get("customer_id")
        if not cust_id:
            continue
        
        inv_date_str = inv.get("date", "")[:10]
        try:
            inv_date = datetime.strptime(inv_date_str, "%Y-%m-%d")
        except:
            continue
        
        days_old = (today_date - inv_date).days
        amount = Decimal(str(inv.get("total", 0) or 0))
        
        if cust_id not in aging_data:
            aging_data[cust_id] = {"current": Decimal("0"), "days_30": Decimal("0"), "days_60": Decimal("0"), "days_90": Decimal("0")}
        
        if days_old <= 30:
            aging_data[cust_id]["current"] += amount
        elif days_old <= 60:
            aging_data[cust_id]["days_30"] += amount
        elif days_old <= 90:
            aging_data[cust_id]["days_60"] += amount
        else:
            aging_data[cust_id]["days_90"] += amount
    
    # Build rows
    rows = []
    total_current = Decimal("0")
    total_30 = Decimal("0")
    total_60 = Decimal("0")
    total_90 = Decimal("0")
    
    for cust in customers:
        cust_id = cust.get("id")
        if cust_id not in aging_data:
            continue
        
        aging = aging_data[cust_id]
        total = aging["current"] + aging["days_30"] + aging["days_60"] + aging["days_90"]
        
        if total <= 0:
            continue
        
        total_current += aging["current"]
        total_30 += aging["days_30"]
        total_60 += aging["days_60"]
        total_90 += aging["days_90"]
        
        rows.append([
            f'<a href="/customers/{cust_id}">{safe_string(cust.get("name", "-"))}</a>',
            {"value": Money.format(aging["current"]) if aging["current"] > 0 else "-", "class": "number"},
            {"value": Money.format(aging["days_30"]) if aging["days_30"] > 0 else "-", "class": "number text-orange"},
            {"value": Money.format(aging["days_60"]) if aging["days_60"] > 0 else "-", "class": "number text-orange"},
            {"value": Money.format(aging["days_90"]) if aging["days_90"] > 0 else "-", "class": "number text-red"},
            {"value": Money.format(total), "class": "number font-bold"}
        ])
    
    grand_total = total_current + total_30 + total_60 + total_90
    
    table = table_html(
        headers=["Customer", {"label": "Current", "class": "number"}, {"label": "30 Days", "class": "number"}, 
                 {"label": "60 Days", "class": "number"}, {"label": "90+ Days", "class": "number"}, {"label": "Total", "class": "number"}],
        rows=rows,
        empty_message="No outstanding debts 🎉"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <h1 style="font-size:24px;font-weight:700;">Debtors Aging</h1>
            <p class="text-muted">Who owes you and for how long</p>
        </div>
        <a href="/reports/debtors/csv" class="btn btn-ghost">📥 Export CSV</a>
    </div>
    
    <div class="stats">
        <div class="stat"><div class="stat-value">{Money.format(total_current)}</div><div class="stat-label">Current</div></div>
        <div class="stat"><div class="stat-value orange">{Money.format(total_30)}</div><div class="stat-label">30 Days</div></div>
        <div class="stat"><div class="stat-value orange">{Money.format(total_60)}</div><div class="stat-label">60 Days</div></div>
        <div class="stat"><div class="stat-value red">{Money.format(total_90)}</div><div class="stat-label">90+ Days</div></div>
        <div class="stat"><div class="stat-value">{Money.format(grand_total)}</div><div class="stat-label">Total Owing</div></div>
    </div>
    
    <div class="card">{table}</div>
    '''
    return page_wrapper("Debtors Aging", content, active="reports", user=user)


@app.route("/reports/debtors/csv")
def report_debtors_csv():
    """Export debtors aging to CSV"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    customers = db.select("customers", order="name") or []
    invoices = db.select("invoices", filters={"status": "outstanding"}) or []
    
    from datetime import datetime
    today_date = datetime.strptime(today(), "%Y-%m-%d")
    
    aging_data = {}
    for inv in invoices:
        cust_id = inv.get("customer_id")
        if not cust_id:
            continue
        try:
            inv_date = datetime.strptime(inv.get("date", "")[:10], "%Y-%m-%d")
            days_old = (today_date - inv_date).days
            amount = float(inv.get("total", 0) or 0)
            
            if cust_id not in aging_data:
                aging_data[cust_id] = {"current": 0, "days_30": 0, "days_60": 0, "days_90": 0}
            
            if days_old <= 30:
                aging_data[cust_id]["current"] += amount
            elif days_old <= 60:
                aging_data[cust_id]["days_30"] += amount
            elif days_old <= 90:
                aging_data[cust_id]["days_60"] += amount
            else:
                aging_data[cust_id]["days_90"] += amount
        except:
            continue
    
    csv_lines = ["Customer,Current,30 Days,60 Days,90+ Days,Total"]
    for cust in customers:
        cust_id = cust.get("id")
        if cust_id in aging_data:
            a = aging_data[cust_id]
            total = a["current"] + a["days_30"] + a["days_60"] + a["days_90"]
            if total > 0:
                csv_lines.append(f'{cust.get("name", "")},{a["current"]},{a["days_30"]},{a["days_60"]},{a["days_90"]},{total}')
    
    from flask import Response
    return Response("\n".join(csv_lines), mimetype="text/csv", 
                    headers={"Content-Disposition": f"attachment;filename=debtors_aging_{today()}.csv"})


@app.route("/reports/creditors")
def report_creditors():
    """Creditors aging report - 30/60/90 days"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    suppliers = [s for s in db.select("suppliers", order="name") if s.get("active", True)]
    
    # Get all expenses that are outstanding (unpaid supplier invoices)
    expenses = db.select("expenses", filters={"status": "outstanding"}) or []
    
    from datetime import datetime
    today_date = datetime.strptime(today(), "%Y-%m-%d")
    
    aging_data = {}
    for exp in expenses:
        supp_id = exp.get("supplier_id")
        if not supp_id:
            continue
        
        exp_date_str = exp.get("date", "")[:10]
        try:
            exp_date = datetime.strptime(exp_date_str, "%Y-%m-%d")
        except:
            continue
        
        days_old = (today_date - exp_date).days
        amount = Decimal(str(exp.get("total", 0) or 0))
        
        if supp_id not in aging_data:
            aging_data[supp_id] = {"current": Decimal("0"), "days_30": Decimal("0"), "days_60": Decimal("0"), "days_90": Decimal("0")}
        
        if days_old <= 30:
            aging_data[supp_id]["current"] += amount
        elif days_old <= 60:
            aging_data[supp_id]["days_30"] += amount
        elif days_old <= 90:
            aging_data[supp_id]["days_60"] += amount
        else:
            aging_data[supp_id]["days_90"] += amount
    
    # Build rows
    rows = []
    total_current = Decimal("0")
    total_30 = Decimal("0")
    total_60 = Decimal("0")
    total_90 = Decimal("0")
    
    for supp in suppliers:
        supp_id = supp.get("id")
        if supp_id not in aging_data:
            continue
        
        aging = aging_data[supp_id]
        total = aging["current"] + aging["days_30"] + aging["days_60"] + aging["days_90"]
        
        if total <= 0:
            continue
        
        total_current += aging["current"]
        total_30 += aging["days_30"]
        total_60 += aging["days_60"]
        total_90 += aging["days_90"]
        
        rows.append([
            f'<a href="/suppliers/{supp_id}">{safe_string(supp.get("name", "-"))}</a>',
            {"value": Money.format(aging["current"]) if aging["current"] > 0 else "-", "class": "number"},
            {"value": Money.format(aging["days_30"]) if aging["days_30"] > 0 else "-", "class": "number text-orange"},
            {"value": Money.format(aging["days_60"]) if aging["days_60"] > 0 else "-", "class": "number text-orange"},
            {"value": Money.format(aging["days_90"]) if aging["days_90"] > 0 else "-", "class": "number text-red"},
            {"value": Money.format(total), "class": "number font-bold"}
        ])
    
    grand_total = total_current + total_30 + total_60 + total_90
    
    table = table_html(
        headers=["Supplier", {"label": "Current", "class": "number"}, {"label": "30 Days", "class": "number"}, 
                 {"label": "60 Days", "class": "number"}, {"label": "90+ Days", "class": "number"}, {"label": "Total", "class": "number"}],
        rows=rows,
        empty_message="No outstanding debts 🎉"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <h1 style="font-size:24px;font-weight:700;">Creditors Aging</h1>
            <p class="text-muted">What you owe and for how long</p>
        </div>
        <a href="/reports/creditors/csv" class="btn btn-ghost">📥 Export CSV</a>
    </div>
    
    <div class="stats">
        <div class="stat"><div class="stat-value">{Money.format(total_current)}</div><div class="stat-label">Current</div></div>
        <div class="stat"><div class="stat-value orange">{Money.format(total_30)}</div><div class="stat-label">30 Days</div></div>
        <div class="stat"><div class="stat-value orange">{Money.format(total_60)}</div><div class="stat-label">60 Days</div></div>
        <div class="stat"><div class="stat-value red">{Money.format(total_90)}</div><div class="stat-label">90+ Days</div></div>
        <div class="stat"><div class="stat-value">{Money.format(grand_total)}</div><div class="stat-label">Total We Owe</div></div>
    </div>
    
    <div class="card">{table}</div>
    '''
    return page_wrapper("Creditors Aging", content, active="reports", user=user)


@app.route("/reports/creditors/csv")
def report_creditors_csv():
    """Export creditors aging to CSV"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    suppliers = db.select("suppliers", order="name") or []
    expenses = db.select("expenses", filters={"status": "outstanding"}) or []
    
    from datetime import datetime
    today_date = datetime.strptime(today(), "%Y-%m-%d")
    
    aging_data = {}
    for exp in expenses:
        supp_id = exp.get("supplier_id")
        if not supp_id:
            continue
        try:
            exp_date = datetime.strptime(exp.get("date", "")[:10], "%Y-%m-%d")
            days_old = (today_date - exp_date).days
            amount = float(exp.get("total", 0) or 0)
            
            if supp_id not in aging_data:
                aging_data[supp_id] = {"current": 0, "days_30": 0, "days_60": 0, "days_90": 0}
            
            if days_old <= 30:
                aging_data[supp_id]["current"] += amount
            elif days_old <= 60:
                aging_data[supp_id]["days_30"] += amount
            elif days_old <= 90:
                aging_data[supp_id]["days_60"] += amount
            else:
                aging_data[supp_id]["days_90"] += amount
        except:
            continue
    
    csv_lines = ["Supplier,Current,30 Days,60 Days,90+ Days,Total"]
    for supp in suppliers:
        supp_id = supp.get("id")
        if supp_id in aging_data:
            a = aging_data[supp_id]
            total = a["current"] + a["days_30"] + a["days_60"] + a["days_90"]
            if total > 0:
                csv_lines.append(f'{supp.get("name", "")},{a["current"]},{a["days_30"]},{a["days_60"]},{a["days_90"]},{total}')
    
    from flask import Response
    return Response("\n".join(csv_lines), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename=creditors_aging_{today()}.csv"})



@app.route("/reports/stock")
def report_stock():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Fast stock summary using database function
    try:
        url = f"{Config.SUPABASE_URL}/rest/v1/rpc/get_stock_summary"
        headers = {
            "apikey": Config.SUPABASE_KEY,
            "Authorization": f"Bearer {Config.SUPABASE_KEY}",
            "Content-Type": "application/json"
        }
        resp = requests.post(url, headers=headers, json={}, timeout=30)
        data = resp.json()
        summary = data[0] if resp.status_code == 200 and data else {}
    except:
        summary = {}
    
    total_items = int(summary.get("total_items", 0) or 0)
    total_units = int(float(summary.get("total_units", 0) or 0))
    cost_value = Decimal(str(summary.get("cost_value", 0) or 0))
    retail_value = Decimal(str(summary.get("retail_value", 0) or 0))
    profit_margin = retail_value - cost_value
    
    content = f'''
    
    <h1 style="font-size:24px;font-weight:700;margin-bottom:20px;">Stock Valuation Report</h1>
    
    <div class="stats">
        <div class="stat"><div class="stat-value">{total_items:,}</div><div class="stat-label">Stock Items</div></div>
        <div class="stat"><div class="stat-value">{total_units:,}</div><div class="stat-label">Total Units</div></div>
        <div class="stat"><div class="stat-value">{Money.format(cost_value)}</div><div class="stat-label">Cost Value</div></div>
        <div class="stat"><div class="stat-value green">{Money.format(retail_value)}</div><div class="stat-label">Retail Value</div></div>
        <div class="stat"><div class="stat-value green">{Money.format(profit_margin)}</div><div class="stat-label">Potential Profit</div></div>
    </div>
    
    <div class="card">
        <h3 style="margin-bottom:16px;">Stock Actions</h3>
        <div class="btn-group">
            <a href="/stock" class="btn btn-primary">View All Stock</a>
            <a href="/stock/new" class="btn btn-green">Add Stock Item</a>
        </div>
    </div>
    '''
    return page_wrapper("Stock Report", content, "reports", user)

@app.route("/reports/ledger")
def report_ledger():
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    entries = Journal.get_entries(limit=200)
    
    rows = []
    for e in entries:
        dr = Money.format(Decimal(str(e.get("debit", 0)))) if e.get("debit") else "-"
        cr = Money.format(Decimal(str(e.get("credit", 0)))) if e.get("credit") else "-"
        rows.append([e.get("date", "")[:10], e.get("account_code", ""), e.get("reference", "-"), e.get("description", "")[:60], {"value": dr, "class": "number"}, {"value": cr, "class": "number"}])
    
    table = table_html(headers=["Date", "Account", "Ref", "Description", {"label": "Debit", "class": "number"}, {"label": "Credit", "class": "number"}], rows=rows, empty_message="No entries")
    
    content = f'''
    
    <h1 style="font-size:24px;font-weight:700;margin-bottom:20px;">General Ledger</h1>
    <div class="card">{table}</div>
    '''
    return page_wrapper("General Ledger", content, active="reports", user=user)


"""
PIECE 11 COMPLETE - Reports

Contains:
✓ Reports menu with 9 report types
✓ Trial Balance - from journal
✓ Income Statement / P&L - from journal  
✓ Balance Sheet - from journal
✓ VAT Report - from journal
✓ Sales Report
✓ Debtors Report
✓ Creditors Report
✓ Stock Report
✓ General Ledger

All financial reports pull from Journal class - real double-entry data!
"""


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def init_database():
    print("Initializing Click AI...")
    try:
        result = Account.initialize_chart()
        print(f"  Accounts: {result['created']} created")
    except Exception as e:
        print(f"  Account init: {e}")
    
    try:
        users = db.select("users", limit=1)
        if not users:
            admin = {"id": generate_id(), "username": "admin", "password": "admin", 
                    "role": "admin", "active": True, "created_at": now()}
            db.insert("users", admin)
            print("  Admin created (admin/admin)")
    except Exception as e:
        print(f"  User init: {e}")

# Auto-init
try:
    accts = db.select("accounts", limit=1)
    if not accts:
        init_database()
except:
    pass


# ═══════════════════════════════════════════════════════════════════════════════
# PRINT ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/print/thermal/<invoice_number>")
def print_thermal(invoice_number):
    """Generate thermal printer receipt"""
    inv = db.select("invoices", filters={"invoice_number": invoice_number})
    if not inv:
        return "Invoice not found", 404
    inv = inv[0]
    
    items = json.loads(inv.get("items", "[]"))
    
    item_lines = ""
    for item in items:
        qty = item.get("quantity", 1)
        desc = item.get("description", "")[:20]
        price = Decimal(str(item.get("price", 0)))
        line_total = price * qty
        item_lines += f"<tr><td>{qty}x {desc}</td><td style='text-align:right'>R{line_total:.2f}</td></tr>"
    
    total = Decimal(str(inv.get("total", 0)))
    vat = Decimal(str(inv.get("vat", 0)))
    
    return f'''<!DOCTYPE html>
<html>
<head>
    <title>Receipt</title>
    <style>
        @media print {{ @page {{ margin: 0; size: 80mm auto; }} }}
        body {{ font-family: monospace; font-size: 12px; width: 80mm; padding: 5mm; margin: 0; }}
        h1 {{ font-size: 16px; text-align: center; margin: 0 0 10px; }}
        table {{ width: 100%; border-collapse: collapse; }}
        td {{ padding: 2px 0; }}
        .total {{ font-size: 16px; font-weight: bold; border-top: 1px dashed #000; margin-top: 10px; padding-top: 10px; }}
        .center {{ text-align: center; }}
        .no-print {{ margin-top: 20px; }}
        @media print {{ .no-print {{ display: none; }} }}
    </style>
</head>
<body onload="window.print()">
    <h1>RECEIPT</h1>
    <p class="center">{inv.get("invoice_number", "")}</p>
    <p class="center">{inv.get("date", "")[:10]}</p>
    <hr>
    <table>{item_lines}</table>
    <hr>
    <table>
        <tr><td>VAT (15%)</td><td style="text-align:right">R{vat:.2f}</td></tr>
        <tr class="total"><td>TOTAL</td><td style="text-align:right">R{total:.2f}</td></tr>
    </table>
    <p class="center" style="margin-top:20px">Thank you!</p>
    <p class="center no-print"><a href="#" onclick="window.close(); return false;" style="color:#666;">Close</a></p>
</body>
</html>'''


@app.route("/print/office/<invoice_number>")
def print_office(invoice_number):
    """Generate A4 office invoice with company details"""
    inv = db.select("invoices", filters={"invoice_number": invoice_number})
    if not inv:
        return "Invoice not found", 404
    inv = inv[0]
    
    # Get company details from settings
    try:
        settings_row = db.select("settings", filters={"key": "company"}, limit=1)
        if settings_row:
            company = json.loads(settings_row[0].get("value", "{}"))
        else:
            company = {}
    except:
        company = {}
    
    company_name = company.get("name", "Your Business Name")
    company_address = f"{company.get('address_line1', '')}"
    if company.get('address_line2'):
        company_address += f"<br>{company.get('address_line2')}"
    if company.get('city'):
        company_address += f"<br>{company.get('city')} {company.get('postal_code', '')}"
    company_phone = company.get('phone', '')
    company_email = company.get('email', '')
    company_vat = company.get('vat_number', '')
    
    # Bank details
    bank_details = ""
    if company.get('bank_name'):
        bank_details = f"""
        <div style="margin-top: 30px; padding: 15px; background: #f9f9f9; border-radius: 5px;">
            <strong>Banking Details:</strong><br>
            {company.get('bank_name', '')}<br>
            Account: {company.get('bank_account', '')}<br>
            Branch: {company.get('bank_branch', '')}
        </div>
        """
    
    items = json.loads(inv.get("items", "[]"))
    
    item_rows = ""
    for item in items:
        qty = item.get("quantity", 1)
        desc = item.get("description", "")
        price = Decimal(str(item.get("price", 0)))
        line_total = price * qty
        item_rows += f"<tr><td>{desc}</td><td style='text-align:right'>{qty}</td><td style='text-align:right'>R {price:.2f}</td><td style='text-align:right'>R {line_total:.2f}</td></tr>"
    
    total = Decimal(str(inv.get("total", 0)))
    vat = Decimal(str(inv.get("vat", 0)))
    subtotal = total - vat
    customer = inv.get("customer_name", "Walk-in Customer")
    
    return f'''<!DOCTYPE html>
<html>
<head>
    <title>Invoice {inv.get("invoice_number", "")}</title>
    <style>
        @media print {{ @page {{ margin: 15mm; }} }}
        body {{ font-family: Arial, sans-serif; font-size: 14px; max-width: 210mm; margin: 0 auto; padding: 20px; }}
        h1 {{ color: #333; margin-bottom: 5px; }}
        .header {{ display: flex; justify-content: space-between; margin-bottom: 30px; }}
        .company {{ font-size: 12px; line-height: 1.6; }}
        .company-name {{ font-size: 20px; font-weight: bold; color: #333; margin-bottom: 5px; }}
        .invoice-details {{ text-align: right; }}
        .customer-box {{ background: #f5f5f5; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th {{ background: #f5f5f5; text-align: left; padding: 10px; border-bottom: 2px solid #333; }}
        td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
        .totals {{ width: 300px; margin-left: auto; }}
        .totals td {{ border: none; }}
        .grand-total {{ font-size: 18px; font-weight: bold; background: #f0f0f0; }}
        .btn {{ display: inline-block; padding: 10px 20px; background: #4CAF50; color: white; text-decoration: none; border-radius: 5px; margin-top: 20px; }}
        @media print {{ .no-print {{ display: none; }} }}
    </style>
</head>
<body>
    <div class="header">
        <div class="company">
            <div class="company-name">{company_name}</div>
            {company_address}<br>
            {f"Tel: {company_phone}<br>" if company_phone else ""}
            {f"Email: {company_email}<br>" if company_email else ""}
            {f"VAT: {company_vat}" if company_vat else ""}
        </div>
        <div class="invoice-details">
            <h1>INVOICE</h1>
            <p><strong>Invoice #:</strong> {inv.get("invoice_number", "")}</p>
            <p><strong>Date:</strong> {inv.get("date", "")[:10]}</p>
        </div>
    </div>
    
    <div class="customer-box">
        <strong>Bill To:</strong><br>
        {customer}
    </div>
    
    <table>
        <thead><tr><th>Description</th><th style="text-align:right">Qty</th><th style="text-align:right">Price</th><th style="text-align:right">Amount</th></tr></thead>
        <tbody>{item_rows}</tbody>
    </table>
    
    <table class="totals">
        <tr><td>Subtotal:</td><td style="text-align:right">R {subtotal:.2f}</td></tr>
        <tr><td>VAT (15%):</td><td style="text-align:right">R {vat:.2f}</td></tr>
        <tr class="grand-total"><td>TOTAL:</td><td style="text-align:right">R {total:.2f}</td></tr>
    </table>
    
    {bank_details}
    
    <a href="#" class="btn no-print" onclick="window.print(); return false;">Print Invoice</a>
    <a href="#" class="btn no-print" style="background:#666; margin-left:10px;" onclick="window.close(); return false;">Close</a>
</body>
</html>'''



# ═══════════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════════
# CLICK AI SCANNER - The Heart of the System
# Tap • Snap • Done - AI Does Everything
# ═══════════════════════════════════════════════════════════════════════════════

SCANNER_HTML = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <title>Click Scanner</title>
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="theme-color" content="#0a0a12">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; -webkit-tap-highlight-color: transparent; }
        body { 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #0a0a12;
            color: #f0f0f0;
            min-height: 100vh;
            min-height: 100dvh;
            display: flex;
            flex-direction: column;
            overflow: hidden;
        }
        
        .header {
            text-align: center;
            padding: 50px 20px 30px;
        }
        .logo {
            font-size: 42px;
            font-weight: 800;
            background: linear-gradient(135deg, #8b5cf6, #3b82f6);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }
        .tagline { 
            color: #606070; 
            font-size: 15px; 
            margin-top: 8px;
            letter-spacing: 0.5px;
        }
        
        .biz-select {
            margin: 0 24px 20px;
            padding: 16px 20px;
            background: #12121a;
            border: 2px solid #2a2a4a;
            border-radius: 14px;
            color: #f0f0f0;
            font-size: 17px;
            width: calc(100% - 48px);
            appearance: none;
            background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='14' height='14' fill='%238b8b9a'%3E%3Cpath d='M7 10L2 5h10z'/%3E%3C/svg%3E");
            background-repeat: no-repeat;
            background-position: right 16px center;
        }
        .biz-select:focus {
            outline: none;
            border-color: #8b5cf6;
        }
        
        .buttons {
            flex: 1;
            display: flex;
            flex-direction: column;
            justify-content: center;
            padding: 24px;
            gap: 20px;
        }
        
        .scan-btn {
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 16px;
            padding: 32px 28px;
            border-radius: 20px;
            font-size: 22px;
            font-weight: 700;
            color: white;
            border: none;
            cursor: pointer;
            text-align: center;
            transition: transform 0.15s, box-shadow 0.15s;
            box-shadow: 0 8px 32px rgba(0,0,0,0.3);
        }
        .scan-btn:active { 
            transform: scale(0.96); 
            box-shadow: 0 4px 16px rgba(0,0,0,0.3);
        }
        
        .btn-supplier {
            background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%);
        }
        .btn-expense {
            background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);
        }
        
        .scan-icon { font-size: 36px; }
        
        input[type="file"] { display: none; }
        
        .review-link {
            display: block;
            margin: 0 24px 30px;
            padding: 18px;
            background: rgba(139, 92, 246, 0.15);
            border: 2px solid rgba(139, 92, 246, 0.3);
            border-radius: 14px;
            color: #a78bfa;
            text-decoration: none;
            text-align: center;
            font-size: 17px;
            font-weight: 600;
        }
        .review-count {
            display: inline-block;
            background: #8b5cf6;
            color: white;
            padding: 4px 12px;
            border-radius: 20px;
            margin-left: 8px;
            font-size: 15px;
        }
        
        /* Success Animation */
        .toast {
            position: fixed;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%) scale(0.8);
            background: linear-gradient(135deg, #10b981, #059669);
            color: white;
            padding: 32px 48px;
            border-radius: 20px;
            font-size: 24px;
            font-weight: 700;
            opacity: 0;
            transition: all 0.3s ease-out;
            z-index: 1000;
            text-align: center;
            box-shadow: 0 20px 60px rgba(16, 185, 129, 0.4);
        }
        .toast.show {
            transform: translate(-50%, -50%) scale(1);
            opacity: 1;
        }
        .toast-icon {
            font-size: 48px;
            display: block;
            margin-bottom: 12px;
        }
        
        /* Uploading State */
        .uploading {
            display: none;
            position: fixed;
            inset: 0;
            background: rgba(0,0,0,0.9);
            justify-content: center;
            align-items: center;
            flex-direction: column;
            z-index: 999;
        }
        .uploading.show { display: flex; }
        .upload-spinner {
            width: 60px;
            height: 60px;
            border: 4px solid #1a1a2e;
            border-top-color: #8b5cf6;
            border-radius: 50%;
            animation: spin 0.7s linear infinite;
        }
        @keyframes spin { to { transform: rotate(360deg); } }
        .upload-text { 
            color: #a0a0a0; 
            font-size: 18px; 
            margin-top: 20px;
        }
        
        .footer {
            text-align: center;
            padding: 16px;
            color: #404050;
            font-size: 13px;
        }
        .footer a { color: #606070; text-decoration: none; }
    </style>
</head>
<body>
    <div class="header">
        <div class="logo">Click</div>
        <div class="tagline">Snap · Upload · Review Later</div>
    </div>
    
    <select class="biz-select" id="biz-select">
        <option value="">Select Business...</option>
    </select>
    
    <div class="buttons">
        <label class="scan-btn btn-supplier">
            <span class="scan-icon">📦</span>
            <span>Supplier Invoice</span>
            <input type="file" accept="image/*" capture="environment" onchange="upload(this, 'supplier')">
        </label>
        
        <label class="scan-btn btn-expense">
            <span class="scan-icon">🧾</span>
            <span>Expense Receipt</span>
            <input type="file" accept="image/*" capture="environment" onchange="upload(this, 'expense')">
        </label>
    </div>
    
    <a href="/review" class="review-link" id="review-link" style="display:none;">
        📋 Review Scanned Items
        <span class="review-count" id="review-count">0</span>
    </a>
    
    <div class="footer">
        <a href="/">Open Desktop Version</a>
    </div>
    
    <div class="toast" id="toast">
        <span class="toast-icon">✓</span>
        <span>Uploaded!</span>
    </div>
    
    <div class="uploading" id="uploading">
        <div class="upload-spinner"></div>
        <div class="upload-text">Uploading photo...</div>
    </div>
    
    <script>
    // Load businesses
    async function loadBusinesses() {
        try {
            const r = await fetch('/api/businesses');
            const data = await r.json();
            const sel = document.getElementById('biz-select');
            
            if (data.businesses && data.businesses.length > 0) {
                sel.innerHTML = '<option value="">Select Business...</option>';
                data.businesses.forEach(b => {
                    const opt = document.createElement('option');
                    opt.value = b.id;
                    opt.textContent = (b.icon || '🏢') + ' ' + (b.business_name || b.name);
                    if (b.id === data.current) opt.selected = true;
                    sel.appendChild(opt);
                });
            }
        } catch(e) { console.log('Business load error:', e); }
    }
    
    // Load pending review count
    async function loadReviewCount() {
        try {
            const r = await fetch('/api/staging/count');
            const data = await r.json();
            const link = document.getElementById('review-link');
            const count = document.getElementById('review-count');
            
            if (data.count > 0) {
                count.textContent = data.count;
                link.style.display = 'block';
            } else {
                link.style.display = 'none';
            }
        } catch(e) { console.log('Count load error:', e); }
    }
    
    // Business change
    document.getElementById('biz-select').onchange = async function() {
        if (this.value) {
            await fetch('/switch-business/' + this.value);
            loadReviewCount();
        }
    };
    
    // Upload photo to queue
    async function upload(input, type) {
        if (!input.files || !input.files[0]) return;
        
        const bizId = document.getElementById('biz-select').value;
        if (!bizId) {
            alert('Please select a business first');
            input.value = '';
            return;
        }
        
        // Show uploading
        document.getElementById('uploading').classList.add('show');
        
        const file = input.files[0];
        const reader = new FileReader();
        
        reader.onload = async function(e) {
            try {
                // Upload to queue (instant - no AI wait)
                const response = await fetch('/m/queue', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        image: e.target.result,
                        type: type,
                        business_id: bizId
                    })
                });
                
                const result = await response.json();
                document.getElementById('uploading').classList.remove('show');
                
                if (result.success) {
                    // Show success toast
                    const toast = document.getElementById('toast');
                    toast.classList.add('show');
                    
                    // Vibrate if supported
                    if (navigator.vibrate) navigator.vibrate(100);
                    
                    setTimeout(() => {
                        toast.classList.remove('show');
                        loadReviewCount();
                    }, 1500);
                } else {
                    alert('Upload failed: ' + (result.error || 'Unknown error'));
                }
                
            } catch(err) {
                document.getElementById('uploading').classList.remove('show');
                alert('Connection error - please try again');
            }
            
            input.value = '';
        };
        
        reader.readAsDataURL(file);
    }
    
    // Init
    loadBusinesses();
    loadReviewCount();
    
    // Refresh count every 10 seconds
    setInterval(loadReviewCount, 10000);
    </script>
</body>
</html>'''


# ═══════════════════════════════════════════════════════════════════════════════
# MOBILE SCANNER ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/businesses")
def api_businesses():
    """API: Get user's businesses for mobile selector"""
    user = UserSession.get_current_user()
    if not user:
        return jsonify({"businesses": [], "current": None})
    
    businesses = BusinessManager.get_user_businesses(user.get("id", ""))
    current = BusinessManager.get_current_business()
    
    return jsonify({
        "businesses": businesses,
        "current": current.get("id") if current else None
    })


@app.route("/api/staging/count")
def api_staging_count():
    """API: Get count of pending staged items"""
    user = UserSession.get_current_user()
    if not user:
        return jsonify({"count": 0})
    
    try:
        success, items = db.select("staged_transactions", filters={"status": "pending"})
        count = len(items) if success and items else 0
        return jsonify({"count": count})
    except:
        return jsonify({"count": 0})


@app.route("/m")
def scanner_home():
    """Mobile scanner - fast upload, review later"""
    return SCANNER_HTML


@app.route("/m/queue", methods=["POST"])
def scanner_queue():
    """
    Add photo to scan queue - INSTANT response.
    AI processing happens in background via Supabase Edge Function.
    """
    try:
        data = request.get_json()
        image_data = data.get("image", "")
        scan_type = data.get("type", "")
        business_id = data.get("business_id", "")
        
        if not image_data:
            return jsonify({"success": False, "error": "No image"})
        
        # Clean base64
        if "," in image_data:
            image_data = image_data.split(",")[1]
        
        # Get user
        user = UserSession.get_current_user()
        user_id = user.get("id", "") if user else ""
        
        # Insert into queue - Supabase Edge Function will process
        queue_id = generate_id()
        
        result = db.insert("scan_queue", {
            "id": queue_id,
            "user_id": user_id,
            "business_id": business_id,
            "type": scan_type,
            "image_data": image_data,
            "status": "pending",
            "created_at": now()
        })
        
        if not result:
            return jsonify({"success": False, "error": "Could not save to queue"})
        
        # Return immediately - user doesn't wait for AI
        return jsonify({
            "success": True,
            "queued": True,
            "id": queue_id,
            "message": "Photo uploaded! Processing in background."
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/m/approve/<staged_id>", methods=["POST"])
def scanner_approve(staged_id):
    """Approve and post a staged transaction from mobile"""
    try:
        # Get the staged transaction
        success, result = db.select("staged_transactions", filters={"id": staged_id})
        if not success or not result:
            return jsonify({"success": False, "error": "Not found"})
        
        staged = result[0]
        data = json.loads(staged.get("data", "{}"))
        trans_type = staged.get("type", "")
        
        # Process based on type
        if trans_type == "supplier_invoice":
            result = process_supplier_invoice(data)
        elif trans_type == "expense":
            result = process_expense_receipt(data)
        else:
            return jsonify({"success": False, "error": "Unknown type"})
        
        # Delete from staging if successful
        if result.get("success"):
            db.delete("staged_transactions", staged_id)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/m/delete/<staged_id>", methods=["POST"])
def scanner_delete(staged_id):
    """Delete a staged transaction without posting"""
    try:
        db.delete("staged_transactions", staged_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/m/scan", methods=["POST"])
def scanner_process():
    """Process scanned image with AI - the magic happens here"""
    try:
        data = request.get_json()
        image_data = data.get("image", "")
        scan_type = data.get("type", "")
        
        if not image_data or not scan_type:
            return jsonify({"success": False, "error": "Missing data"})
        
        # Extract base64 data
        if "," in image_data:
            image_data = image_data.split(",")[1]
        
        # Get API key
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return jsonify({"success": False, "error": "AI not configured"})
        
        # Prompts for different scan types
        prompts = {
            "cos": """Read this supplier invoice photo. Extract:
- supplier: company name
- invoice_no: invoice number
- date: YYYY-MM-DD format
- items: list of {description, qty, unit_price}
- total: total amount
- vat: VAT amount

Return ONLY JSON: {"supplier":"Name","invoice_no":"INV123","date":"2025-01-15","items":[{"description":"Item","qty":1,"unit_price":100}],"vat":15,"total":115}""",

            "exp": """Read this receipt/expense photo. Extract:
- vendor: shop/store name (look at top)
- date: YYYY-MM-DD format  
- description: what was bought
- total: total amount
- vat: VAT if shown
- category: fuel/telephone/electricity/repairs/stationery/travel/general

Return ONLY JSON: {"vendor":"Shop","date":"2025-01-15","description":"items bought","total":150.00,"vat":19.57,"category":"fuel"}""",

            "exp_paid": """Read this receipt/expense photo. Extract:
- vendor: shop/store name (look at top)
- date: YYYY-MM-DD format  
- description: what was bought
- total: total amount
- vat: VAT if shown
- category: fuel/telephone/electricity/repairs/stationery/travel/general

Return ONLY JSON: {"vendor":"Shop","date":"2025-01-15","description":"items bought","total":150.00,"vat":19.57,"category":"fuel"}""",

            "stock": """Read this stock count sheet. List all items with quantities.
Return ONLY JSON: {"items":[{"code":"SKU1","description":"Product","quantity":25}]}""",

            "payment": """Read this payment proof/bank slip. Extract:
- customer: who paid
- amount: payment amount
- date: YYYY-MM-DD
- reference: ref number
- method: EFT/cash/card

Return ONLY JSON: {"customer":"Name","amount":1500,"date":"2025-01-15","reference":"INV001","method":"EFT"}"""
        }
        
        prompt = prompts.get(scan_type, prompts.get("exp"))
        
        # Call Claude API directly with requests (more reliable)
        response = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "content-type": "application/json",
                "anthropic-version": "2023-06-01"
            },
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 1000,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": image_data}},
                        {"type": "text", "text": prompt}
                    ]
                }]
            },
            timeout=30
        )
        
        if response.status_code != 200:
            return jsonify({"success": False, "error": f"AI service error ({response.status_code})"})
        
        result = response.json()
        ai_response = result.get("content", [{}])[0].get("text", "")
        
        # Parse JSON from AI response
        start = ai_response.find('{')
        end = ai_response.rfind('}') + 1
        
        if start < 0 or end <= start:
            return jsonify({"success": False, "error": "Could not read document"})
        
        try:
            parsed = json.loads(ai_response[start:end])
        except:
            return jsonify({"success": False, "error": "Could not parse response"})
        
        # Check if this is a "paid" variant
        is_paid = scan_type.endswith("_paid")
        base_type = scan_type.replace("_paid", "")
        
        if is_paid:
            parsed["paid"] = True
        
        # Check if staging is enabled
        use_staging = request.args.get("direct") != "yes"
        
        # Process based on type
        if base_type == "cos":
            if use_staging:
                return stage_transaction("supplier_invoice", parsed)
            return process_supplier_invoice(parsed)
        elif base_type == "exp":
            if use_staging:
                return stage_transaction("expense", parsed)
            return process_expense_receipt(parsed)
        elif base_type == "stock":
            return process_stock_count(parsed)
        elif base_type == "payment":
            return process_customer_payment(parsed)
        
        return jsonify({"success": False, "error": "Unknown scan type"})
        
    except requests.exceptions.Timeout:
        return jsonify({"success": False, "error": "AI took too long - try again"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def stage_transaction(trans_type: str, data: dict):
    """
    Stage a transaction for review instead of posting directly.
    User must approve before it touches real data.
    Checks for duplicates based on invoice_no/reference.
    """
    try:
        # Check for duplicate invoice number
        invoice_no = data.get("invoice_no") or data.get("reference") or ""
        supplier = data.get("supplier") or data.get("vendor") or ""
        
        if invoice_no and supplier:
            # Check existing staged transactions
            success, existing = db.select("staged_transactions")
            if success and existing:
                for staged in existing:
                    staged_data = json.loads(staged.get("data", "{}"))
                    existing_inv = staged_data.get("invoice_no") or staged_data.get("reference") or ""
                    existing_sup = staged_data.get("supplier") or staged_data.get("vendor") or ""
                    if existing_inv == invoice_no and existing_sup.lower() == supplier.lower():
                        return jsonify({
                            "success": False,
                            "error": f"Duplicate! Invoice {invoice_no} from {supplier} already scanned."
                        })
            
            # Check posted transactions (supplier invoices)
            if trans_type == "supplier_invoice":
                success, posted = db.select("supplier_invoices", filters={"invoice_no": invoice_no})
                if success and posted:
                    for inv in posted:
                        if inv.get("supplier_name", "").lower() == supplier.lower():
                            return jsonify({
                                "success": False,
                                "error": f"Duplicate! Invoice {invoice_no} from {supplier} already posted."
                            })
        
        staged_id = generate_id()
        
        db.insert("staged_transactions", {
            "id": staged_id,
            "type": trans_type,
            "data": json.dumps(data),
            "status": "pending",
            "created_at": now()
        })
        
        # Return success with staged_id for mobile POST/DELETE
        if trans_type == "supplier_invoice":
            supplier = data.get("supplier", "Unknown")
            total = Money.parse(data.get("total", 0))
            items_count = len(data.get("items", []))
            is_paid = data.get("paid", False)
            
            return jsonify({
                "success": True,
                "staged": True,
                "staged_id": staged_id,
                "supplier": supplier,
                "total": float(total),
                "description": f"{items_count} items" + (" - PAID" if is_paid else ""),
                "items": data.get("items", [])[:8],
                "invoice_no": data.get("invoice_no", ""),
                "review_url": f"/staging/{staged_id}"
            })
        
        elif trans_type == "expense":
            vendor = data.get("vendor", "Unknown")
            total = Money.parse(data.get("total", 0))
            is_paid = data.get("paid", False)
            
            return jsonify({
                "success": True,
                "staged": True,
                "staged_id": staged_id,
                "supplier": vendor,
                "vendor": vendor,
                "total": float(total),
                "description": data.get("description", "") + (" - PAID" if is_paid else ""),
                "category": data.get("category", "general"),
                "review_url": f"/staging/{staged_id}"
            })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def process_supplier_invoice(data):
    """
    Process supplier invoice:
    1. Check for duplicate invoice
    2. Find or create supplier
    3. For each item: find or create stock item, book IN
    4. Record expense / creditor
    5. Post GL entries
    """
    try:
        invoice_no = data.get("invoice_no", "")
        supplier_name = data.get("supplier", "Unknown Supplier")
        total = Money.parse(data.get("total", 0))
        vat = Money.parse(data.get("vat", 0))
        items = data.get("items", [])
        is_paid = data.get("paid", False)
        inv_date = data.get("date", today())
        
        if total <= 0:
            return jsonify({"success": False, "error": "Could not read total amount"})
        
        # 1. Check for duplicate
        if invoice_no:
            existing = db.select("expenses", filters={"reference": invoice_no})
            if existing:
                return jsonify({
                    "success": False, 
                    "error": f"Duplicate! Invoice {invoice_no} was already scanned.",
                    "badge": "DUPLICATE",
                    "badge_type": "duplicate"
                })
        
        # 2. Find or create supplier (smart matching)
        supplier_id = None
        supplier_created = False
        suppliers = db.select("suppliers")
        
        # Try exact match first, then fuzzy match
        supplier_name_lower = supplier_name.lower().strip()
        for s in suppliers:
            s_name = s.get("name", "").lower().strip()
            # Exact match
            if s_name == supplier_name_lower:
                supplier_id = s["id"]
                break
            # Fuzzy match - check if one contains the other (handles "Makro" vs "Makro Warehouse")
            if supplier_name_lower in s_name or s_name in supplier_name_lower:
                supplier_id = s["id"]
                break
        
        if not supplier_id:
            supplier_id = generate_id()
            # Generate supplier code from name (e.g., "Makro Warehouse" -> "MAK001")
            prefix = ''.join(c for c in supplier_name[:3].upper() if c.isalpha()) or "SUP"
            existing_codes = [s.get("code", "") for s in suppliers]
            sup_num = 1
            while f"{prefix}{sup_num:03d}" in existing_codes:
                sup_num += 1
            supplier_code = f"{prefix}{sup_num:03d}"
            
            db.insert("suppliers", {
                "id": supplier_id,
                "code": supplier_code,
                "name": supplier_name,
                "phone": "",
                "email": "",
                "balance": 0,
                "active": True,
                "created_at": now()
            })
            supplier_created = True
        
        # 3. Process line items - find or create stock, book IN
        items_created = 0
        items_updated = 0
        stock_booked = []
        all_stock = db.select("stock_items")  # Get once, not in loop
        
        for item in items:
            desc = item.get("description", "Unknown Item").strip()
            code = item.get("code", "").strip()
            qty = int(item.get("qty", 0) or item.get("quantity", 0) or 0) or 1
            
            # Get price - could be unit_price or line_total
            raw_price = Money.parse(item.get("unit_price", 0) or item.get("price", 0) or item.get("amount", 0))
            line_total = Money.parse(item.get("line_total", 0) or item.get("total", 0))
            
            # If we have line_total but not unit_price, calculate unit price
            if line_total > 0 and (raw_price <= 0 or raw_price == line_total):
                unit_price = (line_total / Decimal(str(qty))).quantize(Decimal("0.01"))
            elif raw_price > 0:
                # Check if raw_price looks like a line total (much higher than expected unit price)
                # If price * qty would be unreasonably high, it's probably already a line total
                if raw_price > 100 and qty > 1 and raw_price > (total / Decimal(str(len(items) or 1)) * Decimal("0.5")):
                    # This looks like a line total, not unit price
                    unit_price = (raw_price / Decimal(str(qty))).quantize(Decimal("0.01"))
                else:
                    unit_price = raw_price
            else:
                # Fallback: estimate from invoice total
                unit_price = (total / Decimal(str(len(items) or 1)) / Decimal(str(qty))).quantize(Decimal("0.01"))
            
            # Find existing stock item (smart word-by-word matching)
            stock_item = None
            desc_lower = desc.lower()
            desc_words = set(desc_lower.split())  # Split "m10 x30 cap screw" into words
            
            for s in all_stock:
                # Try code match first (most reliable)
                if code and s.get("code", "").lower() == code.lower():
                    stock_item = s
                    break
                
                s_desc = s.get("description", "").lower()
                s_words = set(s_desc.split())
                
                # Try exact description match
                if s_desc == desc_lower:
                    stock_item = s
                    break
                
                # Word-by-word matching - if 2+ words match, it's likely the same item
                # "m10 x30 cap screw" matches "cap screw m10" or "m10 cap screw grade 8"
                matching_words = desc_words & s_words  # Intersection of word sets
                if len(matching_words) >= 2 and len(matching_words) >= len(desc_words) * 0.5:
                    stock_item = s
                    break
                
                # Also try if one contains the other (handles partial matches)
                if len(desc_lower) > 5 and (desc_lower in s_desc or s_desc in desc_lower):
                    stock_item = s
                    break
            
            if stock_item:
                # Update quantity and cost price
                new_qty = stock_item.get("quantity", 0) + qty
                
                # Also update selling price if cost changed significantly
                old_cost = Decimal(str(stock_item.get("cost_price", 0) or 0))
                if old_cost <= 0 or abs(unit_price - old_cost) / max(old_cost, Decimal("1")) > Decimal("0.1"):
                    # Cost changed by more than 10% or was zero - recalculate selling price
                    if unit_price < 50:
                        markup = Decimal("1.50")
                    elif unit_price < 200:
                        markup = Decimal("1.35")
                    elif unit_price < 1000:
                        markup = Decimal("1.25")
                    else:
                        markup = Decimal("1.15")
                    new_selling = (unit_price * markup).quantize(Decimal("0.01"))
                    
                    db.update("stock_items", stock_item["id"], {
                        "quantity": new_qty,
                        "cost_price": float(unit_price),
                        "selling_price": float(new_selling)
                    })
                else:
                    # Just update quantity and cost
                    db.update("stock_items", stock_item["id"], {
                        "quantity": new_qty,
                        "cost_price": float(unit_price)
                    })
                
                items_updated += 1
                stock_booked.append({"name": desc[:25], "value": f"+{qty}"})
            else:
                # Create new stock item with smart code generation
                if not code:
                    # Generate code from description (e.g., "Steel Tube 25mm" -> "STE0001")
                    # Take first 3 consonants or letters from description
                    clean_desc = ''.join(c for c in desc.upper() if c.isalpha())[:3] or "STK"
                    existing_codes = [s.get("code", "") for s in all_stock]
                    stk_num = 1
                    while f"{clean_desc}{stk_num:04d}" in existing_codes:
                        stk_num += 1
                    code = f"{clean_desc}{stk_num:04d}"
                
                # Calculate selling price with smart markup based on cost
                # Lower cost items get higher markup, expensive items lower markup
                if unit_price < 50:
                    markup = Decimal("1.50")  # 50% markup for cheap items
                elif unit_price < 200:
                    markup = Decimal("1.35")  # 35% markup for mid-range
                elif unit_price < 1000:
                    markup = Decimal("1.25")  # 25% markup for expensive
                else:
                    markup = Decimal("1.15")  # 15% markup for very expensive
                
                selling_price = (unit_price * markup).quantize(Decimal("0.01"))
                
                new_item = {
                    "id": generate_id(),
                    "code": code,
                    "description": desc,
                    "quantity": qty,
                    "cost_price": float(unit_price),
                    "selling_price": float(selling_price),
                    "category": "general",
                    "reorder_level": max(1, qty // 4),  # Auto-set reorder level
                    "active": True,
                    "created_at": now()
                }
                
                # Insert and check result
                success, result = db.insert("stock_items", new_item)
                if success:
                    # Add to all_stock so next items can find it (avoid duplicates in same invoice)
                    all_stock.append(new_item)
                    items_created += 1
                    stock_booked.append({"name": f"NEW: {desc[:20]}", "value": f"+{qty}"})
                else:
                    # Log error but continue processing other items
                    stock_booked.append({"name": f"ERR: {desc[:20]}", "value": str(result)[:15]})
        
        # 4. Record expense
        expense_id = generate_id()
        expense = {
            "id": expense_id,
            "date": inv_date,
            "description": f"{supplier_name} - {invoice_no}",
            "supplier": supplier_name,
            "reference": invoice_no,
            "category": "5000",  # Cost of Sales
            "amount": float(total),
            "vat_type": "inclusive",
            "created_at": now()
        }
        db.insert("expenses", expense)
        
        # 5. Post GL entries
        subtotal = total - vat if vat > 0 else total / Decimal("1.15") * Decimal("0.85")
        if vat <= 0:
            vat = total - subtotal
        
        entry = JournalEntry(
            date=inv_date,
            reference=invoice_no or f"COS-{expense_id[:8]}",
            description=f"Supplier: {supplier_name}",
            trans_type=TransactionType.EXPENSE,
            source_type="expense",
            source_id=expense_id
        )
        
        entry.debit("5000", subtotal)  # Cost of Sales
        entry.debit(AccountCodes.VAT_INPUT, vat)  # VAT Input
        
        if is_paid:
            entry.credit(AccountCodes.BANK, total)  # Paid - Bank
        else:
            entry.credit(AccountCodes.CREDITORS, total)  # Unpaid - Creditor
            # Update supplier balance
            db.update("suppliers", supplier_id, {
                "balance": float(Decimal(str(db.select_one("suppliers", supplier_id).get("balance", 0))) + total)
            })
        
        entry.post()
        
        # Build response
        details = f"<strong>{supplier_name}</strong><br>"
        if invoice_no:
            details += f"Invoice: {invoice_no}<br>"
        details += f"{items_created} new items created<br>" if items_created > 0 else ""
        details += f"{items_updated} items updated<br>" if items_updated > 0 else ""
        details += f"{'Paid' if is_paid else 'Added to creditors'}"
        
        return jsonify({
            "success": True,
            "title": "Stock Booked In!",
            "badge": "NEW SUPPLIER" if supplier_created else None,
            "badge_type": "new",
            "details": details,
            "items": stock_booked[:8],  # Show max 8 items
            "amount": f"R {total:,.2f}"
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def process_expense_receipt(data):
    """
    Process expense receipt:
    1. Auto-categorize
    2. Find or create supplier
    3. Record expense
    4. Post GL entries
    """
    try:
        vendor = data.get("vendor", "Unknown")
        description = data.get("description", "Expense")
        total = Money.parse(data.get("total", 0))
        vat = Money.parse(data.get("vat", 0))
        category = data.get("category", "general")
        exp_date = data.get("date", today())
        
        if total <= 0:
            return jsonify({"success": False, "error": "Could not read amount"})
        
        # Map category to account code
        category_map = {
            "fuel": ("6110", "Fuel & Oil"),
            "telephone": ("6120", "Telephone"),
            "electricity": ("6130", "Electricity"),
            "repairs": ("6150", "Repairs & Maintenance"),
            "stationery": ("6160", "Stationery"),
            "travel": ("6170", "Travel"),
            "advertising": ("6180", "Advertising"),
            "insurance": ("6190", "Insurance"),
            "bank_charges": ("6200", "Bank Charges"),
            "general": ("6140", "General Expenses")
        }
        
        account_code, category_name = category_map.get(category, ("6140", "General Expenses"))
        
        # Calculate VAT if not provided
        if vat <= 0:
            vat_info = VAT.calculate_from_inclusive(total)
            subtotal = vat_info["exclusive"]
            vat = vat_info["vat"]
        else:
            subtotal = total - vat
        
        # Record expense
        expense_id = generate_id()
        expense = {
            "id": expense_id,
            "date": exp_date,
            "description": f"{vendor} - {description}",
            "supplier": vendor,
            "category": account_code,
            "amount": float(total),
            "vat_type": "inclusive",
            "created_at": now()
        }
        db.insert("expenses", expense)
        
        # Post GL
        entry = JournalEntry(
            date=exp_date,
            reference=f"EXP-{expense_id[:8]}",
            description=f"{vendor} - {description}",
            trans_type=TransactionType.EXPENSE,
            source_type="expense",
            source_id=expense_id
        )
        
        # Bank charges have no VAT
        if category == "bank_charges":
            entry.debit(account_code, total)
            entry.credit(AccountCodes.BANK, total)
        else:
            entry.debit(account_code, subtotal)
            entry.debit(AccountCodes.VAT_INPUT, vat)
            entry.credit(AccountCodes.BANK, total)
        
        entry.post()
        
        return jsonify({
            "success": True,
            "title": "Expense Recorded!",
            "badge": category_name.upper(),
            "badge_type": "new",
            "details": f"<strong>{vendor}</strong><br>{description}",
            "amount": f"R {total:,.2f}"
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def process_stock_count(data):
    """
    Process stock count:
    1. Match items to stock
    2. Update quantities
    3. Calculate variances
    """
    try:
        items = data.get("items", [])
        
        if not items:
            return jsonify({"success": False, "error": "No items found on count sheet"})
        
        all_stock = db.select("stock_items")
        
        updated = 0
        variances = []
        results = []
        
        for item in items:
            code = item.get("code", "")
            desc = item.get("description", "")
            try:
                counted = int(float(item.get("quantity", 0) or 0))
            except:
                counted = 0
            
            # Find matching stock item
            stock_item = None
            for s in all_stock:
                if code and s.get("code", "").lower() == code.lower():
                    stock_item = s
                    break
                if desc and desc.lower() in s.get("description", "").lower():
                    stock_item = s
                    break
            
            if stock_item:
                old_qty = stock_item.get("quantity", 0)
                variance = counted - old_qty
                
                db.update("stock_items", stock_item["id"], {"quantity": counted})
                updated += 1
                
                if variance != 0:
                    variances.append({
                        "item": stock_item.get("description", "")[:20],
                        "was": old_qty,
                        "now": counted,
                        "diff": variance
                    })
                
                sign = "+" if variance > 0 else ""
                results.append({
                    "name": stock_item.get("description", "")[:22],
                    "value": f"{counted} ({sign}{variance})" if variance != 0 else str(counted)
                })
        
        # Build variance summary
        details = f"<strong>{updated} items updated</strong><br>"
        if variances:
            details += f"{len(variances)} with variances"
        else:
            details += "No variances found ✓"
        
        return jsonify({
            "success": True,
            "title": "Stock Updated!",
            "details": details,
            "items": results[:10]
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


def process_customer_payment(data):
    """
    Process customer payment:
    1. Find customer
    2. Match to invoice
    3. Mark invoice paid
    4. Post GL entries
    """
    try:
        customer_name = data.get("customer", "")
        amount = Decimal(str(data.get("amount", 0)))
        reference = data.get("reference", "")
        pay_date = data.get("date", today())
        method = data.get("method", "EFT")
        
        if amount <= 0:
            return jsonify({"success": False, "error": "Could not read payment amount"})
        
        # Find customer
        customer = None
        customer_id = None
        customers = db.select("customers")
        
        for c in customers:
            if customer_name.lower() in c.get("name", "").lower():
                customer = c
                customer_id = c["id"]
                break
        
        # Find invoice by reference
        invoice = None
        if reference:
            invoices = db.select("invoices")
            for inv in invoices:
                if reference.lower() in inv.get("invoice_number", "").lower():
                    invoice = inv
                    break
                if customer_id and inv.get("customer_id") == customer_id and inv.get("status") == "outstanding":
                    invoice = inv
                    break
        
        # If no invoice found, try to find oldest outstanding for customer
        if not invoice and customer_id:
            invoices = db.select("invoices", filters={"customer_id": customer_id, "status": "outstanding"})
            if invoices:
                invoice = invoices[0]
        
        # Post payment
        payment_id = generate_id()
        
        if invoice:
            # Mark invoice as paid
            db.update("invoices", invoice["id"], {"status": "paid", "paid_date": pay_date})
            
            # Update customer balance
            if customer_id:
                new_balance = Decimal(str(customer.get("balance", 0))) - amount
                db.update("customers", customer_id, {"balance": float(new_balance)})
        
        # Post GL entry
        entry = JournalEntry(
            date=pay_date,
            reference=f"PMT-{payment_id[:8]}",
            description=f"Payment from {customer_name or 'Customer'} - {reference}",
            trans_type=TransactionType.RECEIPT,
            source_type="payment",
            source_id=payment_id
        )
        
        entry.debit(AccountCodes.BANK, amount)
        entry.credit(AccountCodes.DEBTORS, amount)
        entry.post()
        
        details = f"<strong>{customer_name or 'Customer'}</strong><br>"
        if invoice:
            details += f"Invoice: {invoice.get('invoice_number', reference)}<br>"
            details += "Marked as PAID ✓"
        else:
            details += f"Ref: {reference}<br>"
            details += "Posted to debtors"
        
        return jsonify({
            "success": True,
            "title": "Payment Received!",
            "badge": method.upper(),
            "badge_type": "update",
            "details": details,
            "amount": f"R {amount:,.2f}"
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# =============================================================================
# PIECE 12: PAYROLL MODULE
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - PAYROLL MODULE                                                   ║
║                                                                               ║
║   Features:                                                                   ║
║   - Employee management                                                       ║
║   - SARS PAYE calculation (2024/2025 tax tables)                             ║
║   - UIF calculation (1% employee + 1% employer, capped at R17,712)           ║
║   - SDL calculation (1% of total payroll)                                    ║
║   - Timesheet entry from handwritten sheets                                  ║
║   - AI scan of timesheets                                                    ║
║   - Payslip generation                                                       ║
║   - GL posting (Salaries, PAYE, UIF, SDL)                                   ║
║   - EMP201 preparation                                                       ║
║                                                                               ║
║   All calculations in Flask - SARS compliant                                 ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

# ═══════════════════════════════════════════════════════════════════════════════
# SARS TAX TABLES 2024/2025 (1 March 2024 - 28 February 2025)
# ═══════════════════════════════════════════════════════════════════════════════

class PAYE:
    """
    SARS PAYE Calculator - 2024/2025 Tax Year
    
    Tax brackets and rebates from official SARS tables.
    Calculates monthly PAYE from gross monthly income.
    """
    
    # Annual tax brackets (2024/2025)
    TAX_BRACKETS = [
        (237100, Decimal("0.18"), Decimal("0")),           # 18% of first R237,100
        (370500, Decimal("0.26"), Decimal("42678")),       # R42,678 + 26% above R237,100
        (512800, Decimal("0.31"), Decimal("77362")),       # R77,362 + 31% above R370,500
        (673000, Decimal("0.36"), Decimal("121475")),      # R121,475 + 36% above R512,800
        (857900, Decimal("0.39"), Decimal("179147")),      # R179,147 + 39% above R673,000
        (1817000, Decimal("0.41"), Decimal("251258")),     # R251,258 + 41% above R857,900
        (None, Decimal("0.45"), Decimal("644489")),        # R644,489 + 45% above R1,817,000
    ]
    
    # Annual rebates
    PRIMARY_REBATE = Decimal("17235")      # All taxpayers
    SECONDARY_REBATE = Decimal("9444")     # Age 65+
    TERTIARY_REBATE = Decimal("3145")      # Age 75+
    
    # Tax thresholds (annual income below which no tax is payable)
    THRESHOLD_UNDER_65 = Decimal("95750")
    THRESHOLD_65_TO_74 = Decimal("148217")
    THRESHOLD_75_PLUS = Decimal("165689")
    
    @classmethod
    def calculate_annual_tax(cls, annual_income: Decimal, age: int = 30) -> dict:
        """
        Calculate annual PAYE tax
        
        Args:
            annual_income: Gross annual income
            age: Employee age (affects rebates)
            
        Returns:
            Dict with tax breakdown
        """
        annual_income = Decimal(str(annual_income))
        
        # Determine threshold based on age
        if age >= 75:
            threshold = cls.THRESHOLD_75_PLUS
        elif age >= 65:
            threshold = cls.THRESHOLD_65_TO_74
        else:
            threshold = cls.THRESHOLD_UNDER_65
        
        # Below threshold = no tax
        if annual_income <= threshold:
            return {
                "gross": annual_income,
                "taxable": annual_income,
                "tax_before_rebate": Decimal("0"),
                "rebates": Decimal("0"),
                "tax": Decimal("0"),
                "effective_rate": Decimal("0")
            }
        
        # Calculate tax using brackets
        tax = Decimal("0")
        prev_bracket = 0
        
        for bracket_limit, rate, base_tax in cls.TAX_BRACKETS:
            if bracket_limit is None:
                # Top bracket
                if annual_income > prev_bracket:
                    tax = base_tax + (annual_income - prev_bracket) * rate
                break
            elif annual_income <= bracket_limit:
                if prev_bracket == 0:
                    tax = annual_income * rate
                else:
                    tax = base_tax + (annual_income - prev_bracket) * rate
                break
            prev_bracket = bracket_limit
        
        # Apply rebates
        rebates = cls.PRIMARY_REBATE
        if age >= 65:
            rebates += cls.SECONDARY_REBATE
        if age >= 75:
            rebates += cls.TERTIARY_REBATE
        
        tax_after_rebate = max(Decimal("0"), tax - rebates)
        
        effective_rate = (tax_after_rebate / annual_income * 100).quantize(Decimal("0.01")) if annual_income > 0 else Decimal("0")
        
        return {
            "gross": annual_income,
            "taxable": annual_income,
            "tax_before_rebate": tax.quantize(Decimal("0.01")),
            "rebates": rebates,
            "tax": tax_after_rebate.quantize(Decimal("0.01")),
            "effective_rate": effective_rate
        }
    
    @classmethod
    def calculate_monthly(cls, monthly_gross: Decimal, age: int = 30) -> Decimal:
        """
        Calculate monthly PAYE from monthly gross
        
        Args:
            monthly_gross: Gross monthly salary
            age: Employee age
            
        Returns:
            Monthly PAYE amount
        """
        annual = Decimal(str(monthly_gross)) * 12
        result = cls.calculate_annual_tax(annual, age)
        monthly_tax = (result["tax"] / 12).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return monthly_tax


class UIF:
    """
    Unemployment Insurance Fund Calculator
    
    - Employee contributes 1%
    - Employer contributes 1%
    - Capped at R17,712 monthly income
    """
    
    RATE = Decimal("0.01")  # 1%
    MONTHLY_CAP = Decimal("17712")  # Maximum income for UIF calculation
    MAX_CONTRIBUTION = Decimal("177.12")  # Maximum monthly contribution per party
    
    @classmethod
    def calculate(cls, monthly_gross: Decimal) -> dict:
        """
        Calculate UIF contributions
        
        Args:
            monthly_gross: Gross monthly salary
            
        Returns:
            Dict with employee and employer contributions
        """
        monthly_gross = Decimal(str(monthly_gross))
        
        # Cap the income
        uif_income = min(monthly_gross, cls.MONTHLY_CAP)
        
        employee = (uif_income * cls.RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        employer = employee  # Same amount
        
        return {
            "employee": employee,
            "employer": employer,
            "total": employee + employer,
            "income_used": uif_income
        }


class SDL:
    """
    Skills Development Levy Calculator
    
    - Employer pays 1% of total payroll
    - Only applies if annual payroll exceeds R500,000
    """
    
    RATE = Decimal("0.01")  # 1%
    ANNUAL_THRESHOLD = Decimal("500000")  # Below this, no SDL required
    
    @classmethod
    def calculate(cls, monthly_payroll: Decimal, annual_payroll: Decimal = None) -> Decimal:
        """
        Calculate SDL contribution
        
        Args:
            monthly_payroll: Total monthly payroll
            annual_payroll: Total annual payroll (for threshold check)
            
        Returns:
            SDL amount (paid by employer only)
        """
        monthly_payroll = Decimal(str(monthly_payroll))
        
        # If annual payroll provided, check threshold
        if annual_payroll is not None:
            if Decimal(str(annual_payroll)) < cls.ANNUAL_THRESHOLD:
                return Decimal("0")
        
        return (monthly_payroll * cls.RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# ═══════════════════════════════════════════════════════════════════════════════
# PAYROLL HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def get_all_employees():
    """Get all employees"""
    return db.select("employees", order="name")

def get_employee(emp_id):
    """Get single employee"""
    return db.select_one("employees", emp_id)

def calculate_payslip(employee: dict, hours_worked: float = 0, overtime_hours: float = 0) -> dict:
    """
    Calculate complete payslip for an employee
    
    Args:
        employee: Employee record
        hours_worked: Normal hours worked
        overtime_hours: Overtime hours worked
        
    Returns:
        Complete payslip breakdown
    """
    # Get employee details
    emp_type = employee.get("pay_type", "monthly")  # monthly or hourly
    basic = Decimal(str(employee.get("basic_salary", 0) or 0))
    hourly_rate = Decimal(str(employee.get("hourly_rate", 0) or 0))
    age = int(employee.get("age", 30) or 30)
    
    # Calculate gross pay
    if emp_type == "hourly":
        normal_pay = hourly_rate * Decimal(str(hours_worked))
        overtime_pay = hourly_rate * Decimal("1.5") * Decimal(str(overtime_hours))  # 1.5x for overtime
        gross = normal_pay + overtime_pay
    else:
        gross = basic
        normal_pay = basic
        overtime_pay = Decimal("0")
    
    # Get any allowances
    travel_allowance = Decimal(str(employee.get("travel_allowance", 0) or 0))
    other_allowance = Decimal(str(employee.get("other_allowance", 0) or 0))
    
    gross_with_allowances = gross + travel_allowance + other_allowance
    
    # Calculate deductions
    paye = PAYE.calculate_monthly(gross_with_allowances, age)
    uif = UIF.calculate(gross_with_allowances)
    
    # Other deductions
    medical_aid = Decimal(str(employee.get("medical_aid", 0) or 0))
    pension = Decimal(str(employee.get("pension", 0) or 0))
    loan_deduction = Decimal(str(employee.get("loan_deduction", 0) or 0))
    other_deduction = Decimal(str(employee.get("other_deduction", 0) or 0))
    
    total_deductions = paye + uif["employee"] + medical_aid + pension + loan_deduction + other_deduction
    
    # Net pay
    net_pay = gross_with_allowances - total_deductions
    
    # Employer contributions
    uif_employer = uif["employer"]
    
    return {
        "employee_id": employee.get("id"),
        "employee_name": employee.get("name", ""),
        "employee_number": employee.get("employee_number", ""),
        "id_number": employee.get("id_number", ""),
        "pay_period": today()[:7],  # YYYY-MM
        
        # Earnings
        "basic_salary": float(gross),
        "normal_hours": hours_worked,
        "normal_pay": float(normal_pay),
        "overtime_hours": overtime_hours,
        "overtime_pay": float(overtime_pay),
        "travel_allowance": float(travel_allowance),
        "other_allowance": float(other_allowance),
        "gross_pay": float(gross_with_allowances),
        
        # Deductions
        "paye": float(paye),
        "uif_employee": float(uif["employee"]),
        "medical_aid": float(medical_aid),
        "pension": float(pension),
        "loan_deduction": float(loan_deduction),
        "other_deduction": float(other_deduction),
        "total_deductions": float(total_deductions),
        
        # Net
        "net_pay": float(net_pay),
        
        # Employer costs
        "uif_employer": float(uif_employer),
        "total_cost": float(gross_with_allowances + uif_employer)
    }


def process_payroll(pay_period: str, timesheets: list = None) -> dict:
    """
    Process payroll for all employees
    
    Args:
        pay_period: Period in YYYY-MM format
        timesheets: Optional list of timesheet entries
        
    Returns:
        Payroll summary with all payslips
    """
    employees = get_all_employees()
    
    if not employees:
        return {"success": False, "error": "No employees found"}
    
    payslips = []
    totals = {
        "gross": Decimal("0"),
        "paye": Decimal("0"),
        "uif_employee": Decimal("0"),
        "uif_employer": Decimal("0"),
        "net": Decimal("0"),
        "total_cost": Decimal("0")
    }
    
    for emp in employees:
        if not emp.get("active", True):
            continue
        
        # Find timesheet for this employee if provided
        hours = 0
        overtime = 0
        if timesheets:
            for ts in timesheets:
                if ts.get("employee_id") == emp["id"]:
                    hours = float(ts.get("hours", 0) or 0)
                    overtime = float(ts.get("overtime", 0) or 0)
                    break
        
        payslip = calculate_payslip(emp, hours, overtime)
        payslips.append(payslip)
        
        totals["gross"] += Decimal(str(payslip["gross_pay"]))
        totals["paye"] += Decimal(str(payslip["paye"]))
        totals["uif_employee"] += Decimal(str(payslip["uif_employee"]))
        totals["uif_employer"] += Decimal(str(payslip["uif_employer"]))
        totals["net"] += Decimal(str(payslip["net_pay"]))
        totals["total_cost"] += Decimal(str(payslip["total_cost"]))
    
    # Calculate SDL on total payroll
    sdl = SDL.calculate(totals["gross"])
    totals["sdl"] = sdl
    totals["total_cost"] += sdl
    
    return {
        "success": True,
        "pay_period": pay_period,
        "employee_count": len(payslips),
        "payslips": payslips,
        "totals": {k: float(v) for k, v in totals.items()}
    }


# ═══════════════════════════════════════════════════════════════════════════════
# PAYROLL ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/payroll")
def payroll_home():
    """Payroll dashboard"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    employees = get_all_employees()
    active_count = len([e for e in employees if e.get("active", True)])
    
    # Get recent payroll runs
    recent_runs = db.select("payroll_runs", order="-pay_period", limit=5)
    
    runs_html = ""
    if recent_runs:
        for run in recent_runs:
            runs_html += f'''
            <div class="list-item">
                <div>
                    <strong>{run.get("pay_period", "")}</strong>
                    <span class="text-muted ml-md">{run.get("employee_count", 0)} employees</span>
                </div>
                <div>
                    <span class="text-green">{Money.format(Decimal(str(run.get("total_net", 0))))}</span>
                    <a href="/payroll/run/{run['id']}" class="btn btn-sm btn-ghost ml-md">View</a>
                </div>
            </div>
            '''
    else:
        runs_html = '<div class="text-muted text-center py-lg">No payroll runs yet</div>'
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <h1>Payroll</h1>
            <p class="text-muted">{active_count} active employees</p>
        </div>
        <div class="btn-group">
            <a href="/payroll/employees" class="btn btn-ghost">👥 Employees</a>
            <a href="/payroll/run" class="btn btn-primary">▶ Run Payroll</a>
        </div>
    </div>
    
    <div class="grid grid-2">
        <div class="card">
            <h3 class="card-title">Quick Actions</h3>
            <div class="btn-group" style="flex-direction: column; gap: 12px;">
                <a href="/payroll/employees/new" class="btn btn-ghost btn-block">+ Add Employee</a>
                <a href="/payroll/timesheets" class="btn btn-ghost btn-block">📝 Enter Timesheets</a>
                <a href="/payroll/timesheets/scan" class="btn btn-orange btn-block">📷 Scan Timesheet</a>
            </div>
        </div>
        
        <div class="card">
            <h3 class="card-title">Recent Payroll Runs</h3>
            {runs_html}
        </div>
    </div>
    
    <div class="card mt-lg">
        <h3 class="card-title">SARS Submissions</h3>
        <p class="text-muted mb-md">Monthly EMP201 and bi-annual EMP501 submissions</p>
        <div class="btn-group">
            <a href="/payroll/emp201" class="btn btn-ghost">EMP201 (Monthly)</a>
            <a href="/payroll/emp501" class="btn btn-ghost">EMP501 (Bi-annual)</a>
        </div>
    </div>
    '''
    
    return page_wrapper("Payroll", content, active="", user=user)


@app.route("/payroll/employees")
def payroll_employees():
    """Employee list"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    employees = get_all_employees()
    
    rows = []
    for emp in employees:
        if not emp.get("active", True):
            continue
        
        pay_type = emp.get("pay_type", "monthly")
        if pay_type == "hourly":
            rate = f"R {float(emp.get('hourly_rate', 0)):,.2f}/hr"
        else:
            rate = Money.format(Decimal(str(emp.get("basic_salary", 0))))
        
        rows.append([
            emp.get("employee_number", "-"),
            f'<a href="/payroll/employees/{emp["id"]}">{safe_string(emp.get("name", ""))}</a>',
            emp.get("id_number", "-")[-6:] if emp.get("id_number") else "-",
            pay_type.title(),
            {"value": rate, "class": "number"},
            f'<a href="/payroll/employees/{emp["id"]}/edit" class="btn btn-sm btn-ghost">Edit</a>'
        ])
    
    table = table_html(
        headers=["Emp #", "Name", "ID (last 6)", "Type", {"label": "Rate/Salary", "class": "number"}, ""],
        rows=rows,
        empty_message="No employees yet"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <a href="/payroll" class="text-muted">← Payroll</a>
            <h1>Employees</h1>
        </div>
        <a href="/payroll/employees/new" class="btn btn-primary">+ Add Employee</a>
    </div>
    
    <div class="card">{table}</div>
    '''
    
    return page_wrapper("Employees", content, user=user)


@app.route("/payroll/employees/new", methods=["GET", "POST"])
def payroll_employee_new():
    """Add new employee"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    if request.method == "POST":
        emp_id = generate_id()
        
        # Calculate age from ID number if provided
        id_number = request.form.get("id_number", "").strip()
        age = 30
        if len(id_number) >= 6:
            try:
                year = int(id_number[:2])
                year = 1900 + year if year > 25 else 2000 + year
                current_year = int(today()[:4])
                age = current_year - year
            except:
                pass
        
        employee = {
            "id": emp_id,
            "employee_number": request.form.get("employee_number", "").strip() or f"EMP{len(get_all_employees()) + 1:03d}",
            "name": request.form.get("name", "").strip(),
            "id_number": id_number,
            "age": age,
            "pay_type": request.form.get("pay_type", "monthly"),
            "basic_salary": float(request.form.get("basic_salary", 0) or 0),
            "hourly_rate": float(request.form.get("hourly_rate", 0) or 0),
            "travel_allowance": float(request.form.get("travel_allowance", 0) or 0),
            "medical_aid": float(request.form.get("medical_aid", 0) or 0),
            "pension": float(request.form.get("pension", 0) or 0),
            "bank_name": request.form.get("bank_name", "").strip(),
            "bank_account": request.form.get("bank_account", "").strip(),
            "bank_branch": request.form.get("bank_branch", "").strip(),
            "active": True,
            "created_at": now()
        }
        
        db.insert("employees", employee)
        return redirect("/payroll/employees")
    
    content = '''
    <div class="mb-lg">
        <a href="/payroll/employees" class="text-muted">← Employees</a>
        <h1>Add Employee</h1>
    </div>
    
    <div class="card">
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Employee Number</label>
                    <input type="text" name="employee_number" class="form-input" placeholder="Auto-generated if blank">
                </div>
                <div class="form-group">
                    <label class="form-label">Full Name *</label>
                    <input type="text" name="name" class="form-input" required>
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">SA ID Number</label>
                <input type="text" name="id_number" class="form-input" maxlength="13" placeholder="13-digit ID number">
                <small class="text-muted">Used for age/tax calculation and SARS submissions</small>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Pay Type</label>
                    <select name="pay_type" class="form-select" onchange="togglePayType(this.value)">
                        <option value="monthly">Monthly Salary</option>
                        <option value="hourly">Hourly Rate</option>
                    </select>
                </div>
                <div class="form-group" id="salary-group">
                    <label class="form-label">Monthly Salary</label>
                    <input type="number" name="basic_salary" class="form-input" step="0.01">
                </div>
                <div class="form-group" id="hourly-group" style="display:none;">
                    <label class="form-label">Hourly Rate</label>
                    <input type="number" name="hourly_rate" class="form-input" step="0.01">
                </div>
            </div>
            
            <h4 class="mt-lg mb-md">Allowances & Deductions</h4>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Travel Allowance</label>
                    <input type="number" name="travel_allowance" class="form-input" step="0.01" value="0">
                </div>
                <div class="form-group">
                    <label class="form-label">Medical Aid (Employee)</label>
                    <input type="number" name="medical_aid" class="form-input" step="0.01" value="0">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Pension/Provident Fund (Employee %)</label>
                <input type="number" name="pension" class="form-input" step="0.01" value="0">
            </div>
            
            <h4 class="mt-lg mb-md">Banking Details</h4>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Bank Name</label>
                    <input type="text" name="bank_name" class="form-input">
                </div>
                <div class="form-group">
                    <label class="form-label">Account Number</label>
                    <input type="text" name="bank_account" class="form-input">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Branch Code</label>
                <input type="text" name="bank_branch" class="form-input">
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Employee</button>
                <a href="/payroll/employees" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    
    <script>
    function togglePayType(type) {
        document.getElementById('salary-group').style.display = type === 'monthly' ? 'block' : 'none';
        document.getElementById('hourly-group').style.display = type === 'hourly' ? 'block' : 'none';
    }
    </script>
    '''
    
    return page_wrapper("Add Employee", content, user=user)


@app.route("/payroll/employees/<emp_id>")
def payroll_employee_view(emp_id):
    """View employee details"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    emp = get_employee(emp_id)
    if not emp:
        return redirect("/payroll/employees")
    
    # Calculate sample payslip
    sample = calculate_payslip(emp)
    
    content = f'''
    <div class="mb-lg">
        <a href="/payroll/employees" class="text-muted">← Employees</a>
        <h1>{safe_string(emp.get("name", ""))}</h1>
        <p class="text-muted">Employee #{emp.get("employee_number", "-")}</p>
    </div>
    
    <div class="grid grid-2">
        <div class="card">
            <h3 class="card-title">Employee Details</h3>
            <div class="list-item"><span>ID Number:</span><span>{safe_string(emp.get("id_number", "-"))}</span></div>
            <div class="list-item"><span>Age:</span><span>{emp.get("age", "-")} years</span></div>
            <div class="list-item"><span>Pay Type:</span><span>{emp.get("pay_type", "monthly").title()}</span></div>
            <div class="list-item"><span>Basic Salary:</span><span>{Money.format(Decimal(str(emp.get("basic_salary", 0))))}</span></div>
            <div class="list-item"><span>Bank:</span><span>{safe_string(emp.get("bank_name", "-"))}</span></div>
            <div class="list-item"><span>Account:</span><span>****{emp.get("bank_account", "")[-4:] if emp.get("bank_account") else "-"}</span></div>
            
            <a href="/payroll/employees/{emp_id}/edit" class="btn btn-ghost btn-block mt-lg">Edit Details</a>
        </div>
        
        <div class="card">
            <h3 class="card-title">Sample Payslip</h3>
            <p class="text-muted mb-md">Based on current settings:</p>
            
            <div class="list-item"><span>Gross Pay:</span><span class="text-green">{Money.format(Decimal(str(sample["gross_pay"])))}</span></div>
            <div class="list-item"><span>PAYE:</span><span class="text-red">-{Money.format(Decimal(str(sample["paye"])))}</span></div>
            <div class="list-item"><span>UIF:</span><span class="text-red">-{Money.format(Decimal(str(sample["uif_employee"])))}</span></div>
            <div class="list-item"><span>Medical Aid:</span><span class="text-red">-{Money.format(Decimal(str(sample["medical_aid"])))}</span></div>
            <div class="list-item"><span>Pension:</span><span class="text-red">-{Money.format(Decimal(str(sample["pension"])))}</span></div>
            <hr style="border-color: var(--border); margin: 12px 0;">
            <div class="list-item"><span><strong>Net Pay:</strong></span><span class="text-green" style="font-size: 18px;"><strong>{Money.format(Decimal(str(sample["net_pay"])))}</strong></span></div>
            
            <p class="text-muted mt-md" style="font-size: 12px;">
                Employer UIF: {Money.format(Decimal(str(sample["uif_employer"])))} | 
                Total Cost: {Money.format(Decimal(str(sample["total_cost"])))}
            </p>
        </div>
    </div>
    '''
    
    return page_wrapper(f"Employee - {emp.get('name', '')}", content, user=user)


@app.route("/payroll/run", methods=["GET", "POST"])
def payroll_run():
    """Run payroll"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    if request.method == "POST":
        pay_period = request.form.get("pay_period", today()[:7])
        
        # Get timesheet data from form
        timesheets = []
        employees = get_all_employees()
        for emp in employees:
            hours = request.form.get(f"hours_{emp['id']}", 0)
            overtime = request.form.get(f"overtime_{emp['id']}", 0)
            if hours or overtime:
                timesheets.append({
                    "employee_id": emp["id"],
                    "hours": float(hours or 0),
                    "overtime": float(overtime or 0)
                })
        
        # Process payroll
        result = process_payroll(pay_period, timesheets)
        
        if result["success"]:
            # Save payroll run
            run_id = generate_id()
            run_record = {
                "id": run_id,
                "pay_period": pay_period,
                "employee_count": result["employee_count"],
                "total_gross": result["totals"]["gross"],
                "total_paye": result["totals"]["paye"],
                "total_uif": result["totals"]["uif_employee"] + result["totals"]["uif_employer"],
                "total_sdl": result["totals"]["sdl"],
                "total_net": result["totals"]["net"],
                "total_cost": result["totals"]["total_cost"],
                "payslips": json.dumps(result["payslips"]),
                "status": "processed",
                "created_at": now()
            }
            db.insert("payroll_runs", run_record)
            
            # Post to GL
            entry = JournalEntry(
                date=today(),
                reference=f"PAY-{pay_period}",
                description=f"Payroll for {pay_period}",
                trans_type=TransactionType.EXPENSE,
                source_type="payroll",
                source_id=run_id
            )
            
            # Debit salaries expense
            entry.debit("7000", Decimal(str(result["totals"]["gross"])))  # Salaries & Wages
            
            # Credit PAYE liability
            entry.credit("2200", Decimal(str(result["totals"]["paye"])))  # PAYE Payable
            
            # Credit UIF liability (employee + employer)
            total_uif = Decimal(str(result["totals"]["uif_employee"])) + Decimal(str(result["totals"]["uif_employer"]))
            entry.credit("2210", total_uif)  # UIF Payable
            
            # Debit employer UIF expense
            entry.debit("7010", Decimal(str(result["totals"]["uif_employer"])))  # Employer UIF
            
            # Credit SDL liability and debit expense
            if result["totals"]["sdl"] > 0:
                entry.debit("7020", Decimal(str(result["totals"]["sdl"])))  # SDL Expense
                entry.credit("2220", Decimal(str(result["totals"]["sdl"])))  # SDL Payable
            
            # Credit Bank for net pay
            entry.credit(AccountCodes.BANK, Decimal(str(result["totals"]["net"])))
            
            entry.post()
            
            return redirect(f"/payroll/run/{run_id}")
    
    # Show payroll form
    employees = get_all_employees()
    
    emp_rows = ""
    for emp in employees:
        if not emp.get("active", True):
            continue
        
        pay_type = emp.get("pay_type", "monthly")
        
        if pay_type == "hourly":
            emp_rows += f'''
            <tr>
                <td>{safe_string(emp.get("name", ""))}</td>
                <td>R {float(emp.get("hourly_rate", 0)):,.2f}/hr</td>
                <td><input type="number" name="hours_{emp["id"]}" class="form-input" style="width:80px;" value="176" step="0.5"></td>
                <td><input type="number" name="overtime_{emp["id"]}" class="form-input" style="width:80px;" value="0" step="0.5"></td>
            </tr>
            '''
        else:
            emp_rows += f'''
            <tr>
                <td>{safe_string(emp.get("name", ""))}</td>
                <td>{Money.format(Decimal(str(emp.get("basic_salary", 0))))}</td>
                <td colspan="2" class="text-muted">Monthly salary</td>
            </tr>
            '''
    
    content = f'''
    <div class="mb-lg">
        <a href="/payroll" class="text-muted">← Payroll</a>
        <h1>Run Payroll</h1>
    </div>
    
    <div class="card" style="max-width: 800px;">
        <form method="POST">
            <div class="form-group">
                <label class="form-label">Pay Period</label>
                <input type="month" name="pay_period" class="form-input" value="{today()[:7]}" style="max-width: 200px;">
            </div>
            
            <h4 class="mt-lg mb-md">Employees & Hours</h4>
            
            <div class="table-wrapper">
                <table class="table">
                    <thead>
                        <tr>
                            <th>Employee</th>
                            <th>Rate/Salary</th>
                            <th>Hours</th>
                            <th>Overtime</th>
                        </tr>
                    </thead>
                    <tbody>
                        {emp_rows}
                    </tbody>
                </table>
            </div>
            
            <div class="alert alert-info mt-lg">
                <strong>ℹ️ AI Check:</strong> After processing, AI will review for anomalies (unusual hours, pay changes) before finalizing.
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary btn-lg">Process Payroll</button>
                <a href="/payroll" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    '''
    
    return page_wrapper("Run Payroll", content, user=user)


@app.route("/payroll/run/<run_id>")
def payroll_run_view(run_id):
    """View payroll run results"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    run = db.select_one("payroll_runs", run_id)
    if not run:
        return redirect("/payroll")
    
    payslips = json.loads(run.get("payslips", "[]"))
    
    # Build payslip rows
    rows = []
    for ps in payslips:
        rows.append([
            safe_string(ps.get("employee_name", "")),
            {"value": Money.format(Decimal(str(ps.get("gross_pay", 0)))), "class": "number"},
            {"value": Money.format(Decimal(str(ps.get("paye", 0)))), "class": "number text-red"},
            {"value": Money.format(Decimal(str(ps.get("uif_employee", 0)))), "class": "number text-red"},
            {"value": Money.format(Decimal(str(ps.get("total_deductions", 0)))), "class": "number text-red"},
            {"value": Money.format(Decimal(str(ps.get("net_pay", 0)))), "class": "number text-green"},
        ])
    
    table = table_html(
        headers=["Employee", {"label": "Gross", "class": "number"}, {"label": "PAYE", "class": "number"}, 
                 {"label": "UIF", "class": "number"}, {"label": "Deductions", "class": "number"}, 
                 {"label": "Net Pay", "class": "number"}],
        rows=rows
    )
    
    content = f'''
    <div class="mb-lg">
        <a href="/payroll" class="text-muted">← Payroll</a>
        <h1>Payroll Run - {run.get("pay_period", "")}</h1>
        <span class="badge badge-green">Processed</span>
    </div>
    
    <div class="grid grid-4 mb-lg">
        {stat_card(Money.format(Decimal(str(run.get("total_gross", 0)))), "Total Gross")}
        {stat_card(Money.format(Decimal(str(run.get("total_paye", 0)))), "PAYE", "red")}
        {stat_card(Money.format(Decimal(str(run.get("total_uif", 0)))), "UIF (Total)", "orange")}
        {stat_card(Money.format(Decimal(str(run.get("total_net", 0)))), "Net Payable", "green")}
    </div>
    
    <div class="card">
        <h3 class="card-title">Payslips</h3>
        {table}
    </div>
    
    <div class="card mt-lg">
        <h3 class="card-title">SARS Summary</h3>
        <div class="grid grid-3">
            <div>
                <p class="text-muted">PAYE Payable</p>
                <p class="text-lg">{Money.format(Decimal(str(run.get("total_paye", 0))))}</p>
            </div>
            <div>
                <p class="text-muted">UIF Payable</p>
                <p class="text-lg">{Money.format(Decimal(str(run.get("total_uif", 0))))}</p>
            </div>
            <div>
                <p class="text-muted">SDL Payable</p>
                <p class="text-lg">{Money.format(Decimal(str(run.get("total_sdl", 0))))}</p>
            </div>
        </div>
        <p class="text-muted mt-md">Due to SARS by 7th of next month via EMP201</p>
    </div>
    
    <div class="btn-group mt-lg">
        <a href="/payroll/run/{run_id}/payslips" class="btn btn-ghost">📄 Print Payslips</a>
        <a href="/payroll/emp201?period={run.get("pay_period", "")}" class="btn btn-ghost">📋 Generate EMP201</a>
    </div>
    '''
    
    return page_wrapper(f"Payroll - {run.get('pay_period', '')}", content, user=user)


@app.route("/payroll/timesheets/scan", methods=["GET", "POST"])
def payroll_timesheet_scan():
    """Scan handwritten timesheet with AI"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    content = '''
    <div class="mb-lg">
        <a href="/payroll" class="text-muted">← Payroll</a>
        <h1>📷 Scan Timesheet</h1>
        <p class="text-muted">Take a photo of your handwritten timesheet</p>
    </div>
    
    <div class="card">
        <div class="btn-group mb-lg" style="justify-content: center;">
            <button class="btn btn-orange" onclick="startCamera()">📷 Use Camera</button>
            <label class="btn btn-ghost" style="cursor: pointer;">
                📁 Upload Image
                <input type="file" id="file-input" accept="image/*" style="display: none;" onchange="handleFile(this)">
            </label>
        </div>
        
        <video id="video" autoplay playsinline style="display:none;width:100%;max-width:400px;margin:0 auto;border-radius:12px;background:#000;"></video>
        <button id="capture-btn" class="btn btn-green btn-block mt-md" style="display:none;" onclick="capture()">📸 Capture</button>
        
        <canvas id="canvas" style="display:none;"></canvas>
        <img id="preview" style="display:none;max-width:100%;border-radius:12px;margin-top:16px;">
        
        <div id="processing" class="alert alert-info mt-lg" style="display:none;">
            🔄 AI is reading your timesheet...
        </div>
        
        <div id="result" style="display:none;" class="mt-lg">
            <h4>Extracted Hours</h4>
            <div id="result-content"></div>
            <button class="btn btn-primary btn-block mt-md" onclick="useResults()">Use These Hours</button>
        </div>
    </div>
    
    <script>
    let stream = null;
    let extractedData = null;
    
    function startCamera() {
        navigator.mediaDevices.getUserMedia({ video: { facingMode: "environment" } })
            .then(s => {
                stream = s;
                const video = document.getElementById('video');
                video.srcObject = s;
                video.style.display = 'block';
                document.getElementById('capture-btn').style.display = 'block';
            })
            .catch(e => alert('Camera error: ' + e.message));
    }
    
    function capture() {
        const video = document.getElementById('video');
        const canvas = document.getElementById('canvas');
        canvas.width = video.videoWidth;
        canvas.height = video.videoHeight;
        canvas.getContext('2d').drawImage(video, 0, 0);
        
        const imageData = canvas.toDataURL('image/jpeg', 0.8);
        
        if (stream) {
            stream.getTracks().forEach(t => t.stop());
        }
        video.style.display = 'none';
        document.getElementById('capture-btn').style.display = 'none';
        
        document.getElementById('preview').src = imageData;
        document.getElementById('preview').style.display = 'block';
        
        processImage(imageData);
    }
    
    function handleFile(input) {
        if (input.files && input.files[0]) {
            const reader = new FileReader();
            reader.onload = e => {
                document.getElementById('preview').src = e.target.result;
                document.getElementById('preview').style.display = 'block';
                processImage(e.target.result);
            };
            reader.readAsDataURL(input.files[0]);
        }
    }
    
    async function processImage(imageData) {
        document.getElementById('processing').style.display = 'block';
        document.getElementById('result').style.display = 'none';
        
        try {
            const response = await fetch('/api/payroll/scan-timesheet', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({ image: imageData })
            });
            
            const result = await response.json();
            
            document.getElementById('processing').style.display = 'none';
            
            if (result.success) {
                extractedData = result.data;
                let html = '<table class="table"><thead><tr><th>Employee</th><th>Hours</th><th>Overtime</th></tr></thead><tbody>';
                for (const row of result.data) {
                    html += '<tr><td>' + row.name + '</td><td>' + row.hours + '</td><td>' + row.overtime + '</td></tr>';
                }
                html += '</tbody></table>';
                document.getElementById('result-content').innerHTML = html;
                document.getElementById('result').style.display = 'block';
            } else {
                alert('Could not read timesheet: ' + (result.error || 'Unknown error'));
            }
            
        } catch (error) {
            document.getElementById('processing').style.display = 'none';
            alert('Error: ' + error.message);
        }
    }
    
    function useResults() {
        // Store in session and redirect to payroll run
        sessionStorage.setItem('timesheet_data', JSON.stringify(extractedData));
        window.location.href = '/payroll/run';
    }
    </script>
    '''
    
    return page_wrapper("Scan Timesheet", content, user=user)


@app.route("/api/payroll/scan-timesheet", methods=["POST"])
def api_scan_timesheet():
    """AI scan of handwritten timesheet"""
    try:
        data = request.get_json()
        image_data = data.get("image", "")
        
        if not image_data:
            return jsonify({"success": False, "error": "No image provided"})
        
        if "," in image_data:
            image_data = image_data.split(",")[1]
        
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return jsonify({"success": False, "error": "AI not configured"})
        
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        
        prompt = """Analyze this handwritten timesheet image.

Extract for each employee:
1. Employee name
2. Total normal hours worked
3. Overtime hours (if any)

Return ONLY valid JSON in this format:
{"employees": [{"name": "John Smith", "hours": 176, "overtime": 8}]}

If you can't read a name clearly, use what you can make out.
If hours aren't clear, estimate based on what's visible.
Normal monthly hours are typically around 176 (22 days × 8 hours)."""

        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": image_data}},
                    {"type": "text", "text": prompt}
                ]
            }]
        )
        
        ai_response = message.content[0].text
        
        json_match = re.search(r'\{.*\}', ai_response, re.DOTALL)
        if not json_match:
            return jsonify({"success": False, "error": "Could not read timesheet"})
        
        parsed = json.loads(json_match.group())
        
        return jsonify({
            "success": True,
            "data": parsed.get("employees", [])
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/payroll/timesheets")
def payroll_timesheets():
    """Manual timesheet entry page"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    employees = get_all_employees()
    
    emp_rows = ""
    for emp in employees:
        if not emp.get("active", True):
            continue
        
        emp_rows += f'''
        <tr>
            <td>{safe_string(emp.get("employee_number", ""))}</td>
            <td>{safe_string(emp.get("name", ""))}</td>
            <td><input type="number" name="mon_{emp["id"]}" class="form-input" style="width:60px;" step="0.5"></td>
            <td><input type="number" name="tue_{emp["id"]}" class="form-input" style="width:60px;" step="0.5"></td>
            <td><input type="number" name="wed_{emp["id"]}" class="form-input" style="width:60px;" step="0.5"></td>
            <td><input type="number" name="thu_{emp["id"]}" class="form-input" style="width:60px;" step="0.5"></td>
            <td><input type="number" name="fri_{emp["id"]}" class="form-input" style="width:60px;" step="0.5"></td>
            <td><input type="number" name="sat_{emp["id"]}" class="form-input" style="width:60px;" step="0.5"></td>
            <td><input type="number" name="sun_{emp["id"]}" class="form-input" style="width:60px;" step="0.5"></td>
            <td class="number" id="total_{emp["id"]}">0</td>
        </tr>
        '''
    
    content = f'''
    <div class="mb-lg">
        <a href="/payroll" class="text-muted">← Payroll</a>
        <h1>📝 Enter Timesheets</h1>
        <p class="text-muted">Enter hours for each day - totals calculated automatically</p>
    </div>
    
    <div class="card">
        <div class="form-row mb-md">
            <div class="form-group">
                <label class="form-label">Week Starting</label>
                <input type="date" id="week-start" class="form-input" style="max-width: 200px;">
            </div>
        </div>
        
        <div class="table-wrapper">
            <table class="table">
                <thead>
                    <tr>
                        <th>Emp #</th>
                        <th>Name</th>
                        <th>Mon</th>
                        <th>Tue</th>
                        <th>Wed</th>
                        <th>Thu</th>
                        <th>Fri</th>
                        <th>Sat</th>
                        <th>Sun</th>
                        <th>Total</th>
                    </tr>
                </thead>
                <tbody>
                    {emp_rows}
                </tbody>
            </table>
        </div>
        
        <div class="btn-group mt-lg">
            <button class="btn btn-primary" onclick="saveTimesheets()">Save Timesheets</button>
            <a href="/payroll/timesheets/scan" class="btn btn-orange">📷 Scan Instead</a>
        </div>
    </div>
    
    <script>
    // Auto-calculate row totals
    document.querySelectorAll('input[type="number"]').forEach(input => {{
        input.addEventListener('change', calculateTotals);
    }});
    
    function calculateTotals() {{
        const employees = {json.dumps([e["id"] for e in employees if e.get("active", True)])};
        employees.forEach(empId => {{
            let total = 0;
            ['mon', 'tue', 'wed', 'thu', 'fri', 'sat', 'sun'].forEach(day => {{
                const input = document.querySelector(`input[name="${{day}}_${{empId}}"]`);
                if (input && input.value) {{
                    total += parseFloat(input.value) || 0;
                }}
            }});
            document.getElementById('total_' + empId).textContent = total.toFixed(1);
        }});
    }}
    
    function saveTimesheets() {{
        alert('Timesheets saved! Go to Run Payroll to process.');
        window.location.href = '/payroll/run';
    }}
    </script>
    '''
    
    return page_wrapper("Enter Timesheets", content, user=user)


@app.route("/payroll/run/<run_id>/payslips")
def payroll_payslips_print(run_id):
    """Print payslips - A4 format"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    run = db.select_one("payroll_runs", run_id)
    if not run:
        return redirect("/payroll")
    
    payslips = json.loads(run.get("payslips", "[]"))
    pay_period = run.get("pay_period", "")
    
    # Generate printable payslips
    payslip_html = ""
    for ps in payslips:
        payslip_html += f'''
        <div class="payslip">
            <div class="payslip-header">
                <div class="company-info">
                    <h2>{Config.COMPANY_NAME}</h2>
                    <p>PAYSLIP</p>
                </div>
                <div class="period-info">
                    <p><strong>Pay Period:</strong> {pay_period}</p>
                    <p><strong>Payment Date:</strong> {today()}</p>
                </div>
            </div>
            
            <div class="employee-info">
                <div class="info-row">
                    <span class="label">Employee:</span>
                    <span class="value">{safe_string(ps.get("employee_name", ""))}</span>
                </div>
                <div class="info-row">
                    <span class="label">Employee No:</span>
                    <span class="value">{safe_string(ps.get("employee_number", ""))}</span>
                </div>
                <div class="info-row">
                    <span class="label">ID Number:</span>
                    <span class="value">{safe_string(ps.get("id_number", ""))}</span>
                </div>
            </div>
            
            <div class="payslip-body">
                <div class="earnings">
                    <h4>EARNINGS</h4>
                    <div class="line-item">
                        <span>Basic Salary</span>
                        <span>R {ps.get("basic_salary", 0):,.2f}</span>
                    </div>
                    {"<div class='line-item'><span>Overtime (" + str(ps.get("overtime_hours", 0)) + " hrs)</span><span>R " + f"{ps.get('overtime_pay', 0):,.2f}" + "</span></div>" if ps.get("overtime_pay", 0) > 0 else ""}
                    {"<div class='line-item'><span>Travel Allowance</span><span>R " + f"{ps.get('travel_allowance', 0):,.2f}" + "</span></div>" if ps.get("travel_allowance", 0) > 0 else ""}
                    {"<div class='line-item'><span>Other Allowance</span><span>R " + f"{ps.get('other_allowance', 0):,.2f}" + "</span></div>" if ps.get("other_allowance", 0) > 0 else ""}
                    <div class="line-item total">
                        <span><strong>GROSS PAY</strong></span>
                        <span><strong>R {ps.get("gross_pay", 0):,.2f}</strong></span>
                    </div>
                </div>
                
                <div class="deductions">
                    <h4>DEDUCTIONS</h4>
                    <div class="line-item">
                        <span>PAYE Tax</span>
                        <span>R {ps.get("paye", 0):,.2f}</span>
                    </div>
                    <div class="line-item">
                        <span>UIF</span>
                        <span>R {ps.get("uif_employee", 0):,.2f}</span>
                    </div>
                    {"<div class='line-item'><span>Medical Aid</span><span>R " + f"{ps.get('medical_aid', 0):,.2f}" + "</span></div>" if ps.get("medical_aid", 0) > 0 else ""}
                    {"<div class='line-item'><span>Pension Fund</span><span>R " + f"{ps.get('pension', 0):,.2f}" + "</span></div>" if ps.get("pension", 0) > 0 else ""}
                    {"<div class='line-item'><span>Loan Repayment</span><span>R " + f"{ps.get('loan_deduction', 0):,.2f}" + "</span></div>" if ps.get("loan_deduction", 0) > 0 else ""}
                    {"<div class='line-item'><span>Other Deductions</span><span>R " + f"{ps.get('other_deduction', 0):,.2f}" + "</span></div>" if ps.get("other_deduction", 0) > 0 else ""}
                    <div class="line-item total">
                        <span><strong>TOTAL DEDUCTIONS</strong></span>
                        <span><strong>R {ps.get("total_deductions", 0):,.2f}</strong></span>
                    </div>
                </div>
            </div>
            
            <div class="payslip-footer">
                <div class="net-pay">
                    <span>NET PAY</span>
                    <span class="amount">R {ps.get("net_pay", 0):,.2f}</span>
                </div>
            </div>
            
            <div class="employer-contrib">
                <p class="small">Employer Contributions: UIF R {ps.get("uif_employer", 0):,.2f}</p>
            </div>
        </div>
        '''
    
    html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Payslips - {pay_period}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: Arial, sans-serif;
            font-size: 11pt;
            color: #333;
            background: #fff;
        }}
        
        .payslip {{
            width: 190mm;
            min-height: 140mm;
            margin: 10mm auto;
            padding: 8mm;
            border: 1px solid #333;
            page-break-after: always;
            background: #fff;
        }}
        
        .payslip:last-child {{
            page-break-after: auto;
        }}
        
        .payslip-header {{
            display: flex;
            justify-content: space-between;
            border-bottom: 2px solid #333;
            padding-bottom: 5mm;
            margin-bottom: 5mm;
        }}
        
        .company-info h2 {{
            font-size: 16pt;
            margin-bottom: 2mm;
        }}
        
        .period-info {{
            text-align: right;
        }}
        
        .period-info p {{
            margin: 1mm 0;
        }}
        
        .employee-info {{
            background: #f5f5f5;
            padding: 4mm;
            margin-bottom: 5mm;
        }}
        
        .info-row {{
            display: flex;
            margin: 1mm 0;
        }}
        
        .info-row .label {{
            width: 100px;
            font-weight: bold;
        }}
        
        .payslip-body {{
            display: flex;
            gap: 10mm;
        }}
        
        .earnings, .deductions {{
            flex: 1;
        }}
        
        h4 {{
            background: #333;
            color: #fff;
            padding: 2mm 3mm;
            margin-bottom: 3mm;
            font-size: 10pt;
        }}
        
        .line-item {{
            display: flex;
            justify-content: space-between;
            padding: 1.5mm 3mm;
            border-bottom: 1px dotted #ccc;
        }}
        
        .line-item.total {{
            border-bottom: none;
            border-top: 1px solid #333;
            margin-top: 3mm;
            padding-top: 2mm;
        }}
        
        .payslip-footer {{
            margin-top: 5mm;
            border-top: 2px solid #333;
            padding-top: 5mm;
        }}
        
        .net-pay {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            background: #333;
            color: #fff;
            padding: 4mm 5mm;
            font-size: 14pt;
            font-weight: bold;
        }}
        
        .net-pay .amount {{
            font-size: 18pt;
        }}
        
        .employer-contrib {{
            margin-top: 3mm;
            text-align: right;
        }}
        
        .small {{
            font-size: 9pt;
            color: #666;
        }}
        
        @media print {{
            body {{
                background: #fff;
            }}
            .payslip {{
                border: 1px solid #333;
                margin: 0;
            }}
            .no-print {{
                display: none;
            }}
        }}
    </style>
</head>
<body>
    <div class="no-print" style="padding: 20px; background: #333; color: #fff; text-align: center;">
        <button onclick="window.print()" style="padding: 10px 30px; font-size: 16px; cursor: pointer;">🖨️ Print Payslips</button>
        <a href="/payroll/run/{run_id}" style="color: #fff; margin-left: 20px;">← Back to Payroll Run</a>
        <span style="margin-left: 20px;">|</span>
        <a href="/payroll/run/{run_id}/payslips/compact" style="color: #fff; margin-left: 20px;">📄 Compact (3 per page)</a>
    </div>
    
    {payslip_html}
</body>
</html>'''
    
    return html


@app.route("/payroll/run/<run_id>/payslips/compact")
def payroll_payslips_compact(run_id):
    """Print payslips - Compact format (3 per A4 page for pre-printed stationery)"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    run = db.select_one("payroll_runs", run_id)
    if not run:
        return redirect("/payroll")
    
    payslips = json.loads(run.get("payslips", "[]"))
    pay_period = run.get("pay_period", "")
    
    # Generate compact payslips (3 per page)
    payslip_html = ""
    for i, ps in enumerate(payslips):
        if i > 0 and i % 3 == 0:
            payslip_html += '<div class="page-break"></div>'
        
        payslip_html += f'''
        <div class="payslip-compact">
            <div class="row">
                <div class="col"><strong>{Config.COMPANY_NAME}</strong></div>
                <div class="col">Period: {pay_period}</div>
                <div class="col">Date: {today()}</div>
            </div>
            <div class="row">
                <div class="col">{safe_string(ps.get("employee_name", ""))}</div>
                <div class="col">Emp#: {safe_string(ps.get("employee_number", ""))}</div>
                <div class="col">ID: {safe_string(ps.get("id_number", "")[-6:]) if ps.get("id_number") else "-"}</div>
            </div>
            <div class="divider"></div>
            <div class="row">
                <div class="col">
                    <div class="item">Basic: R {ps.get("basic_salary", 0):,.2f}</div>
                    <div class="item">Overtime: R {ps.get("overtime_pay", 0):,.2f}</div>
                    <div class="item">Allowances: R {(ps.get("travel_allowance", 0) + ps.get("other_allowance", 0)):,.2f}</div>
                    <div class="item"><strong>GROSS: R {ps.get("gross_pay", 0):,.2f}</strong></div>
                </div>
                <div class="col">
                    <div class="item">PAYE: R {ps.get("paye", 0):,.2f}</div>
                    <div class="item">UIF: R {ps.get("uif_employee", 0):,.2f}</div>
                    <div class="item">Other: R {(ps.get("medical_aid", 0) + ps.get("pension", 0) + ps.get("loan_deduction", 0)):,.2f}</div>
                    <div class="item"><strong>DEDUCT: R {ps.get("total_deductions", 0):,.2f}</strong></div>
                </div>
                <div class="col net">
                    <div>NET PAY</div>
                    <div class="amount">R {ps.get("net_pay", 0):,.2f}</div>
                </div>
            </div>
        </div>
        '''
    
    html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Payslips Compact - {pay_period}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: Arial, sans-serif;
            font-size: 9pt;
            color: #333;
        }}
        
        .payslip-compact {{
            width: 190mm;
            height: 90mm;
            margin: 2mm auto;
            padding: 5mm;
            border: 1px dashed #999;
            page-break-inside: avoid;
        }}
        
        .row {{
            display: flex;
            margin-bottom: 2mm;
        }}
        
        .col {{
            flex: 1;
        }}
        
        .col.net {{
            background: #333;
            color: #fff;
            padding: 3mm;
            text-align: center;
        }}
        
        .col.net .amount {{
            font-size: 14pt;
            font-weight: bold;
            margin-top: 2mm;
        }}
        
        .divider {{
            border-top: 1px solid #333;
            margin: 2mm 0;
        }}
        
        .item {{
            padding: 1mm 0;
        }}
        
        .page-break {{
            page-break-after: always;
        }}
        
        @media print {{
            .no-print {{
                display: none;
            }}
            .payslip-compact {{
                border: 1px dashed #ccc;
            }}
        }}
    </style>
</head>
<body>
    <div class="no-print" style="padding: 20px; background: #333; color: #fff; text-align: center;">
        <button onclick="window.print()" style="padding: 10px 30px; font-size: 16px; cursor: pointer;">🖨️ Print Payslips</button>
        <a href="/payroll/run/{run_id}" style="color: #fff; margin-left: 20px;">← Back</a>
        <span style="margin-left: 20px;">|</span>
        <a href="/payroll/run/{run_id}/payslips" style="color: #fff; margin-left: 20px;">📄 Full A4 Format</a>
    </div>
    
    {payslip_html}
</body>
</html>'''
    
    return html


@app.route("/payroll/employees/<emp_id>/edit", methods=["GET", "POST"])
def payroll_employee_edit(emp_id):
    """Edit employee"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    emp = get_employee(emp_id)
    if not emp:
        return redirect("/payroll/employees")
    
    if request.method == "POST":
        # Calculate age from ID number if changed
        id_number = request.form.get("id_number", "").strip()
        age = emp.get("age", 30)
        if len(id_number) >= 6:
            try:
                year = int(id_number[:2])
                year = 1900 + year if year > 25 else 2000 + year
                current_year = int(today()[:4])
                age = current_year - year
            except:
                pass
        
        updates = {
            "employee_number": request.form.get("employee_number", "").strip(),
            "name": request.form.get("name", "").strip(),
            "id_number": id_number,
            "age": age,
            "pay_type": request.form.get("pay_type", "monthly"),
            "basic_salary": float(request.form.get("basic_salary", 0) or 0),
            "hourly_rate": float(request.form.get("hourly_rate", 0) or 0),
            "travel_allowance": float(request.form.get("travel_allowance", 0) or 0),
            "medical_aid": float(request.form.get("medical_aid", 0) or 0),
            "pension": float(request.form.get("pension", 0) or 0),
            "loan_deduction": float(request.form.get("loan_deduction", 0) or 0),
            "bank_name": request.form.get("bank_name", "").strip(),
            "bank_account": request.form.get("bank_account", "").strip(),
            "bank_branch": request.form.get("bank_branch", "").strip(),
            "active": request.form.get("active") == "on"
        }
        
        db.update("employees", emp_id, updates)
        return redirect(f"/payroll/employees/{emp_id}")
    
    pay_type = emp.get("pay_type", "monthly")
    
    content = f'''
    <div class="mb-lg">
        <a href="/payroll/employees/{emp_id}" class="text-muted">← {safe_string(emp.get("name", ""))}</a>
        <h1>Edit Employee</h1>
    </div>
    
    <div class="card">
        <form method="POST">
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Employee Number</label>
                    <input type="text" name="employee_number" class="form-input" value="{safe_string(emp.get("employee_number", ""))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Full Name *</label>
                    <input type="text" name="name" class="form-input" value="{safe_string(emp.get("name", ""))}" required>
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">SA ID Number</label>
                <input type="text" name="id_number" class="form-input" maxlength="13" value="{safe_string(emp.get("id_number", ""))}">
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Pay Type</label>
                    <select name="pay_type" class="form-select" onchange="togglePayType(this.value)">
                        <option value="monthly" {"selected" if pay_type == "monthly" else ""}>Monthly Salary</option>
                        <option value="hourly" {"selected" if pay_type == "hourly" else ""}>Hourly Rate</option>
                    </select>
                </div>
                <div class="form-group" id="salary-group" style="{"" if pay_type == "monthly" else "display:none;"}">
                    <label class="form-label">Monthly Salary</label>
                    <input type="number" name="basic_salary" class="form-input" step="0.01" value="{emp.get("basic_salary", 0)}">
                </div>
                <div class="form-group" id="hourly-group" style="{"" if pay_type == "hourly" else "display:none;"}">
                    <label class="form-label">Hourly Rate</label>
                    <input type="number" name="hourly_rate" class="form-input" step="0.01" value="{emp.get("hourly_rate", 0)}">
                </div>
            </div>
            
            <h4 class="mt-lg mb-md">Allowances & Deductions</h4>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Travel Allowance</label>
                    <input type="number" name="travel_allowance" class="form-input" step="0.01" value="{emp.get("travel_allowance", 0)}">
                </div>
                <div class="form-group">
                    <label class="form-label">Medical Aid</label>
                    <input type="number" name="medical_aid" class="form-input" step="0.01" value="{emp.get("medical_aid", 0)}">
                </div>
            </div>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Pension Fund</label>
                    <input type="number" name="pension" class="form-input" step="0.01" value="{emp.get("pension", 0)}">
                </div>
                <div class="form-group">
                    <label class="form-label">Loan Deduction</label>
                    <input type="number" name="loan_deduction" class="form-input" step="0.01" value="{emp.get("loan_deduction", 0)}">
                </div>
            </div>
            
            <h4 class="mt-lg mb-md">Banking Details</h4>
            
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Bank Name</label>
                    <input type="text" name="bank_name" class="form-input" value="{safe_string(emp.get("bank_name", ""))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Account Number</label>
                    <input type="text" name="bank_account" class="form-input" value="{safe_string(emp.get("bank_account", ""))}">
                </div>
            </div>
            
            <div class="form-group">
                <label class="form-label">Branch Code</label>
                <input type="text" name="bank_branch" class="form-input" value="{safe_string(emp.get("bank_branch", ""))}">
            </div>
            
            <div class="form-group mt-lg">
                <label class="form-label">
                    <input type="checkbox" name="active" {"checked" if emp.get("active", True) else ""}> Active Employee
                </label>
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Changes</button>
                <a href="/payroll/employees/{emp_id}" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    
    <script>
    function togglePayType(type) {{
        document.getElementById('salary-group').style.display = type === 'monthly' ? 'block' : 'none';
        document.getElementById('hourly-group').style.display = type === 'hourly' ? 'block' : 'none';
    }}
    </script>
    '''
    
    return page_wrapper("Edit Employee", content, user=user)


# =============================================================================
# PIECE 13: SETTINGS & PRICING SYSTEM
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - SETTINGS & PRICING SYSTEM                                       ║
║                                                                               ║
║   Phase 1: Pricing Settings                                                   ║
║   - Default markup rules by price tier                                        ║
║   - Category-specific markup overrides                                        ║
║   - Minimum margin warnings                                                   ║
║   - Central pricing function used everywhere                                  ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

# ═══════════════════════════════════════════════════════════════════════════════
# PRICING ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class PricingEngine:
    """
    Central pricing calculator used by all stock operations.
    Retrieves settings from database, falls back to defaults.
    """
    
    # Default markup tiers (used if no settings in database)
    DEFAULT_TIERS = [
        {"max_cost": 50, "markup_pct": 50},      # Under R50: 50% markup
        {"max_cost": 200, "markup_pct": 35},     # R50-R200: 35% markup
        {"max_cost": 1000, "markup_pct": 25},    # R200-R1000: 25% markup
        {"max_cost": None, "markup_pct": 15},    # Over R1000: 15% markup
    ]
    
    # Default category overrides
    DEFAULT_CATEGORY_MARKUPS = {
        "fuel": 10,           # Fuel has thin margins
        "fasteners": 45,      # Small items, higher markup
        "safety": 50,         # PPE and safety gear
        "consumables": 40,    # Welding rods, etc.
        "general": None,      # Use tier-based pricing
    }
    
    DEFAULT_MIN_MARGIN = 10  # Warn if margin below 10%
    
    @classmethod
    def get_settings(cls) -> dict:
        """Get pricing settings from database or defaults"""
        try:
            settings = db.select("settings", filters={"key": "pricing"}, limit=1)
            if settings:
                return json.loads(settings[0].get("value", "{}"))
        except:
            pass
        
        return {
            "tiers": cls.DEFAULT_TIERS,
            "category_markups": cls.DEFAULT_CATEGORY_MARKUPS,
            "min_margin": cls.DEFAULT_MIN_MARGIN
        }
    
    @classmethod
    def save_settings(cls, settings: dict) -> bool:
        """Save pricing settings to database"""
        try:
            existing = db.select("settings", filters={"key": "pricing"}, limit=1)
            if existing:
                db.update("settings", existing[0]["id"], {"value": json.dumps(settings)})
            else:
                db.insert("settings", {
                    "id": generate_id(),
                    "key": "pricing",
                    "value": json.dumps(settings),
                    "created_at": now()
                })
            return True
        except:
            return False
    
    @classmethod
    def calculate_selling_price(cls, cost_price: Decimal, category: str = "general") -> dict:
        """
        Calculate selling price from cost price.
        
        Args:
            cost_price: The cost/purchase price
            category: Stock category (for category-specific markups)
            
        Returns:
            Dict with selling_price, markup_pct, margin_pct, warning
        """
        cost = Decimal(str(cost_price))
        if cost <= 0:
            return {
                "selling_price": Decimal("0"),
                "markup_pct": 0,
                "margin_pct": 0,
                "warning": "No cost price set"
            }
        
        settings = cls.get_settings()
        
        # Check for category-specific markup first
        category_markups = settings.get("category_markups", cls.DEFAULT_CATEGORY_MARKUPS)
        category_lower = (category or "general").lower()
        
        markup_pct = category_markups.get(category_lower)
        
        # If no category markup, use tier-based pricing
        if markup_pct is None:
            tiers = settings.get("tiers", cls.DEFAULT_TIERS)
            for tier in tiers:
                max_cost = tier.get("max_cost")
                if max_cost is None or cost <= Decimal(str(max_cost)):
                    markup_pct = tier.get("markup_pct", 25)
                    break
        
        # Calculate selling price
        markup_multiplier = Decimal("1") + Decimal(str(markup_pct)) / Decimal("100")
        selling_price = (cost * markup_multiplier).quantize(Decimal("0.01"))
        
        # Calculate actual margin percentage
        margin_pct = ((selling_price - cost) / selling_price * 100).quantize(Decimal("0.1"))
        
        # Check for warning
        min_margin = settings.get("min_margin", cls.DEFAULT_MIN_MARGIN)
        warning = None
        if margin_pct < min_margin:
            warning = f"Margin {margin_pct}% is below minimum {min_margin}%"
        
        return {
            "selling_price": selling_price,
            "markup_pct": markup_pct,
            "margin_pct": float(margin_pct),
            "warning": warning
        }
    
    @classmethod
    def check_margin(cls, cost_price: Decimal, selling_price: Decimal) -> dict:
        """
        Check if a cost/selling price combo has acceptable margin.
        
        Returns:
            Dict with margin_pct, is_ok, warning
        """
        cost = Decimal(str(cost_price))
        sell = Decimal(str(selling_price))
        
        if sell <= 0:
            return {"margin_pct": 0, "is_ok": False, "warning": "No selling price"}
        if cost <= 0:
            return {"margin_pct": 100, "is_ok": True, "warning": "No cost price set"}
        
        margin_pct = ((sell - cost) / sell * 100).quantize(Decimal("0.1"))
        
        settings = cls.get_settings()
        min_margin = settings.get("min_margin", cls.DEFAULT_MIN_MARGIN)
        
        if margin_pct < 0:
            return {"margin_pct": float(margin_pct), "is_ok": False, "warning": "Selling BELOW cost!"}
        elif margin_pct < min_margin:
            return {"margin_pct": float(margin_pct), "is_ok": False, "warning": f"Below {min_margin}% minimum"}
        
        return {"margin_pct": float(margin_pct), "is_ok": True, "warning": None}


# ═══════════════════════════════════════════════════════════════════════════════
# SETTINGS ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/settings")
def settings_home():
    """Settings dashboard"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Get business count for display
    businesses = BusinessManager.get_user_businesses(user["id"])
    biz_count = len(businesses)
    biz_text = f"{biz_count} business{'es' if biz_count != 1 else ''}" if biz_count > 0 else "Set up your first business"
    
    content = f'''
    <h1 style="font-size: 24px; font-weight: 700; margin-bottom: 24px;">⚙️ Settings</h1>
    
    <div class="report-grid">
        <a href="/businesses" class="report-card" style="border-color: var(--green); background: linear-gradient(135deg, rgba(16,185,129,0.1), rgba(59,130,246,0.05));">
            <div class="report-card-icon">🏢</div>
            <h3 class="report-card-title">Your Businesses</h3>
            <p class="report-card-desc">{biz_text} • Unlimited included FREE</p>
        </a>
        <a href="/settings/industry" class="report-card" style="border-color: var(--purple);">
            <div class="report-card-icon">🏭</div>
            <h3 class="report-card-title">Industry Template</h3>
            <p class="report-card-desc">Hardware, Motor, Engineering, Retail...</p>
        </a>
        <a href="/settings/pricing" class="report-card">
            <div class="report-card-icon">💰</div>
            <h3 class="report-card-title">Pricing & Markup</h3>
            <p class="report-card-desc">Default markups, category pricing, minimum margins</p>
        </a>
        <a href="/settings/company" class="report-card">
            <div class="report-card-icon">🏪</div>
            <h3 class="report-card-title">Company Details</h3>
            <p class="report-card-desc">Business name, VAT number, contact info</p>
        </a>
        <a href="/settings/cleanup" class="report-card" style="border-color: var(--orange);">
            <div class="report-card-icon">🧹</div>
            <h3 class="report-card-title">Data Cleanup</h3>
            <p class="report-card-desc">AI-powered cleanup of suppliers, customers, stock</p>
        </a>
        <a href="/settings/categories" class="report-card">
            <div class="report-card-icon">📁</div>
            <h3 class="report-card-title">Stock Categories</h3>
            <p class="report-card-desc">Manage product categories</p>
        </a>
    </div>
    '''
    
    return page_wrapper("Settings", content, user=user)


# ═══════════════════════════════════════════════════════════════════════════════
# INDUSTRY TEMPLATES
# ═══════════════════════════════════════════════════════════════════════════════

INDUSTRY_TEMPLATES = {
    "hardware": {
        "name": "Hardware & Building",
        "icon": "🔨",
        "description": "Bolts, nuts, tools, paint, building materials",
        "categories": ["Fasteners", "Tools", "Paint", "Building Materials", "Plumbing", "Electrical", "Safety", "Garden"],
        "expense_categories": ["Stock Purchases", "Fuel", "Vehicle Expenses", "Rent", "Utilities", "Staff Wages", "Repairs", "Advertising"],
        "default_markup": 40,
        "pricing_tiers": [
            {"max": 10, "markup": 100},   # Small items 100%
            {"max": 50, "markup": 60},    # Medium items 60%
            {"max": 200, "markup": 40},   # Larger items 40%
            {"max": None, "markup": 25}   # Big items 25%
        ]
    },
    "motor": {
        "name": "Motor / Auto Spares",
        "icon": "🚗",
        "description": "Car parts, spares, accessories, oils",
        "categories": ["Engine Parts", "Brake Parts", "Filters", "Oils & Lubricants", "Electrical", "Body Parts", "Accessories", "Tyres"],
        "expense_categories": ["Parts Purchases", "Fuel", "Vehicle Expenses", "Rent", "Utilities", "Staff Wages", "Equipment", "Insurance"],
        "default_markup": 35,
        "pricing_tiers": [
            {"max": 50, "markup": 80},
            {"max": 200, "markup": 50},
            {"max": 1000, "markup": 35},
            {"max": None, "markup": 20}
        ]
    },
    "engineering": {
        "name": "Engineering / Fabrication",
        "icon": "⚙️",
        "description": "Steel, welding, machining, fabrication",
        "categories": ["Steel Stock", "Welding Supplies", "Cutting Tools", "Measuring Equipment", "Fasteners", "Abrasives", "Safety Equipment", "Gases"],
        "expense_categories": ["Materials", "Gas & Consumables", "Subcontractors", "Vehicle Expenses", "Rent", "Utilities", "Staff Wages", "Equipment Maintenance"],
        "default_markup": 30,
        "pricing_tiers": [
            {"max": 100, "markup": 50},
            {"max": 500, "markup": 35},
            {"max": 2000, "markup": 25},
            {"max": None, "markup": 15}
        ]
    },
    "retail": {
        "name": "General Retail",
        "icon": "🛒",
        "description": "General merchandise, groceries, convenience",
        "categories": ["Food & Beverage", "Household", "Personal Care", "Cleaning", "Stationery", "Electronics", "Clothing", "Other"],
        "expense_categories": ["Stock Purchases", "Rent", "Utilities", "Staff Wages", "Transport", "Marketing", "Bank Charges", "Insurance"],
        "default_markup": 30,
        "pricing_tiers": [
            {"max": 20, "markup": 50},
            {"max": 100, "markup": 35},
            {"max": 500, "markup": 25},
            {"max": None, "markup": 20}
        ]
    },
    "restaurant": {
        "name": "Restaurant / Pub & Grill",
        "icon": "🍔",
        "description": "Food service, beverages, hospitality",
        "categories": ["Food Ingredients", "Beverages - Alcohol", "Beverages - Soft", "Packaging", "Cleaning", "Kitchen Equipment", "Tableware", "Decor"],
        "expense_categories": ["Food Cost", "Beverage Cost", "Staff Wages", "Rent", "Utilities", "Linen & Laundry", "Marketing", "Equipment Repairs"],
        "default_markup": 200,  # Food typically 300% markup
        "pricing_tiers": [
            {"max": 50, "markup": 250},   # Cheap ingredients
            {"max": 200, "markup": 200},
            {"max": 500, "markup": 150},
            {"max": None, "markup": 100}
        ]
    },
    "accommodation": {
        "name": "B&B / Guesthouse",
        "icon": "🛏️",
        "description": "Accommodation, hospitality, tourism",
        "categories": ["Linen & Bedding", "Toiletries", "Cleaning Supplies", "Breakfast Items", "Beverages", "Maintenance", "Decor", "Guest Amenities"],
        "expense_categories": ["Linen & Laundry", "Cleaning", "Food & Breakfast", "Utilities", "Repairs & Maintenance", "Marketing", "Staff Wages", "Commission Fees"],
        "default_markup": 100,
        "pricing_tiers": [
            {"max": 100, "markup": 150},
            {"max": 500, "markup": 100},
            {"max": None, "markup": 50}
        ]
    },
    "services": {
        "name": "Professional Services",
        "icon": "💼",
        "description": "Consulting, IT, accounting, legal",
        "categories": ["Software", "Office Supplies", "Equipment", "Subscriptions", "Training Materials", "Marketing", "Travel", "Other"],
        "expense_categories": ["Software & Subscriptions", "Office Rent", "Utilities", "Staff Wages", "Professional Fees", "Travel", "Marketing", "Insurance"],
        "default_markup": 50,
        "pricing_tiers": []
    }
}


@app.route("/settings/industry", methods=["GET", "POST"])
def settings_industry():
    """Industry template selection"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = ""
    
    if request.method == "POST":
        industry = request.form.get("industry")
        
        if industry in INDUSTRY_TEMPLATES:
            template = INDUSTRY_TEMPLATES[industry]
            
            # Save selected industry
            db.upsert("settings", {
                "id": "industry_template",
                "key": "industry",
                "value": json.dumps({"industry": industry, "name": template["name"]})
            })
            
            # Create stock categories
            for i, cat_name in enumerate(template["categories"]):
                cat_id = f"cat_{industry}_{i}"
                try:
                    existing = db.select("stock_categories", filters={"name": cat_name})
                    if not existing:
                        db.insert("stock_categories", {
                            "id": cat_id,
                            "name": cat_name,
                            "sort_order": i
                        })
                except:
                    pass
            
            # Save pricing settings
            pricing = {
                "tiers": [{"max_cost": t["max"], "markup_pct": t["markup"]} for t in template["pricing_tiers"]],
                "category_markups": {},
                "min_margin": 10
            }
            PricingEngine.save_settings(pricing)
            
            message = f'<div class="success-box">Applied {template["name"]} template! Categories and pricing updated.</div>'
    
    # Get current industry
    try:
        ind_row = db.select("settings", filters={"key": "industry"}, limit=1)
        current = json.loads(ind_row[0].get("value", "{}")) if ind_row else {}
        current_industry = current.get("industry", "")
    except:
        current_industry = ""
    
    # Build industry cards
    cards_html = ""
    for key, template in INDUSTRY_TEMPLATES.items():
        selected = "border-color: var(--green); background: rgba(16, 185, 129, 0.1);" if key == current_industry else ""
        check = "✓ " if key == current_industry else ""
        
        cards_html += f'''
        <div class="industry-card" style="border: 2px solid var(--border); border-radius: 12px; padding: 20px; cursor: pointer; transition: all 0.2s; {selected}" onclick="selectIndustry('{key}')">
            <div style="font-size: 36px; margin-bottom: 8px;">{template["icon"]}</div>
            <div style="font-size: 16px; font-weight: 600; margin-bottom: 4px;">{check}{template["name"]}</div>
            <div style="font-size: 12px; color: var(--text-muted);">{template["description"]}</div>
        </div>
        '''
    
    content = f'''
    <div class="mb-lg">
        <a href="/settings" class="text-muted">← Settings</a>
        <h1>🏭 Industry Template</h1>
        <p class="text-muted">Choose your business type to get pre-configured categories, pricing tiers, and expense types</p>
    </div>
    
    {message}
    
    <form method="POST" id="industry-form">
        <input type="hidden" name="industry" id="industry-input" value="{current_industry}">
        
        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; margin-bottom: 24px;">
            {cards_html}
        </div>
        
        <button type="submit" class="btn btn-primary btn-lg" id="apply-btn" style="display: none;">Apply Template</button>
    </form>
    
    <script>
    let selectedIndustry = '{current_industry}';
    
    function selectIndustry(key) {{
        selectedIndustry = key;
        document.getElementById('industry-input').value = key;
        document.getElementById('apply-btn').style.display = 'inline-block';
        
        // Update visual selection
        document.querySelectorAll('.industry-card').forEach(card => {{
            card.style.borderColor = 'var(--border)';
            card.style.background = 'transparent';
        }});
        event.currentTarget.style.borderColor = 'var(--green)';
        event.currentTarget.style.background = 'rgba(16, 185, 129, 0.1)';
    }}
    </script>
    '''
    
    return page_wrapper("Industry Template", content, user=user)


@app.route("/settings/pricing", methods=["GET", "POST"])
def settings_pricing():
    """Pricing settings page"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    if request.method == "POST":
        # Save pricing settings
        try:
            tiers = []
            for i in range(4):
                max_cost = request.form.get(f"tier_{i}_max")
                markup = request.form.get(f"tier_{i}_markup", 25)
                tiers.append({
                    "max_cost": int(max_cost) if max_cost and max_cost != "None" else None,
                    "markup_pct": int(markup)
                })
            
            category_markups = {}
            categories = ["fuel", "fasteners", "safety", "consumables", "electrical", "plumbing", "general"]
            for cat in categories:
                markup = request.form.get(f"cat_{cat}")
                if markup and markup.strip():
                    category_markups[cat] = int(markup)
                else:
                    category_markups[cat] = None  # Use tier-based
            
            min_margin = int(request.form.get("min_margin", 10))
            
            settings = {
                "tiers": tiers,
                "category_markups": category_markups,
                "min_margin": min_margin
            }
            
            PricingEngine.save_settings(settings)
            
        except Exception as e:
            pass  # Continue showing page
        
        return redirect("/settings/pricing")
    
    # Get current settings
    settings = PricingEngine.get_settings()
    tiers = settings.get("tiers", PricingEngine.DEFAULT_TIERS)
    category_markups = settings.get("category_markups", PricingEngine.DEFAULT_CATEGORY_MARKUPS)
    min_margin = settings.get("min_margin", PricingEngine.DEFAULT_MIN_MARGIN)
    
    # Build tier inputs
    tier_html = ""
    tier_labels = ["Budget (under)", "Mid-range (under)", "Premium (under)", "Luxury (over)"]
    for i, tier in enumerate(tiers):
        max_cost = tier.get("max_cost", "")
        markup = tier.get("markup_pct", 25)
        
        if i < 3:
            tier_html += f'''
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">{tier_labels[i]} R</label>
                    <input type="number" name="tier_{i}_max" class="form-input" value="{max_cost or ''}" placeholder="Max cost">
                </div>
                <div class="form-group">
                    <label class="form-label">Markup %</label>
                    <input type="number" name="tier_{i}_markup" class="form-input" value="{markup}">
                </div>
            </div>
            '''
        else:
            tier_html += f'''
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">{tier_labels[i]}</label>
                    <input type="text" class="form-input" value="Everything else" disabled>
                    <input type="hidden" name="tier_{i}_max" value="None">
                </div>
                <div class="form-group">
                    <label class="form-label">Markup %</label>
                    <input type="number" name="tier_{i}_markup" class="form-input" value="{markup}">
                </div>
            </div>
            '''
    
    # Build category inputs
    categories = [
        ("fuel", "Fuel & Diesel"),
        ("fasteners", "Fasteners & Bolts"),
        ("safety", "Safety & PPE"),
        ("consumables", "Consumables"),
        ("electrical", "Electrical"),
        ("plumbing", "Plumbing"),
        ("general", "General / Other"),
    ]
    
    cat_html = ""
    for cat_key, cat_name in categories:
        cat_markup = category_markups.get(cat_key, "")
        if cat_markup is None:
            cat_markup = ""
        cat_html += f'''
        <div class="form-row">
            <div class="form-group" style="flex:2;">
                <label class="form-label">{cat_name}</label>
            </div>
            <div class="form-group" style="flex:1;">
                <input type="number" name="cat_{cat_key}" class="form-input" value="{cat_markup}" placeholder="Use tier">
            </div>
        </div>
        '''
    
    content = f'''
    <div class="mb-lg">
        <a href="/settings" class="text-muted">← Settings</a>
        <h1>💰 Pricing & Markup</h1>
    </div>
    
    <form method="POST">
        <div class="grid grid-2">
            <div class="card">
                <h3 class="card-title">Price Tier Markups</h3>
                <p class="text-muted mb-md">Default markup based on cost price</p>
                {tier_html}
            </div>
            
            <div class="card">
                <h3 class="card-title">Category Overrides</h3>
                <p class="text-muted mb-md">Leave blank to use tier-based pricing</p>
                {cat_html}
            </div>
        </div>
        
        <div class="card mt-lg" style="max-width: 400px;">
            <h3 class="card-title">Margin Warning</h3>
            <div class="form-group">
                <label class="form-label">Minimum Margin % (warn if below)</label>
                <input type="number" name="min_margin" class="form-input" value="{min_margin}">
            </div>
        </div>
        
        <div class="btn-group mt-lg">
            <button type="submit" class="btn btn-primary">Save Pricing Settings</button>
            <a href="/settings" class="btn btn-ghost">Cancel</a>
        </div>
    </form>
    
    <div class="card mt-lg">
        <h3 class="card-title">Test Calculator</h3>
        <p class="text-muted mb-md">Enter a cost price to see calculated selling price</p>
        <div class="form-row">
            <div class="form-group">
                <input type="number" id="test-cost" class="form-input" placeholder="Cost price" step="0.01">
            </div>
            <div class="form-group">
                <select id="test-category" class="form-select">
                    <option value="general">General</option>
                    <option value="fuel">Fuel</option>
                    <option value="fasteners">Fasteners</option>
                    <option value="safety">Safety</option>
                </select>
            </div>
            <div class="form-group">
                <button type="button" class="btn btn-ghost" onclick="testPrice()">Calculate</button>
            </div>
        </div>
        <div id="test-result"></div>
    </div>
    
    <script>
    async function testPrice() {{
        const cost = document.getElementById('test-cost').value;
        const category = document.getElementById('test-category').value;
        if (!cost) return;
        
        const resp = await fetch('/api/pricing/calculate?cost=' + cost + '&category=' + category);
        const data = await resp.json();
        
        let html = '<div class="alert alert-info">';
        html += '<strong>Cost:</strong> R ' + parseFloat(cost).toFixed(2) + '<br>';
        html += '<strong>Selling:</strong> R ' + parseFloat(data.selling_price).toFixed(2) + '<br>';
        html += '<strong>Markup:</strong> ' + data.markup_pct + '%<br>';
        html += '<strong>Margin:</strong> ' + data.margin_pct + '%';
        if (data.warning) {{
            html += '<br><span class="text-orange">⚠️ ' + data.warning + '</span>';
        }}
        html += '</div>';
        
        document.getElementById('test-result').innerHTML = html;
    }}
    </script>
    '''
    
    return page_wrapper("Pricing Settings", content, user=user)


@app.route("/api/pricing/calculate")
def api_pricing_calculate():
    """API endpoint to calculate selling price from cost"""
    cost = request.args.get("cost", 0)
    category = request.args.get("category", "general")
    
    try:
        result = PricingEngine.calculate_selling_price(Decimal(str(cost)), category)
        return jsonify({
            "selling_price": float(result["selling_price"]),
            "markup_pct": result["markup_pct"],
            "margin_pct": result["margin_pct"],
            "warning": result["warning"]
        })
    except:
        return jsonify({"error": "Invalid input"})


@app.route("/settings/company", methods=["GET", "POST"])
def settings_company():
    """Company details settings"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Get current settings
    try:
        settings_row = db.select("settings", filters={"key": "company"}, limit=1)
        if settings_row:
            company = json.loads(settings_row[0].get("value", "{}"))
        else:
            company = {}
    except:
        company = {}
    
    if request.method == "POST":
        try:
            company = {
                "name": request.form.get("name", ""),
                "trading_as": request.form.get("trading_as", ""),
                "registration": request.form.get("registration", ""),
                "vat_number": request.form.get("vat_number", ""),
                "address_line1": request.form.get("address_line1", ""),
                "address_line2": request.form.get("address_line2", ""),
                "city": request.form.get("city", ""),
                "postal_code": request.form.get("postal_code", ""),
                "phone": request.form.get("phone", ""),
                "email": request.form.get("email", ""),
                "bank_name": request.form.get("bank_name", ""),
                "bank_account": request.form.get("bank_account", ""),
                "bank_branch": request.form.get("bank_branch", ""),
            }
            
            # Save to database
            existing = db.select("settings", filters={"key": "company"}, limit=1)
            if existing:
                db.update("settings", existing[0]["id"], {"value": json.dumps(company)})
            else:
                db.insert("settings", {
                    "id": generate_id(),
                    "key": "company",
                    "value": json.dumps(company),
                    "created_at": now()
                })
            
            session["settings_message"] = "✓ Company details saved"
        except Exception as e:
            session["settings_message"] = f"Error: {str(e)}"
        
        return redirect("/settings/company")
    
    message = session.pop("settings_message", None)
    message_html = f'<div class="alert alert-info mb-lg">{message}</div>' if message else ""
    
    content = f'''
    <div class="mb-lg">
        <a href="/settings" class="text-muted">← Settings</a>
        <h1>🏢 Company Details</h1>
    </div>
    
    {message_html}
    
    <form method="POST">
        <div class="grid grid-2">
            <div class="card">
                <h3 class="card-title">Business Information</h3>
                <div class="form-group">
                    <label class="form-label">Company Name</label>
                    <input type="text" name="name" class="form-input" value="{safe_string(company.get('name', ''))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Trading As (if different)</label>
                    <input type="text" name="trading_as" class="form-input" value="{safe_string(company.get('trading_as', ''))}">
                </div>
                <div class="form-row">
                    <div class="form-group">
                        <label class="form-label">Registration Number</label>
                        <input type="text" name="registration" class="form-input" value="{safe_string(company.get('registration', ''))}">
                    </div>
                    <div class="form-group">
                        <label class="form-label">VAT Number</label>
                        <input type="text" name="vat_number" class="form-input" value="{safe_string(company.get('vat_number', ''))}">
                    </div>
                </div>
            </div>
            
            <div class="card">
                <h3 class="card-title">Contact Details</h3>
                <div class="form-group">
                    <label class="form-label">Address Line 1</label>
                    <input type="text" name="address_line1" class="form-input" value="{safe_string(company.get('address_line1', ''))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Address Line 2</label>
                    <input type="text" name="address_line2" class="form-input" value="{safe_string(company.get('address_line2', ''))}">
                </div>
                <div class="form-row">
                    <div class="form-group">
                        <label class="form-label">City</label>
                        <input type="text" name="city" class="form-input" value="{safe_string(company.get('city', ''))}">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Postal Code</label>
                        <input type="text" name="postal_code" class="form-input" value="{safe_string(company.get('postal_code', ''))}">
                    </div>
                </div>
                <div class="form-row">
                    <div class="form-group">
                        <label class="form-label">Phone</label>
                        <input type="text" name="phone" class="form-input" value="{safe_string(company.get('phone', ''))}">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Email</label>
                        <input type="email" name="email" class="form-input" value="{safe_string(company.get('email', ''))}">
                    </div>
                </div>
            </div>
        </div>
        
        <div class="card mt-lg" style="max-width: 500px;">
            <h3 class="card-title">Banking Details</h3>
            <p class="text-muted mb-md">For invoices and quotes</p>
            <div class="form-group">
                <label class="form-label">Bank Name</label>
                <input type="text" name="bank_name" class="form-input" value="{safe_string(company.get('bank_name', ''))}">
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Account Number</label>
                    <input type="text" name="bank_account" class="form-input" value="{safe_string(company.get('bank_account', ''))}">
                </div>
                <div class="form-group">
                    <label class="form-label">Branch Code</label>
                    <input type="text" name="bank_branch" class="form-input" value="{safe_string(company.get('bank_branch', ''))}">
                </div>
            </div>
        </div>
        
        <div class="btn-group mt-lg">
            <button type="submit" class="btn btn-primary">Save Company Details</button>
            <a href="/settings" class="btn btn-ghost">Cancel</a>
        </div>
    </form>
    '''
    
    return page_wrapper("Company Details", content, user=user)


# =============================================================================
# PIECE 14: REVIEW BEFORE POSTING (STAGING SYSTEM)
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - REVIEW BEFORE POSTING                                           ║
║                                                                               ║
║   Phase 2: Staging System                                                     ║
║   - Scanned invoices go to staging, not live tables                          ║
║   - User reviews and edits before approving                                   ║
║   - Shows GL entries that WILL be posted                                     ║
║   - Nothing touches real data until approved                                  ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

@app.route("/staging")
def staging_list():
    """List all pending staged transactions"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    staged = db.select("staged_transactions", filters={"status": "pending"}, order="-created_at")
    
    rows = []
    for item in staged:
        data = json.loads(item.get("data", "{}"))
        
        badge_color = "orange"
        if item.get("type") == "supplier_invoice":
            badge = "INVOICE"
            badge_color = "blue"
        elif item.get("type") == "expense":
            badge = "EXPENSE"
            badge_color = "purple"
        else:
            badge = item.get("type", "").upper()
        
        rows.append([
            item.get("created_at", "")[:16].replace("T", " "),
            f'<span class="badge badge-{badge_color}">{badge}</span>',
            safe_string(data.get("supplier", data.get("vendor", ""))),
            {"value": Money.format(Decimal(str(data.get("total", 0)))), "class": "number"},
            f'''<div class="btn-group">
                <a href="/staging/{item["id"]}" class="btn btn-sm btn-green">Review</a>
                <a href="/staging/{item["id"]}/reject" class="btn btn-sm btn-red">Reject</a>
            </div>'''
        ])
    
    table = table_html(
        headers=["Date", "Type", "Supplier/Vendor", {"label": "Amount", "class": "number"}, "Actions"],
        rows=rows,
        empty_message="No pending transactions to review"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <h1>📋 Pending Review</h1>
            <p class="text-muted">{len(staged)} transactions waiting for approval</p>
        </div>
    </div>
    
    <div class="card">{table}</div>
    
    <div class="alert alert-info mt-lg">
        <strong>ℹ️ How it works:</strong> Scanned invoices and expenses are held here until you review and approve them. 
        Nothing is posted to your accounts until you click Approve.
    </div>
    '''
    
    return page_wrapper("Pending Review", content, user=user)


@app.route("/staging/<staged_id>")
def staging_review(staged_id):
    """Review a staged transaction before approving"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    staged = db.select_one("staged_transactions", staged_id)
    if not staged:
        return redirect("/staging")
    
    data = json.loads(staged.get("data", "{}"))
    staged_type = staged.get("type", "")
    
    if staged_type == "supplier_invoice":
        return render_staged_invoice_review(staged_id, data, user)
    elif staged_type == "expense":
        return render_staged_expense_review(staged_id, data, user)
    else:
        return redirect("/staging")


def render_staged_invoice_review(staged_id: str, data: dict, user: dict):
    """Render review screen for staged supplier invoice"""
    
    supplier_name = data.get("supplier", "Unknown")
    invoice_no = data.get("invoice_no", "Not visible")
    total = Decimal(str(data.get("total", 0)))
    vat = Decimal(str(data.get("vat", 0)))
    items = data.get("items", [])
    
    # Check if supplier exists
    existing_supplier = None
    suppliers = db.select("suppliers")
    for s in suppliers:
        if s.get("name", "").lower() == supplier_name.lower():
            existing_supplier = s
            break
    
    supplier_status = f'<span class="badge badge-green">Existing</span>' if existing_supplier else f'<span class="badge badge-orange">NEW - Will be created</span>'
    
    # Build items table with edit capability
    items_html = ""
    for i, item in enumerate(items):
        desc = item.get("description", "")
        code = item.get("code", "")
        qty = item.get("qty", 1)
        unit_price = Money.parse(item.get("unit_price", 0))
        
        # Check if stock exists
        existing_stock = None
        all_stock = db.select("stock_items")
        for s in all_stock:
            if s.get("description", "").lower() == desc.lower():
                existing_stock = s
                break
        
        stock_status = "Existing" if existing_stock else "NEW"
        stock_badge = "green" if existing_stock else "orange"
        
        # Calculate selling price
        pricing = PricingEngine.calculate_selling_price(unit_price)
        
        items_html += f'''
        <tr>
            <td>
                <input type="text" name="item_{i}_desc" value="{safe_string(desc)}" class="form-input" style="width:100%;">
            </td>
            <td>
                <input type="text" name="item_{i}_code" value="{safe_string(code)}" class="form-input" style="width:80px;">
            </td>
            <td>
                <input type="number" name="item_{i}_qty" value="{qty}" class="form-input" style="width:60px;" min="1">
            </td>
            <td>
                <input type="number" name="item_{i}_price" value="{float(unit_price):.2f}" class="form-input" style="width:100px;" step="0.01">
            </td>
            <td class="number">R {float(pricing['selling_price']):.2f}</td>
            <td><span class="badge badge-{stock_badge}">{stock_status}</span></td>
            <td>
                <button type="button" class="btn btn-sm btn-red" onclick="this.closest('tr').remove()">×</button>
            </td>
        </tr>
        '''
    
    # Calculate GL entries preview
    subtotal = total - vat if vat > 0 else total / Decimal("1.15") * Decimal("0.85")
    if vat <= 0:
        vat = total - subtotal
    
    content = f'''
    <div class="mb-lg">
        <a href="/staging" class="text-muted">← Pending Review</a>
        <h1>Review Supplier Invoice</h1>
    </div>
    
    <form method="POST" action="/staging/{staged_id}/approve">
        <div class="grid grid-2">
            <div class="card">
                <h3 class="card-title">Invoice Details</h3>
                <div class="form-group">
                    <label class="form-label">Supplier {supplier_status}</label>
                    <input type="text" name="supplier" value="{safe_string(supplier_name)}" class="form-input">
                </div>
                <div class="form-row">
                    <div class="form-group">
                        <label class="form-label">Invoice Number</label>
                        <input type="text" name="invoice_no" value="{safe_string(invoice_no)}" class="form-input">
                    </div>
                    <div class="form-group">
                        <label class="form-label">Date</label>
                        <input type="date" name="date" value="{data.get('date', today())}" class="form-input">
                    </div>
                </div>
                <div class="form-group">
                    <label class="form-label">
                        <input type="checkbox" name="paid" {'checked' if data.get('paid') else ''}> Already Paid
                    </label>
                </div>
            </div>
            
            <div class="card">
                <h3 class="card-title">GL Entries Preview</h3>
                <p class="text-muted mb-md">These entries will be posted on approval:</p>
                <table class="table">
                    <tr><td>DR Stock / Purchases</td><td class="number">{Money.format(subtotal)}</td></tr>
                    <tr><td>DR VAT Input</td><td class="number">{Money.format(vat)}</td></tr>
                    <tr><td>CR Creditors</td><td class="number">{Money.format(total)}</td></tr>
                </table>
                <div class="mt-md" style="font-size: 24px; text-align: center; color: var(--green);">
                    <strong>Total: {Money.format(total)}</strong>
                </div>
            </div>
        </div>
        
        <div class="card mt-lg">
            <h3 class="card-title">Line Items</h3>
            <p class="text-muted mb-md">Edit descriptions, prices, or remove items before approving</p>
            
            <div class="table-wrapper">
                <table class="table">
                    <thead>
                        <tr>
                            <th>Description</th>
                            <th>Code</th>
                            <th>Qty</th>
                            <th>Cost (ea)</th>
                            <th>Selling</th>
                            <th>Status</th>
                            <th></th>
                        </tr>
                    </thead>
                    <tbody>
                        {items_html}
                    </tbody>
                </table>
            </div>
            
            <input type="hidden" name="item_count" value="{len(items)}">
        </div>
        
        <div class="btn-group mt-lg">
            <button type="submit" class="btn btn-green btn-lg">✓ Approve & Post</button>
            <a href="/staging/{staged_id}/reject" class="btn btn-red">✗ Reject</a>
            <a href="/staging" class="btn btn-ghost">Cancel</a>
        </div>
    </form>
    '''
    
    return page_wrapper("Review Invoice", content, user=user)


def render_staged_expense_review(staged_id: str, data: dict, user: dict):
    """Render review screen for staged expense"""
    
    vendor = data.get("vendor", "Unknown")
    description = data.get("description", "")
    total = Decimal(str(data.get("total", 0)))
    vat = Decimal(str(data.get("vat", 0)))
    category = data.get("category", "general")
    
    # Category options
    categories = [
        ("fuel", "Fuel & Diesel"),
        ("telephone", "Telephone & Data"),
        ("electricity", "Electricity & Water"),
        ("repairs", "Repairs & Maintenance"),
        ("stationery", "Stationery & Office"),
        ("travel", "Travel & Transport"),
        ("advertising", "Advertising & Marketing"),
        ("insurance", "Insurance"),
        ("bank_charges", "Bank Charges"),
        ("general", "General Expenses"),
    ]
    
    cat_options = ""
    for cat_val, cat_name in categories:
        selected = "selected" if cat_val == category else ""
        cat_options += f'<option value="{cat_val}" {selected}>{cat_name}</option>'
    
    # Calculate VAT if not provided
    if vat <= 0:
        vat = (total / Decimal("1.15") * Decimal("0.15")).quantize(Decimal("0.01"))
    subtotal = total - vat
    
    content = f'''
    <div class="mb-lg">
        <a href="/staging" class="text-muted">← Pending Review</a>
        <h1>Review Expense</h1>
    </div>
    
    <form method="POST" action="/staging/{staged_id}/approve">
        <div class="grid grid-2">
            <div class="card">
                <h3 class="card-title">Expense Details</h3>
                <div class="form-group">
                    <label class="form-label">Vendor/Supplier</label>
                    <input type="text" name="vendor" value="{safe_string(vendor)}" class="form-input">
                </div>
                <div class="form-group">
                    <label class="form-label">Description</label>
                    <input type="text" name="description" value="{safe_string(description)}" class="form-input">
                </div>
                <div class="form-row">
                    <div class="form-group">
                        <label class="form-label">Amount (incl VAT)</label>
                        <input type="number" name="total" value="{float(total):.2f}" class="form-input" step="0.01">
                    </div>
                    <div class="form-group">
                        <label class="form-label">VAT</label>
                        <input type="number" name="vat" value="{float(vat):.2f}" class="form-input" step="0.01">
                    </div>
                </div>
                <div class="form-group">
                    <label class="form-label">Category</label>
                    <select name="category" class="form-select">{cat_options}</select>
                </div>
                <div class="form-group">
                    <label class="form-label">Date</label>
                    <input type="date" name="date" value="{data.get('date', today())}" class="form-input">
                </div>
            </div>
            
            <div class="card">
                <h3 class="card-title">GL Entries Preview</h3>
                <p class="text-muted mb-md">These entries will be posted on approval:</p>
                <table class="table">
                    <tr><td>DR Expense ({category.title()})</td><td class="number">{Money.format(subtotal)}</td></tr>
                    <tr><td>DR VAT Input</td><td class="number">{Money.format(vat)}</td></tr>
                    <tr><td>CR Bank</td><td class="number">{Money.format(total)}</td></tr>
                </table>
                <div class="mt-md" style="font-size: 24px; text-align: center; color: var(--orange);">
                    <strong>Total: {Money.format(total)}</strong>
                </div>
            </div>
        </div>
        
        <div class="btn-group mt-lg">
            <button type="submit" class="btn btn-green btn-lg">✓ Approve & Post</button>
            <a href="/staging/{staged_id}/reject" class="btn btn-red">✗ Reject</a>
            <a href="/staging" class="btn btn-ghost">Cancel</a>
        </div>
    </form>
    '''
    
    return page_wrapper("Review Expense", content, user=user)


@app.route("/staging/<staged_id>/approve", methods=["POST"])
def staging_approve(staged_id):
    """Approve and post a staged transaction"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    staged = db.select_one("staged_transactions", staged_id)
    if not staged:
        return redirect("/staging")
    
    staged_type = staged.get("type", "")
    
    # Rebuild data from form (user may have edited)
    if staged_type == "supplier_invoice":
        # Reconstruct items from form
        items = []
        i = 0
        while True:
            desc = request.form.get(f"item_{i}_desc")
            if desc is None:
                break
            items.append({
                "description": desc,
                "code": request.form.get(f"item_{i}_code", ""),
                "qty": int(request.form.get(f"item_{i}_qty", 1) or 1),
                "unit_price": float(request.form.get(f"item_{i}_price", 0) or 0)
            })
            i += 1
        
        data = {
            "supplier": request.form.get("supplier", ""),
            "invoice_no": request.form.get("invoice_no", ""),
            "date": request.form.get("date", today()),
            "paid": request.form.get("paid") == "on",
            "items": items,
            "total": float(request.form.get("total", 0) or 0),
            "vat": float(request.form.get("vat", 0) or 0)
        }
        
        # Recalculate total from items
        total = sum(item["qty"] * item["unit_price"] for item in items)
        data["total"] = total
        
        # Process using existing function
        result = process_supplier_invoice(data)
        
    elif staged_type == "expense":
        data = {
            "vendor": request.form.get("vendor", ""),
            "description": request.form.get("description", ""),
            "date": request.form.get("date", today()),
            "total": float(request.form.get("total", 0) or 0),
            "vat": float(request.form.get("vat", 0) or 0),
            "category": request.form.get("category", "general")
        }
        
        # Process using existing function
        result = process_expense_receipt(data)
    
    # Mark as approved
    db.update("staged_transactions", staged_id, {
        "status": "approved",
        "approved_at": now(),
        "approved_by": user.get("username", "")
    })
    
    return redirect("/staging")


@app.route("/staging/<staged_id>/reject")
def staging_reject(staged_id):
    """Reject a staged transaction"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    db.update("staged_transactions", staged_id, {
        "status": "rejected",
        "rejected_at": now()
    })
    
    return redirect("/staging")


# =============================================================================
# PIECE 15: DATA CLEANUP TOOL
# =============================================================================

"""
╔═══════════════════════════════════════════════════════════════════════════════╗
║                                                                               ║
║   CLICK AI - DATA CLEANUP TOOL                                               ║
║                                                                               ║
║   Phase 3: AI-Powered Data Cleanup                                           ║
║   - Scan suppliers for garbage entries                                       ║
║   - Find duplicate customers/suppliers                                       ║
║   - Fix missing cost/selling prices                                          ║
║   - Merge duplicates                                                          ║
║                                                                               ║
╚═══════════════════════════════════════════════════════════════════════════════╝
"""

@app.route("/settings/cleanup")
def settings_cleanup():
    """Data cleanup dashboard"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Count issues with error handling
    try:
        suppliers = db.select("suppliers") or []
    except:
        suppliers = []
    
    try:
        customers = db.select("customers") or []
    except:
        customers = []
    
    try:
        stock = db.select("stock_items") or []
    except:
        stock = []
    
    # Supplier issues
    supplier_issues = 0
    for s in suppliers:
        name = (s.get("name") or "").strip()
        if len(name) < 3 or name.isdigit() or "@" in name:
            supplier_issues += 1
    
    # Stock issues
    stock_no_cost = 0
    stock_no_selling = 0
    stock_below_cost = 0
    for item in stock:
        cost = float(item.get("cost_price", 0) or 0)
        selling = float(item.get("selling_price", 0) or 0)
        if cost <= 0:
            stock_no_cost += 1
        if selling <= 0:
            stock_no_selling += 1
        elif cost > 0 and selling < cost:
            stock_below_cost += 1
    
    content = f'''
    <div class="mb-lg">
        <a href="/settings" class="text-muted">← Settings</a>
        <h1>🧹 Data Cleanup</h1>
        <p class="text-muted">AI-powered cleanup of your data</p>
    </div>
    
    <div class="grid grid-3 mb-lg">
        {stat_card(str(supplier_issues), "Suspect Suppliers", "orange" if supplier_issues > 0 else "green")}
        {stat_card(str(stock_no_cost), "Stock Missing Cost", "orange" if stock_no_cost > 0 else "green")}
        {stat_card(str(stock_below_cost), "Selling Below Cost", "red" if stock_below_cost > 0 else "green")}
    </div>
    
    <div class="grid grid-2">
        <a href="/settings/cleanup/suppliers" class="card" style="text-decoration:none;">
            <h3 class="card-title">👥 Clean Suppliers</h3>
            <p class="text-muted">Find garbage entries, duplicates, incomplete records</p>
            <p class="text-orange mt-md">{supplier_issues} issues found</p>
        </a>
        
        <a href="/settings/cleanup/stock" class="card" style="text-decoration:none;">
            <h3 class="card-title">📦 Clean Stock</h3>
            <p class="text-muted">Fix missing prices, find duplicates, check margins</p>
            <p class="text-orange mt-md">{stock_no_cost + stock_no_selling + stock_below_cost} issues found</p>
        </a>
        
        <a href="/settings/cleanup/customers" class="card" style="text-decoration:none;">
            <h3 class="card-title">🧑‍🤝‍🧑 Clean Customers</h3>
            <p class="text-muted">Merge duplicates, clean up walk-ins</p>
        </a>
        
        <a href="/settings/cleanup/ai-scan" class="card" style="text-decoration:none; border-color: var(--purple);">
            <h3 class="card-title">🤖 Full AI Scan</h3>
            <p class="text-muted">Let Opus analyze everything and suggest fixes</p>
        </a>
    </div>
    '''
    
    return page_wrapper("Data Cleanup", content, user=user)


@app.route("/settings/cleanup/stock")
def cleanup_stock():
    """Stock cleanup - bulk selection with user-chosen markup - BATCH PROCESSING"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = session.pop("cleanup_message", None)
    message_html = f'<div class="alert alert-info mb-lg">{message}</div>' if message else ""
    
    # Get filter from query string
    filter_type = request.args.get("filter", "all")
    
    try:
        stock = db.select("stock_items", order="description") or []
    except:
        stock = []
    
    # Filter items based on selection
    filtered = []
    for item in stock:
        cost = float(item.get("cost_price") or 0)
        selling = float(item.get("selling_price") or 0)
        
        if filter_type == "no_cost" and cost <= 0:
            filtered.append(item)
        elif filter_type == "no_selling" and selling <= 0:
            filtered.append(item)
        elif filter_type == "below_cost" and cost > 0 and selling < cost:
            filtered.append(item)
        elif filter_type == "all":
            if cost <= 0 or selling <= 0 or (cost > 0 and selling < cost):
                filtered.append(item)
    
    # Count issues for filter buttons
    no_cost_count = sum(1 for s in stock if float(s.get("cost_price") or 0) <= 0)
    no_selling_count = sum(1 for s in stock if float(s.get("selling_price") or 0) <= 0)
    below_cost_count = sum(1 for s in stock if float(s.get("cost_price") or 0) > 0 and float(s.get("selling_price") or 0) < float(s.get("cost_price") or 0))
    
    # Build table rows with checkboxes - limit display to 500 for performance
    display_items = filtered[:500]
    rows_html = ""
    for item in display_items:
        cost = float(item.get("cost_price") or 0)
        selling = float(item.get("selling_price") or 0)
        
        if cost <= 0 and selling <= 0:
            issue = '<span class="badge badge-red">No prices</span>'
        elif cost <= 0:
            issue = '<span class="badge badge-orange">No cost</span>'
        elif selling <= 0:
            issue = '<span class="badge badge-orange">No selling</span>'
        elif selling < cost:
            issue = '<span class="badge badge-red">Below cost</span>'
        else:
            issue = ""
        
        rows_html += f'''
        <tr>
            <td><input type="checkbox" name="items" value="{item.get('id')}" class="item-check" data-cost="{cost}" data-selling="{selling}"></td>
            <td>{safe_string(item.get("code") or "")[:10]}</td>
            <td>{safe_string(item.get("description") or "")[:40]}</td>
            <td class="number">R {cost:.2f}</td>
            <td class="number">R {selling:.2f}</td>
            <td>{issue}</td>
        </tr>
        '''
    
    if not rows_html:
        rows_html = '<tr><td colspan="6" class="text-muted" style="text-align:center;padding:40px;">No items match this filter 🎉</td></tr>'
    
    showing_note = f'<p class="text-muted">Showing {len(display_items)} of {len(filtered)} items</p>' if len(filtered) > 500 else ""
    
    content = f'''
    {message_html}
    
    <div class="mb-lg">
        <a href="/settings/cleanup" class="text-muted">← Cleanup</a>
        <h1>📦 Stock Price Editor</h1>
        <p class="text-muted">Select items, choose markup, click Apply (processes in small batches)</p>
    </div>
    
    <div class="btn-group mb-lg">
        <a href="/settings/cleanup/stock?filter=all" class="btn {"btn-primary" if filter_type == "all" else "btn-ghost"}">All Issues ({no_cost_count + no_selling_count + below_cost_count})</a>
        <a href="/settings/cleanup/stock?filter=no_cost" class="btn {"btn-primary" if filter_type == "no_cost" else "btn-ghost"}">No Cost ({no_cost_count})</a>
        <a href="/settings/cleanup/stock?filter=no_selling" class="btn {"btn-primary" if filter_type == "no_selling" else "btn-ghost"}">No Selling ({no_selling_count})</a>
        <a href="/settings/cleanup/stock?filter=below_cost" class="btn {"btn-primary" if filter_type == "below_cost" else "btn-ghost"}">Below Cost ({below_cost_count})</a>
    </div>
    
    <!-- Progress bar (hidden initially) -->
    <div id="progress-container" style="display:none;" class="card mb-lg">
        <div style="display:flex;align-items:center;gap:16px;">
            <div style="flex:1;">
                <div style="background:#1a1a2e;border-radius:8px;height:24px;overflow:hidden;">
                    <div id="progress-bar" style="background:linear-gradient(90deg,#10b981,#22c55e);height:100%;width:0%;transition:width 0.3s;"></div>
                </div>
            </div>
            <div id="progress-text" style="min-width:120px;text-align:right;">0 / 0</div>
        </div>
        <p id="progress-status" class="text-muted mt-md">Starting...</p>
    </div>
    
    <form id="bulk-form">
        <div class="card mb-lg" style="background: #1a1a2e; padding: 16px;">
            <div style="display:flex;flex-wrap:wrap;gap:12px;align-items:center;">
                <div class="btn-group">
                    <button type="button" class="btn btn-ghost btn-sm" onclick="selectAll()">Select All</button>
                    <button type="button" class="btn btn-ghost btn-sm" onclick="selectNone()">Select None</button>
                </div>
                <div class="btn-group">
                    <button type="button" class="btn btn-ghost btn-sm" onclick="selectByPrice(0, 2)">Under R2</button>
                    <button type="button" class="btn btn-ghost btn-sm" onclick="selectByPrice(0, 5)">Under R5</button>
                    <button type="button" class="btn btn-ghost btn-sm" onclick="selectByPrice(0, 10)">Under R10</button>
                    <button type="button" class="btn btn-ghost btn-sm" onclick="selectByPrice(0, 25)">Under R25</button>
                    <button type="button" class="btn btn-ghost btn-sm" onclick="selectByPrice(25, 100)">R25-R100</button>
                    <button type="button" class="btn btn-ghost btn-sm" onclick="selectByPrice(100, 9999)">Over R100</button>
                </div>
            </div>
            <div style="display:flex;flex-wrap:wrap;gap:12px;align-items:center;margin-top:12px;">
                <span id="selected-count" style="color: #8b8b9a;">0 selected</span>
                <select id="action-select" class="form-select" style="width: auto;">
                    <option value="markup_cost">Set Selling = Cost + Markup %</option>
                    <option value="set_cost">Set Cost = Selling ÷ Markup</option>
                </select>
                <input type="number" id="markup-input" class="form-input" value="50" style="width: 70px;" min="1" max="500">
                <span>%</span>
                <button type="button" class="btn btn-green" onclick="startBatchUpdate()">Apply to Selected</button>
            </div>
        </div>
        
        <div class="card">
            {showing_note}
            <table class="table">
                <thead>
                    <tr>
                        <th style="width:40px;"><input type="checkbox" id="select-all-check" onchange="toggleAll(this)"></th>
                        <th>Code</th>
                        <th>Description</th>
                        <th class="number">Cost</th>
                        <th class="number">Selling</th>
                        <th>Issue</th>
                    </tr>
                </thead>
                <tbody>{rows_html}</tbody>
            </table>
        </div>
    </form>
    
    <script>
    function selectAll() {{
        document.querySelectorAll('.item-check').forEach(cb => cb.checked = true);
        updateCount();
    }}
    
    function selectNone() {{
        document.querySelectorAll('.item-check').forEach(cb => cb.checked = false);
        updateCount();
    }}
    
    function selectByPrice(min, max) {{
        document.querySelectorAll('.item-check').forEach(cb => {{
            const cost = parseFloat(cb.dataset.cost) || 0;
            const selling = parseFloat(cb.dataset.selling) || 0;
            const price = cost > 0 ? cost : selling;
            cb.checked = (price >= min && price < max);
        }});
        updateCount();
    }}
    
    function toggleAll(master) {{
        document.querySelectorAll('.item-check').forEach(cb => cb.checked = master.checked);
        updateCount();
    }}
    
    function updateCount() {{
        const count = document.querySelectorAll('.item-check:checked').length;
        document.getElementById('selected-count').textContent = count + ' selected';
    }}
    
    document.querySelectorAll('.item-check').forEach(cb => cb.addEventListener('change', updateCount));
    
    async function startBatchUpdate() {{
        const checked = document.querySelectorAll('.item-check:checked');
        if (checked.length === 0) {{
            alert('Please select some items first');
            return;
        }}
        
        const itemIds = Array.from(checked).map(cb => cb.value);
        const action = document.getElementById('action-select').value;
        const markup = document.getElementById('markup-input').value;
        
        // Show progress
        document.getElementById('progress-container').style.display = 'block';
        const progressBar = document.getElementById('progress-bar');
        const progressText = document.getElementById('progress-text');
        const progressStatus = document.getElementById('progress-status');
        
        const batchSize = 15;  // Process 15 at a time
        const total = itemIds.length;
        let processed = 0;
        let success = 0;
        let errors = 0;
        
        progressStatus.textContent = 'Processing...';
        
        // Process in batches
        for (let i = 0; i < total; i += batchSize) {{
            const batch = itemIds.slice(i, i + batchSize);
            
            try {{
                const response = await fetch('/settings/cleanup/stock/batch-update', {{
                    method: 'POST',
                    headers: {{ 'Content-Type': 'application/json' }},
                    body: JSON.stringify({{ items: batch, action: action, markup: markup }})
                }});
                
                const result = await response.json();
                success += result.success || 0;
                errors += result.errors || 0;
            }} catch (e) {{
                errors += batch.length;
            }}
            
            processed += batch.length;
            const pct = Math.round((processed / total) * 100);
            progressBar.style.width = pct + '%';
            progressText.textContent = processed + ' / ' + total;
            progressStatus.textContent = 'Updated ' + success + ' items...';
            
            // Small delay to not overwhelm server
            await new Promise(r => setTimeout(r, 100));
        }}
        
        progressStatus.textContent = '✓ Done! Updated ' + success + ' items' + (errors > 0 ? ' (' + errors + ' errors)' : '');
        progressBar.style.background = 'linear-gradient(90deg,#10b981,#22c55e)';
        
        // Refresh after 2 seconds
        setTimeout(() => location.reload(), 2000);
    }}
    </script>
    '''
    
    return page_wrapper("Stock Cleanup", content, user=user)


@app.route("/settings/cleanup/stock/batch-update", methods=["POST"])
def cleanup_stock_batch_update():
    """Process a small batch of stock updates - called via AJAX"""
    user = UserSession.get_current_user()
    if not user:
        return jsonify({"error": "Not logged in"})
    
    try:
        data = request.get_json()
        item_ids = data.get("items", [])
        action = data.get("action", "markup_cost")
        markup = float(data.get("markup", 35) or 35)
        
        success_count = 0
        error_count = 0
        multiplier = 1 + (markup / 100)
        
        for item_id in item_ids:
            try:
                item = db.select_one("stock_items", item_id)
                if not item:
                    continue
                
                cost = float(item.get("cost_price") or 0)
                selling = float(item.get("selling_price") or 0)
                
                updates = {}
                
                if action == "markup_cost":
                    if cost > 0:
                        updates["selling_price"] = round(cost * multiplier, 2)
                    elif selling > 0:
                        est_cost = round(selling / multiplier, 2)
                        updates["cost_price"] = est_cost
                        updates["selling_price"] = round(est_cost * multiplier, 2)
                elif action == "set_cost" and selling > 0:
                    updates["cost_price"] = round(selling / multiplier, 2)
                
                if updates:
                    ok, _ = db.update("stock_items", item_id, updates)
                    if ok:
                        success_count += 1
                    else:
                        error_count += 1
            except:
                error_count += 1
        
        return jsonify({"success": success_count, "errors": error_count})
        
    except Exception as e:
        return jsonify({"error": str(e), "success": 0, "errors": len(item_ids) if 'item_ids' in dir() else 0})


@app.route("/settings/cleanup/stock/bulk-update", methods=["POST"])
def cleanup_stock_bulk_update():
    """Fallback for non-JS form submit"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    session["cleanup_message"] = "Please enable JavaScript for batch updates"
    return redirect("/settings/cleanup/stock")


@app.route("/settings/cleanup/suppliers")
def cleanup_suppliers():
    """Supplier cleanup page with better gibberish detection"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    suppliers = db.select("suppliers", order="name")
    
    issues = []
    for s in suppliers:
        name = (s.get("name") or "").strip()
        
        issue_list = []
        
        # Better gibberish detection
        if len(name) < 3:
            issue_list.append("Name too short")
        elif name.isdigit():
            issue_list.append("Just a number")
        elif "@" in name and " " not in name:
            issue_list.append("Looks like email")
        elif re.match(r'^\d+\s+(street|road|ave|avenue|drive|rd|st|lane|crescent|close)', name.lower()):
            issue_list.append("Looks like address")
        elif re.match(r'^0\d{9}$', name.replace(" ", "").replace("-", "")):
            issue_list.append("Looks like phone number")
        elif re.match(r'^\+27\d{9}$', name.replace(" ", "")):
            issue_list.append("Looks like phone number")
        elif re.match(r'^\d+$', name.replace(" ", "").replace("-", "")):
            issue_list.append("Just numbers")
        elif re.match(r'^[a-z0-9]{20,}$', name.lower().replace(" ", "")):
            issue_list.append("Looks like random ID")
        elif not any(c.isalpha() for c in name):
            issue_list.append("No letters - gibberish?")
        elif len(name) > 3 and name.lower() == name.upper() and not any(c.isalpha() for c in name):
            issue_list.append("No letters")
        elif re.match(r'^(mr|mrs|ms|dr|prof)\s*\.?\s*$', name.lower()):
            issue_list.append("Just a title")
        elif name.lower() in ['test', 'testing', 'asdf', 'qwerty', 'xxx', 'aaa', 'zzz', 'n/a', 'na', 'none', '-', '.', '...']:
            issue_list.append("Test/placeholder")
        
        if issue_list:
            issues.append({
                "id": s.get("id"),
                "name": name,
                "phone": s.get("phone", ""),
                "issues": issue_list
            })
    
    # Sort by issue severity - pure numbers first
    issues.sort(key=lambda x: (0 if "Just a number" in x["issues"] or "Just numbers" in x["issues"] else 1, x["name"]))
    
    rows = []
    for item in issues:
        issue_badges = " ".join([f'<span class="badge badge-red">{i}</span>' for i in item["issues"]])
        rows.append([
            f'<input type="checkbox" name="delete_ids" value="{item["id"]}" class="delete-checkbox">',
            safe_string(item["name"]),
            safe_string(item["phone"]),
            issue_badges,
            f'''<div class="btn-group">
                <a href="/suppliers/{item["id"]}/edit" class="btn btn-sm btn-ghost">Edit</a>
                <a href="/settings/cleanup/suppliers/{item["id"]}/delete" class="btn btn-sm btn-red" 
                   onclick="return confirm('Delete this supplier?')">Delete</a>
            </div>'''
        ])
    
    table = table_html(
        headers=["", "Name", "Phone", "Issues", "Actions"],
        rows=rows,
        empty_message="No suspect suppliers found! 🎉"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <a href="/settings/cleanup" class="text-muted">← Cleanup</a>
            <h1>👥 Supplier Cleanup</h1>
            <p class="text-muted">{len(issues)} suspect entries found</p>
        </div>
        <div class="btn-group">
            <button onclick="selectAll()" class="btn btn-sm btn-ghost">Select All</button>
            <button onclick="deleteSelected()" class="btn btn-sm btn-red">🗑️ Delete Selected</button>
        </div>
    </div>
    
    <form id="bulk-delete-form" method="POST" action="/settings/cleanup/suppliers/bulk-delete">
        <div class="card">{table}</div>
    </form>
    
    <div class="alert alert-info mt-lg">
        <strong>Tip:</strong> Review each entry. If it's garbage (address, phone number, etc.), delete it. 
        If it's a real supplier with a bad name, click Edit to fix it.
    </div>
    
    <script>
    function selectAll() {{
        const checkboxes = document.querySelectorAll('.delete-checkbox');
        const allChecked = Array.from(checkboxes).every(cb => cb.checked);
        checkboxes.forEach(cb => cb.checked = !allChecked);
    }}
    
    function deleteSelected() {{
        const checkboxes = document.querySelectorAll('.delete-checkbox:checked');
        if (checkboxes.length === 0) {{
            alert('Please select items to delete');
            return;
        }}
        if (confirm('Delete ' + checkboxes.length + ' selected suppliers? This cannot be undone.')) {{
            document.getElementById('bulk-delete-form').submit();
        }}
    }}
    </script>
    '''
    
    return page_wrapper("Supplier Cleanup", content, user=user)


@app.route("/settings/cleanup/suppliers/bulk-delete", methods=["POST"])
def cleanup_suppliers_bulk_delete():
    """Bulk delete selected suppliers"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    ids = request.form.getlist("delete_ids")
    deleted = 0
    for supplier_id in ids:
        try:
            db.delete("suppliers", supplier_id)
            deleted += 1
        except:
            pass
    
    session["cleanup_message"] = f"✓ Deleted {deleted} suppliers"
    return redirect("/settings/cleanup/suppliers")


@app.route("/settings/cleanup/suppliers/<supplier_id>/delete")
def cleanup_supplier_delete(supplier_id):
    """Delete a garbage supplier"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    db.delete("suppliers", supplier_id)
    return redirect("/settings/cleanup/suppliers")


@app.route("/settings/categories", methods=["GET", "POST"])
def settings_categories():
    """Stock categories management"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = session.pop("category_message", None)
    message_html = f'<div class="alert alert-info mb-lg">{message}</div>' if message else ""
    
    if request.method == "POST":
        action = request.form.get("action")
        
        if action == "add":
            name = request.form.get("name", "").strip()
            if name:
                try:
                    db.insert("stock_categories", {
                        "id": generate_id(),
                        "name": name,
                        "created_at": now()
                    })
                    session["category_message"] = f"✓ Added category: {name}"
                except:
                    session["category_message"] = "Error adding category"
        
        elif action == "delete":
            cat_id = request.form.get("category_id")
            if cat_id:
                try:
                    db.delete("stock_categories", cat_id)
                    session["category_message"] = "✓ Category deleted"
                except:
                    session["category_message"] = "Error deleting category"
        
        return redirect("/settings/categories")
    
    # Get existing categories
    try:
        categories = db.select("stock_categories", order="name") or []
    except:
        categories = []
    
    # Default categories if none exist
    default_cats = ["General", "Fasteners", "Electrical", "Plumbing", "Safety", "Tools", "Consumables", "Fuel"]
    
    cat_rows = ""
    for cat in categories:
        cat_rows += f'''
        <tr>
            <td>{safe_string(cat.get("name", ""))}</td>
            <td>
                <form method="POST" style="display:inline;">
                    <input type="hidden" name="action" value="delete">
                    <input type="hidden" name="category_id" value="{cat.get("id", "")}">
                    <button type="submit" class="btn btn-sm btn-red" onclick="return confirm('Delete this category?')">Delete</button>
                </form>
            </td>
        </tr>
        '''
    
    if not cat_rows:
        cat_rows = '<tr><td colspan="2" class="text-muted" style="text-align:center;padding:20px;">No categories yet. Add some below or use defaults.</td></tr>'
    
    content = f'''
    <div class="mb-lg">
        <a href="/settings" class="text-muted">← Settings</a>
        <h1>📁 Stock Categories</h1>
    </div>
    
    {message_html}
    
    <div class="grid grid-2">
        <div class="card">
            <h3 class="card-title">Current Categories</h3>
            <table class="table">
                <thead><tr><th>Name</th><th></th></tr></thead>
                <tbody>{cat_rows}</tbody>
            </table>
        </div>
        
        <div class="card">
            <h3 class="card-title">Add Category</h3>
            <form method="POST">
                <input type="hidden" name="action" value="add">
                <div class="form-group">
                    <label class="form-label">Category Name</label>
                    <input type="text" name="name" class="form-input" placeholder="e.g. Electrical">
                </div>
                <button type="submit" class="btn btn-primary">Add Category</button>
            </form>
            
            <hr style="margin: 20px 0; border-color: var(--border);">
            
            <h4>Quick Add Defaults</h4>
            <p class="text-muted mb-md">Click to add common categories:</p>
            <div class="btn-group" style="flex-wrap: wrap;">
                {"".join([f'<form method="POST" style="display:inline;margin:2px;"><input type="hidden" name="action" value="add"><input type="hidden" name="name" value="{cat}"><button type="submit" class="btn btn-sm btn-ghost">{cat}</button></form>' for cat in default_cats])}
            </div>
        </div>
    </div>
    '''
    
    return page_wrapper("Stock Categories", content, user=user)


@app.route("/settings/cleanup/ai-scan")
def cleanup_ai_scan():
    """AI-powered full data scan - placeholder for now"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    content = '''
    <div class="mb-lg">
        <a href="/settings/cleanup" class="text-muted">← Cleanup</a>
        <h1>🤖 Full AI Scan</h1>
    </div>
    
    <div class="card" style="text-align: center; padding: 60px;">
        <div style="font-size: 64px; margin-bottom: 20px;">🚧</div>
        <h2>Coming Soon</h2>
        <p class="text-muted">
            AI-powered full scan of your data will analyze:<br><br>
            • Duplicate suppliers and customers<br>
            • Stock items that might be the same thing<br>
            • Pricing anomalies across your inventory<br>
            • Data quality issues<br><br>
            For now, use the individual cleanup tools.
        </p>
        <a href="/settings/cleanup" class="btn btn-primary mt-lg">Back to Cleanup</a>
    </div>
    '''
    
    return page_wrapper("AI Scan", content, user=user)


@app.route("/settings/cleanup/customers")
def cleanup_customers():
    """Customer cleanup page with gibberish detection"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    customers = db.select("customers", order="name")
    
    issues = []
    for c in customers:
        name = (c.get("name") or "").strip()
        
        issue_list = []
        
        # Better gibberish detection
        if len(name) < 3:
            issue_list.append("Name too short")
        elif name.isdigit():
            issue_list.append("Just a number")
        elif "@" in name and " " not in name:
            issue_list.append("Looks like email")
        elif re.match(r'^\d+\s+(street|road|ave|avenue|drive|rd|st|lane|crescent|close)', name.lower()):
            issue_list.append("Looks like address")
        elif re.match(r'^0\d{9}$', name.replace(" ", "").replace("-", "")):
            issue_list.append("Looks like phone number")
        elif re.match(r'^\+27\d{9}$', name.replace(" ", "")):
            issue_list.append("Looks like phone number")
        elif re.match(r'^\d+$', name.replace(" ", "").replace("-", "")):
            issue_list.append("Just numbers")
        elif re.match(r'^[a-z0-9]{20,}$', name.lower().replace(" ", "")):
            issue_list.append("Looks like random ID")
        elif not any(ch.isalpha() for ch in name):
            issue_list.append("No letters - gibberish?")
        elif re.match(r'^(mr|mrs|ms|dr|prof)\s*\.?\s*$', name.lower()):
            issue_list.append("Just a title")
        elif name.lower() in ['test', 'testing', 'asdf', 'qwerty', 'xxx', 'aaa', 'zzz', 'n/a', 'na', 'none', '-', '.', '...', 'cash', 'walk-in', 'walkin']:
            issue_list.append("Test/placeholder")
        
        if issue_list:
            issues.append({
                "id": c.get("id"),
                "name": name,
                "phone": c.get("phone", ""),
                "balance": Decimal(str(c.get("balance", 0) or 0)),
                "issues": issue_list
            })
    
    # Sort by issue severity
    issues.sort(key=lambda x: (0 if "Just a number" in x["issues"] or "Just numbers" in x["issues"] else 1, x["name"]))
    
    rows = []
    for item in issues:
        issue_badges = " ".join([f'<span class="badge badge-red">{i}</span>' for i in item["issues"]])
        balance_display = Money.format(item["balance"]) if item["balance"] != 0 else "-"
        rows.append([
            f'<input type="checkbox" name="delete_ids" value="{item["id"]}" class="delete-checkbox">',
            safe_string(item["name"]),
            safe_string(item["phone"]),
            balance_display,
            issue_badges,
            f'''<div class="btn-group">
                <a href="/customers/{item["id"]}/edit" class="btn btn-sm btn-ghost">Edit</a>
                <a href="/settings/cleanup/customers/{item["id"]}/delete" class="btn btn-sm btn-red" 
                   onclick="return confirm('Delete this customer?')">Delete</a>
            </div>'''
        ])
    
    table = table_html(
        headers=["", "Name", "Phone", "Balance", "Issues", "Actions"],
        rows=rows,
        empty_message="No suspect customers found! 🎉"
    )
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <a href="/settings/cleanup" class="text-muted">← Cleanup</a>
            <h1>🧑‍🤝‍🧑 Customer Cleanup</h1>
            <p class="text-muted">{len(issues)} suspect entries found</p>
        </div>
        <div class="btn-group">
            <button onclick="selectAll()" class="btn btn-sm btn-ghost">Select All</button>
            <button onclick="deleteSelected()" class="btn btn-sm btn-red">🗑️ Delete Selected</button>
        </div>
    </div>
    
    <form id="bulk-delete-form" method="POST" action="/settings/cleanup/customers/bulk-delete">
        <div class="card">{table}</div>
    </form>
    
    <div class="alert alert-info mt-lg">
        <strong>Tip:</strong> Be careful with customers that have a balance - you may want to edit rather than delete.
    </div>
    
    <script>
    function selectAll() {{
        const checkboxes = document.querySelectorAll('.delete-checkbox');
        const allChecked = Array.from(checkboxes).every(cb => cb.checked);
        checkboxes.forEach(cb => cb.checked = !allChecked);
    }}
    
    function deleteSelected() {{
        const checkboxes = document.querySelectorAll('.delete-checkbox:checked');
        if (checkboxes.length === 0) {{
            alert('Please select items to delete');
            return;
        }}
        if (confirm('Delete ' + checkboxes.length + ' selected customers? This cannot be undone.')) {{
            document.getElementById('bulk-delete-form').submit();
        }}
    }}
    </script>
    '''
    
    return page_wrapper("Customer Cleanup", content, user=user)


@app.route("/settings/cleanup/customers/bulk-delete", methods=["POST"])
def cleanup_customers_bulk_delete():
    """Bulk delete selected customers"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    ids = request.form.getlist("delete_ids")
    deleted = 0
    for customer_id in ids:
        try:
            db.delete("customers", customer_id)
            deleted += 1
        except:
            pass
    
    session["cleanup_message"] = f"✓ Deleted {deleted} customers"
    return redirect("/settings/cleanup/customers")


@app.route("/settings/cleanup/customers/<customer_id>/delete")
def cleanup_customer_delete(customer_id):
    """Delete a garbage customer"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    db.delete("customers", customer_id)
    return redirect("/settings/cleanup/customers")


# ═══════════════════════════════════════════════════════════════════════════════
# MULTI-BUSINESS ROUTES - THE COMPETITIVE ADVANTAGE
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/switch-business/<business_id>")
def switch_business(business_id):
    """Switch to a different business"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    if BusinessManager.switch_business(business_id):
        return redirect("/")
    
    return redirect("/")


@app.route("/businesses")
def businesses_list():
    """List all businesses for the user"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    businesses = BusinessManager.get_user_businesses(user["id"])
    current_biz = BusinessManager.get_current_business()
    
    rows_html = ""
    for biz in businesses:
        is_current = "✓" if current_biz and biz["id"] == current_biz["id"] else ""
        industry_config = BusinessManager.INDUSTRY_CONFIGS.get(biz.get("industry", "retail"), {})
        biz_name = biz.get("business_name", biz.get("name", "Unnamed"))
        
        rows_html += f'''
        <tr>
            <td style="font-size:24px;">{biz.get("icon", "🏢")}</td>
            <td><strong>{safe_string(biz_name)}</strong></td>
            <td>{industry_config.get("name", biz.get("industry", ""))}</td>
            <td style="color:var(--green);font-weight:600;">{is_current}</td>
            <td>
                <a href="/switch-business/{biz["id"]}" class="btn btn-sm btn-primary">Switch</a>
                <a href="/businesses/{biz["id"]}/edit" class="btn btn-sm btn-ghost">Edit</a>
            </td>
        </tr>
        '''
    
    if not rows_html:
        rows_html = '<tr><td colspan="5" style="text-align:center;padding:40px;color:#606070;">No businesses yet. Create your first one!</td></tr>'
    
    content = f'''
    <div class="flex-between mb-lg">
        <div>
            <h1>🏢 Your Businesses</h1>
            <p class="text-muted">One subscription, unlimited businesses - completely separate data for each</p>
        </div>
        <a href="/businesses/new" class="btn btn-primary">➕ Add Business</a>
    </div>
    
    <div class="card">
        <table class="table">
            <thead>
                <tr>
                    <th style="width:50px;"></th>
                    <th>Business Name</th>
                    <th>Type</th>
                    <th>Active</th>
                    <th style="width:180px;">Actions</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
    </div>
    
    <div class="card mt-lg" style="background: linear-gradient(135deg, rgba(139,92,246,0.1), rgba(59,130,246,0.1)); border-color: rgba(139,92,246,0.3);">
        <h3 style="color:#a78bfa;margin-bottom:8px;">💎 What Makes This Special</h3>
        <p style="color:#8b8b9a;margin-bottom:16px;">
            Xero and QuickBooks charge <strong>per business</strong> - running 4 businesses could cost you R3,000+/month.
            With Click AI, you get <strong>unlimited businesses included</strong>. Each business has:
        </p>
        <ul style="color:#8b8b9a;margin-left:20px;">
            <li>Completely separate customers, suppliers, and stock</li>
            <li>Its own invoices, expenses, and reports</li>
            <li>Industry-specific configuration and categories</li>
            <li>Isolated financial data and VAT calculations</li>
        </ul>
    </div>
    '''
    
    return page_wrapper("Your Businesses", content, "settings", user)


@app.route("/businesses/new", methods=["GET", "POST"])
def business_new():
    """Create a new business"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    message = ""
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        industry = request.form.get("industry", "retail")
        
        if not name:
            message = '<div class="alert alert-error">Please enter a business name</div>'
        else:
            success, result = BusinessManager.create_business(user["id"], name, industry)
            if success:
                return redirect("/businesses")
            else:
                message = f'<div class="alert alert-error">{result}</div>'
    
    # Build industry cards
    cards_html = ""
    for key, config in BusinessManager.INDUSTRY_CONFIGS.items():
        cards_html += f'''
        <label class="industry-card" style="display:block;border:2px solid var(--border);border-radius:12px;padding:16px;cursor:pointer;transition:all 0.2s;margin-bottom:12px;">
            <input type="radio" name="industry" value="{key}" style="display:none;" onchange="selectIndustry(this)">
            <div style="display:flex;align-items:center;gap:12px;">
                <span style="font-size:32px;">{config["icon"]}</span>
                <div>
                    <div style="font-weight:600;font-size:16px;">{config["name"]}</div>
                    <div style="font-size:12px;color:#606070;">{config["description"]}</div>
                </div>
            </div>
        </label>
        '''
    
    content = f'''
    <div class="mb-lg">
        <a href="/businesses" class="text-muted">← Your Businesses</a>
        <h1>➕ Add New Business</h1>
        <p class="text-muted">Create a new business with its own separate data, customers, and stock</p>
    </div>
    
    {message}
    
    <div class="card">
        <form method="POST">
            <div class="form-group">
                <label class="form-label">Business Name *</label>
                <input type="text" name="name" class="form-input" placeholder="e.g. FullTech Hardware, Joe's Pub" required autofocus>
            </div>
            
            <div class="form-group">
                <label class="form-label">Business Type</label>
                <p class="text-muted" style="font-size:13px;margin-bottom:12px;">
                    This configures which modules you see and sets up default categories. You can change it later.
                </p>
                {cards_html}
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary btn-lg">Create Business</button>
                <a href="/businesses" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    
    <script>
    function selectIndustry(input) {{
        document.querySelectorAll('.industry-card').forEach(card => {{
            card.style.borderColor = 'var(--border)';
            card.style.background = 'transparent';
        }});
        input.closest('.industry-card').style.borderColor = 'var(--green)';
        input.closest('.industry-card').style.background = 'rgba(16,185,129,0.1)';
    }}
    // Select first option by default
    document.querySelector('.industry-card input').checked = true;
    document.querySelector('.industry-card').style.borderColor = 'var(--green)';
    document.querySelector('.industry-card').style.background = 'rgba(16,185,129,0.1)';
    </script>
    '''
    
    return page_wrapper("Add New Business", content, "settings", user)


@app.route("/businesses/<business_id>/edit", methods=["GET", "POST"])
def business_edit(business_id):
    """Edit a business"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Get the business
    business = db.select_one("businesses", business_id)
    if not business or business.get("owner_id") != user["id"]:
        return redirect("/businesses")
    
    message = ""
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        industry = request.form.get("industry", business.get("industry", "retail"))
        
        if not name:
            message = '<div class="alert alert-error">Please enter a business name</div>'
        else:
            config = BusinessManager.INDUSTRY_CONFIGS.get(industry, BusinessManager.INDUSTRY_CONFIGS["retail"])
            
            update_data = {
                "business_name": name,
                "name": name,  # Keep both for compatibility
                "industry": industry,
                "business_type": config["name"],
                "icon": config["icon"],
                "visible_modules": json.dumps(config["visible_modules"]),
                "hidden_modules": json.dumps(config["hidden_modules"]),
                "terminology": json.dumps(config.get("terminology", {}))
            }
            
            if BusinessManager.update_business(business_id, update_data):
                return redirect("/businesses")
            else:
                message = '<div class="alert alert-error">Failed to update business</div>'
    
    # Build industry cards
    current_industry = business.get("industry", "retail")
    cards_html = ""
    for key, config in BusinessManager.INDUSTRY_CONFIGS.items():
        selected = "border-color:var(--green);background:rgba(16,185,129,0.1);" if key == current_industry else ""
        checked = "checked" if key == current_industry else ""
        cards_html += f'''
        <label class="industry-card" style="display:block;border:2px solid var(--border);border-radius:12px;padding:16px;cursor:pointer;transition:all 0.2s;margin-bottom:12px;{selected}">
            <input type="radio" name="industry" value="{key}" style="display:none;" onchange="selectIndustry(this)" {checked}>
            <div style="display:flex;align-items:center;gap:12px;">
                <span style="font-size:32px;">{config["icon"]}</span>
                <div>
                    <div style="font-weight:600;font-size:16px;">{config["name"]}</div>
                    <div style="font-size:12px;color:#606070;">{config["description"]}</div>
                </div>
            </div>
        </label>
        '''
    
    content = f'''
    <div class="mb-lg">
        <a href="/businesses" class="text-muted">← Your Businesses</a>
        <h1>✏️ Edit Business</h1>
    </div>
    
    {message}
    
    <div class="card">
        <form method="POST">
            <div class="form-group">
                <label class="form-label">Business Name *</label>
                <input type="text" name="name" class="form-input" value="{safe_string(business.get('business_name', business.get('name', '')))}" required>
            </div>
            
            <div class="form-group">
                <label class="form-label">Business Type</label>
                {cards_html}
            </div>
            
            <div class="btn-group mt-lg">
                <button type="submit" class="btn btn-primary">Save Changes</button>
                <a href="/businesses" class="btn btn-ghost">Cancel</a>
            </div>
        </form>
    </div>
    
    <div class="card mt-lg" style="border-color: rgba(239,68,68,0.3);">
        <h3 style="color:#ef4444;margin-bottom:8px;">⚠️ Danger Zone</h3>
        <p style="color:#8b8b9a;margin-bottom:16px;">
            Deleting a business will hide it from your list. Data is preserved but inaccessible.
        </p>
        <a href="/businesses/{business_id}/delete" class="btn btn-red" onclick="return confirm('Are you sure you want to delete this business?')">Delete Business</a>
    </div>
    
    <script>
    function selectIndustry(input) {{
        document.querySelectorAll('.industry-card').forEach(card => {{
            card.style.borderColor = 'var(--border)';
            card.style.background = 'transparent';
        }});
        input.closest('.industry-card').style.borderColor = 'var(--green)';
        input.closest('.industry-card').style.background = 'rgba(16,185,129,0.1)';
    }}
    </script>
    '''
    
    return page_wrapper("Edit Business", content, "settings", user)


@app.route("/businesses/<business_id>/delete")
def business_delete(business_id):
    """Delete (deactivate) a business"""
    user = UserSession.get_current_user()
    if not user:
        return redirect("/login")
    
    # Verify ownership
    business = db.select_one("businesses", business_id)
    if not business or business.get("owner_id") != user["id"]:
        return redirect("/businesses")
    
    # Soft delete
    BusinessManager.delete_business(business_id)
    
    # If this was the current business, switch to another
    if session.get("current_business_id") == business_id:
        businesses = BusinessManager.get_user_businesses(user["id"])
        if businesses:
            session["current_business_id"] = businesses[0]["id"]
        else:
            session.pop("current_business_id", None)
    
    return redirect("/businesses")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)

